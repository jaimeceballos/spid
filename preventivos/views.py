 #!/usr/bin/python
 # -*- coding: iso-8859-15 -*-
from __future__ import unicode_literals
import re
from preventivos.models import *
from preventivos.forms import *
from django.core import serializers
from django.contrib.auth.models import Group,Permission,User
from django.contrib.admin.models import LogEntry
from django.template.context_processors import csrf
import smtplib
import http.client,httplib2
import urllib.request
from xmlrpc import client
from xml.dom.minidom import parseString
import xml
import json
from dicttoxml import dicttoxml
from httplib2 import Http
from urllib.parse import urlencode
from base64 import b64encode
import base64
import dict2xml
import html2text
from django.http import QueryDict
from django.core.mail import send_mail
from email.mime.text import MIMEText
from django.template import Context, Template, RequestContext
from django.template.loader import get_template,render_to_string
from openpyxl import Workbook
from django.core.exceptions import ObjectDoesNotExist,MultipleObjectsReturned
from django.utils.encoding import smart_bytes, smart_text
import locale 
locale.setlocale(locale.LC_ALL, ('es_AR', 'utf8'))
from django.http import HttpResponse,HttpResponseRedirect, HttpResponse,Http404, HttpResponseBadRequest
from django.shortcuts import render, render_to_response,get_object_or_404
from django.contrib.auth import authenticate, login,logout
from django.contrib import auth
from datetime import date,timedelta
import datetime
from time import strptime
from decorators.auth import group_required
import json as simplejson
from django.utils import timezone
from django.utils.html import strip_tags
from django.utils.safestring import mark_safe,SafeString
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required,permission_required
from django.db import transaction,IntegrityError,connection
from django.urls import reverse
# set up the environment using the settings module
#from django.core.management import setup_environ
from django.conf import settings
from spid import settings
import sys,os, re, calendar
import random,datetime,time
from django.utils.encoding import smart_text
#imports para la paginacion
from django.core.paginator import Paginator, InvalidPage, EmptyPage
from django.db.models.signals  import  post_save,post_delete
from formtools.wizard.views import SessionWizardView
from django.template import Context, loader
from haystack.query import SearchQuerySet
from django.db.models import Q
from django.utils.translation import ugettext
#from wkhtmltopdf.views import PDFResponse, PDFTemplateView, PDFTemplateResponse
from django.forms.utils import ErrorList
from xml.sax import make_parser, SAXException
from xml.sax.handler import feature_namespaces
from django.utils.dateparse import parse_datetime
import pytz
from django.db import connections
from django.core.mail import EmailMultiAlternatives




def normalize_query(query_string,
                                        findterms=re.compile(r'"([^"]+)"|(\S+)').findall,
                                        normspace=re.compile(r'\s{2,}').sub):
        ''' Splits the query string in invidual keywords, getting rid of unecessary spaces
                and grouping quoted words together.
                Example:

                >>> normalize_query('  some random  words "with   quotes  " and   spaces')
                ['some', 'random', 'words', 'with quotes', 'and', 'spaces']

        '''
        return [normspace(' ', (t[0] or t[1]).strip()) for t in findterms(query_string)]

def get_query(query_string, search_fields):
        ''' Returns a query, that is a combination of Q objects. That combination
                aims to search keywords within a model by testing the given search fields.

        '''
        query = None # Query to search for every search term
        terms = normalize_query(query_string)
        for term in terms:
                or_query = None # Query to search for a given term in each field
                for field_name in search_fields:
                        q = Q(**{"%s__icontains" % field_name: term})
                        if or_query is None:
                                or_query = q
                        else:
                                or_query = or_query | q
                if query is None:
                        query = or_query
                else:
                        query = query & or_query

        return query

def inicio(request):
    user = request.user
    profile = user.userprofile
    destino = "%s / %s" % (profile.depe,profile.ureg)
    state = user.groups.values_list('name', flat=True)
    no_enviados = False                                           #inicializa una variable de no enviados
    #verifica si el usuario es preventor
    if Actuantes.objects.filter(funcion__gt=1,documento=user.username):
        no_enviados = obtener_cantidad_no_enviados(request)         #obtiene la cantidad de preventivos no enviados
        no_autorizados = obtener_cantidad_no_autorizados(request)     #obtiene la cantidad de preventivos no autorizados
        radio_user = False                                            #crea una variable radio user para identificar si el usuario pertenece a una radiocabecera
    #verifica si dentro de los grupos del usuario se encuentra radio
    if user.groups.filter(name='radio'):
        radio_user = True                                       #pone en tru radio user
    autorizados = 0                                               #establece un contador de preventivos autorizados en 0
    no_autorizados = 0
    #si es usuario de radiocabecera
    if radio_user:
        preventivos = ""
        dependencias = ""
        dependencias = Dependencias.objects.filter(ciudad = user.userprofile.depe.ciudad )
        preventivos = Preventivos.objects.filter(dependencia__in=dependencias,fecha_autorizacion__isnull=False,fecha_envio__isnull = True)            #para esas dependencias obtiene los preventivos autorizados no enviados
        #si hay preventivos no enviados
        if preventivos.count() > 0:
            autorizados = preventivos.count()                     #obtiene la cantidad de preventivos autorizados para enviar
    form = DependenciasForm()       #formulario de dependencias
    formpass = CambiarPassForm()    #formulario de cambio de contraseña
    request.session['state']=state                                #si es correcto carga en la sesion la variable estado
    request.session['destino']=destino                            #carga en la sesion la variable destino
    return render(request, './index1.html', {'form':form,'state':state, 'destino': destino,'formpass':formpass,'no_enviados':no_enviados,'no_autorizados':no_autorizados,'radio_user':radio_user,'autorizados':autorizados})

def search_caratula(request):
    if request.method == "POST":
        search_text = request.POST['search_text']
    else:
        search_text =''
    #preventivos=Preventivos.objects.filter(caratula__contains=search_text)
    preventivos = SearchQuerySet().autocomplete(content_auto=request.POST.get('search_text',''))
    return render(request,'search_html',{'preventivos':preventivos})

def page_not_found(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    return render(request,'./error404.html',{'state':state, 'destino': destino})

def server_error(request):
 return render(request, '500.html')


FORMS = [("nro", PrimerForm),
                 ("actores", SegundoForm),
                 ("autoridades", TerceroForm),
                 ("confirmation", FinForm)]

TEMPLATES = {"nro": "template0.html",
                         "actores": "template1.html",
                         "autoridades": "template2.html",
                         "confirmation": "template3.html"}


class preventivos(SessionWizardView):


        def __init__(self, *args, **kwargs):
                super(preventivos, self).__init__(*args, **kwargs)


        def get_template_names(self):

                return [TEMPLATES[self.steps.current]]

        def get_context_data(self, form, **kwargs):
            context = super(preventivos, self).get_context_data(form=form, **kwargs)
            if self.steps.current:
                context.update({'state': self.request.session.get('state')})
                context.update({'destino': self.request.session.get('destino')})
                return context


        def get_form(self, step=None, data=None, files=None):

                form = super(preventivos, self).get_form(step, data, files)

                if self.request.user.userprofile.depe.descripcion == 'RADIO CABECERA-PM':
                                depe = self.request.user.userprofile.ureg
                                ureg= UnidadesRegionales.objects.filter(descripcion__contains='MADRYN')
                                if depe is None:
                                     form.errors['__all__'] = form.error_class(["Regrese a la pantalla anterior e Ingrese todos los Datos Obligatorios"])

                                depe= Dependencias.objects.filter(unidades_regionales_id__exact=ureg)

                                form.fields['unidad'].queryset = ureg
                                if step=='nro':
                                     form.fields['dependencia'].queryset = depe

                if  self.request.user.userprofile.depe.descripcion == 'RADIO CABECERA-TW':
                                depe = self.request.user.userprofile.ureg

                                ureg= UnidadesRegionales.objects.filter(descripcion__contains='TRELEW')
                                if depe is None:
                                     form.errors['__all__'] = form.error_class(["Regrese a la pantalla anterior e Ingrese todos los Datos Obligatorios"])

                                depe= Dependencias.objects.filter(unidades_regionales_id__exact=ureg)
                                form.fields['unidad'].queryset = ureg
                                if step=='nro':
                                     form.fields['dependencia'].queryset = depe

                if  self.request.user.userprofile.depe.descripcion == 'RADIO CABECERA-ESQ':
                                depe = self.request.user.userprofile.ureg

                                ureg= UnidadesRegionales.objects.filter(descripcion__contains='ESQUEL')
                                if depe is None:
                                     form.errors['__all__'] = form.error_class(["Regrese a la pantalla anterior e Ingrese todos los Datos Obligatorios"])

                                depe= Dependencias.objects.filter(unidades_regionales_id__exact=ureg)
                                form.fields['unidad'].queryset = ureg
                                if step=='nro':
                                     form.fields['dependencia'].queryset = depe

                if  self.request.user.userprofile.depe.descripcion == 'RADIO CABECERA-CR':
                                depe = self.request.user.userprofile.ureg

                                ureg= UnidadesRegionales.objects.filter(descripcion__contains='COMODORO RIVADAVIA')
                                if depe is None:
                                     form.errors['__all__'] = form.error_class(["Regrese a la pantalla anterior e Ingrese todos los Datos Obligatorios"])

                                depe= Dependencias.objects.filter(unidades_regionales_id__exact=ureg)
                                form.fields['unidad'].queryset = ureg
                                if step=='nro':
                                     form.fields['dependencia'].queryset = depe

                if  self.request.user.userprofile.depe.descripcion == 'CENTRAL RADIO':
                                depe = self.request.user.userprofile.ureg
                                if step is None:
                                     step=self.steps.current
                                ureg= UnidadesRegionales.objects.exclude(descripcion__icontains='AREA')
                                ureg= ureg.exclude(descripcion__icontains='INVESTIGACION')
                                #depes= Dependencias.objects.all
                                #depes= Dependencias.objects.filter(unidades_regionales_id__exact=ureg)
                                form.fields['unidad'].queryset = ureg
                                #form.fields['dependencia'].queryset = depes

                ### realizar el paso en caso de que el usuario sea de investigaciones
                #depe=self.request.user.userprofile.depe
                if step == "actores":

                            if self.request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or  'RADIO' in self.request.user.userprofile.depe.descripcion:

                                depe = self.get_cleaned_data_for_step('nro')['dependencia']
                                nro = self.get_cleaned_data_for_step('nro')['nro']
                                anio= self.get_cleaned_data_for_step('nro')['anio']
                                existen= Preventivos.objects.filter(dependencia__exact=depe.id,nro__exact=nro,anio__exact=anio).values('nro')
                                if existen:
                                     form.errors['__all__'] = form.error_class(["Preventivo Existente. regrese a la pantalla anterior."])
                                     #form.fields['nro'].initial = 0
                                     #form.fields['anio'].initial = 0



                                if depe is None:
                                     form.errors['__all__'] = form.error_class(["Regrese a la pantalla anterior e Ingrese todos los Datos Obligatorios"])


                                id_depe=Dependencias.objects.filter(descripcion__exact=depe).values('id')
                                actuante=Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=1) | Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=3)
                                preventor=Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=2) | Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=3)

                                if actuante is None:

                                    form.fields['actuante'].initial = data.get('actuante')
                                    form.fields['actuante'].widget.attrs['readonly'] = True
                                    form.fields['actuante'].widget.attrs['disabled'] = True
                                    form.fields['preventor'].initial = data.get('preventor')
                                    form.fields['preventor'].widget.attrs['readonly'] = True
                                    form.fields['preventor'].widget.attrs['disabled'] = True
                                else:

                                    form.fields['actuante'].queryset = actuante
                                    form.fields['preventor'].queryset = preventor


                            if self.request.user.userprofile.depe.descripcion!="Jefatura" and self.request.user.userprofile.depe.descripcion != 'INVESTIGACIONES' and  self.request.user.userprofile.depe.descripcion != 'CENTRAL RADIO':

                                if 'RADIO CABECERA' in self.request.user.userprofile.depe.descripcion:
                                     depe=self.get_cleaned_data_for_step('nro')['dependencia']
                                else:
                                     depe=self.request.user.userprofile.depe
                                id_depe=Dependencias.objects.filter(descripcion__exact=depe).values('id')

                                actuante=Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=1) | Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=3)
                                preventor=Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=2) | Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=3)

                                if actuante is None:

                                    form.fields['actuante'].initial = data.get('actuante')
                                    form.fields['actuante'].widget.attrs['readonly'] = True
                                    form.fields['actuante'].widget.attrs['disabled'] = True
                                    form.fields['preventor'].initial = data.get('preventor')
                                    form.fields['preventor'].widget.attrs['readonly'] = True
                                    form.fields['preventor'].widget.attrs['disabled'] = True
                                else:

                                    form.fields['actuante'].queryset = actuante
                                    form.fields['preventor'].queryset = preventor


                if step == 'autoridades':
                         if self.request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or  self.request.user.userprofile.depe.descripcion == 'CENTRAL RADIO':
                                id_ciudad = Dependencias.objects.get(id=self.get_cleaned_data_for_step('nro')['dependencia'].id).ciudad_id
                         else:
                                id_ciudad=self.request.user.userprofile.depe.ciudad.id
                         #id_ciudad=self.request.user.userprofile.depe.ciudad.id
                         if self.request.user.userprofile.depe!="Jefatura":
                                form.fields['autoridades'].queryset=RefCiudades.objects.get(id=id_ciudad).ciu_autori.all()


                if step=='confirmation':
                     if self.request.user.userprofile.depe.descripcion == 'INVESTIGACIONES'  or  'RADIO' in self.request.user.userprofile.depe.descripcion:
                                depe = self.get_cleaned_data_for_step('nro')['dependencia']
                                #id_ciudad = Dependencias.objects.get(id=self.get_cleaned_data_for_step('nro')['dependencia'].id).ciudad_id

                     else:
                             depe=self.request.user.userprofile.depe
                             #id_ciudad=self.request.user.userprofile.depe.ciudad.id

                     id_depe=Dependencias.objects.filter(descripcion__exact=depe).values('id')

                     form.fields['fecha_denuncia'].initial= self.get_cleaned_data_for_step('nro')['fecha_denuncia']
                     form.fields['caratula'].initial =  self.get_cleaned_data_for_step('nro')['caratula']
                     form.fields['fecha_denuncia'].widget.attrs['readonly'] = True
                     form.fields['caratula'].widget.attrs['readonly'] = True
                     form.fields['actuante'].queryset = Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=1) | Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=3)
                     form.fields['preventor'].queryset= Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=2) | Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=3)
                     #form.fields['actuante'].initial= self.get_cleaned_data_for_step('actores')['actuante'].id
                     #form.fields['preventor'].initial =  self.get_cleaned_data_for_step('actores')['preventor'].id
                     id_ciudad=Dependencias.objects.filter(descripcion__exact=depe).values('ciudad')
                     form.fields['autoridades'].queryset=RefCiudades.objects.get(id=id_ciudad).ciu_autori.all()
                     autoridades=self.get_cleaned_data_for_step('autoridades')['autoridades']
                     autoridad=[]
                     for seleccion in autoridades:
                             ids=int(RefAutoridad.objects.get(descripcion=seleccion).id)
                             autoridad.append(ids)

                     form.fields['autoridades'].initial=autoridad
                     form.fields['autoridades'].widget.attrs["onclick"] = False
                return form


        def done(self, form_list, **kwargs):
                form=form_list[len(form_list)-1]
                pre=Preventivos()
                fecha_denuncia=form.cleaned_data['fecha_denuncia']
                caratula=form.cleaned_data['caratula']
                actuante= form.cleaned_data['actuante']
                preventor=  form.cleaned_data['preventor']
                autoridades=form.cleaned_data['autoridades']
                if self.request.user.userprofile.depe.descripcion == 'INVESTIGACIONES'  or  'RADIO' in self.request.user.userprofile.depe.descripcion:
                    dependencia = self.get_cleaned_data_for_step('nro')['dependencia']
                    nro = self.get_cleaned_data_for_step('nro')['nro']
                    anio= self.get_cleaned_data_for_step('nro')['anio']

                    existen= Preventivos.objects.filter(dependencia__exact=dependencia.id,nro__exact=nro,anio__exact=anio).values('nro')
                    if existen:
                         form.errors['__all__'] = form.error_class(["Preventivo Existente. regrese a la pantalla inicial de carga de preventivo."])
                    else:
                         pre.nro=nro
                         pre.anio=anio


                else:
                    depe=self.request.user.userprofile.depe

                    dependencia=Dependencias.objects.get(id__exact=depe.id)

                    existen= Preventivos.objects.filter(dependencia__exact=dependencia,anio__exact=date.today().year).values('nro')
                    #genera nro de preventivos por dependencia

                    if len(existen)==0:
                         nro=1
                    else:
                         nro=existen[len(existen)-1]['nro']+1






                pre.fecha_carga=datetime.datetime.now()
                pre.caratula=caratula
                pre.nro=nro
                pre.anio=date.today().year
                pre.fecha_denuncia=fecha_denuncia
                pre.actuante=actuante
                pre.preventor=preventor
                pre.dependencia=dependencia

                pre.save()
                jefe = False
                for grabauto in form.cleaned_data['autoridades']:
                        if 'JEFE DE POLICIA' == grabauto.descripcion:
                            jefe = True
                        pre.autoridades.add(int(RefAutoridad.objects.get(descripcion=grabauto).id))
                if not jefe:
                    pre.autoridades.add(int(RefAutoridad.objects.get(descripcion='JEFE DE POLICIA').id))

                idprev=pre.id

                return HttpResponseRedirect('./first/%s/' % idprev)

@login_required
@group_required(["policia","investigaciones",'radio'])
def obtener_datosfirst(request,idprev):
    state= request.session.get('state')
    destino= request.session.get('destino')
    preventivo = Preventivos.objects.get(id = idprev)
    #idprev=preventivo
    depe= preventivo.dependencia
    voyper=False
    errors=[]
    continua='no'
    grabo='no'
    delito=''
    motivo=''
    idhec=''
    boton='no'
    delitos = []
    notienePer=False
    idper=0
    mostrar=''
    descripcionhecho=''
    fecha_desde=''
    fecha_hasta=''
    if request.POST.get("continua")=='Continuar' or request.POST.get("continua")=='Agregar' or request.POST.get("continuar")=='Agregar':
         form=HechosForm(request.POST, request.FILES)



         if form.is_valid():
            continua="si"

            hec=Hechos()

            if request.POST.get("continua")=='Agregar' and request.POST.get("id"):



                 hecho = Hechos.objects.get(id=request.POST.get('id'))
                 idhec=hecho.id
                 fd=request.POST.get('fecha_denuncia')
                 fde=request.POST.get('fecha_desde')
                 fha=request.POST.get('fecha_hasta')

                 if fde>fd or fha>fd:
                    errors.append('La Fecha de Denuncia nunca puede ser menor a la Fecha y Hora del Hecho sucedido')
                    continua="no"
                 else:

                    if request.POST.get('delito'):

                     hechoDelito = HechosDelito()
                     hechoDelito.hechos = hecho
                     hechoDelito.refdelito = RefDelito.objects.get(id = request.POST.get('delito'))

                     if request.POST.get('modos'):
                        hechoDelito.refmodoshecho = RefModosHecho.objects.get(id = request.POST.get('modos'))
                     if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or 'RADIO' in request.user.userprofile.depe.descripcion:

                         try:
                             idhec=hecho.id
                             busco=HechosDelito.objects.filter(refdelito_id = request.POST.get('delito'),hechos_id=idhec)
                             listadelitos=HechosDelito.objects.filter(hechos_id=idhec)
                             if busco:
                                 errors.append('Delito que intenta agregar ya fue cargado')
                             else:
                                 hechoDelito.save()

                         except IntegrityError:
                             errors.append('Delito que intenta agregar ya fue cargado')


                         delitos =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)
                         continua='si'

                     else:
                        errors.append('No se puede modificar preventivos de otras dependencias.')

                    else:
                     continua="no"



                 if 'MUJER' in request.user.userprofile.depe.descripcion and request.user.userprofile.depe==depe:
                    tipodelito=RefDelito.objects.get(id = request.POST.get('delito'))

                    for delis in listadelitos:
                       delitus=str(delis)
                       if 'VIOLENCIA FAMILIAR' in delitus or 'Violencia Familiar' in delitus:

                          boton='si'
            else:

             motivo=form.cleaned_data['motivo']
             fecha_desde=form.cleaned_data['fecha_desde']
             fecha_hasta=form.cleaned_data['fecha_hasta']
             fd=request.POST.get('fecha_denuncia')
             fde=request.POST.get('fecha_desde')
             fha=request.POST.get('fecha_hasta')

             if fde>fd or fha>fd:
                    errors.append('La Fecha de Denuncia nunca puede ser menor a la Fecha y Hora del Hecho sucedido')
                    continua="no"
             else:
                if request.POST.get('delito'):

                    hec.motivo=motivo
                    hec.fecha_desde=fecha_desde
                    hec.fecha_hasta=fecha_hasta
                    hec.fecha_carga=datetime.datetime.now()
                    hec.descripcion=''
                    hec.preventivo_id=idprev

                    if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or 'RADIO' in  request.user.userprofile.depe.descripcion:
                        try:
                            hec.save()
                        except IntegrityError:
                            continua="si"

                    else:
                         errors.append('No se puede modificar preventivos de otras dependencias.')

                    if request.POST.get("continua")=='Agregar' and request.POST.get('id'):
                         hecho = Hechos.objects.get(id=request.POST.get('id'))
                         idhec=hecho.id
                         hechoDelito = HechosDelito()
                         hechoDelito.hechos = Hechos.objects.get(id=request.POST.get('id'))
                         descripcionhecho=request.POST.get('descrihecho')
                         continua='no'
                         grabo='si'

                    else:
                         idhec=hec.id
                         hechoDelito = HechosDelito()
                         hechoDelito.hechos = hec

                    hechoDelito.refdelito = RefDelito.objects.get(id = request.POST.get('delito'))

                    if request.POST.get('modos'):
                     hechoDelito.refmodoshecho = RefModosHecho.objects.get(id = request.POST.get('modos'))

                    if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or 'RADIO' in request.user.userprofile.depe.descripcion:
                        try:

                            hechoDelito.save()
                            delitos =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)

                            continua='si'

                        except IntegrityError:
                            errores=True

                            idhecs=Hechos.objects.filter(preventivo_id=idprev).values('id')
                            idhe=Hechos.objects.get(id=idhecs)
                            delitos =HechosDelito.objects.filter(hechos = idhecs,borrado__isnull=True)
                            if idhe:

                                 form=HechosForm(instance=idhe)
                                 idhec=idhe.id
                            continua='si'
                else:
                    continua="no"

                if 'MUJER' in request.user.userprofile.depe.descripcion and request.user.userprofile.depe==depe:
                    tipodelito=RefDelito.objects.get(id = request.POST.get('delito'))

                    if 'VIOLENCIA FAMILIAR' in tipodelito.descripcion or 'Violencia Familiar' in tipodelito.descripcion:
                       boton='si'

         else:

             if request.POST.get("continua")=='Agregar':
                 hecho = Hechos.objects.get(id=request.POST.get('id'))
                 idhec=hecho.id
                 if request.POST.get('delito'):
                    continua="si"
                    hechoDelito = HechosDelito()
                    hechoDelito.hechos = hecho
                    hechoDelito.refdelito = RefDelito.objects.get(id = request.POST.get('delito'))
                    if request.POST.get('modos'):
                     hechoDelito.refmodoshecho = RefModosHecho.objects.get(id = request.POST.get('modos'))
                    try:
                     hechoDelito.save()

                    except IntegrityError:

                        errors.append('Delito que intenta agregar ya fue cargado')


                    delitos =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)
                 else:
                    if 'None' not in request.POST.get('id'):
                     continua="si"
                     hecho = Hechos.objects.get(id=request.POST.get('id'))
                     idhec=hecho.id
                     delitos =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)
                 if request.POST.get('descrihecho'):
                        hecho.descripcion=request.POST.get('descrihecho')
                 motivo=request.POST.get('motivo')
                 fecha_desde=request.POST.get('fecha_desde')
                 fecha_hasta=request.POST.get('fecha_hasta')

    else:

            if request.POST.get("grabar")=='Guardar Hecho' or request.POST.get("grabar")=='Modificar':

                 form=HechosForm(request.POST, request.FILES)
                 hecho = Hechos.objects.get(id=request.POST.get('id'))

                 idhec=hecho.id
                 delitos =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)



                 hecho.descripcion=request.POST.get('descrihecho').encode('utf-8', 'xmlcharrefreplace')
                 hecho.descripcion=strip_tags(hecho.descripcion)
                 hecho.descripcion=hecho.descripcion.replace('&nbsp','')
                 hecho.descripcion=hecho.descripcion.strip()
                 hecho.descripcion=request.POST.get('descrihecho').strip()
                 if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or 'RADIO' in request.user.userprofile.depe.descripcion:

                            hecho.save()
                            idhec=hecho.id


                 motivo=request.POST.get('motivo')
                 fecha_desde=request.POST.get('fecha_desde')
                 fecha_hasta=request.POST.get('fecha_hasta')
                 form=Hechos.objects.get(id=idhec)


                 grabo='fin'

            else:

                if request.POST.get("borrar"):
                 continua="si"
                 form=HechosForm(request.POST, request.FILES)
                 hecho = Hechos.objects.get(id=request.POST.get('id'))
                 hechoDelito = HechosDelito()
                 try:
                     hechoDelito = HechosDelito.objects.filter(id=request.POST.get("borrar"))
                     try:
                            hechoDelito.delete()
                            descripcionhecho=request.POST.get('descrihecho')
                     except IntegrityError:
                            errorsd=True
                 except IntegrityError:
                     errorsd=True

                 idhec=hecho.id
                 delitos =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)


                 motivo=request.POST.get('motivo')
                 fecha_desde=request.POST.get('fecha_desde')
                 fecha_hasta=request.POST.get('fecha_hasta')
                 if 'MUJER' in request.user.userprofile.depe.descripcion and request.user.userprofile.depe==depe:
                    for delis in delitos:
                       delitus=str(delis)
                       if 'VIOLENCIA FAMILIAR' in delitus or 'Violencia Familiar' in delitus:

                          boton='si'
                else:
                    data = 0
                    try:
                        data=Preventivos.objects.get(id=idprev).hecho.id
                    except:
                        print('no tiene hecho')
        
                    if data>0:
                     if  Hechos.objects.get(id=data).involu.all():

                            datosinvo=Hechos.objects.get(id=Preventivos.objects.get(id=idprev).hecho.id).involu.all()
                            notienePer= True
                     if not notienePer:
                         state= request.session.get('state')
                         destino= request.session.get('destino')
                         return render(request,'./errorsinper.html',{'state':state, 'destino': destino})

                     idhecs=Hechos.objects.filter(preventivo_id=idprev).values('id')
                     idhe=Hechos.objects.get(id=idhecs)
                     delito =HechosDelito.objects.filter(hechos = idhecs,borrado__isnull=True)

                     if idhe:

                                 form=HechosForm(instance=idhe)
                                 idhec=idhe.id
                                 fecha_desde=idhe.fecha_desde
                                 fecha_hasta=idhe.fecha_hasta
                                 descripcionhecho=idhe.descripcion
                                 continua='des'
                     else:
                            si = Hechos()
                            form=HechosForm(instance=si)
                            idhec=''
                    else:
                            si = Hechos()
                            form=HechosForm(instance=si)
                            idhec=''

    ftiposdelitos=DelitoForm()
    modos=RefModosHechoForm()
    datos=Preventivos.objects.get(id=idprev)
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    autoridades= datos.autoridades.values_list('descripcion',flat=True)
    dependencia=datos.dependencia.descripcion
    unidadreg=datos.dependencia.unidades_regionales.descripcion
    fecha_autorizacion=datos.fecha_autorizacion
    info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
    'caratula':caratula,'descripcionhecho':descripcionhecho,'boton':boton,
    'actuante':actuante,'fecha_autorizacion':fecha_autorizacion,
    'preventor':preventor,'idprev':idprev,'delito':delito,'fecha_desde':fecha_desde,'fecha_hasta':fecha_hasta,
    'autoridades':autoridades,'idper':idper,'mostrar':mostrar,
    'errors': errors, 'grabo':grabo,'dependencia':dependencia,'unidadreg':unidadreg,
    'state':state, 'continua':continua,'delitos':delitos,'motivo':motivo,'depe':depe,
    'destino': destino,'form':form,'ftiposdelitos':ftiposdelitos,'modos':modos,'idhec':idhec}
    return render(request,'./templateidp.html',info)





def nologin(request):
        logout(request)
        try:
                del request.session['state']
                del request.session['destino']
                request.session.flush()
        except KeyError:
                pass
                state = "Ud. se desconecto del Sistema"

        return render(request, 'index.html', {'state':state})

@login_required
@permission_required('user.is_staff')
def control(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors=[]
    listacontrol=[]
    control=False
    lista = UserProfile.objects.all()
    user_groups = Group.objects.all()
    form = GroupForm(request.POST)
    return render(request,'./controluser.html',{'control':control,'listacontrol':listacontrol,'form':form,'user_groups':user_groups,'grupos':grupos,'lista':lista,'state':state, 'errors':errors,'destino': destino})


@login_required
@permission_required('user.is_staff')
def reporactivity(request, user):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    usuarios=""
    control=False
    if user.isnumeric():
            if request.POST.get('cancelar')=="Cancelar":
                 return HttpResponseRedirect('../')
            else:
                    usuarios = User.objects.get(username=user)

                    usua = User.objects.filter(username=user).values()
                    idem = usuarios.userprofile
                    dnis = usuarios.username
                    apellidos=usuarios.last_name
                    nombres=usuarios.first_name
                    lastlog=usuarios.last_login
                    fechajob=usuarios.date_joined
                    if lastlog==fechajob:
                         reenvio=True
                    ids=usuarios.id
                    formnew = UserForm(instance=usuarios)
                    form = UserProfileForm(instance=idem)

                    lista = UserProfile.objects.filter(user=user)
                    datos=UserProfile.objects.get(user=ids)
                    listacontrol=Registrouser.objects.all().filter(user_id=usuarios.id).order_by('-fecha',)

                    control=True
                    ure=datos.ureg
                    #lista = User.objects.all()
                    return render(request,'./controluser.html', {'control':control,'listacontrol':listacontrol,'lista':lista,'usuarios':usuarios,'form':form,'formnew':formnew,'errors': errors,'state':state, 'destino': destino})
    else:
        lista = UserProfile.objects.all()
        user_groups = Group.objects.all()
        form = GroupForm(request.POST)
        return render(request,'./controluser.html',{'control':control,'form':form,'user_groups':user_groups,'grupos':grupos,'lista':lista,'state':state, 'errors':errors,'destino': destino})



@login_required
def inicial(request):
    ciudades = ""
    destino = "%s / %s" % (request.user.userprofile.depe,request.user.userprofile.ureg)
    state = request.user.groups.values_list('name', flat=True)

    no_enviados = False
    if Actuantes.objects.filter(funcion__gt=1,documento=request.user.username):
        no_enviados = obtener_cantidad_no_enviados(request)
    no_autorizados = obtener_cantidad_no_autorizados(request)
    usuario = request.user
    radio_user = False
    autorizados = 0
    if usuario.groups.filter(name='radio'):
        radio_user = True
    if radio_user:

        dependencias = Dependencias.objects.filter(ciudad = usuario.userprofile.depe.ciudad )
        preventivos = Preventivos.objects.filter(dependencia__in=dependencias,fecha_autorizacion__isnull=False,fecha_envio__isnull = True).order_by('-id')
        if preventivos.count() > 0:
            autorizados = preventivos.count()
    return render(request,'./index1.html',{'state':state, 'destino': destino,'no_enviados':no_enviados,'no_autorizados':no_autorizados,'radio_user':radio_user,'autorizados':autorizados})

#la funcion en donde se guardan los grupos de usuarios que son ingresados en usuarios
@login_required
@permission_required('user.is_staff')
def ngrupos(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    grupos = ""
    errors=[]
    if request.POST.get('grabar')=="Grabar":
            form = GroupForm(request.POST)
            grupo = request.POST.get('name')
            if not grupo:
                        errors.append('Ingrese Grupo de pertenencia de Usuario')
                        return HttpResponseRedirect('.')
            else:
                             if not(len(grupo)>=4 and len(grupo)< 80):
                                 errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
                             else:
                                         if form.is_valid():
                                             form.save()
                                             formp = UserForm(request.POST)
                                             form = GroupForm(request.POST)
                                             return HttpResponseRedirect('../user/new/')
                                         else:
                                             errors.append('El Grupo de Usuario que Ud. intenta grabar ya existe')
            lista = UserProfile.objects.all()
            user_groups = Group.objects.all()
            form = GroupForm(request.POST)
            return render(request,'./newuser.html',{'form':form,'user_groups':user_groups,'grupos':grupos,'lista':lista,'state':state, 'errors':errors,'destino': destino})
    else:
         formp = UserForm()
         form  = GroupForm()
         lista = UserProfile.objects.all()
         user_groups = Group.objects.all()
         return render(request,'./newuser.html',{'form':form,'formp':formp,'user_groups':user_groups,'grupos':grupos,'errors': errors,'lista':lista,'state':state, 'destino': destino})


@login_required
@permission_required('user.is_staff','administrador')
def user_create(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    form = UserCreateForm()
    return render(request,"./user_create.html",{'form':form})

@login_required
@permission_required('user.is_staff','administrador')
def user_edit(request):
    state   = request.session.get('state')
    destino = request.session.get('destino')
    form    = UserCreateForm()
    return render(request,"./user_edit.html",{'form':form})

@login_required
def user_edit_destino(request,usuario):
    if request.is_ajax():
        usuario = User.objects.get(id=usuario).username
        form=""
        actuante = True
        try:
            actuantes   = Actuantes.objects.get(documento=usuario)
            form        = ActuantesForm(instance=actuantes)

        except:
            profile = User.objects.get(username=usuario).userprofile
            form    = UserProfileForm(instance=profile)
            actuante = False
        finally:
            return render(request,"./user_edit-destino.html",{'form':form,'actuante':actuante,'usuario':usuario})
    return HttpResponseBadRequest()

def user_edit_save_destino(request,usuario):
    if request.is_ajax():
        msg = ""
        if request.method == 'POST':
            funcion     = ""
            unidad      = ""
            dependencia = ""
            actuante = False
            esActuante=""
            if request.POST['actuante'] == 'True':
                funcion = request.POST['funcion']
                unidad = request.POST['unidadreg_id']
                dependencia = request.POST['dependencia_id']
                jerarquia = request.POST['jerarquia_id']
                try:
                    actuante = Actuantes.objects.get(documento = usuario)
                except:
                    actuante = Actuantes()
                    actuante.documento       = usuario
                    actuante.apeynombres     = "%s, %s" % (User.objects.get(username=usuario).last_name,User.objects.get(username=usuario).first_name)
                    actuante.persona_id      = Personas.objects.get(nro_doc=usuario)

                actuante.funcion = funcion
                actuante.unidadreg_id = UnidadesRegionales.objects.get(id=unidad)
                actuante.dependencia_id = Dependencias.objects.get(id=dependencia)
                actuante.jerarquia_id = RefJerarquias.objects.get(id=jerarquia)
            else:
                unidad = request.POST['ureg']
                dependencia = request.POST['depe']
            profile = User.objects.get(username = usuario).userprofile
            profile.depe = Dependencias.objects.get(id=dependencia)
            profile.ureg = UnidadesRegionales.objects.get(id=unidad)
            try:
                if actuante:
                    actuante.save()
                profile.save()
                return HttpResponse('<div class="col-md-12"><h2>Usuario modificado correctamente.</h2></div>')
            except Exception as e:
                 msg = '<div class="col-md-12"><h2>'+ e +'</h2></div>'
    return HttpResponseBadRequest(msg)

def user_edit_actuante(request,usuario):
    msg=""
    if request.is_ajax():
        form = ActuantesForm()
        grupos = User.objects.get(username = usuario).groups.filter(Q(name__icontains="policia")|Q(name__icontains="jefes"))
        if grupos:
            try:
                actuante = Actuantes.objects.get(documento=usuario)
                form = ActuantesForm(instance=actuante)
            except:
                data ={}
                data['documento']       = usuario
                data['apeynombres']     = "%s, %s" % (User.objects.get(username=usuario).last_name,User.objects.get(username=usuario).first_name)
                data['persona_id']      = Personas.objects.get(nro_doc=usuario)
                data['unidadreg_id']    = User.objects.get(username=usuario).userprofile.ureg
                data['dependencia_id']  = User.objects.get(username=usuario).userprofile.depe
                form = ActuantesForm(initial=data)
                return render(request,"./user_edit-destino.html",{'form':form,'actuante':True,'usuario':usuario})
        else:
            msg = "<h4> Verifique los roles asignados al usuario.</h4>"
    return HttpResponseBadRequest(msg)



@login_required
@transaction.atomic
@permission_required('user.is_staff','administrador')
def user_create_save(request):
    if request.is_ajax:
        if request.method == 'POST':
            form = UserCreateForm(request.POST)
            msg=""
            if form.is_valid():
                if not User.objects.filter(email = form.cleaned_data['email']):

                    persona     = Personas()
                    usuario     = User()
                    profile     = UserProfile()
                    personal    = Personal()
                    documento   = form.cleaned_data['documento']
                    password    = User.objects.make_random_password(length=10)
                    dependencia = Dependencias.objects.get(id=form.cleaned_data['lugar_trabajo_id'])
                    try:
                        persona = Personas.objects.get(nro_doc = documento)
                        personal = Personal.objects.get(persona_id = persona.id)
                    except:
                        persona.nro_doc = documento
                    finally:
                        persona.nombres         = form.cleaned_data['nombre'].upper()
                        persona.apellidos       = form.cleaned_data['apellido'].upper()
                        persona.fecha_nac       = form.cleaned_data['fecha_nacimiento']
                        persona.ciudad_nac      = RefCiudades.objects.get(id = form.cleaned_data['ciudad_nacimiento_id'])
                        persona.sexo_id         = RefSexo.objects.get(id=form.cleaned_data['sexo'])
                        persona.estado_civil    = RefEstadosciv.objects.get(id=form.cleaned_data['estados_civiles'])
                        persona.ciudad_res      = RefCiudades.objects.get(id = form.cleaned_data['ciudad_residencia_id'])
                        persona.tipo_doc        = RefTipoDocumento.objects.get(descripcion='DNI')
                        persona.ocupacion       = RefOcupacion.objects.get(descripcion = 'EMPLEADO POLICIAL')

                    usuario.set_password(password)
                    usuario.username        = documento
                    usuario.email           = form.cleaned_data['email']
                    usuario.first_name      = persona.nombres
                    usuario.last_name       = persona.apellidos
                    usuario.is_active       = form.cleaned_data['activo']
                    usuario.is_staff        = False
                    usuario.is_superuser    = False

                    try:
                        persona.save()
                        personal.persona_id     = persona
                        if not personal.id:
                            personal.credencial = 0
                            personal.legajo = personal.nro_cuenta_bco = personal.nro_seros = ""
                        personal.save()
                        usuario.save()
                        for rol in form.cleaned_data['grupos']:
                            usuario.groups.add(Group.objects.get(id=rol))
                        profile         = usuario.userprofile
                        profile.depe    = dependencia
                        profile.ureg    = dependencia.unidades_regionales
                        profile.save()
                        if form.cleaned_data['funcion'] !="":
                            actuante = Actuantes()
                            actuante.funcion        = form.cleaned_data['funcion']
                            actuante.documento      = persona.nro_doc
                            actuante.apeynombres    = "%s, %s" % (persona.apellidos,persona.nombres)
                            actuante.jerarquia_id   =  RefJerarquias.objects.get(id=form.cleaned_data['jerarquia_id'])
                            actuante.persona_id     = persona
                            actuante.unidadreg_id   = dependencia.unidades_regionales
                            actuante.dependencia_id = dependencia
                            actuante.save()
                        enviar_correo_usuario(usuario,password)
                    except Exception as e:
                        print(e)
                        return HttpResponseBadRequest("No se pudo realizar la operacion solicitada.")
                else:
                    msg = "El Email ingresado ya esta siendo utilizado. Por favor indique otro."
                    return HttpResponseBadRequest(msg)
            else:
                return render(request,"./user_create.html",{'form':form})



    return HttpResponse('<div class="col-md-12"><h2>Usuario creado correctamente.</h2></div>')

@login_required
@permission_required('user.is_staff','administrador')
def user_reenviar_mail(request,usuario):
    if request.is_ajax():
        enviar_correo_usuario(User.objects.get(id=usuario),"")
        return HttpResponse('<h4>Se reenviaron los datos al usuario correctamente.</h4>')
    return HttpResponseBadRequest()

@login_required
@permission_required('user.is_staff','administrador')
def user_modificar_mail(request,usuario):
    if request.is_ajax():
        if request.method == 'POST':
            mail = request.POST['mailChange']
            user = ""
            try:
                user = User.objects.get(email=mail).id
            except Exception as e:
                user = usuario

            if usuario != user:
                return HttpResponseBadRequest("<h4>El correo ya esta siendo utilizado por otro usuario</h4>")
            else:
                usuario = User.objects.get(id=usuario)
                usuario.email = mail
                usuario.save()
                return HttpResponse("<h4>Correo modificado correctamente.</h4>")
    return HttpResponseBadRequest("<h4>No se puede realizar la operacion. Vuelva a intentarlo mas tarde</h4>")

@login_required
@permission_required('user.is_staff','administrador')
def user_activar(request,usuario):
    if request.is_ajax():
        usuario = User.objects.get(id=usuario)
        usuario.is_active = True
        usuario.save()
        return HttpResponse("<h4>Usuario Activado.</h4>")
    return HttpResponseBadRequest("<h4>No se puede realizar la operacion. Vuelva a intentarlo mas tarde</h4>")

@login_required
@permission_required('user.is_staff','administrador')
def user_roles(request,usuario):
    if request.is_ajax():
        usuario = User.objects.get(id = usuario)
        form = UserGroupsForm(instance = usuario)
        return render(request,'./user_edit-roles.html',{'form':form,'usuario':usuario})
    return HttpResponseBadRequest()

@login_required
@permission_required('user.is_staff','administrador')
def user_roles_save(request):
    if request.is_ajax():
        if request.method == 'POST':
            form = UserGroupsForm(request.POST)
            if form.is_valid():
                usuario = User.objects.get(id=request.POST['usuario'])
                usuario.is_staff = form.cleaned_data['is_staff']
                usuario.save()
                usuario.groups.set(form.cleaned_data['groups'])
                return HttpResponse('<h4>La operacion se realizo con exito.</h4>')
    return HttpResponseBadRequest("<h4>No se puede realizar la operacion. Vuelva a intentarlo mas tarde</h4>")

def enviar_correo_usuario(usuario,password):
    """Envia el correo de nuevo usuario, recibe
    el usuario y el password sin haber sido hasheado,
    si recibe una cadena vacia genera un nuevo password"""

    if password == "":
        password    = User.objects.make_random_password(length=10)
        usuario.set_password(password)
        usuario.save()
    roles = []
    for rol in usuario.groups.all():
        roles.append(rol.name)
    profile = usuario.userprofile
    subject, from_email, to = 'Asunto : Usuario y Password - SPID' ,'divsistemasjp@policia.chubut.gov.ar',usuario.email
    text_content = ("Este email es creado por Div. Sistemas Informaticos Rw."\
    "<br>Ureg. : %s <br> Dependencia : %s <br> Usuario: %s <br> Password: %s <br><strong> Grupo Usuarios : %s </strong>"\
    "<br>  <strong>Link Sistema :</strong> <a href='policia.chubut.gov.ar/spid/'>SPID</a><br>Por cualquier consulta y/o reclamos al:\n\n\n<br> "\
    "email: divsistemasjp@policia.chubut.gov.ar.-"% (profile.ureg.descripcion,profile.depe.descripcion,usuario.username,str(password),str(roles)))

    msg = EmailMultiAlternatives(subject,text_content,from_email, [to])
    msg.attach_alternative(text_content,'text/html')
    profile.last_login = 1
    profile.save()
    try:
         msg.send(fail_silently=False)
         return True
    except Exception as e:
        return False

@login_required
@transaction.atomic
@permission_required('user.is_staff','administrador')
def new_user(request):
        state= request.session.get('state')
        destino= request.session.get('destino')
        errors=[]
        usuarios=""
        estado=''
        ocupacion=''
        listap=[]
        esvisita=''
        visitaes=False
        dni=request.POST.get('username')
        create_user =True
        return render(request,'./newuser.html', {'state':state, 'destino': destino,'create_user':create_user})

@login_required
@transaction.atomic
@permission_required('user.is_staff')
def gruposperm(request):
        state= request.session.get('state')
        destino= request.session.get('destino')
        errors=[]
        usuarios=""
        estado=''
        ocupacion=''
        grupos=''
        form=''
        vector=[]
        listapg={}
        form = GroupForm()


        if request.POST.get('groups'):

             formnew=UserForm(request.POST)
             gr=Group.objects.get(id=request.POST.get('groups'))
             formnew.fields['groups'].initial=gr
             listapg=gr.permissions.filter(group=gr)
             grupos=gr

             filtrado= User.objects.filter(groups=grupos)
             if filtrado:
                    formnew=UserForm(instance=filtrado[0])
        else:

                formnew=UserForm()




        if request.POST.get('dar')=="Asignar":
            form = GroupForm(request.POST)

            if form.is_valid():
                 grupo = request.POST.get('groups')
                 group = Group.objects.get(id=grupo)
                 for estan in form.cleaned_data['permissions']:
                     group.permissions.add(estan)

                 formnew.fields['groups'].initial=group
                 listapg=group.permissions.all()

                 grupos=group

                 filtrado= User.objects.filter(groups=grupos)
                 if filtrado:
                    #insta=User.objects.get(groups=grupos)
                    formnew=UserForm(instance=filtrado[0])

        if request.POST.get('group_permissions'):
            form = GroupForm(request.POST)

            if form.is_valid():
                 grupo = request.POST.get('groups')
                 group = Group.objects.get(id=grupo)
                 idpermiso=request.POST.get('group_permissions')

                 group.permissions.remove(idpermiso)

                 formnew.fields['groups'].initial=group
                 listapg=group.permissions.all()

                 grupos=group

                 filtrado= User.objects.filter(groups=grupos)
                 if filtrado:
                    insta=User.objects.get(groups=grupos)
                    formnew=UserForm(instance=insta)






        return render(request,'./gruposper.html', {'grupos':grupos,'listapg':listapg,'usuarios':usuarios,'form':form,'formnew':formnew,'errors': errors,'state':state, 'destino': destino})

"""
grupo de funciones que
controla el ABM de grupos
de usuarios SPID
"""
#la funcion en donde se actualiza y/o elimina los grupos ingresados
@login_required
@permission_required('user.is_staff')
def grupos(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    grupos = ""
    errors=[]
    if request.POST.get('grabar')=="Grabar":
            form = GroupForm(request.POST)
            grupo = Group()
            grupos = request.POST.get('name')
            if not grupos:
                        errors.append('Ingrese Grupo de Usuario')
            else:
                             if not(len(grupos)>=4 and len(grupos)< 80):
                                         errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
                             else:

                                             if form.is_valid():
                                                grupo.name = grupos
                                                try:
                                                    grupo.save()
                                                except IntegrityError:
                                                    errors.append('Grupo de Usuario ya existe')
                                                lista = Group.objects.all()

                                             else:
                                                 errors.append('El Grupo de Usuario que Ud. intenta grabar ya existe')

            form = GroupForm()
            lista = Group.objects.all()
            return render(request,'./grupusers.html',{'form':form,'grupos':grupos,'errors': errors,'lista':lista,'state':state, 'destino': destino})
    else:

         form = GroupForm()
         lista = Group.objects.all()

         return render(request,'./grupusers.html',{'form':form,'grupos':grupos,'errors': errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se actualiza y/o elimina los paises
@login_required
@permission_required('user.is_staff')
def grupusers(request, idgr):
    state= request.session.get('state')
    destino= request.session.get('destino')
    grupos = ""
    errors=[]
    if request.POST.get('cancelar')=="Cancelar":
         form = GroupForm(request.POST)
         lista = Group.objects.all()
         return render(request,'./grupusers.html',{'form':form,'grupos':grupos,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":
             try:
                 Group.objects.get(id=idgr).delete()
                 noborro=False
             except IntegrityError:
                 noborro=True
             form = GroupForm(request.POST)
             lista = Group.objects.all()
             if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
             else:
                 return HttpResponseRedirect('../')
        else:
            if request.POST.get('modifica')=='Actualizar':
                            grupo = Group.objects.get(id=idgr)
                            form = GroupForm(request.POST, request.FILES)
                            if form.is_valid():
                                 name =  request.POST.get('name')
                                 grupo.name = name
                                 grupo.save()

                            else:
                                errors.append('Error ya existe un grupo de Usuario con esa referencia')
    form = GroupForm(request.POST)
    lista = Group.objects.all()
    grupos= Group.objects.get(id=idgr)


    return render(request,'./grupusers.html',{'form':form,'grupos':grupos,'errors':errors,'lista':lista,'state':state, 'destino': destino})


#la funcion en donde se guardan los paises que son ingresados en provincia
@login_required
@permission_required('user.is_staff')
def npais(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    ciudades = ""
    errors=[]
    if request.POST.get('grabar')=="Grabar":
            form = PaisesForm(request.POST)
            pais = request.POST.get('descripcion')
            if not pais:
                        errors.append('Ingrese pais')
                        return HttpResponseRedirect('.')
            else:
                             if not(len(pais)>=4 and len(pais)< 45):
                                 errors.append('El dato ingresado debe tener entre 4 y 45 caracteres')
                             else:
                                         if form.is_valid():
                                             form.save()
                                             formp = ProvinciasForm(request.POST)
                                             form = PaisesForm(request.POST)
                                             return HttpResponseRedirect('../provincias/')
                                         else:
                                             errors.append('El Pais que Ud. intenta grabar ya existe')
            lista = RefProvincia.objects.all()
            combo = RefPaises.objects.all()
            form = PaisesForm(request.POST)
            return render(request,'./provincias.html',{'formp':formp,'combo':combo,'ciudades':ciudades,'lista':lista,'state':state, 'errors':errors,'destino': destino})
    else:
         formp = ProvinciasForm()
         form  = PaisesForm()
         lista = RefProvincia.objects.all()
         combo = RefPaises.objects.all()
         return render(request,'./provincias.html',{'form':form,'formp':formp,'combo':combo,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})


@login_required
def obtener_barrios(request,idcit):
                data = request.POST
                barrios = RefBarrios.objects.filter(ciudad = idcit)
                data = serializers.serialize("json", barrios)
                return HttpResponse(data, content_type='application/json')

@login_required
def obtener_calle(request,idcit):
                data = request.POST
                calles = RefCalles.objects.filter(ciudad = idcit)
                data = serializers.serialize("json", calles)
                return HttpResponse(data, content_type='application/json')

@login_required
def obtener_ciudades(request,pais):
                data = request.POST
                ciudades = RefCiudades.objects.filter(pais= pais)
                data = serializers.serialize("json", ciudades)
                return HttpResponse(data, content_type='application/json')

@login_required
def obtener_modos(request,idd):
                data = request.POST
                modos = RefModosHecho.objects.filter(delito = idd)
                data = serializers.serialize("json", modos)

                return HttpResponse(data, content_type='application/json')

@login_required
def obtener_subtiposa(request,uso):
                data = request.POST
                subtiposa = RefSubtiposa.objects.filter(tipo = uso)
                data = serializers.serialize("json", subtiposa)
                return HttpResponse(data, content_type='application/json')

@login_required
def obtener_delitos(request,idtd):
                data = request.POST
                delitos = RefDelito.objects.filter(tipo_delito = idtd)
                data = serializers.serialize("json", delitos)
                return HttpResponse(data, content_type='application/json')


@login_required
def obtener_dependencias(request,ure):
                data = request.POST
                dependencias = Dependencias.objects.filter(unidades_regionales_id = ure)
                data = serializers.serialize("json", dependencias)
                return HttpResponse(data, content_type='application/json')

@login_required
def obtener_provincia(request,pais):
                data = request.POST
                provincias = RefProvincia.objects.filter(pais = pais)
                data = serializers.serialize("json", provincias)
                return HttpResponse(data, content_type='application/json')

@login_required
def obtener_departa(request,prvi):
                data = request.POST
                depa = RefDepartamentos.objects.filter(provincia = prvi)
                data = serializers.serialize("json", depa)
                return HttpResponse(data, content_type='application/json')

@login_required
@group_required(["policia","investigaciones","visita","radio"])
def ciudadesadd(request):
    ciudades = ""
    departa = ""
    errors=[]
    state= request.session.get('state')
    destino= request.session.get('destino')

    if request.POST.get('grabarciu')=="Grabar":
         formc = CiudadesForm(request.POST, request.FILES)
         descripcion = request.POST.get('descripcion')
         pais = request.POST.get('pais')
         if not descripcion or not pais:
                 errors.append('Seleccione un Pais y una referencia de Ciudad que pertenezca a ese Pais')
         else:
                         if not(len(descripcion)>=4 and len(descripcion)< 80):
                                         errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
                         else:
                                                 formc = CiudadesForm(request.POST, request.FILES)

                                                 if formc.is_valid():
                                                         formc.save()
                                                         errors.append('Datos guardados')
                                                         form=PersonasForm()
                                                         return render(request, './ciudadadd.html', {'errors': errors,'form':form,})

                                                         #return HttpResponse('')
                                                 else:
                                                            errors.append('La ciudad que UD. desea Guardar ya Existe. Verifique')

         formc = CiudadesForm()
         formd = DepartamentosForm()
         formp = ProvinciasForm()
         lista = RefCiudades.objects.all()
         return render(request,'./ciudadadd.html', {'formd':formd,'formp':formp,'formc:':formc,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})
    else:
         formc = []
         formd = []
         formp = ProvinciasForm()
         lista = RefCiudades.objects.all()

    return render(request,'./ciudadadd.html', {'formd':formd,'formp':formp,'formc:':formc,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})


@login_required
@permission_required('user.is_staff')
def ciudades(request):
    ciudades = ""
    departa = ""
    errors=[]
    state= request.session.get('state')
    destino= request.session.get('destino')
    if request.POST.get('grabar')=="Grabar":
         formc = CiudadesForm(request.POST, request.FILES)
         descripcion = request.POST.get('descripcion')
         pais = request.POST.get('pais')
         provincia = request.POST.get('provincia')
         departamento = request.POST.get('departamento')
         if not descripcion or not pais:
                 errors.append('Ingrese una referencia(Pais / Provincia / Dpto ) a la que pertenece la ciudad')
         else:
                         if not(len(descripcion)>=4 and len(descripcion)< 80):
                                         errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
                         else:
                                                 formc = CiudadesForm(request.POST, request.FILES)

                                                 if formc.is_valid():
                                                         formc.save()
                                                         return HttpResponseRedirect('.')
                                                 else:
                                                            errors.append('La ciudad que UD. desea Guardar ya Existe. Verifique')

         formc = CiudadesForm()
         formd = DepartamentosForm()
         formp = ProvinciasForm()
         lista = RefCiudades.objects.all()
         return render(request,'./cities.html', {'formd':formd,'formp':formp,'formc:':formc,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})
    else:
         formc = []
         formd = []
         formp = ProvinciasForm()
         lista = RefCiudades.objects.all()

    return render(request,'./cities.html', {'formd':formd,'formp':formp,'formc:':formc,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})
@login_required
@permission_required('user.is_staff')
def ciudad(request, idciu):
    state= request.session.get('state')
    destino= request.session.get('destino')
    departa = ""
    errors = []
    if request.POST.get('cancelar')=="Cancelar":
         formc = CiudadesForm()
         formd = DepartamentosForm()
         formp = ProvinciasForm()
         lista = RefCiudades.objects.all()
         return render(request,'./cities.html', {'formd':formd,'formp':formp,'formc:':formc,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":
            try:
             RefCiudades.objects.get(id=idciu).delete()
             noborro=False
            except IntegrityError:
             noborro=True

            formc = CiudadesForm()
            formd = DepartamentosForm()
            formp = ProvinciasForm()
            lista = RefCiudades.objects.all()
            if noborro:
                return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
                return HttpResponseRedirect('../../')
        else:
            if request.POST.get('modifica')=='Actualizar':
                            ciuda = RefCiudades.objects.get(id=idciu)
                            formc = CiudadesForm(request.POST, request.FILES)

                            if formc.is_valid():

                                 descripcion = formc.cleaned_data['descripcion']
                                 departamento = formc.cleaned_data['departamento']
                                 provincia = formc.cleaned_data['provincia']
                                 pais = formc.cleaned_data['pais']
                                 ciuda.descripcion = descripcion
                                 ciuda.pais = pais
                                 ciuda.provincia = provincia
                                 ciuda.departamento = departamento
                                 ciuda.save()
                                 return HttpResponseRedirect('../../')
                            else:
                                 errors.append('Verifique los Datos. Error.')
                                 return HttpResponseRedirect('../../',{'errors':errors})
            else:
                ciudad= RefCiudades.objects.get(id=idciu)
                prov= RefCiudades.objects.filter(provincia__isnull=False)
                dep = RefCiudades.objects.filter(departamento__isnull=False)
                if ciudad in prov:
                     formd = DepartamentosForm(instance=ciudad)
                     formc = CiudadesForm(instance=ciudad)
                else:
                     formd=[]
                     formc=[]
                formp = ProvinciasForm(instance=ciudad)
                lista = RefCiudades.objects.all()
                return render(request,'./cities.html', {'formd':formd,'formp':formp,'ciudad':ciudad,'formc':formc,'errors': errors,'lista':lista,'state':state, 'destino': destino})

@login_required
@permission_required('user.is_staff')
def departamentos(request):
    departa = ""
    errors= []
    state= request.session.get('state')
    destino= request.session.get('destino')
    if request.POST.get('grabar')=="Grabar":
         formd = DepartamentosForm(request.POST, request.FILES)
         descripcion = request.POST.get('descripcion')
         provincia = request.POST.get('provincia')
         if not descripcion:
                 errors.append('Ingrese departamento')
         else:
                         if not(len(descripcion)>=4 and len(descripcion)< 45):
                                         errors.append('El dato ingresado debe tener entre 4 y 45 caracteres')
                         else:
                                         if formd.is_valid():
                                             formd.save()
                                             return HttpResponseRedirect('.')
                                         else:
                                             errors.append('El Departamento ya existe!!!.Verifique que el ingreso de datos sea correcto y completo')

         dpto = RefProvincia.objects.filter(descripcion__contains='CHUBUT').values('id')
         provincia = RefProvincia.objects.get(id=dpto)
         formd = DepartamentosForm()
         lista = RefDepartamentos.objects.all()
         combo = RefPaises.objects.all()
         return render(request,'./departamentos.html',{'formd':formd,'provincia':provincia,'combo':combo,'departa':departa,'errors': errors,'lista':lista,'state':state, 'destino': destino})
    else:
         dpto = RefProvincia.objects.filter(descripcion__contains='CHUBUT').values('id')
         if not dpto:
                errors.append('error')
                provincia =""
         else:
                provincia = RefProvincia.objects.get(id=dpto)

         formd = DepartamentosForm()
         lista = RefDepartamentos.objects.all()
         combo = RefProvincia.objects.all()
         return render(request,'./departamentos.html',{'formd':formd,'provincia':provincia,'combo':combo,'departa':departa,'errors': errors,'lista':lista,'state':state, 'destino': destino})

@login_required
@permission_required('user.is_staff')
def depto(request, iddepto):
    state= request.session.get('state')
    destino= request.session.get('destino')
    departa = ""
    errors = []
    if request.POST.get('cancelar')=="Cancelar":
         dpto = RefProvincia.objects.filter(descripcion__contains='CHUBUT').values('id')
         provincia = RefProvincia.objects.get(id=dpto)
         formd = DepartamentosForm()
         lista = RefDepartamentos.objects.all()
         combo = RefProvincia.objects.all()
         return render(request,'./departamentos.html',{'formd':formd,'provincia':provincia,'combo':combo,'departa':departa,'errors': errors,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":
            try:
             RefDepartamentos.objects.get(id=iddepto).delete()
             noborro=False
            except IntegrityError:
             noborro=True

            dpto = RefProvincia.objects.filter(descripcion__contains='CHUBUT').values('id')
            provincia = RefProvincia.objects.get(id=dpto)
            formd = DepartamentosForm()
            lista = RefDepartamentos.objects.all()
            combo = RefProvincia.objects.all()

            if noborro:
                return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
                return HttpResponseRedirect('../')
        else:
            if request.POST.get('modifica')=='Actualizar':
                            dptos = RefDepartamentos.objects.get(id=iddepto)
                            formd = DepartamentosForm(request.POST, request.FILES)
                            if formd.is_valid():
                                 descripcion = formd.cleaned_data['descripcion']
                                 provincia = formd.cleaned_data['provincia']
                                 dptos.descripcion = descripcion
                                 dptos.save()
                                 dpto = RefProvincia.objects.filter(descripcion__contains='CHUBUT').values('id')
                                 provincia = RefProvincia.objects.get(id=dpto)
                                 formd = DepartamentosForm()
                                 lista = RefDepartamentos.objects.all()
                                 combo = RefProvincia.objects.all()
                                 return HttpResponseRedirect('../')
            else:
                dpto = RefProvincia.objects.filter(descripcion__contains='CHUBUT').values('id')
                departa = RefDepartamentos.objects.get(id=iddepto)
                provincia = RefProvincia.objects.get(id=dpto)
                formd = DepartamentosForm()
                lista = RefDepartamentos.objects.all()
                combo = RefProvincia.objects.all()
                return render(request,'./departamentos.html',{'formd':formd,'provincia':provincia,'combo':combo,'departa':departa,'errors': errors,'lista':lista,'state':state, 'destino': destino})

@login_required
@permission_required('user.is_staff')
def provincias(request):
    ciudades = ""
    errors= []
    state= request.session.get('state')
    destino= request.session.get('destino')
    if request.POST.get('grabar')=="Grabar":
         formp = ProvinciasForm(request.POST)
         form = PaisesForm(request.POST)
         descripcion = request.POST.get('descripcion')
         pais = request.POST.get('pais')
         if not descripcion:
                 errors.append('Ingrese provincia')
         else:
                        if not(len(descripcion)>=4 and len(descripcion)< 45):
                                         errors.append('El dato ingresado debe tener entre 4 y 45 caracteres')
                        else:
                                         if formp.is_valid():
                                             pais = formp.cleaned_data['pais']
                                             formp.save()
                                             return HttpResponseRedirect('.')
                                         else:
                                                errors.append('Datos inválidos. Verifique que el ingreso de datos sea correcto y completo')
         lista = RefProvincia.objects.all()
         combo = RefPaises.objects.all()
         form = PaisesForm(request.POST)
         return render(request,'./provincias.html',{'formp':formp,'combo':combo,'ciudades':ciudades,'lista':lista,'state':state, 'errors':errors,'destino': destino})
    else:
         formp = ProvinciasForm()
         form  = PaisesForm()
         lista = RefProvincia.objects.all()
         combo = RefPaises.objects.all()
         return render(request,'./provincias.html',{'form':form,'formp':formp,'combo':combo,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})

@login_required
@permission_required('user.is_staff')
def provi(request, idpcia):
    state= request.session.get('state')
    destino= request.session.get('destino')
    ciudades = ""
    errors = []
    if request.POST.get('cancelar')=="Cancelar":
         formp = ProvinciasForm()
         form  = PaisesForm()
         lista = RefProvincia.objects.all()
         combo = RefPaises.objects.all()
         return render(request,'./provincias.html',{'form':form,'ciudades':ciudades,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":
            try:
             RefProvincia.objects.get(id=idpcia).delete()
             noborro=False
            except IntegrityError:
             noborro=True
            form = PaisesForm(request.POST)
            formp = ProvinciasForm()
            combo = RefPaises.objects.all()
            lista= RefProvincia.objects.all()

            if noborro:
             return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
             return HttpResponseRedirect('../')
        else:
            if request.POST.get('modifica')=='Actualizar':
                            provi = RefProvincia.objects.get(id=idpcia)
                            formp = ProvinciasForm(request.POST, request.FILES)
                            if formp.is_valid():
                                 descripcion = formp.cleaned_data['descripcion']
                                 pais = formp.cleaned_data['pais']
                                 provi.descripcion = descripcion
                                 provi.pais = pais
                                 provi.save()
                                 form = PaisesForm(request.POST)
                                 formp = ProvinciasForm()
                                 lista = RefPaises.objects.all()
                                 ciudades= RefProvincia.objects.all()
                                 return HttpResponseRedirect('../')
            else:
                ciudades= RefProvincia.objects.get(id=idpcia)
                formp = ProvinciasForm(instance=ciudades)
                form  = PaisesForm()
                lista = RefProvincia.objects.all()

                return render(request,'./provincias.html',{'form':form,'formp':formp,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se guardan los paises que son ingresados en provincia
@login_required
@permission_required('user.is_staff')
def npais(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    ciudades = ""
    errors=[]
    if request.POST.get('grabar')=="Grabar":
            form = PaisesForm(request.POST)
            pais = request.POST.get('descripcion')
            if not pais:
                        errors.append('Ingrese pais')
                        return HttpResponseRedirect('.')
            else:
                             if not(len(pais)>=4 and len(pais)< 45):
                                 errors.append('El dato ingresado debe tener entre 4 y 45 caracteres')
                             else:
                                         if form.is_valid():
                                             form.save()
                                             formp = ProvinciasForm(request.POST)
                                             form = PaisesForm(request.POST)
                                             return HttpResponseRedirect('../provincias/')
                                         else:
                                             errors.append('El Pais que Ud. intenta grabar ya existe')
            lista = RefProvincia.objects.all()
            combo = RefPaises.objects.all()
            form = PaisesForm(request.POST)
            return render(request,'./provincias.html',{'formp':formp,'combo':combo,'ciudades':ciudades,'lista':lista,'state':state, 'errors':errors,'destino': destino})
    else:
         formp = ProvinciasForm()
         form  = PaisesForm()
         lista = RefProvincia.objects.all()
         combo = RefPaises.objects.all()
         return render(request,'./provincias.html',{'form':form,'formp':formp,'combo':combo,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})



#la funcion en donde se actualiza y/o elimina los paises ingresados
@login_required
@permission_required('user.is_staff')
def pais(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    ciudades = ""
    errors=[]
    if request.POST.get('grabar')=="Grabar":
            form = PaisesForm(request.POST)
            pais = request.POST.get('descripcion')
            if not pais:
                        errors.append('Ingrese pais')
            else:
                             if not(len(pais)>=4 and len(pais)< 45):
                                         errors.append('El dato ingresado debe tener entre 4 y 45 caracteres')
                             else:
                                             if form.is_valid():
                                                form.save()
                                                lista = RefPaises.objects.all()
                                                return HttpResponseRedirect('.')
                                             else:
                                                 errors.append('El Pais que Ud. intenta grabar ya existe')

            form = PaisesForm()
            lista = RefPaises.objects.all()
            return render(request,'./paises.html',{'form':form,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})
    else:

         form = PaisesForm()
         lista = RefPaises.objects.all()
         return render(request,'./paises.html',{'form':form,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se actualiza y/o elimina los paises
@login_required
@permission_required('user.is_staff')
def paise(request, idlista):
    state= request.session.get('state')
    destino= request.session.get('destino')
    ciudades = ""
    errors=[]
    if request.POST.get('cancelar')=="Cancelar":
         form = PaisesForm(request.POST)
         lista = RefPaises.objects.all()
         return render(request,'./paises.html',{'form':form,'ciudades':ciudades,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":

            try:
             RefPaises.objects.get(id=idlista).delete()
             noborro=False
            except IntegrityError:
             noborro=True

            form = PaisesForm(request.POST)
            lista = RefPaises.objects.all()
            return render(request,'./paises.html',{'errors':errors,'form':form,'ciudades':ciudades,'lista':lista,'state':state, 'destino': destino})

        else:
            if request.POST.get('modifica')=='Actualizar':
                            pais = RefPaises.objects.get(id=idlista)
                            form = PaisesForm(request.POST, request.FILES)
                            if form.is_valid():
                                 descripcion = form.cleaned_data['descripcion']
                                 pais.descripcion = descripcion
                                 pais.save()
                                 form = PaisesForm(request.POST)
                                 lista = RefPaises.objects.all()
                                 return HttpResponseRedirect('../')
            else:
                form = PaisesForm(request.POST)
                lista = RefPaises.objects.all()
                ciudades= RefPaises.objects.get(id=idlista)
    #.values_list('descripcion', flat=False).distinct()
    return render(request,'./paises.html',{'form':form,'ciudades':ciudades,'lista':lista,'state':state, 'destino': destino})


#la funcion en donde se cargan los tipos de lugares
@login_required
@permission_required('user.is_staff')
def lugares(request):
        state= request.session.get('state')
        destino= request.session.get('destino')
        lugar = ""
        errors=[]
        if request.POST.get('grabar')=="Grabar":
            form = LugaresForm(request.POST)
            lugar = request.POST.get('descripcion')
            if not lugar:
                        errors.append('Ingrese Tipo de Lugar')
            else:
                             if not(len(lugar)>=4 and len(lugar)< 100):
                                         errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
                             else:
                                             if form.is_valid():
                                                form.save()
                                                lista = RefLugares.objects.all()
                                                return HttpResponseRedirect('.')
                                             else:
                                                 errors.append('La Referencia de Lugar que Ud. intenta grabar ya existe')

            form = LugaresForm()
            lista = RefLugares.objects.all()
            return render(request,'./lugares.html',{'form':form,'lista':lista,'state':state,'errors': errors,'destino': destino,'lugar':lugar})
        else:
         form = LugaresForm()
         lista = RefLugares.objects.all()
         return render(request,'./lugares.html',{'form':form,'lista':lista,'state':state,'errors': errors,'destino': destino,'lugar':lugar})

#la funcion en donde se actualiza y/o elimina los tipos de lugares en donde ocurrio los hechos
@login_required
@permission_required('user.is_staff')
def nlugares(request, idlugar):
    state= request.session.get('state')
    destino= request.session.get('destino')
    lugar = ""
    if request.POST.get('cancelar')=="Cancelar":
         form = LugaresForm(request.POST)
         lista = RefLugares.objects.all()
         return render(request,'./lugares.html',{'form':form,'lugar':lugar,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":
             try:
                 RefLugares.objects.get(id=idlugar).delete()
                 noborro=False
             except IntegrityError:
                 noborro=True

             form = LugaresForm(request.POST)
             lista = RefLugares.objects.all()
             if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
             else:
                 return HttpResponseRedirect('../')
        else:
            if request.POST.get('modifica')=='Actualizar':
                            lugar = RefLugares.objects.get(id=idlugar)
                            form = LugaresForm(request.POST, request.FILES)
                            if form.is_valid():
                                 descripcion = form.cleaned_data['descripcion']
                                 lugar.descripcion = descripcion
                                 lugar.save()
                                 form = LugaresForm(request.POST)
                                 lista = RefLugares.objects.all()
                                 return HttpResponseRedirect('../')
                            else:
                                 errors.append('Verifique los Datos. Error.')
            else:
                form = LugaresForm(request.POST)
                lista = RefLugares.objects.all()
                lugar= RefLugares.objects.get(id=idlugar)

    return render(request,'./lugares.html',{'form':form,'lugar':lugar,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se carga el template y se guarda los tipos de viviendas
@login_required
@permission_required('user.is_staff')
def hogares(request):
        state= request.session.get('state')
        destino= request.session.get('destino')
        hogar = ""
        errors=[]
        if request.POST.get('grabar')=="Grabar":
            form = HogaresForm(request.POST)
            casa = request.POST.get('descripcion')
            hogar=RefHogares()
            if not casa:
                        errors.append('Ingrese Tipo de Hogar')
            else:
                             if not(len(casa)>=4 and len(casa)<100):
                                         errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
                             else:
                                             if form.is_valid():

                                                hogar.descripcion = casa
                                                hogar.save()
                                                lista = RefHogares.objects.all()
                                                return HttpResponseRedirect('.')
                                             else:
                                                 errors.append('La Referencia de Hogar / Casa que Ud. intenta grabar ya existe')

            form = HogaresForm()
            lista = RefHogares.objects.all()
            return render(request,'./hogares.html',{'form':form,'lista':lista,'state':state,'errors': errors,'destino': destino,'hogar':hogar})
        else:
         form = HogaresForm()
         lista = RefHogares.objects.all()
         return render(request,'./hogares.html',{'form':form,'lista':lista,'hogar':hogar, 'state':state,'errors': errors,'destino': destino,'hogar':hogar})

#la funcion en donde se actualiza y/o elimina un tipos de viviendas
@login_required
@permission_required('user.is_staff')
def nhogares(request, idipv):
    state= request.session.get('state')
    destino= request.session.get('destino')
    hogar = ""
    errors=[]
    if request.POST.get('cancelar')=="Cancelar":
         form = HogaresForm(request.POST)
         lista = RefHogares.objects.all()
         return render(request,'./hogares.html',{'form':form,'hogar':hogar,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":
            try:
             RefHogares.objects.get(id=idipv).delete()
             noborro=False
            except IntegrityError:
             noborro=True

            form = HogaresForm(request.POST)
            lista = RefHogares.objects.all()
            if noborro:
                return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
             return HttpResponseRedirect('../')
        else:

            if request.POST.get('modifica')=='Actualizar':
                            hogares = RefHogares.objects.get(id=idipv)
                            form = HogaresForm(request.POST, request.FILES)

                            if form.is_valid():
                                 descripcion = form.cleaned_data['descripcion']
                                 hogares.descripcion = descripcion
                                 hogares.save()
                                 form = HogaresForm(request.POST)
                                 lista = RefHogares.objects.all()
                                 return HttpResponseRedirect('../')
            else:
                form = HogaresForm(request.POST)
                lista = RefHogares.objects.all()
                hogar= RefHogares.objects.get(id=idipv)


    return render(request,'./hogares.html',{'form':form,'hogar':hogar,'lista':lista,'state':state, 'destino': destino})

#condiciones climaticas
@login_required
@permission_required('user.is_staff')
def climas(request):
        state= request.session.get('state')
        destino= request.session.get('destino')
        climas = ""
        errors=[]
        if request.POST.get('grabar')=="Grabar":
            form = CondclimasForm(request.POST)
            condi = request.POST.get('descripcion')
            if not condi:
                        errors.append('Ingrese Condicion climatica')
            else:
                             if not(len(condi)>=4 and len(condi)< 150):
                                         errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
                             else:
                                             if form.is_valid():
                                                form.save()
                                                lista = RefCondclimas.objects.all()
                                                return HttpResponseRedirect('.')
                                             else:
                                                 errors.append('La Referencia de Clima que Ud. intenta grabar ya existe')

            form = CondclimasForm()
            lista = RefCondclimas.objects.all()
            return render(request,'./climas.html',{'form':form,'lista':lista,'state':state,'errors': errors,'destino': destino,'climas':climas})
        else:
         form = CondclimasForm()
         lista = RefCondclimas.objects.all()
         return render(request,'./climas.html',{'form':form,'lista':lista,'state':state,'errors': errors,'destino': destino,'climas':climas})

#la funcion en donde se actualiza y/o elimina los tipos de climas referente a un hecho delictivo
@login_required
@permission_required('user.is_staff')
def nclimas(request, idcli):
    state= request.session.get('state')
    destino= request.session.get('destino')
    climas = ""
    if request.POST.get('cancelar')=="Cancelar":
         form = CondclimasForm(request.POST)
         lista = RefCondclimas.objects.all()
         return render(request,'./climas.html',{'form':form,'climas':climas,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":
            try:
             RefCondclimas.objects.get(id=idcli).delete()
             noborro=False
            except IntegrityError:
             noborro=True
            form = CondclimasForm(request.POST)
            lista = RefCondclimas.objects.all()
            if noborro:
                return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
                return HttpResponseRedirect('../')
        else:
            if request.POST.get('modifica')=='Actualizar':
                            climas = RefCondclimas.objects.get(id=idcli)
                            form = CondclimasForm(request.POST, request.FILES)
                            if form.is_valid():
                                 descripcion = form.cleaned_data['descripcion']
                                 climas.descripcion = descripcion
                                 climas.save()
                                 form = CondclimasForm(request.POST)
                                 lista = RefCondclimas.objects.all()
                                 return HttpResponseRedirect('../')
            else:
                form = CondclimasForm(request.POST)
                lista = RefCondclimas.objects.all()
                climas= RefCondclimas.objects.get(id=idcli)

    return render(request,'./climas.html',{'form':form,'climas':climas,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se carga el template de unidades regionales y se graba
@login_required
@permission_required('user.is_staff')
def unidades(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    unidades = ""
    if request.POST.get('grabar')=='Grabar':
        form = UnidadesForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        ciudad = request.POST.get('ciudad')
        if not descripcion:
            errors.append('Ingrese Unidad Regional')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<80):
                    errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
             else:
                    if ciudad == 'Seleccione ciudad':
                        errors.append('Debe seleccionar una ciudad')
                    else:

                        if form.is_valid():
                            form.save()
                            lista = UnidadesRegionales.objects.all()
                            return HttpResponseRedirect('.')
                        else:
                            errors.append('La Unidad que intenta grabar ya existe')
    form = UnidadesForm()
    lista = UnidadesRegionales.objects.all()
    ciudades = RefCiudades.objects.filter(provincia__descripcion__contains = 'CHUBUT').values('id','descripcion')
    return render(request,'./unidades.html',{'ciudades':ciudades,'unidades':unidades,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se actualiza y/o elimina una Unidad Regional
@login_required
@permission_required('user.is_staff')
def unidad(request, idUnidad):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    unidades = ""
    if request.POST.get('cancelar')=='Cancelar':
        form = UnidadesForm(instance = unidades)
        lista = UnidadesRegionales.objects.all()
        return render(request,'./unidades.html',{'unidades':unidades,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('modifica')=='Actualizar':
            unidad = UnidadesRegionales.objects.get(id = idUnidad)
            form = UnidadesForm(request.POST, request.FILES)
            if form.is_valid():
                descripcion = form.cleaned_data['descripcion']
                ciudad = form.cleaned_data['ciudad']
                unidad.descripcion = descripcion
                unidad.ciudad = ciudad
                unidad.save()
                return HttpResponseRedirect('../')
            else:
                return HttpResponseRedirect('../')
        if request.POST.get('borrar')=='Borrar':
            try:
                UnidadesRegionales.objects.get(id=idUnidad).delete()
                noborro=False
            except IntegrityError:
                noborro=True

            form = UnidadesForm()
            lista = UnidadesRegionales.objects.all()
            if noborro:
                return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
                return HttpResponseRedirect('../')
    unidades = UnidadesRegionales.objects.get(id = idUnidad)
    form = UnidadesForm(instance = unidades)
    lista = UnidadesRegionales.objects.all()
    return render(request,'./unidades.html',{'unidades':unidades,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se carga el template de ingreso de dependencias
@login_required
@permission_required('user.is_staff')
def dependencias(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    depes = ""
    if request.POST.get('grabar')=='Grabar':
        form = DependenciasForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        ciudad = request.POST.get('ciudad')
        unidad = request.POST.get('unidades_regionales')
        if not descripcion:
            errors.append('Ingrese el nombre de la dependencia')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<80):
                    errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
             else:
                    if not ciudad:
                        errors.append('Debe seleccionar una ciudad')
                    else:
                        if not unidad:
                            errors.append('Debe seleccionar una unidad regional')
                        else:
                            if form.is_valid():
                                form.save()
                                lista = Dependencias.objects.all()
                                return HttpResponseRedirect('.')
                            else:
                                errors.append('La Dependencias que intenta grabar ya existe')
    form = DependenciasForm()
    lista = Dependencias.objects.all()
    return render(request,'./depe.html',{'depes':depes,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se actualiza y/o elimina las dependencias segun unidad regional
@login_required
@permission_required('user.is_staff')
def dependencia(request, idDepe):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    depes = ""
    if request.POST.get('cancelar')=='Cancelar':
        form = DependenciasForm(instance = depes)
        lista = Dependencias.objects.all()
        return render(request,'./unidades.html',{'depes':depes,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('modifica')=='Actualizar':
            depe = Dependencias.objects.get(id = idDepe)
            form = DependenciasForm(request.POST, request.FILES)
            if form.is_valid():
                descripcion = form.cleaned_data['descripcion']
                ciudad = form.cleaned_data['ciudad']
                unidad = form.cleaned_data['unidades_regionales']
                depe.descripcion = descripcion
                depe.ciudad = ciudad
                depe.unidades_regionales=unidad
                depe.save()
                return HttpResponseRedirect('../')
            else:
                return HttpResponseRedirect('../')
        if request.POST.get('borrar')=='Borrar':
            try:
                 Dependencias.objects.get(id=idDepe).delete()
                 noborro=False
            except IntegrityError:
                 noborro=True
            form = DependenciasForm()
            lista = Dependencias.objects.all()
            if noborro:
                return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
             return HttpResponseRedirect('../')

    depes = Dependencias.objects.get(id = idDepe)
    form = DependenciasForm(instance = depes)
    lista = Dependencias.objects.all()
    return render(request,'./depe.html',{'depes':depes,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se carga el template para el ingreso de tipos de personas involucradas
@login_required
@permission_required('user.is_staff')
def personasi(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    invo = ""
    errors=[]
    if request.POST.get('grabar')=="Grabar":
            form = PeopleForm(request.POST)
            invo = request.POST.get('descripcion')
            if not invo:
                        errors.append('Ingrese Tipo de Persona Involucrada')
            else:
                            if not(len(invo)>=4 and len(invo)< 150):
                                         errors.append('El dato ingresado debe tener entre 4 y 150 caracteres')
                            else:
                                             if form.is_valid():
                                                form.save()
                                                lista = RefPeople.objects.all()
                                                return HttpResponseRedirect('.')
                                             else:
                                                 errors.append('El Tipo de persona involucrada que Ud. intenta grabar ya existe')

            form = PeopleForm()
            lista = RefPeople.objects.all()
            return render(request,'./peopleenv.html',{'form':form,'invo':invo,'errors': errors,'lista':lista,'state':state, 'destino': destino})
    else:

         form = PeopleForm()
         lista = RefPeople.objects.all()
         return render(request,'./peopleenv.html',{'form':form,'invo':invo,'errors': errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se actualiza y/o elimina un tipo de persona involucrada
@login_required
@permission_required('user.is_staff')
def npersonasi(request, idple):
    state= request.session.get('state')
    destino= request.session.get('destino')
    invo = ""
    if request.POST.get('cancelar')=="Cancelar":
         form = PeopleForm(request.POST)
         lista = RefPeople.objects.all()
         return render(request,'./peopleenv.html',{'form':form,'invo':invo,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":
             try:
                 RefPeople.objects.get(id=idple).delete()
                 noborro=False
             except ObjectDoesNotExist:
                noborro=True

             form = PeopleForm(request.POST)
             lista = RefPeople.objects.all()
             if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
             else:
                 return HttpResponseRedirect('../')
        else:
            if request.POST.get('modifica')=='Actualizar':
                            invo = RefPeople.objects.get(id=idple)
                            form = PeopleForm(request.POST, request.FILES)
                            if form.is_valid():
                                 descripcion = form.cleaned_data['descripcion']
                                 invo.descripcion = descripcion
                                 invo.save()
                                 form = PeopleForm(request.POST)
                                 lista = RefPeople.objects.all()
                                 return HttpResponseRedirect('../')
            else:
                form = PeopleForm(request.POST)
                lista = RefPeople.objects.all()
                invo= RefPeople.objects.get(id=idple)
    return render(request,'./peopleenv.html',{'form':form,'invo':invo,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se guardan los paises que son ingresados en provincia
@login_required
@permission_required('user.is_staff')
def ntipodelito(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    ciudades = ""
    errors=[]
    if request.POST.get('grabar')=="Grabar":
            form = TipoDelitosForm(request.POST)
            descripcion = request.POST.get('descripcion')
            if not descripcion:
                        errors.append('Ingrese Tipos de Delitos')
                        return HttpResponseRedirect('.')
            else:
                             if not(len(descripcion)>=4 and len(descripcion)< 50):
                                 errors.append('El dato ingresado debe tener entre 4 y 50 caracteres')
                             else:
                                         if form.is_valid():
                                             form.save()
                                             formp = DelitoForm(request.POST)
                                             form = TipoDelitosForm(request.POST)
                                             return HttpResponseRedirect('../delitos/')
                                         else:
                                             errors.append('El Tipo de Delito que Ud. intenta grabar ya existe')
            lista = RefDelito.objects.all()
            combo = RefTipoDelitos.objects.all()
            form = TipoDelitosForm(request.POST)
            return render(request,'./delito.html',{'formp':formp,'combo':combo,'ciudades':ciudades,'lista':lista,'state':state, 'errors':errors,'destino': destino})
    else:
         formp = DelitoForm()
         form  = TipoDelitosForm()
         lista = RefDelito.objects.all()
         combo = RefTipoDelitos.objects.all()
         return render(request,'./delito.html',{'form':form,'formp':formp,'combo':combo,'ciudades':ciudades,'errors': errors,'lista':lista,'state':state, 'destino': destino})


@login_required
@permission_required('user.is_staff')
def tipodelitos(request):
    state = request.session.get('state')
    destino = request.session.get('destino')
    errors = []
    tdel =""
    if request.POST.get('grabar')=='Grabar':
        form = TipoDelitosForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        if not descripcion:

            errors.append('Debe Ingresar una descripcion para tipo de delito')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<50):
                errors.append('El dato ingresado debe tener entre 4 y 50 caracteres')
             else:
                if form.is_valid():
                    form.save()
                    lista = RefTipoDelitos.objects.all()
                    return HttpResponseRedirect('.')
                else:
                    errors.append('Verifique la informacion a cargar y vuelva a intentar')
    form = TipoDelitosForm()
    lista = RefTipoDelitos.objects.all()
    return render(request,'./tipodel.html',{'errors':errors,'tdel':tdel,'lista':lista,'state':state,'destino':destino})

@login_required
@permission_required('user.is_staff')
def tipodelito(request,idtipo):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    tdel = ""
    if request.POST.get('cancelar')=='Cancelar':
        form = TipoDelitosForm(instance = tdel)
        lista = RefTipoDelitos.objects.all()
        return render(request,'./tipodel.html',{'tdel':tdel,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('modifica')=='Actualizar':
            tipodel = RefTipoDelitos.objects.get(id = idtipo)
            form = TipoDelitosForm(request.POST, request.FILES)
            if form.is_valid():
                descripcion = form.cleaned_data['descripcion']
                tipodel.descripcion = descripcion
                tipodel.save()
                return HttpResponseRedirect('../')
            else:
                return HttpResponseRedirect('../')
        if request.POST.get('borrar')=='Borrar':
         try:
            RefTipoDelitos.objects.get(id=idtipo).delete()
            noborro=False
         except IntegrityError:
            noborro=True

         form = TipoDelitosForm()
         lista = RefTipoDelitos.objects.all()

         if noborro:
                return HttpResponseRedirect('/nosepudoborrarlorequerido')
         else:
                return HttpResponseRedirect('../')
    tdel = RefTipoDelitos.objects.get(id = idtipo)
    form = TipoDelitosForm(instance = tdel)
    lista = RefTipoDelitos.objects.all()
    return render(request,'./tipodel.html',{'tdel':tdel,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})

@login_required
@permission_required('user.is_staff')
def delitos(request):
    state = request.session.get('state')
    destino = request.session.get('destino')
    errors = []
    delito =""
    if request.POST.get('grabar')=='Grabar':
        form = DelitoForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        if not descripcion:
            errors.append('Debe Ingresar una descripcion para tipo de delito')
        else:
            if not(len(descripcion)>=4 and len(descripcion)< 50):
                errors.append('El dato ingresado debe tener entre 4 y 50 caracteres')
            else:
                if form.is_valid():
                    form.save()
                    lista = RefTipoDelitos.objects.all()
                    return HttpResponseRedirect('.')
                else:
                    errors.append('Verifique la informacion a cargar y vuelva a intentar')
    form = DelitoForm()

    lista = RefDelito.objects.all()
    return render(request,'./delito.html',{'form':form,'errors':errors,'delito':delito,'lista':lista,'state':state,'destino':destino})

@login_required
@permission_required('user.is_staff')
def delito(request,idelito):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    delito = ""
    if request.POST.get('cancelar')=='Cancelar':
        form = DelitoForm(instance = delito)
        lista = RefDelito.objects.all()
        return render(request,'./delito.html',{'delito':delito,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('modifica')=='Actualizar':
            deli = RefDelito.objects.get(id = idelito)
            form = DelitoForm(request.POST, request.FILES)
            if form.is_valid():
                descripcion = form.cleaned_data['descripcion']
                tipo_delito = form.cleaned_data['tipo_delito']
                deli.descripcion = descripcion
                deli.tipo_delito = tipo_delito
                deli.save()
                return HttpResponseRedirect('../')
            else:
                return HttpResponseRedirect('../')
        if request.POST.get('borrar')=='Borrar':
            try:
                RefDelito.objects.get(id=idelito).delete()
                noborro=False
            except IntegrityError:
                noborro=True
            form = DelitoForm()
            lista = RefDelito.objects.all()
            if noborro:
                return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
                return HttpResponseRedirect('../')
    delito = RefDelito.objects.get(id = idelito)
    form = DelitoForm(instance = delito)
    lista = RefDelito.objects.all()
    return render(request,'./delito.html',{'delito':delito,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})


@login_required
@permission_required('user.is_staff')
def jobs(request):
    state = request.session.get('state')
    destino = request.session.get('destino')
    errors =[]
    job = ""
    if request.POST.get('grabar') == 'Grabar':
        form = JobsForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        if not descripcion:
            errors.append('El Campo descripcion es obligatorio')
        else:
                if not(len(descripcion)>=4 and len(descripcion)<50):
                    errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
                else:
                 if form.is_valid():
                    form.save()
                    lista = RefOcupacion.objects.all()
                    return HttpResponseRedirect('.')
                 else:
                    errors.append('Verifique la informacion a cargar y vuelva a intentar')
    form = JobsForm
    lista = RefOcupacion.objects.all()
    return render(request,'./jobs.html',{'form':form,'errors':errors,'job':job,'lista':lista,'state':state,'destino':destino})




@login_required
@permission_required('user.is_staff')
def jobselected(request,idjob):
        state = request.session.get('state')
        destino = request.session.get('destino')
        errors = []
        job = ""
        if request.POST.get('cancelar') == 'Cancelar':
            form = JobsForm(instance = job)
            lista = RefOcupacion.objects.all()
            return render(request,'./jobs.html',{'form':form,'job':job,'errors':errors,'lista':lista,'state':state,'destino':destino})
        else:
            if request.POST.get('modifica')  == 'Actualizar':
                jobs = RefOcupacion.objects.get(id = idjob)
                form = JobsForm(request.POST, request.FILES)
                if form.is_valid():
                    descripcion = form.cleaned_data['descripcion']
                    jobs.descripcion = descripcion
                    jobs.save()
                    return HttpResponseRedirect('../')
                else:
                    return HttpResponseRedirect('../')
            if request.POST.get('borrar') == 'Borrar':
                try:
                    RefOcupacion.objects.get(id = idjob).delete()
                    noborro=False
                except IntegrityError:
                    noborro=True

                form = JobsForm()
                lista = RefOcupacion.objects.all()
                if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
                else:
                 return HttpResponseRedirect('../')
        job = RefOcupacion.objects.get(id = idjob)
        form = JobsForm(instance = job)
        lista = RefOcupacion.objects.all()
        return render(request,'./jobs.html',{'form':form,'job':job,'errors':errors,'lista':lista,'state':state,'destino':destino})

#funcion que llama al template de marcas
@login_required
@permission_required('user.is_staff')
def marcas(request):
    state = request.session.get('state')
    destino = request.session.get('destino')
    errors = []
    marcas =""
    if request.POST.get('grabar')=='Grabar':
        form = TrademarkForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        if not descripcion:
            errors.append('Debe Ingresar una descripcion para una Marca registrada')
        else:
            if not(len(descripcion)>=2 and len(descripcion) < 100):
                errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
            else:
                if form.is_valid():
                    form.save()
                    lista = RefTrademark.objects.all()
                    return HttpResponseRedirect('.')
                else:
                    errors.append('Verifique la informacion a cargar y vuelva a intentar')
    form = TrademarkForm()
    lista = RefTrademark.objects.all()
    return render(request,'./marcas.html',{'form':form,'marcas':marcas,'errors':errors,'lista':lista,'state':state,'destino':destino})

@login_required
@permission_required('user.is_staff')
def tradeselec(request,idmars):
        state = request.session.get('state')
        destino = request.session.get('destino')
        errors = []
        marcas = ""
        if request.POST.get('cancelar') == 'Cancelar':
            form =TrademarkForm()
            lista = RefTrademark.objects.all()
            return render(request,'./marcas.html',{'form':form,'marcas':marcas,'errors':errors,'lista':lista,'state':state,'destino':destino})
        else:
            if request.POST.get('modifica')  == 'Actualizar':
                marcas = RefTrademark.objects.get(id = idmars)
                form = TrademarkForm(request.POST, request.FILES)
                if form.is_valid():
                    descripcion = form.cleaned_data['descripcion']
                    marcas.descripcion = descripcion
                    marcas.save()
                    return HttpResponseRedirect('../')
                else:
                    return HttpResponseRedirect('../')
            if request.POST.get('borrar') == 'Borrar':
                try:
                    RefTrademark.objects.get(id = idmars).delete()
                    noborro=False
                except IntegrityError:
                    noborro=True
                form = TrademarkForm()
                lista = RefTrademars.objects.all()
                if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
                else:
                 return HttpResponseRedirect('../')
        marcas = RefTrademark.objects.get(id = idmars)
        form = TrademarkForm(instance = marcas)
        lista = RefTrademark.objects.all()
        return render(request,'./marcas.html',{'form':form,'marcas':marcas,'errors':errors,'lista':lista,'state':state,'destino':destino})

#funcion que llama al template de tipos de armas
@login_required
@permission_required('user.is_staff')
def tiposarmas(request):
    state = request.session.get('state')
    destino = request.session.get('destino')
    errors = []
    tiposa =""
    if request.POST.get('grabar')=='Grabar':
        form = TiposarmasForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        if not descripcion:
            errors.append('Debe Ingresar una descripcion para tipo de Armas')
        else:
            if not(len(descripcion)>=4 and len(descripcion) < 100):
                errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
            else:
                if form.is_valid():
                    form.save()
                    lista = RefTiposarmas.objects.all()
                    return HttpResponseRedirect('.')
                else:
                    errors.append('Verifique la informacion a cargar y vuelva a intentar')
    form = TiposarmasForm()
    lista = RefTiposarmas.objects.all()
    return render(request,'./tiposarms.html',{'form':form,'tiposa':tiposa,'errors':errors,'lista':lista,'state':state,'destino':destino})

@login_required
@permission_required('user.is_staff')
def tiposaselec(request,idta):
        state = request.session.get('state')
        destino = request.session.get('destino')
        errors = []
        tiposa = ""
        if request.POST.get('cancelar') == 'Cancelar':
            form = TiposarmasForm()
            lista = RefTiposarmas.objects.all()
            return render(request,'./tiposarms.html',{'form':form,'tiposa':tiposa,'errors':errors,'lista':lista,'state':state,'destino':destino})
        else:
            if request.POST.get('modifica')  == 'Actualizar':
                tiposa = RefTiposarmas.objects.get(id = idta)
                form =TiposarmasForm(request.POST, request.FILES)
                if form.is_valid():
                    descripcion = form.cleaned_data['descripcion']
                    tiposa.descripcion = descripcion
                    tiposa.save()
                    return HttpResponseRedirect('../')
                else:
                    return HttpResponseRedirect('../')
            if request.POST.get('borrar') == 'Borrar':
                try:
                    RefItems.objects.get(id = idta).delete()
                    noborro=False
                except IntegrityError:
                    noborro=True
                form = TiposarmasForm()
                lista = RefTiposarmas.objects.all()
                if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
                else:
                 return HttpResponseRedirect('../')
        tiposa = RefTiposarmas.objects.get(id = idta)
        form = TiposarmasForm(instance = tiposa)
        lista = RefTiposarmas.objects.all()
        return render(request,'./tiposarms.html',{'form':form,'tiposa':tiposa,'errors':errors,'lista':lista,'state':state,'destino':destino})

#subtipos de armas
@login_required
@permission_required('user.is_staff')
def subtiposarms(request):
    subtiposa = ""
    errors= []
    state= request.session.get('state')
    destino= request.session.get('destino')
    if request.POST.get('grabar')=="Grabar":
         forms = SubtiposaForm(request.POST)
         form = TiposarmasForm(request.POST)
         descripcion = request.POST.get('descripcion')
         tiposa = request.POST.get('tipo')
         if not descripcion:
                 errors.append('Ingrese Sub tipo de armas de fuego')
         else:
                        if not(len(descripcion)>=4 and len(descripcion)< 100):
                                         errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
                        else:
                                         if forms.is_valid():
                                             tiposas = forms.cleaned_data['tipo']
                                             forms.save()
                                             return HttpResponseRedirect('.')
                                         else:
                                                errors.append('Datos inválidos. Verifique que el ingreso de datos sea correcto y completo')
         lista = RefSubtiposa.objects.all()
         combo = RefTiposarmas.objects.all()
         form = TiposarmasForm(request.POST)
         return render(request,'./subtiposarms.html',{'forms':forms,'combo':combo,'subtiposa':subtiposa,'lista':lista,'state':state, 'errors':errors,'destino': destino})
    else:
         forms = SubtiposaForm()
         form  = TiposarmasForm()
         lista = RefSubtiposa.objects.all()
         combo = RefTiposarmas.objects.all()
         return render(request,'./subtiposarms.html',{'form':form,'forms':forms,'combo':combo,'subtiposa':subtiposa,'errors': errors,'lista':lista,'state':state, 'destino': destino})

@login_required
@permission_required('user.is_staff')
def subaselect(request, idsta):
    state= request.session.get('state')
    destino= request.session.get('destino')
    subtiposa = ""
    errors = []
    if request.POST.get('cancelar')=="Cancelar":
        forms = SubtiposaForm()
        form = TiposarmasForm()
        lista = RefSubtiposa.objects.all()
        combo = RefTiposarmas.objects.all()
        return render(request,'./subtiposarms.html',{'form':form,'subtiposa':subtiposa,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":
            try:
             RefSubtiposa.objects.get(id=idsta).delete()
             noborro=False
            except IntegrityError:
             noborro=True
            form = TiposarmasForm(request.POST)
            forms = SubtiposaForm()
            lista = RefTiposarmas.objects.all()
            subtiposa= RefSubtiposa.objects.all()
            if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
                 return HttpResponseRedirect('../')
        else:
            if request.POST.get('modifica')=='Actualizar':
                            subtipo = RefSubtiposa.objects.get(id=idsta)
                            forms =SubtiposaForm(request.POST, request.FILES)
                            if forms.is_valid():
                                 descripcion = forms.cleaned_data['descripcion']
                                 tipo = forms.cleaned_data['tipo']
                                 subtipo.descripcion = descripcion
                                 subtipo.tipo = tipo
                                 subtipo.save()
                                 form = TiposarmasForm(request.POST)
                                 forms = SubtiposaForm()
                                 lista = RefTiposarmas.objects.all()
                                 subtiposa= RefSubtiposa.objects.all()
                                 return HttpResponseRedirect('../')
            else:
                subtiposa= RefSubtiposa.objects.get(id=idsta)
                forms= SubtiposaForm(instance=subtiposa)
                form  = TiposarmasForm()
                lista = RefSubtiposa.objects.all()
            #.values_list('descripcion', flat=False).distinct()
                return render(request,'./subtiposarms.html',{'form':form,'forms':forms,'subtiposa':subtiposa,'errors': errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se guardan los nuevos tipos de armas ingresados en subtipos
@login_required
@permission_required('user.is_staff')
def ntypesa(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    subtiposa = ""
    errors=[]
    if request.POST.get('grabar')=="Grabar":
            form = TiposarmasForm(request.POST)
            tiposa = request.POST.get('descripcion')
            if not tiposa:
                        errors.append('Ingrese tipo de armas de fuego')
                        return HttpResponseRedirect('.')
            else:
                             if not(len(tiposa)>=4 and len(tiposa)< 100):
                                 errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
                             else:
                                         if form.is_valid():
                                             form.save()
                                             forms = SubtiposaForm(request.POST)
                                             form = TiposarmasForm(request.POST)
                                             return HttpResponseRedirect('../subarms/')
                                         else:
                                             return HttpResponseRedirect('./')
            lista = RefSubtiposa.objects.all()
            combo = RefTiposarmas.objects.all()
            form = TiposarmasForm(request.POST)
            return render(request,'./subtiposarms.html',{'forms':forms,'combo':combo,'subtiposa':subtiposa,'lista':lista,'state':state, 'errors':errors,'destino': destino})
    else:
         forms = SubtiposaForm()
         form  = TiposarmasForm()
         lista = RefSubtiposa.objects.all()
         combo = RefTiposarmas.objects.all()
         return render(request,'./subtiposarms.html',{'form':form,'forms':forms,'combo':combo,'subtiposa':subtiposa,'errors': errors,'lista':lista,'state':state, 'destino': destino})

#funcion que llama al template de rubros
@login_required
@permission_required('user.is_staff')
def rubros(request):
    state = request.session.get('state')
    destino = request.session.get('destino')
    errors = []
    rubros =""
    if request.POST.get('grabar')=='Grabar':
        form = ItemForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        if not descripcion:
            errors.append('Debe Ingresar una descripcion para tipo de Rubro')
        else:
            if not(len(descripcion)>=4 and len(descripcion) < 100):
                errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
            else:
                if form.is_valid():
                    form.save()
                    lista = RefItems.objects.all()
                    return HttpResponseRedirect('.')
                else:
                    errors.append('Verifique la informacion a cargar y vuelva a intentar')
    form = ItemForm()
    lista = RefItems.objects.all()
    return render(request,'./rubros.html',{'form':form,'rubros':rubros,'errors':errors,'lista':lista,'state':state,'destino':destino})

@login_required
@permission_required('user.is_staff')
def itemselec(request,idrub):
        state = request.session.get('state')
        destino = request.session.get('destino')
        errors = []
        rubros = ""
        if request.POST.get('cancelar') == 'Cancelar':
            form = ItemForm()
            lista = RefItems.objects.all()
            return render(request,'./rubros.html',{'form':form,'rubros':rubros,'errors':errors,'lista':lista,'state':state,'destino':destino})
        else:
            if request.POST.get('modifica')  == 'Actualizar':
                rubros = RefItems.objects.get(id = idrub)
                form = ItemForm(request.POST, request.FILES)
                if form.is_valid():
                    descripcion = form.cleaned_data['descripcion']
                    rubros.descripcion = descripcion
                    rubros.save()
                    return HttpResponseRedirect('../')
                else:
                    return HttpResponseRedirect('../')
            if request.POST.get('borrar') == 'Borrar':
                try:
                    RefItems.objects.get(id = idrub).delete()
                    noborro=False
                except IntegrityError:
                    noborro=True
                form = ItemForm()
                lista = RefItems.objects.all()
                if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
                else:
                 return HttpResponseRedirect('../')
        rubros = RefItems.objects.get(id = idrub)
        form = ItemForm(instance = rubros)
        lista = RefItems.objects.all()
        return render(request,'./rubros.html',{'form':form,'rubros':rubros,'errors':errors,'lista':lista,'state':state,'destino':destino})

@login_required
@permission_required('user.is_staff')
def categorias(request):
    categorias = ""
    errors= []
    state= request.session.get('state')
    destino= request.session.get('destino')
    if request.POST.get('grabar')=="Grabar":
         formc = CategoryForm(request.POST)
         form = ItemForm(request.POST)
         descripcion = request.POST.get('descripcion')
         rubros = request.POST.get('rubro')
         if not descripcion:
                 errors.append('Ingrese categoria')
         else:
                        if not(len(descripcion)>=4 and len(descripcion)< 100):
                                         errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
                        else:
                                         if formc.is_valid():
                                             rubros = formc.cleaned_data['rubro']
                                             formc.save()
                                             return HttpResponseRedirect('.')
                                         else:
                                                errors.append('Datos inválidos. Verifique que el ingreso de datos sea correcto y completo')
         lista = RefCategory.objects.all()
         combo = RefItems.objects.all()
         form = ItemForm(request.POST)
         return render(request,'./categorias.html',{'formc':formc,'combo':combo,'categorias':categorias,'lista':lista,'state':state, 'errors':errors,'destino': destino})
    else:
         formc = CategoryForm()
         form  = ItemForm()
         lista = RefCategory.objects.all()
         combo = RefItems.objects.all()
         return render(request,'./categorias.html',{'form':form,'formc':formc,'combo':combo,'categorias':categorias,'errors': errors,'lista':lista,'state':state, 'destino': destino})

@login_required
@permission_required('user.is_staff')
def catselect(request, idcat):
    state= request.session.get('state')
    destino= request.session.get('destino')
    categorias = ""
    errors = []
    if request.POST.get('cancelar')=="Cancelar":
         formc = CategoryForm()
         form  = ItemForm()
         lista = RefCategory.objects.all()
         combo = RefItems.objects.all()
         return render(request,'./categorias.html',{'form':form,'categorias':categorias,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":
             try:
                 RefCategory.objects.get(id=idcat).delete()
                 noborro=False
             except IntegrityError:
                 noborro=True
             form = ItemForm(request.POST)
             formc = CategoryForm()
             lista = RefItems.objects.all()
             categorias= RefCategory.objects.all()
             if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
             else:
                 return HttpResponseRedirect('../')
        else:
            if request.POST.get('modifica')=='Actualizar':
                            catego = RefCategory.objects.get(id=idcat)
                            formc = CategoryForm(request.POST, request.FILES)
                            if formc.is_valid():
                                 descripcion = formc.cleaned_data['descripcion']
                                 rubro = formc.cleaned_data['rubro']
                                 catego.descripcion = descripcion
                                 catego.rubro = rubro
                                 catego.save()
                                 form = ItemForm(request.POST)
                                 formc = CategoryForm()
                                 lista = RefItems.objects.all()
                                 categorias= RefCategory.objects.all()
                                 return HttpResponseRedirect('../')
            else:
                categorias= RefCategory.objects.get(id=idcat)
                formc = CategoryForm(instance=categorias)
                form  = ItemForm()
                lista = RefCategory.objects.all()
            #.values_list('descripcion', flat=False).distinct()
                return render(request,'./categorias.html',{'form':form,'formc':formc,'categorias':categorias,'errors': errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se guardan los rubros que son ingresados en categorias
@login_required
@permission_required('user.is_staff')
def nrubros(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    categorias = ""
    errors=[]
    if request.POST.get('grabar')=="Grabar":
            form = ItemForm(request.POST)
            rubro = request.POST.get('descripcion')
            if not rubro:
                        errors.append('Ingrese rubro')
                        return HttpResponseRedirect('.')
            else:
                             if not(len(rubro)>=4 and len(rubro)< 100):
                                 errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
                             else:
                                         if form.is_valid():
                                             form.save()
                                             formc = CategoryForm(request.POST)
                                             form = ItemForm(request.POST)
                                             return HttpResponseRedirect('../category/')
                                         else:
                                             return HttpResponseRedirect('./')
            lista = RefCategory.objects.all()
            combo = RefItems.objects.all()
            form = ItemForm(request.POST)
            return render(request,'./categorias.html',{'formc':formc,'combo':combo,'categorias':categorias,'lista':lista,'state':state, 'errors':errors,'destino': destino})
    else:
         formc = CategoryForm()
         form  = ItemForm()
         lista = RefCategory.objects.all()
         combo = RefItems.objects.all()
         return render(request,'./categorias.html',{'form':form,'formc':formc,'combo':combo,'categorias':categorias,'errors': errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se carga el template de barrios y se graba
@login_required
#@permission_required('user.is_staff')
@group_required(["administrador","policia","investigaciones","radio"])
def barrios(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    barrios = ""
    if request.POST.get('grabar')=='Grabar':
        form = BarriadasForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        ciudad = request.POST.get('ciudad')
        if not descripcion:
            errors.append('Ingrese un Barrio')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<100):
                    errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
             else:
                    if ciudad == 'Seleccione ciudad':
                        errors.append('Debe seleccionar una ciudad')
                    else:


                            if form.is_valid():
                             form.save()
                             lista = RefBarrios.objects.all()
                             return HttpResponseRedirect('.')
                            else:
                             errors.append('Datos Existente. Verifique los datos ingresados.')



    form = BarriadasForm()
    lista = RefBarrios.objects.all()
    formc = ''
    return render(request,'./barrios.html',{'formc':formc,'barrios':barrios,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se carga el template de barrios y se graba
@login_required
#@permission_required('user.is_staff')
@group_required(["administrador","policia","investigaciones","radio"])
def addbarrios(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    barrios = ""
    if request.POST.get('grabar')=='Grabar':
        form = BarriadasForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        ciudad = request.POST.get('ciudad')
        if not descripcion:
            errors.append('Ingrese un Barrio')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<100):
                    errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
             else:
                    if ciudad == 'Seleccione ciudad':
                        errors.append('Debe seleccionar una ciudad')
                    else:


                            if form.is_valid():
                             form.save()
                             lista = RefBarrios.objects.all()
                             errors.append('Datos guardados')
                             return render(request, './barriosadd.html', {'errors': errors})
                            else:
                             errors.append('Datos Existente. Verifique los datos ingresados.')



    form = BarriadasForm()
    lista = RefBarrios.objects.all()
    formc = ''
    return render(request,'./barriosadd.html',{'formc':formc,'barrios':barrios,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se carga el template de ingreso de dependencias
@login_required
#@permission_required('user.is_staff')
@group_required(["administrador","policia","investigaciones","radio"])
def addcalles(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    calles = ""
    if request.POST.get('grabar')=='Grabar':
        form = AddressForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        ciudad = request.POST.get('ciudad')
        if not descripcion:
            errors.append('Ingrese el nombre de la calle')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<150):
                    errors.append('El dato ingresado debe tener entre 4 y 150 caracteres')
             else:
                    if not ciudad:
                        errors.append('Debe seleccionar una ciudad')
                    else:
                         if form.is_valid():
                                form.save()
                                lista = RefCalles.objects.all()
                                errors.append('Datos guardados')
                                return render(request, './callesadd.html', {'errors': errors})
                                #return HttpResponseRedirect('.')
                         else:
                                errors.append('Datos Existentes. Verifique los mismos')
    form = AddressForm()
    lista = RefCalles.objects.all()
    return render(request,'./callesadd.html',{'calles':calles,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})


@login_required
def obtener_calles(request,idci):
                data = request.POST
                calles = RefCalles.objects.filter(ciudad = idci)
                data = serializers.serialize("json", calles)

                return HttpResponse(data, content_type='application/json')


#la funcion en donde se actualiza y/o elimina un Barrio segun la ciudad
@login_required
#@permission_required('user.is_staff')
@group_required(["administrador","policia","investigaciones","radio"])
def nbarrios(request, idbar):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    barrios = ""
    if request.POST.get('cancelar')=='Cancelar':
        form = BarriadasForm()
        lista = RefBarrios.objects.all()
        return render(request,'./barrios.html',{'barrios':barrios,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=='Borrar':
            try:
                RefBarrios.objects.get(id=idbar).delete()
                noborro=False
            except IntegrityError:
                noborro=True
            form = BarriadasForm()
            lista = RefBarrios.objects.all()
            if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
                return HttpResponseRedirect('../')
        else:
            if request.POST.get('modifica')=='Actualizar':
                barrios = RefBarrios.objects.get(id = idbar)
                form = BarriadasForm(request.POST, request.FILES)
                descripcion = request.POST.get('descripcion')
                ciudad = request.POST.get('ciudad')
                calles = request.POST.get('calles')
                if form.is_valid():
                 descripcion = form.cleaned_data['descripcion']
                 ciudad = form.cleaned_data['ciudad']
                 barrios.descripcion = descripcion
                 barrios.ciudad = ciudad
                 barrios.save()
                 return HttpResponseRedirect('../')
                else:
                 errors.append('Datos Existente. Verifique los datos ingresados.')
            else:
                 barrios = RefBarrios.objects.get(id=idbar)
                 form=BarriadasForm(instance=barrios)
                 formc=BarriadasForm(instance=barrios)
                 calles = RefCalles.objects.filter(ciudad =barrios).values('descripcion')
                 lista = RefBarrios.objects.all()
                 return render(request,'./barrios.html',{'calles':calles,'formc':formc,'barrios':barrios,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})


    barrios = RefBarrios.objects.get(id=idbar)
    form=BarriadasForm(instance=barrios)
    formc=BarriadasForm(instance=barrios)
    calles = RefCalles.objects.filter(ciudad =barrios).values('descripcion')
    lista = RefBarrios.objects.all()
    return render(request,'./barrios.html',{'calles':calles,'formc':formc,'barrios':barrios,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se carga el template de ingreso de dependencias
@login_required
#@permission_required('user.is_staff')
@group_required(["administrador","policia","investigaciones","radio"])
def calles(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    calles = ""
    if request.POST.get('grabar')=='Grabar':
        form = AddressForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        ciudad = request.POST.get('ciudad')
        if not descripcion:
            errors.append('Ingrese el nombre de la calle')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<150):
                    errors.append('El dato ingresado debe tener entre 4 y 150 caracteres')
             else:
                    if not ciudad:
                        errors.append('Debe seleccionar una ciudad')
                    else:
                         if form.is_valid():
                                form.save()
                                lista = RefCalles.objects.all()
                                return HttpResponseRedirect('.')
                         else:
                                errors.append('Datos Existentes. Verifique los mismos')
    form = AddressForm()
    lista = RefCalles.objects.all()
    return render(request,'./calles.html',{'calles':calles,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})

#la funcion en donde se actualiza y/o elimina las dependencias segun unidad regional
@login_required
#@permission_required('user.is_staff')
@group_required(["administrador","policia","investigaciones","radio"])
def ncalles(request, idadrs):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors = []
    calles = ""
    if request.POST.get('cancelar')=='Cancelar':
        form = AddressForm(instance = calles)
        lista = RefCalles.objects.all()
        return render(request,'./calles.html',{'calles':calles,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('modifica')=='Actualizar':
            calles = RefCalles.objects.get(id = idadrs)
            form = AddressForm(request.POST, request.FILES)
            if form.is_valid():
                descripcion = form.cleaned_data['descripcion']
                ciudad = form.cleaned_data['ciudad']
                calles.descripcion = descripcion
                calles.ciudad = ciudad
                calles.save()
                return HttpResponseRedirect('../')
            else:
                         errors.append('Datos Existente. Verifique los datos ingresados.')

        if request.POST.get('borrar')=='Borrar':
            try:
                RefCalles.objects.get(id=idadrs).delete()
                noborro=False
            except IntegrityError:
                noborro=True
            form = AddressForm()
            lista = RefCalles.objects.all()
            if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
                return HttpResponseRedirect('../')
    calles = RefCalles.objects.get(id = idadrs)
    form = AddressForm(instance = calles)
    lista = RefCalles.objects.all()
    return render(request,'./calles.html',{'calles':calles,'form':form,'errors':errors,'lista':lista,'state':state, 'destino': destino})

@login_required
@permission_required('user.is_staff')
def autoridades(request):
    state       = request.session.get('state')
    destino     = request.session.get('destino')
    errors      = []
    autoridades = ""
    if request.POST.get('grabar') == 'Grabar':
        form        = AuthoritiesForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        ciudades    = request.POST.get('ciudades')
        email       = request.POST.get('email')
        if not descripcion:
            errors.append('El Campo Descripcion es Obligatorio')
        else:
            if not(len(descripcion)>=4 and len(descripcion)<80):
                errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
            else:
                if not ciudades:
                    errors.append('Debe seleccionar al menos una ciudad')
                else:
                    if not email:
                        errors.append('Debe indicar la direccion de correo electronico')
                    else:
                        if form.is_valid():
                            form.save()
                            lista = RefAutoridad.objects.all()
                            return HttpResponseRedirect('.')
                        else:
                            errors.append('Verifique la informacion a cargar y vuelva a intentar')
    form = AuthoritiesForm()
    formciud = CiudadesForm()
    lista = RefAutoridad.objects.all()
    return render(request,'./authorities.html',{'formciud':formciud,'form':form,'errors':errors,'autoridades':autoridades,'lista':lista,'state':state,'destino':destino})


@login_required
@permission_required('user.is_staff')
def autoridad(request,idaut):
    state = request.session.get('state')
    destino = request.session.get('destino')
    errors = []
    autoridades = ""

    if request.POST.get('cancelar') == 'Cancelar':
        form = AuthoritiesForm()
        lista = RefAutoridad.objects.all()
        return render(request,'./authorities.html',{'form':form,'autoridades':autoridades,'errors':errors,'lista':lista,'state':state,'destino':destino})
    else:

        if request.POST.get('modifica') == 'Actualizar':

            autoridad = RefAutoridad.objects.get(id = idaut)
            form = AuthoritiesForm(request.POST, request.FILES)

            if form.is_valid() or validateEmail(request.POST.get('email')):


                autoridad.descripcion = request.POST.get('descripcion')
                autoridad.email =request.POST.get('email')
                autoridad.ciudades =  form.cleaned_data['ciudades']

                autoridad.save()
                return HttpResponseRedirect('../')
            else:
                return HttpResponseRedirect('../')
        if request.POST.get('borrar') == 'Borrar':
            try:
                RefAutoridad.objects.get(id = idaut).delete()
                noborro=False
            except IntegrityError:
                noborro=True
            form = AuthoritiesForm()
            lista = RefAutoridad.objects.all()
            if noborro:
                return HttpResponseRedirect('/nosepudoborrarlorequerido')
            else:
                return HttpResponseRedirect('../')
    autoridades = RefAutoridad.objects.get(id = idaut)
    form = AuthoritiesForm(instance = autoridades)
    lista = RefAutoridad.objects.all()
    return render(request,'./authorities.html',{'form':form,'autoridades':autoridades,'errors':errors,'lista':lista,'state':state,'destino':destino})

def validateEmail(email):
        if len(email) > 6:
                if re.match(r'\b[\w.-]+@[\w.-]+.\w{2,4}\b', email) != None:

                        return 1
        return 0
#funcion que permite ingresar los actuantes autorizados para preventivos
@login_required
@permission_required('user.is_staff')
def actuantes(request):
    state = request.session.get('state')
    destino = request.session.get('destino')
    dni=''
    apeynom=''
    form = ActuantesForm()
    formj = JerarquiasForm()
    pers = Personas.objects.all()
    lista = Actuantes.objects.all()
    filtro=''
    formp = PersonasForm()
    ciudad = ""
    personas = ""
    errors=''
    return render(request,'./actuantes.html',{'formp':formp,'ciudad':ciudad,'personas':personas,'filtro':filtro,'pers':pers,'formj':formj,'apeynom':apeynom,'dni':dni,'form':form,'errors':errors,'lista':lista,'state':state,'destino':destino})

@login_required
@permission_required('user.is_staff')
def oficiales(request,idact):
    state = request.session.get('state')
    destino = request.session.get('destino')
    errors = []
    apeynom=''
    dni=''
    dnis=''
    botonsi=''

    filtro = Personas.objects.filter(pk=idact).values('nro_doc','apellidos','nombres')
    if request.POST.get('grabar')=='Grabar':
        form = ActuantesForm(request.POST, request.FILES)
        funcion=int(request.POST.get('funcion'))
        personas = Personas.objects.get(id=idact)
        apeynombres='%s, %s'%(personas.apellidos,personas.nombres)
        documento=personas.nro_doc
        persona_id=request.POST.get('persona_id')
        jerarquia_id=request.POST.get('jerarquia_id')
        unidadreg_id=request.POST.get('unidadreg_id')
        dependencia_id=request.POST.get('dependencia_id')
        if not jerarquia_id:
            errors.append('Seleccione Jerarquia')
        else:
            if not dependencia_id:
                 errors.append('Debe seleccionar una dependencia')
            else:

                if form.is_valid():
                    jerarquia_id = form.cleaned_data['jerarquia_id']
                    unidadreg_id = form.cleaned_data['unidadreg_id']
                    dependencia_id = form.cleaned_data['dependencia_id']
                    funcion = int(form.cleaned_data['funcion'])
                    form.jerarquia = jerarquia_id
                    form.unidadreg_id = unidadreg_id
                    form.dependencia_id = dependencia_id
                    form.funcion=funcion
                    form.apeynombres=apeynombres
                    form.documento=documento
                    form.save()
                    form = ActuantesForm()
                    formj = JerarquiasForm()
                    pers = Personas.objects.all()
                    lista = Actuantes.objects.all()
                    dnis=''
                    return render(request,'./actuantes.html',{'botonsi':botonsi,'idact':idact,'filtro':filtro,'pers':pers,'formj':formj,'apeynom':apeynom,'dni':dni,'form':form,'errors':errors,'lista':lista,'state':state,'destino':destino})
                else:
                    errors.append('Datos Existentes. Verifique los mismos')
                    form = ActuantesForm()
                    formj = JerarquiasForm()
                    pers = Personas.objects.all()
                    lista = Actuantes.objects.all()
                    return render(request,'./actuantes.html',{'botonsi':botonsi,'idact':idact,'filtro':filtro,'pers':pers,'formj':formj,'apeynom':apeynom,'dni':dni,'form':form,'errors':errors,'lista':lista,'state':state,'destino':destino})
        form = ActuantesForm()
        formj = JerarquiasForm()
        pers = Personas.objects.all()
        lista = Actuantes.objects.all()
        return render(request,'./actuantes.html',{'botonsi':botonsi,'idact':idact,'filtro':filtro,'pers':pers,'formj':formj,'apeynom':apeynom,'dni':dni,'form':form,'errors':errors,'lista':lista,'state':state,'destino':destino})

    else:
        if request.POST.get('modifica') == 'Actualizar':
            form = ActuantesForm(request.POST, instance=Actuantes.objects.get(persona_id=idact))
            if form.is_valid():
                jerarquia_id = form.cleaned_data['jerarquia_id']
                unidadreg_id = form.cleaned_data['unidadreg_id']
                dependencia_id = form.cleaned_data['dependencia_id']
                funcion = int(form.cleaned_data['funcion'])
                form.jerarquia = jerarquia_id
                form.unidadreg_id = unidadreg_id
                form.dependencia_id = dependencia_id
                form.funcion=funcion
                form.save()
                return HttpResponseRedirect('../')
            else:
                return HttpResponseRedirect('../')
        else:
         #idact es el id de personas para obtener documento, apellidos y nombres
         #filtro en personal ese id de persona para verificar si es personal policial
            if Personal.objects.filter(persona_id=idact).values('persona_id'):
             #si lo encuentra es personal policial seguidamente lo busco en actuantes
             if Actuantes.objects.filter(persona_id=idact):
                identi=Actuantes.objects.get(persona_id=idact)
                datos=ActuantesForm(instance=identi)
                dni=identi.documento
                apeynom=identi.apeynombres
                form = ActuantesForm(instance=identi)
                ure = identi.unidadreg_id
                formd = Dependencias.objects.filter(unidades_regionales_id=ure).values('id','descripcion')
                form.fields['dependencia_id'].queryset=Dependencias.objects.filter(unidades_regionales=identi.unidadreg_id)
                form.fields['dependencia_id'].initial=identi.dependencia_id
                formj = JerarquiasForm()
                pers = Personas.objects.all()
                lista = Actuantes.objects.all()
                dnis='si'
                return render(request,'./actuantes.html',{'formd':formd,'botonsi':botonsi,'dnis':dnis,'datos':datos,'idact':idact,'filtro':filtro,'pers':pers,'formj':formj,'apeynom':apeynom,'dni':dni,'form':form,'errors':errors,'lista':lista,'state':state,'destino':destino})
             else:
                perso=Personal.objects.get(persona_id=idact)
                per=perso.persona_id
                datos=PersonasForm(instance=per)
                personas = Personas.objects.get(id=idact)
                dni = personas.nro_doc
                apeynom='%s, %s'%(personas.apellidos,personas.nombres)
                form = ActuantesForm()
                formj = JerarquiasForm()
                pers = Personas.objects.all()
                lista = Actuantes.objects.all()
                dnis='si'
                botonsi='si'
                return render(request,'./actuantes.html',{'botonsi':botonsi,'dnis':dnis,'datos':datos,'idact':idact,'filtro':filtro,'pers':pers,'formj':formj,'apeynom':apeynom,'dni':dni,'form':form,'errors':errors,'lista':lista,'state':state,'destino':destino})
            else:
             errors.append('La persona seleccionada no es Empleado Policial')
             form = ActuantesForm()
             formj = JerarquiasForm()
             pers = Personas.objects.all()
             lista = Actuantes.objects.all()
             return render(request,'./actuantes.html',{'botonsi':botonsi,'idact':idact,'filtro':filtro,'pers':pers,'formj':formj,'apeynom':apeynom,'dni':dni,'form':form,'errors':errors,'lista':lista,'state':state,'destino':destino})

#abm de personas
@login_required
@transaction.atomic
@group_required(["administrador","policia","investigaciones","radio"])
def personas(request):
    state  = request.session.get('state')
    destino = request.session.get('destino')
    persona = ""
    domicilios = ""
    persona = Personas()
    ciudad = ""
    errors=[]
    lista = Personas.objects.all()
    form = PersonasForm()
    form.fields['tipo_doc'].queryset = RefTipoDocumento.objects.exclude(descripcion__icontains='NO POSEE')
    dom = DomiciliosForm()
    formpa=PadresForm()
    formpr=ProvinciasForm()

    #graba ciudad
    if request.POST.get('grabarciu')=="Grabar":
         formc = CiudadesForm(request.POST, request.FILES)
         descripcion = request.POST.get('descripcion')
         pais = request.POST.get('pais')
         #provincia = request.POST.get('provincia')
         #departamento = request.POST.get('departamento')

         if not descripcion or not pais:
                 errors.append('Ingrese una referencia(Pais) a la que pertenece la ciudad')
         else:
                         if not(len(descripcion)>=4 and len(descripcion)< 80):
                                         errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
                         else:
                                                 formc = CiudadesForm(request.POST, request.FILES)

                                                 if formc.is_valid():
                                                         formc.save()
                                                         return render(request,'./persona.html',{'domicilios':domicilios,'dom':dom,'ciudad':ciudad,'form':form,'personas':persona,'lista':lista,'state':state,'destino':destino})

                                                 else:
                                                            errors.append('La ciudad que UD. desea Guardar ya Existe. Verifique')
    #if request.user.get_all_permissions('preventivos.add_personas' or 'preventivos.change_personas'):
    if request.POST.get('grabar') == 'Grabar':
            form      = PersonasForm(request.POST, request.FILES)
            dom   = DomiciliosForm(request.POST)
            formpa = PadresForm(request.POST)

            if form.is_valid():
                persona.tipo_doc  = form.cleaned_data['tipo_doc']
                persona.nro_doc   = form.cleaned_data['nro_doc']
                persona.apellidos = form.cleaned_data['apellidos']
                persona.nombres   = form.cleaned_data['nombres']
                persona.fecha_nac = form.cleaned_data['fecha_nac']
                persona.ciudad_nac= form.cleaned_data['ciudad_nac']
                persona.pais_nac  = form.cleaned_data['pais_nac']
                persona.ciudad_res= form.cleaned_data['ciudad_res']
                persona.sexo_id   = form.cleaned_data['sexo_id']
                persona.ocupacion = form.cleaned_data['ocupacion']
                persona.cuit      = form.cleaned_data['cuit']
                persona.celular   = form.cleaned_data['celular']
                persona.alias      = form.cleaned_data['alias']
                persona.estado_civil = form.cleaned_data['estado_civil']
                
                try:
                    idpoli=persona.ocupacion
                    refpoli=RefOcupacion.objects.get(descripcion=idpoli)
                    texto=refpoli.descripcion
                except Exception as e:
                    refpoli = RefOcupacion.objects.get(descripcion='SIN DESCRIPCION')


                if dom.is_valid():
                    domicilios = Domicilios()
                    domicilios.barrio_codigo        = dom.cleaned_data['barrio_codigo']
                    domicilios.calle                = dom.cleaned_data['calle']
                    domicilios.altura               = dom.cleaned_data['altura']
                    domicilios.entre                = dom.cleaned_data['entre']
                    domicilios.fecha_desde          = dom.cleaned_data['fecha_desde']
                    domicilios.fecha_hasta          = dom.cleaned_data['fecha_hasta']
                    domicilios.fecha_actualizacion  = dom.cleaned_data['fecha_actualizacion']
                    domicilios.tipos_domicilio      = dom.cleaned_data['tipos_domicilio']
                    domicilios.ref_zona             = dom.cleaned_data['ref_zona']
                    domicilios.departamento         = dom.cleaned_data['departamento']
                    domicilios.piso                 = dom.cleaned_data['piso']
                    domicilios.lote                 = dom.cleaned_data['lote']
                    domicilios.sector               = dom.cleaned_data['sector']
                    domicilios.manzana              = dom.cleaned_data['manzana']


                    persona.save()

                    if refpoli.descripcion.find('POLICI')>=0:
                        policia=Personal()
                        policia.persona_id = persona
                        policia.credencial=0
                        policia.save()

                    domicilios.personas             = persona
                    domicilios.ref_ciudades         = form.cleaned_data['ciudad_res']
                    domicilios.save()

                    if formpa.is_valid():
                        papis=Padres()
                        papis.padre_apellidos = formpa.cleaned_data['padre_apellidos']
                        papis.padre_nombres = formpa.cleaned_data['padre_nombres']
                        papis.madre_apellidos = formpa.cleaned_data['madre_apellidos']
                        papis.madre_nombres = formpa.cleaned_data['madre_nombres']
                        papis.persona = persona
                        papis.save()
                else:
                    ciudad = ""
                    lista = Personas.objects.all()
                    return render(request,'./persona.html',{'domicilios':domicilios,'dom':dom,'ciudad':ciudad,'form':form,'personas':persona,'lista':lista,'state':state,'destino':destino})
            else:
                ciudad = ""
                lista = Personas.objects.all()
                return render(request,'./persona.html',{'domicilios':domicilios,'dom':dom,'ciudad':ciudad,'form':form,'personas':persona,'lista':lista,'state':state,'destino':destino})

    lista = Personas.objects.all()
    form = PersonasForm()
    form.fields['tipo_doc'].queryset = RefTipoDocumento.objects.exclude(descripcion__icontains='NO POSEE')
    dom = DomiciliosForm()
    formpa=PadresForm()
    formpr=ProvinciasForm()
    formcalles= AddressForm()
    formbarrios = BarriadasForm()
    formciu=RefCiudades.objects.all()
    return render(request,'./persona.html',{'formcalles':formcalles,'formbarrios':formbarrios,'formciu':formciu,'formpr':formpr,'errors':errors,'dom':dom,'ciudad':ciudad,'formpa':formpa,'form':form,'personas':personas,'lista':lista,'state':state,'destino':destino})

#abm de personas
@login_required
#@permission_required('user.is_staff')
@group_required(["administrador","policia","investigaciones","radio"])
def newperso(request):
    state       = request.session.get('state')
    destino     = request.session.get('destino')
    errors      = []
    personas = ""
    #graba ciudad
    if request.POST.get('grabarciu')=="Grabar":
         formc = CiudadesForm(request.POST, request.FILES)
         descripcion = request.POST.get('descripcion')
         pais = request.POST.get('pais')
         #provincia = request.POST.get('provincia')
         #departamento = request.POST.get('departamento')

         if not descripcion or not pais:
                 errors.append('Ingrese una referencia(Pais) a la que pertenece la ciudad')
         else:
                         if not(len(descripcion)>=4 and len(descripcion)< 80):
                                         errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
                         else:
                                                 formc = CiudadesForm(request.POST, request.FILES)

                                                 if formc.is_valid():
                                                         formc.save()
                                                         return HttpResponseRedirect('/')
                                                 else:
                                                            errors.append('La ciudad que UD. desea Guardar ya Existe. Verifique')
    if request.POST.get('grabar') == 'Grabar':
                form      = PersonasForm(request.POST, request.FILES)
                if form.is_valid():
                     form.save()
                     lista = Personas.objects.all()
                     return HttpResponseRedirect('.')
    dni=''
    apeynom=''
    form = ActuantesForm()
    formj = JerarquiasForm()
    pers = Personas.objects.all()
    lista = Actuantes.objects.all()
    filtro=''
    formp = PersonasForm()
    formpr = ProvinciasForm()

    ciudad = ""
    personas = ""
    errors=''
    return render(request,'./actuantes.html',{'formpr':formpr,'formp':formp,'ciudad':ciudad,'personas':personas,'filtro':filtro,'pers':pers,'formj':formj,'apeynom':apeynom,'dni':dni,'form':form,'errors':errors,'lista':lista,'state':state,'destino':destino})

@login_required
@transaction.atomic
@group_required(["administrador","policia","investigaciones","radio"])
#@permission_required('user.is_staff')
def persona(request, idper):
    state = request.session.get('state')
    destino = request.session.get('destino')
    errors = []
    personas =""
    siexistepoli=False
    #graba ciudad
    if request.POST.get('grabarciu')=="Grabar":
         formc = CiudadesForm(request.POST, request.FILES)
         descripcion = request.POST.get('descripcion')
         pais = request.POST.get('pais')
         #provincia = request.POST.get('provincia')
         #departamento = request.POST.get('departamento')

         if not descripcion or not pais:
                 errors.append('Ingrese una referencia(Pais) a la que pertenece la ciudad')
         else:
                         if not(len(descripcion)>=4 and len(descripcion)< 80):
                                         errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
                         else:
                                                 formc = CiudadesForm(request.POST, request.FILES)

                                                 if formc.is_valid():
                                                         formc.save()
                                                         return HttpResponseRedirect('.')
                                                 else:
                                                            errors.append('La ciudad que UD. desea Guardar ya Existe. Verifique')
    if request.POST.get('cancelar') == 'Cancelar':
        form = PersonasForm()
        form.fields['tipo_doc'].queryset = RefTipoDocumento.objects.exclude(descripcion__icontains='NO POSEE')
        lista = Personas.objects.all()
        return render(request,'./persona.html',{'form':form, 'personas':personas,'errors':errors,'lista':lista,'state':state,'destino':destino})
    else:
        if request.POST.get('modifica') == 'Actualizar':
            persona = Personas.objects.get(id=idper)
            form = PersonasForm(request.POST, request.FILES)
            dom = DomiciliosForm(request.POST)
            formpa = PadresForm(request.POST)
            #buscar en personal
            findpoli=Personal.objects.filter(persona_id=persona.id)
            if findpoli:
                 siexistepoli=True
            if len(Padres.objects.filter(persona_id=persona.id))>0:
                 papis = Padres.objects.get(persona_id=persona.id)
                 #formpa = PadresForm(instance=papis)
            else:
                 papis = Padres()
            if len(Domicilios.objects.filter(personas = idper)) > 0:
                 domicilios = Domicilios.objects.filter(personas = idper)[0]
                 iddom=domicilios.id
            else:
                    iddom='0'
                    domicilios=Domicilios()

            if form.is_valid() and persona or persona:
                    if persona and not form.is_valid():
                        persona.nro_doc = request.POST.get('nro_doc')
                    else:
                        persona.nro_doc = form.cleaned_data['nro_doc']

                    persona.apellidos = form.cleaned_data['apellidos']
                    persona.nombres = form.cleaned_data['nombres']
                    persona.tipo_doc = form.cleaned_data['tipo_doc']
                    persona.fecha_nac = form.cleaned_data['fecha_nac']
                    persona.pais_nac = form.cleaned_data['pais_nac']
                    persona.ciudad_nac = form.cleaned_data['ciudad_nac']

                    persona.ocupacion = form.cleaned_data['ocupacion']
                    persona.cuit = form.cleaned_data['cuit']
                    persona.celular = form.cleaned_data['celular']
                    persona.ciudad_res = form.cleaned_data['ciudad_res']
                    persona.sexo_id = form.cleaned_data['sexo_id']
                    persona.alias      = form.cleaned_data['alias']
                    persona.estado_civil = form.cleaned_data['estado_civil']

                    try:
                     idpoli=form.cleaned_data['ocupacion']
                     refpoli=RefOcupacion.objects.get(descripcion=idpoli)
                     texto=refpoli.descripcion
                    except Exception as e:
                     refpoli = RefOcupacion.objects.get(descripcion='SIN DESCRIPCION')

                    try:

                         persona.save()
                    except IntegrityError:
                         return HttpResponseRedirect('../../')


                    if refpoli.descripcion.find('POLICI')>=0:

                            encuentra=Personal.objects.filter(persona_id=persona)


                            if not encuentra:

                                 policia=Personal()
                                 policia.persona_id = persona
                                 policia.credencial=0
                                 policia.save()

                    else:
                         if siexistepoli:
                                #borro esa persona en personal
                                borrar=Personal.objects.get(persona_id=persona.id).delete()

                    if dom.is_valid() or iddom!='0':
                        domicilios.barrio_codigo        = dom.cleaned_data['barrio_codigo']
                        domicilios.calle                = dom.cleaned_data['calle']
                        domicilios.altura               = dom.cleaned_data['altura']
                        domicilios.entre                = dom.cleaned_data['entre']
                        domicilios.fecha_desde          = dom.cleaned_data['fecha_desde']
                        domicilios.fecha_hasta          = dom.cleaned_data['fecha_hasta']
                        domicilios.fecha_actualizacion  = dom.cleaned_data['fecha_actualizacion']
                        domicilios.tipos_domicilio      = dom.cleaned_data['tipos_domicilio']
                        domicilios.ref_zona             = dom.cleaned_data['ref_zona']
                        domicilios.departamento         = dom.cleaned_data['departamento']
                        domicilios.piso                 = dom.cleaned_data['piso']
                        domicilios.lote                 = dom.cleaned_data['lote']
                        domicilios.sector               = dom.cleaned_data['sector']
                        domicilios.manzana              = dom.cleaned_data['manzana']
                        domicilios.personas             = persona
                        domicilios.ref_ciudades         = form.cleaned_data['ciudad_res']

                        try:
                            domicilios.save()
                        except IntegrityError:
                            errors.append("Error al intentar Actualizar datos en domicilios")

                        if formpa.is_valid() or persona:
                            papis.padre_apellidos = formpa.cleaned_data['padre_apellidos']
                            papis.padre_nombres = formpa.cleaned_data['padre_nombres']
                            papis.madre_apellidos = formpa.cleaned_data['madre_apellidos']
                            papis.madre_nombres = formpa.cleaned_data['madre_nombres']
                            papis.persona = persona
                            try:
                                papis.save()
                            except IntegrityError:
                                errors.append("Error al intentar Actualizar datos en Padres")

                    return HttpResponseRedirect('../../')
            else:

                 errors.append(form.errors.as_text)

    personas = Personas.objects.get(id = idper)
    form = PersonasForm(instance=personas)
    if personas.tipo_doc.descripcion!="NO POSEE":
         form.fields['tipo_doc'].queryset = RefTipoDocumento.objects.exclude(descripcion__icontains='NO POSEE')

    form.fields['tipo_doc'].initial=personas.tipo_doc
    domicilios = Domicilios()
    dom = DomiciliosForm()

    if len(Padres.objects.filter(persona_id=personas.id))>0:
         papis = Padres.objects.get(persona_id=personas.id)
         formpa = PadresForm(instance=papis)
    else:
         formpa = PadresForm()
    if len(Domicilios.objects.filter(personas = idper)) > 0:
        domicilios = Domicilios.objects.filter(personas = idper)[0]
        dom = DomiciliosForm(instance = domicilios)

    else:
        dom.fields['barrio_codigo'].queryset = RefBarrios.objects.filter(ciudad=personas.ciudad_res)
        dom.fields['calle'].queryset = dom.fields['entre'].queryset= RefCalles.objects.filter(ciudad=personas.ciudad_res)


    formpr = ProvinciasForm()
    if personas.ciudad_res:
        form.fields['pais_res'].initial=personas.ciudad_res.pais

    ciudad = personas.ciudad_nac
    lista = Personas.objects.all()
    return render(request,'./persona.html',{'formpr':formpr,'dom':dom,'domicilios':domicilios,'ciudad':ciudad,'formpa':formpa,'form':form,'personas':personas,'errors':errors,'lista':lista,'state':state,'destino':destino})


@login_required
@permission_required('user.is_staff')
def modus(request):
    modos = ""
    errors= []
    state= request.session.get('state')
    destino= request.session.get('destino')
    if request.POST.get('grabar')=="Grabar":
         formp = RefModosHechoForm(request.POST)
         form = RefDelito(request.POST)
         descripcion = request.POST.get('descripcion')
         refdelito = request.POST.get('refdelito')
         if not descripcion:
                 errors.append('Ingrese modus_operandi')
         else:
                        if not(len(descripcion)>=4 and len(descripcion)< 80):
                                         errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
                        else:

                                         if formp.is_valid():

                                             delito=RefModosHecho()
                                             delito.descripcion=formp.cleaned_data['descripcion']
                                             delito.delito=formp.cleaned_data['delito']
                                             try:
                                                 delito.save()
                                                 return HttpResponseRedirect('.')
                                             except IntegrityError:
                                                    errors.append('Datos inválidos. Ya existe lo que intenta guardar')
                                         else:
                                                errors.append('Datos inválidos. Verifique que el ingreso de datos sea correcto y completo')
         lista = RefModosHecho.objects.all()

         combo = RefDelito.objects.all()
         form  = DelitoForm(request.POST)
         return render(request,'./modusop.html',{'formp':formp,'combo':combo,'modos':modos,'lista':lista,'state':state, 'errors':errors,'destino': destino})
    else:
         formp = RefModosHechoForm()
         form  = DelitoForm()
         lista = RefModosHecho.objects.all()
         combo = RefDelito.objects.all()

         return render(request,'./modusop.html',{'form':form,'formp':formp,'combo':combo,'modos':modos,'errors': errors,'lista':lista,'state':state, 'destino': destino})

@login_required
@permission_required('user.is_staff')
def modos(request, idmod):
    state= request.session.get('state')
    destino= request.session.get('destino')
    modos = ""
    errors = []
    if request.POST.get('cancelar')=="Cancelar":
         formp = RefModosHechoForm()
         form  = DelitoForm()
         lista = RefModosHecho.objects.all()
         combo = RefDelito.objects.all()
         return render(request,'./modusop.html',{'form':form,'modos':modos,'lista':lista,'state':state, 'destino': destino})
    else:
        if request.POST.get('borrar')=="Borrar":
             try:
                 RefModosHecho.objects.get(id=idmod).delete()
                 noborro=False
             except IntegrityError:
                 noborro=True
             form = DelitoForm(request.POST)
             formp = RefModosHechoForm()
             lista = RefModosHecho.objects.all()
             combo = RefDelito.objects.all()
             if noborro:
                 return HttpResponseRedirect('/nosepudoborrarlorequerido')
             else:
                 return HttpResponseRedirect('../')
        else:
            if request.POST.get('modifica')=='Actualizar':
                            hec = RefModosHecho.objects.get(id=idmod)
                            formp = RefModosHechoForm(request.POST, request.FILES)
                            if formp.is_valid():
                                 descripcion = formp.cleaned_data['descripcion']
                                 delito = formp.cleaned_data['delito']
                                 hec.descripcion = descripcion
                                 hec.delito = delito
                                 hec.save()
                                 form = DelitoForm(request.POST)
                                 formp = RefModosHechoForm()
                                 lista = RefModosHecho.objects.all()
                                 modos= RefDelito.objects.all()
                                 return HttpResponseRedirect('../')
            else:
                modos = RefModosHecho.objects.get(id=idmod)
                formp = RefModosHechoForm(instance=modos)
                form  = DelitoForm()
                lista = RefModosHecho.objects.all()
            #.values_list('descripcion', flat=False).distinct()
                return render(request,'./modusop.html',{'form':form,'formp':formp,'modos':modos,'errors': errors,'lista':lista,'state':state, 'destino': destino})

@login_required
@group_required(["policia","investigaciones","visita","radio","judiciales"])
def verprev(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    form=SearchPreveForm()
    anio=''
    nro=''
    fecha_carga=''
    fecha_cargah=''
    caratula=""
    todos=[]
    total='no'
    errors=[]
    unidadregi=''
    jurisdi=''
    peticion        = request.GET.get('peticion')
    if peticion != None and peticion != "":
        nro             = request.GET.get('nro')
        anio            = request.GET.get('anio')
        caratula        = request.GET.get('caratula')
        fecha_carga     = request.GET.get('fecha_carga')
        fecha_cargah    = request.GET.get('fecha_cargah')
        ureg            = request.GET.get('ureg')
        depe            = request.GET.get('depe')

        unidadregi=Dependencias.objects.get(descripcion__contains=request.user.userprofile.depe.descripcion)
        jurisdi=unidadregi.ciudad.descripcion
        todos = Preventivos.objects.all()
        if depe != "" and depe != None:
         todos = todos.filter(dependencia = depe)
        if ureg != "" and ureg != None:
         dependencias = Dependencias.objects.filter(unidades_regionales= ureg)
         todos = todos.filter(dependencia__in = dependencias)
        if nro != "" and nro != None:
         todos = todos.filter(nro = nro)
        if anio != "" and anio != None:
         todos = todos.filter(anio = anio)
        if caratula != "" and caratula != None:
         todos = todos.filter(caratula__icontains = caratula)
        if fecha_carga != "" and fecha_carga != None:
            if fecha_cargah == "" or fecha_cargah == None:
                fecha_cargah =  datetime.datetime.now()
            else:
                fecha_cargah = datetime.datetime.strptime(fecha_cargah+' 00:00:00',"%d/%m/%Y %H:%M:%S")
            fecha_carga = datetime.datetime.strptime(fecha_carga+' 00:00:00',"%d/%m/%Y %H:%M:%S")
            todos = todos.filter(fecha_carga__range = [fecha_carga,fecha_cargah])
        if peticion == 'buscar':
            return render(request,"./listarprev.html",{'todos':todos,'jurisdi':jurisdi})
        else:
            return exportar_listado(request,todos)



    info={'nro':nro,'anio':anio,'fecha_carga':fecha_carga,'fecha_cargah':fecha_cargah,
    'caratula':caratula,'todos':todos,'total':total,'errors':errors,'unidadregi':unidadregi,'jurisdi':jurisdi,
    'state':state,'destino': destino,'form':form}
    return render(request,'./seeprev.html',info)



def exportar_listado(request,todos):
    filadata={}
    filahecho={}
    filadelitos={}
    filalugar={}
    filaperin={}
    filaele={}
    denuncia=''
    allgral=[]
    filagral={}
    ii=4
    fila2=[]
    fila3=[]
    iil=[]
    for datas in todos:

       #if datas.sendwebservice==1:
         preventivo = Preventivos.objects.get(id = datas.id)                    #obtiene el preventivo
         ciudad= preventivo.dependencia.ciudad                                  #obtiene la ciudad
         depe=preventivo.dependencia                                            #obtiene la dependencia
         unireg=depe.unidades_regionales.descripcion                            #obtiene la unidad regional
         filadata={                                                             #prepara los datos de la fila
                'Preventivo Nro.':str(datas.nro)+'/'+str(datas.anio),
                'Unidad Regional':str(datas.dependencia.unidades_regionales),
                'Dependencia':str(datas.dependencia),
                'Localidad':str(ciudad),
                'Caratula':str(datas.caratula.encode("utf-8")),
                'Fecha Denuncia':str(timezone.localtime(datas.fecha_denuncia).strftime("%d/%m/%Y %H:%M:%S"))}
         filagral=filadata                                                      #asigna la fila creada a fila gral
         dhecho=Preventivos.objects.get(id=datas.id).hecho                  #obtiene los hechos relacionados al preventivo
         if dhecho:
          if dhecho.motivo.descripcion==None or dhecho.motivo.descripcion=='': #si no tiene cargado el motivo de denuncia
            motivo='SIN MOTIVO'                                             #le asigna un sin motivo
          else:                                                              # si lo tiene cargado
            motivo=str(dhecho.motivo.descripcion)                           # lo asigna en motivo
          filahecho={                                                        #prepara los datos de la fila hecho
            'Motivo Preventivo':motivo,
            #'Denuncia':denuncia,
            "Dia Hecho":str(timezone.localtime(dhecho.fecha_desde).strftime("%A")),
            "Fecha Desde":str(timezone.localtime(dhecho.fecha_desde).strftime("%d/%m/%Y %H:%M:%S")),
            'Fecha Hasta':str(timezone.localtime(dhecho.fecha_hasta).strftime("%d/%m/%Y %H:%M:%S"))
            }
          filagral.update(filahecho)                                         #actualiza fila gral con fila hecho
          idhec=dhecho.id                                                    #obtiene el id del hecho
          delito =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True) #obtiene los delitos del hecho
          cometidos=[]                                                       #crea un arreglo de cometidos
          hechodeli=""                                                       #crea una variable hechodeli
          for d in delito:                                                   #por cada delito
            cometidos.append(d)                                            # lo agrega a cometidos

          hechodeli=''                                                       #limpia hecho deli
          modus=' SIN MODALIDAD  '                                           #asigna sin modalidad al modus
          for i in cometidos:                                                #para cada delito
            if hechodeli != '':                                            #si hecho deli no esta vacio
                hechodeli=hechodeli+'-'+unicode(str(i).strip(),'UTF-8')    # le agrega el delito separado por guion
            else:                                                          #si esta vacio
                hechodeli=hechodeli+unicode(str(i).strip(),'UTF-8')        #agrega el delito

            if i.refmodoshecho!=None:                                      #si el delito tiene modo
                modus=modus+unicode(str(i.refmodoshecho),'UTF-8')+' - '    # se lo agrega

            filadelitos={                                                  #prepara la filadelitos
                    'Delitos':hechodeli.strip()+modus
                    }

          filagral.update(filadelitos)                                       #a la fila gral le agrega filadelitos

          involuscra=[]                                                         #crea un arreglo de involurados
          eleminvo=[]                                                           #crea un arreglo de elementos
          datosper=""                                                           #crea una variable datosper
          elementos=""                                                          #crea una variable elementos
          involus=Hechos.objects.get(id=idhec).involu.all()                     #obtiene todos los involucrados
          eleinvo=Elementos.objects.filter(hechos=idhec,ampliacion_id__isnull=True,borrado__isnull=True).all() #obtiene todos los elementos involucrados
         else:                                                                      #si no tiene hecho relacionado
             motivo='SIN MOTIVO'                                                # motivo sin motivo
             denuncia='SIN DESCRIPCION'                                         # denuncia sin descripcion
             delis='SIN DELITOS'                                                #delis sin delitos
             moduss='SIN MODO OPERANDI'                                         #modus sin modo
             fechadesde='SIN FECHA'                                             #fechas sin fecha
             fechahasta="SIN FECHA"
             filahecho={                                                        #prepara fila hecho
                'Motivo Preventivo':motivo,
                'Denuncia':denuncia,
                "Dia Hecho":'SIN DESCRIPCION',
                'Fecha Desde':fechadesde,
                'Fecha Hasta':fechahasta
                }
             filagral.update(filahecho)                                         #agrega fila hecho a fila gral

             filadelitos={                                                      #prepara filadelitos
                    'Delitos':delis+moduss
                    }

             filagral.update(filadelitos)                                       #agrega fila delitos a fila gral



         datosgral=""                                                           #datos gral vacio
         lugar=''                                                               # lugar vacio
         lati=''                                                                #latitud vacio
         longi=''                                                               #longitud vacio
         condiciones=''                                                         #condiciones vacio
         perjuridica=''                                                         #persona juridica vacio
         involucrados=0                                                         #involucrados 0
         if len(Lugar.objects.filter(hecho=idhec))>0:                           #si el hecho tiene lugar
             tienelugar=True                                                    #levanta la bandera tiene lugar
             idlugar = Hechos.objects.get(id=idhec).lugar_hecho        #obtiene el id del lugar
             lugar=Hechos.objects.get(id=idhec).lugar_hecho            #obtiene el lugar
             if idlugar.altura==None:                                           #si no tiene altura el lugar
                idlugar.altura=0                                                #lo pone en 0

             lugarhecho={                                                       #prepara el lugar
                    'Lugar':str(idlugar.tipo_lugar),
                    "Zona":str(idlugar.zona),
                    'LugarHecho':idlugar.calle.descripcion+' Nro.: '+str(idlugar.altura)
                    }
             if idlugar.barrio==None:                                           #si no tiene barrio
                lugarbarrio='SIN DESCRIPCION'                                   #le pone sin descripcion
             else:                                                              #si tiene barrio
                 lugarbarrio=unicode(str(idlugar.barrio),'UTF-8')               #lo obtiene formateado a utf-8

             barrio={                                                           #prepara el barrio
                'BarrioHecho':lugarbarrio
                }
             #condiciones= lugar.cond_climaticas.values_list('descripcion',flat=True)
             laticiudad = RefCiudades.objects.get(id=datas.dependencia.ciudad_id) #obtiene la ciudad

             lati=laticiudad.lat                                                #obtiene la latitud
             longi=laticiudad.longi                                             #obtiene la longitud
             filacoorde={                                                       #prepara la fila de coordenadas
                    'Latitud':str(lati),
                    'Longitud':str(longi)
                    }


         filagral.update(lugarhecho)                                            #agrega lugarhecho a fila gral
         filagral.update(barrio)                                                #agrega barrio a fila gral
         filagral.update(filacoorde)                                            #agrega filacoorde a fila gral

         if len(Hechos.objects.get(id=idhec).involu.all())>0:                   #si tiene involucrados
              tienePersonas=True                                                # levanta la bandera tienePersonas
              countinvolus=Hechos.objects.get(id=idhec).involu.all().count()    #obtiene la cantidad de personas involucradas
              rolins=''                                                         #rolins vacio
              for p in Hechos.objects.get(id=idhec).involu.all():               #para cada involucrado
                 bandera=True                                                   #bandera True
                 tienepersona=True                                              #tienePersonas True
                 rolins=rolins+p.roles.descripcion+'-'                          #suma el rol a rolins
                 #involucrados=involucrados+1
                 perinvocondi={                                                 #prepara condicion del involucrado
                        'Franganti':p.infraganti.upper(),
                        'Tentativa':p.tentativa.upper()
                        }
                 filagral.update(perinvocondi)                                  #actualiza fila gral con perinvcondi

              perinvo={                                                         #prepara perinvo
                'Involucrados':rolins
                }
              filagral.update(perinvo)                                          #agrega perinvo filagral
         else:                                                                  # si no tiene personas
             sinrol='SIN DESCRIPCION'                                           #sin rol sin descripcion
             fraganti='NO'                                                      #fraganti no
             tentativa='NO'                                                     #tentativa no
             perinvo= {                                                         #prepara perinvo
                    'Involucrados':sinrol,
                    'Franganti':fraganti,
                    'Tentativa':tentativa
                    }
             perinvo.update(perinvo)                                            #agreaga perinvo a perinvo
             filagral.update(perinvo)                                           #agrega filagral a perinvo

         rotulo='CantidadElementos '+str(len(eleinvo))                          #rotulo = a la cantidad de elementos
         datosgral=''                                                           #datos gral vacio
         obdata=[]                                                              #crea un arreglo obdata
         obdatav=[]                                                             #crea un arreglo obdatav
         deta=''                                                                #deta vacio
         detav=''                                                               #detav vacio
         eleme=''                                                               #eleme vacio
         hay=[]                                                                 #crea un arreglo hay
         i=1                                                                    #crea un contador en 1
         obse=''                                                                #obse vacio
         for eles in eleinvo:                                                   #para cada elemento en eleinvo
                 tieneelementos=True                                            #tieneelementos true
                 obdata=[]
                 obdatav=[]
                 deta=''
                 detav=''

                 if len(Elementosarmas.objects.filter(idelemento=eles.id))>0:   #si tiene armas

                      idar = Elementosarmas.objects.filter(idelemento=eles.id).values('idarma') #obtiene el id
                      tieneelementos=True
                      obdata=Armas.objects.filter(id=idar)                      #obtiene el arma
                      for extra in obdata:                                      #por cada arma
                         titu=' Carateristicas Generales : '                    #obtiene sus caracteristicas
                         tabla=str(extra.subtipos)+'- Tipo/s : \
                                '+str(extra.tipos)+'- Sistema de Disparo : \
                                '+str(extra.sistema_disparo)+'- Marcas : \
                                '+str(extra.marcas)
                         tipos='Calibre : '+str(extra.calibre)+'- Modelo : \
                                '+str(extra.modelo)+'- Nro Serie : \
                                '+str(extra.nro_arma)+'- Propietario : \
                                '+str(extra.nro_doc)+'-'+str(extra.propietario)
                         deta=titu+tabla+tipos                                  #en deta agrega todos los detalles

                 if len(Elementoscars.objects.filter(idelemento=eles.id))>0:    #si tiene vehiculos
                      tieneelementos=True
                      idarv = Elementoscars.objects.filter(idelemento=eles.id).values('idvehiculo') #obtiene el id del vehiculo

                      obdatav=Vehiculos.objects.filter(id=idarv)                #obtiene el vehiculo
                      for extrav in obdatav:                                    #para cada vehiculo
                         tituv=' Carateristicas Generales : '                   #obtiene sus caracteristicas
                         tablav=' Marca/s : '+str(extrav.idmarca)+'- Modelo : '+str(extrav.modelo)+'- Dominio : '+str(extrav.dominio)+'- Año : '+str(extrav.anio)
                         tiposv=' Tipo/s : '+str(extrav.tipov)+'- Nro Chasis : '+str(extrav.nchasis)+'- Nro. Motor : '+str(extrav.nmotor)+'- Propietario : '+str(extrav.nro_doc)+' - '+str(extrav.propietario)
                         detav=tituv+tablav+tiposv                              #agrega las caracteriticas a detav



                 tipo=str(eles.tipo)                                            #obtiene el tipo elemento
                 ampli=''
                 rubro='Elemento/s '+str(eles.tipo)                             #obtiene el rubro
                 rubros='Rubro y Categoria : '+str(eles.rubro)+'-'+str(eles.categoria)  #obtiene el rubro y categoria
                 canti='Cantidad : '+str(eles.cantidad)+'-'+str(eles.unidadmed)     #obtiene la cantidad de elemento y unidad medida
                 if eles.descripcion.strip():
                     obse='Observaciones : '+str(eles.descripcion.encode("utf8")) #codifica a utf-8 la descripcion
                     obse=obse.replace('&NBSP;','')

                 if deta:
                      if detav:
                          eleme=str(i)+'°-'+rubro+' '+rubros+' '+canti+' '+obse+' '+detav+' | '
                      else:
                          eleme=str(i)+'°-'+rubro+' '+rubros+' '+canti+' '+obse+' '+deta+' | '

                 else:
                      if detav:
                          eleme=str(i)+'°-'+rubro+' '+rubros+' '+canti+' '+obse+' '+detav+' | '
                      else:
                          eleme=str(i)+'°-'+rubro+' '+rubros+' '+canti+' '+obse+' '+deta+' | '

                 eleminvo.append(eleme)
                 i=i+1
         elementos=''
         for ja in eleminvo:
             elementos=elementos+ja

         if elementos:
            elementos={'Elementos':elementos}
         else:
            contiene='SIN ELEMENTOS'
            elementos={'Elementos':contiene}

         filagral.update(elementos)                                             #agrega los elementos a fila gral
         fila4=[]                                                               #crea un arreglo fila4
         for key,value in sorted(filagral.items()):                             #para cada elemento en filagral
             fila4.append(value)                                                #agrega el valor en fila4

         f=fila4
         fila3.append(f)                                                        #agrega la fila en fila3
         ii+=1                                                                  #suma uno a ii que inicia en 4
         iil.append(ii)                                                         #agrega ii al arreglo iil


    libro = Workbook()                                                          #crea un nuevo libro
    hoja = libro.get_active_sheet()                                             #crea una nueva hoja
    hoja.title = "Informe Gral"                                                 #le pone titulo a la hoja

    # Ahora, se obtiene las celdas en la cuál se colocará el nombre
    # del campo. como son 8 campos, se necesita 8 celdas
    celda = hoja.cell("B1")
    celda.value=" Elementos recopilados desde la Base de Datos para Estadísticas "
    celda = hoja.cell("B3")
    celda.value=" Tabla con datos obtenidos de Preventivos Enviados e Informados"
    rango_celdas = hoja.range("B4:V4")
    # se crea una tupla con los nombres de los campos
    nombre_campos = "BARRIOHECHO","CARATULA","DELITOS","DEPENDENCIA","DIA DEL HECHO","ELEMENTOS","FECHA DENUNCIA","FECHA DESDE","FECHA HASTA","FRAGANTI","INVOLUCRADOS","LATITUD","LOCALIDAD","LONGITUD","LUGAR","LUGARHECHO","MOTIVO","PREVENTIVO NRO","TENTATIVA","UNIDAD REGIONAL","ZONA"
    # ahora, se asigna cada nombre de campo a cada celda
    for campo in rango_celdas:
         indice = 0  # se crea un contador para acceder a la tupla
         for celda in campo:
                 celda.value = nombre_campos[indice]
                 indice += 1
    longitud=ii
    celdas_datos = hoja.range("B5:V{0}".format(longitud))

    # ahora vamos a dar los valores a las celdas con los datos

    fila=0
    for valuex in fila3:
         
         indice1=0
         for celda in celdas_datos[fila]:
             celda.value = valuex[indice1]
             indice1+=1
         fila += 1



    response = HttpResponse(content_type="application/ms-excel")  # HttpResponse viene del modulo django.http
    nombre_archivo = "informe.xlsx"
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    response["Content-type"] = "application/download"
    libro.save(response)
    return response

@login_required
@group_required(["policia","investigaciones","visita","radio"])
def verhechos(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    form=SearchPreveForm()
    fecha_carga=''
    fecha_cargah=''
    hechos=""
    todosa=[]
    todosna=[]
    total='no'
    errors=[]
    sonauti=0
    nosonauti=0
    auti=''
    todos=[]
    autora=[]
    tipodelito=''
    delito=''
    unidadregi=Dependencias.objects.get(descripcion__contains=request.user.userprofile.depe.descripcion)
    jurisdi=unidadregi.ciudad.descripcion
    if request.POST.get('ver')=='1':
         sonauti=request.POST.get('sonauti')
         nosonauti=request.POST.get('nosonauti')
         hechos=request.POST.get('hechos')
         auti='si'
         if hechos:

             query_string=hechos
             entry_query = get_query(query_string, ['descripcion',])

             filtro=Hechos.objects.all().filter(entry_query).order_by('fecha_carga',)

         if filtro:
                for regs in filtro:
                    autorizados=Preventivos.objects.all().filter(id=regs.preventivo_id,fecha_autorizacion__isnull=False)

                    if autorizados:
                        for hay in autorizados:

                             filtro=Hechos.objects.all().filter(preventivo_id=hay).order_by('fecha_carga',)

                             if filtro not in todos:
                                    todos.append(filtro)
    else:
        if request.POST.get('ver')=='2':
             sonauti=request.POST.get('sonauti')
             nosonauti=request.POST.get('nosonauti')
             hechos=request.POST.get('hechos')
             auti='si'
             if hechos:

                     query_string=hechos
                     entry_query = get_query(query_string, ['descripcion',])

                     filtro=Hechos.objects.all().filter(entry_query).order_by('fecha_carga',)

             if filtro:
                for regs in filtro:
                    autorizados=Preventivos.objects.all().filter(id=regs.preventivo_id,fecha_autorizacion__isnull=True)

                    if autorizados:
                        for hay in autorizados:

                             filtro=Hechos.objects.all().filter(preventivo_id=hay).order_by('fecha_carga',)

                             if filtro not in todos:
                                    todos.append(filtro)

        else:
             if request.POST.get('search')=="Buscar":
                     form=SearchPreveForm(request.POST, request.FILES)
                     hechos=request.POST.get('hechos')

                     if hechos:

                            query_string=hechos
                            entry_query = get_query(query_string, ['descripcion',])

                     filtro=Hechos.objects.all().filter(entry_query).order_by('fecha_carga',)
                     if filtro:
                        for regs in filtro:
                            autorizados=Preventivos.objects.all().filter(id=regs.preventivo_id,fecha_autorizacion__isnull=False)
                            if autorizados:
                                 sonauti+=1
                            else:
                                 nosonauti+=1
                     else:
                         sonauti=0
                         nosonauti=0


                     if filtro not in todosa:
                             todosa.append(filtro)


                     else:
                            errors.append('Por algun error y/o consulta comuniquese con el Administrador. Gracias.-')



             else:
                         todos=''
                         total='si'


    info={'hechos':hechos,'total':total,'errors':errors,'sonauti':sonauti,'nosonauti':nosonauti,'auti':auti,'autora':autora,
    'state':state,'destino': destino,'form':form,'todos':todos,'tipodelito':tipodelito,'delito':delito,'jurisdi':jurisdi}

    return render(request,'./seehec.html',info)


@login_required
@group_required(["policia","investigaciones","visita","radio"])
def verdelitos(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    form=SearchPreveForm()
    fecha_carga=''
    fecha_cargah=''
    hechos=""
    todosa=[]
    todosna=[]
    total='no'
    errors=[]
    sonauti=0
    nosonauti=0
    auti=''
    todos=[]
    autora=[]
    tipodelito=''
    delito=''
    unidadregi=Dependencias.objects.get(descripcion__contains=request.user.userprofile.depe.descripcion)
    jurisdi=unidadregi.ciudad.descripcion
    if request.POST.get('ver')=='1':
         auti='si'

         filtro= HechosDelito.objects.all().filter(refdelito_id=request.POST.get('delito'))
         for regs in filtro:
                 filtroh=Hechos.objects.all().filter(id=regs.hechos_id).values('preventivo_id')
                 autorizados=Preventivos.objects.all().filter(id=filtroh,fecha_autorizacion__isnull=False)
                 if autorizados:
                        filtro=Hechos.objects.all().filter(preventivo_id=autorizados).order_by('fecha_carga',)
                        if filtro not in todos:
                            todos.append(filtro)



         sonauti=request.POST.get('sonauti')
         nosonauti=request.POST.get('nosonauti')
         delito=request.POST.get('delito')
         tipodelito=request.POST.get('tipodelito')
    else:
        if request.POST.get('ver')=='2':
                 auti='si'

                 filtro= HechosDelito.objects.all().filter(refdelito_id=request.POST.get('delito'))
                 for regs in filtro:
                         filtroh=Hechos.objects.all().filter(id=regs.hechos_id).values('preventivo_id')
                         autorizados=Preventivos.objects.all().filter(id=filtroh,fecha_autorizacion__isnull=True)
                         if autorizados:
                                filtro=Hechos.objects.all().filter(preventivo_id=autorizados).order_by('fecha_carga',)
                                if filtro not in todos:
                                    todos.append(filtro)



                 sonauti=request.POST.get('sonauti')
                 nosonauti=request.POST.get('nosonauti')
                 delito=request.POST.get('delito')
                 tipodelito=request.POST.get('tipodelito')

        else:
             if request.POST.get('search')=="Buscar":
                     form=SearchPreveForm(request.POST, request.FILES)
                     tipodelito=request.POST.get('tipodel')
                     delito=request.POST.get('delitoe')

                     if tipodelito:
                            if delito:
                                 filtrod=HechosDelito.objects.all().filter(refdelito_id=delito)
                            else:
                                 filtrod=''
                                 errors.append('Seleccione Delitos')
                            for deli in filtrod:
                                    filtroh=Hechos.objects.all().filter(id=deli.hechos_id).values('preventivo_id')
                                    autorizados=Preventivos.objects.all().filter(id=filtroh,fecha_autorizacion__isnull=False)
                                    noautorizados=Preventivos.objects.all().filter(id=filtroh,fecha_autorizacion__isnull=True)
                                    sonauti+=len(autorizados)
                                    nosonauti+=len(noautorizados)
                     else:
                            errors.append('Seleccione Tipos Delitos')



             else:
                         todos=''
                         total='si'

    ftiposdelitos=DelitoForm()
    info={'hechos':hechos,'total':total,'errors':errors,'sonauti':sonauti,'nosonauti':nosonauti,'auti':auti,'autora':autora,'jurisdi':jurisdi,
    'state':state,'destino': destino,'form':form,'todos':todos,'ftiposdelitos':ftiposdelitos,'tipodelito':tipodelito,'delito':delito,}

    return render(request,'./seedelitos.html',info)

@login_required
@group_required(["policia","investigaciones","radio"])
def pdfs(request,idprev):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors=[]
    continua=''
    tienehecho=False
    tieneper=False
    tienelugar=False
    tieneelementos=False
    elementos=[]
    i=0
    grabo='fin'
    preventivo = Preventivos.objects.get(id = idprev)
    ciudad= preventivo.dependencia.ciudad
    depe=preventivo.dependencia
    unireg=depe.unidades_regionales.descripcion
    image=""
    countinvolus=0
    #Datos del Hecho delicitivo atraves del nro de preventivo
    if len(Hechos.objects.filter(preventivo=idprev))>0:
            hecho = Hechos.objects.get(preventivo=idprev)
            tienehecho=True
            form=HechosForm(instance=hecho)
            ftiposdelitos=DelitoForm()
            motivo=request.POST.get('motivo')
            modos=RefModosHechoForm(instance=hecho)
            descripcion=hecho.descripcion
            idhec=hecho.id
            delito =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)

            cometidos=[]
            hechodeli=""
            for d in delito:
                    cometidos.append(d)

            for i in cometidos:

                     hechodeli=hechodeli+'-'+i.refdelito.descripcion
            #Datos de las Personas involucradas en el hecho
            involuscra=[]
            eleminvo=[]
            datosper=""
            elementos=""

            involus=Hechos.objects.get(id=hecho.id).involu.all()
            eleinvo=Elementos.objects.filter(hechos=hecho.id,ampliacion_id__isnull=True,borrado__isnull=True).all()

            datosgral=""
            lugar=''
            lati=''
            longi=''
            condiciones=''
            perjuridica=''
            if len(Lugar.objects.filter(hecho=idhec))>0:
                tienelugar=True
                lugar = Hechos.objects.get(id=idhec).lugar_hecho
                condiciones= lugar.cond_climaticas.values_list('descripcion',flat=True)
                laticiudad = RefCiudades.objects.get(id=preventivo.dependencia.ciudad_id)
                lati=laticiudad.lat
                longi=laticiudad.longi

            if len(Hechos.objects.get(id=hecho.id).involu.all())>0:
                 tienePersonas=True
                 countinvolus=Hechos.objects.get(id=hecho.id).involu.all().count()
                 for p in Hechos.objects.get(id=hecho.id).involu.all():

                     bandera,personai = funverifica(p.persona.id)
                     if p.menor=='':
                        p.menor="NO"
                     if p.juridica=='si':
                        if p.razon_social!=None:
                          perjuridica=str(p.razon_social)

                        if RefTipoDocumento.objects.get(id=p.cuit_id)!='Null':
                          perjuridica=perjuridica+'-'+str(RefTipoDocumento.objects.get(id=p.cuit_id))

                        if p.nrocuit!=0:
                          perjuridica=perjuridica+'-'+str(p.nrocuit)
                     domi=Personas.objects.get(id=p.persona.id).persodom.all()
                     if domi:
                        for l in Personas.objects.get(id=p.persona.id).persodom.all():
                         #datosgral=str(p.roles)+' - '+str(p)+' '+str(p.persona.tipo_doc)+' :'+str(p.persona.nro_doc)
                         dad=Personas.objects.get(id=p.persona.id).padre.all()

                         if dad:

                                for la in Personas.objects.get(id=p.persona.id).padre.all():



                                    roles='<u>'+str(p.roles)+'</u>'+' : '
                                    if bandera:
                                        if p.juridica=='si':
                                            persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')+('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')
                                        else:
                                            persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')
                                    else:
                                        if p.juridica=='si':
                                            persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+'</dd>')+('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')+('<dd>'+'Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')
                                        else:
                                            persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')

                                    domi='<dd>Reside en : '+str(p.persona.ciudad_res)+',  Domicilio : '+l.calle.descripcion+'  Nro.: '+str(l.altura)+'</dd>'
                                    if la.padre_apellidos or la.padre_nombres or la.madre_apellidos or la.madre_nombres:
                                         padys='<dd>Hijo de : '+str(la.padre_apellidos.encode("utf8"))+', '+str(la.padre_nombres.encode("utf8"))+' y de : '+str(la.madre_apellidos.encode("utf8"))+', '+str(la.madre_nombres.encode("utf8"))+'<br><br></dd>'
                                    else:
                                         padys='<dd>no registra datos de los padres'+'<br></dd>'
                                    datosgral=roles+persona+domi+str(padys)
                         else:
                                 roles='<u>'+str(p.roles)+'</u>'+' : '

                                 if bandera:
                                        if p.juridica=='si':
                                            persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')+('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')
                                        else:
                                            persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')
                                 else:
                                        if p.juridica=='si':
                                            persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+'</dd>')+('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')+('<dd>'+'Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')
                                        else:
                                            persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')

                                 #persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+',  Estado Civil : '+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')
                                 domi='<dd>Reside en : '+str(p.persona.ciudad_res)+',  Domicilio : '+str(l.calle)+'  Nro.: '+str(l.altura)+'</dd>'
                                 padys='<dd>no registra datos de los padres'+'<br></dd>'
                                 datosgral=roles+persona+domi+str(padys)
                         involuscra.append(datosgral)
                     else:

                        roles='<u>'+str(p.roles)+'</u>'+' : '
                        if bandera:
                            if p.juridica=='si':
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')+('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')
                            else:
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')
                        else:
                            if p.juridica=='si':
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+'</dd>')+('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')+('<dd>'+'Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')
                            else:
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')

                        #persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+',  Estado Civil : '+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')
                        domi='<dd>no registra domicilio'+'</dd>'
                        padys='<dd>no registra datos de los padres'+'<br></dd>'
                        datosgral=roles+persona+domi+padys

                        involuscra.append(datosgral)
                 for i in involuscra:
                     datosper=datosper+i

                    #datosper.append(persona)
                 datosgral=''
                 obdata=[]
                 obdatav=[]
                 deta=''
                 detav=''
                 eleme=''
                 hay=[]
                 i=1

                 for eles in eleinvo:
                        tieneelementos=True
                        obdata=[]
                        obdatav=[]
                        deta=''
                        detav=''

                        if len(Elementosarmas.objects.filter(idelemento=eles.id))>0:

                             idar = Elementosarmas.objects.filter(idelemento=eles.id).values('idarma')
                             tieneelementos=True
                             obdata=Armas.objects.filter(id=idar)
                             for extra in obdata:
                                titu='<ul><li><u> Carateristicas Generales : </u></li></ul>'
                                tabla='<dd> '+str(extra.subtipos)+'  ---  Tipo/s : '+str(extra.tipos)+'  ---  Sistema de Disparo : '+str(extra.sistema_disparo)+'   --- Marcas : '+str(extra.marcas)+'</dd>'
                                tipos='<dd> Calibre : '+str(extra.calibre)+'  --- Modelo : '+str(extra.modelo)+'  --- Nro Serie : '+str(extra.nro_arma)+'   ---  Propietario : '+str(extra.nro_doc)+' - '+str(extra.propietario)+'</dd>'
                                deta=titu+tabla+tipos

                        if len(Elementoscars.objects.filter(idelemento=eles.id))>0:
                             tieneelementos=True
                             idarv = Elementoscars.objects.filter(idelemento=eles.id).values('idvehiculo')

                             obdatav=Vehiculos.objects.filter(id=idarv)
                             for extrav in obdatav:
                                tituv='<ul><li><u> Carateristicas Generales : </u></li></ul>'
                                tablav='<dd> Marca/s : '+str(extrav.idmarca)+'  ---   Modelo : '+str(extrav.modelo)+'  ---  Dominio : '+str(extrav.dominio)+'   ---  Año : '+str(extrav.anio)+'</dd>'
                                tiposv='<dd> Tipo/s : '+str(extrav.tipov)+' ---  Nro Chasis : '+str(extrav.nchasis)+' ---  Nro. Motor : '+str(extrav.nmotor)+'</dd>'+'<dd> Propietario : '+str(extrav.nro_doc)+' - '+str(extrav.propietario)+'</dd>'
                                detav=tituv+tablav+tiposv



                        tipo='<br><dd><u>'+str(eles.tipo)+'</u></dd>'
                        ampli=''
                        rubro=' Elemento/s '+str(eles.tipo)
                        rubros='Rubro y Categoria :'+str(eles.rubro)+' --- '+str(eles.categoria)
                        canti=' Cantidad : '+str(eles.cantidad)+' --- '+str(eles.unidadmed)
                        obse=' Observaciones : '+str(eles.descripcion.encode("utf8"))
                        obse=obse.replace('&NBSP;','')
                        if deta:
                             if detav:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+detav+'<br>'
                             else:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+deta+'<br>'

                        else:
                             if detav:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+detav+'<br>'
                             else:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+deta+'<br>'

                        eleminvo.append(eleme)
                        i=i+1

                 for ja in eleminvo:
                        elementos=elementos+ja


            #datos del preventivos
            datos=Preventivos.objects.get(id=idprev)
            nro=datos.nro
            anio=datos.anio
            fecha_denuncia=datos.fecha_denuncia
            fecha_carga=datos.fecha_carga
            fecha_cierre=datos.fecha_cierre
            caratula=datos.caratula
            actuante=datos.actuante
            preventor=datos.preventor
            autoridades= datos.autoridades.values_list('descripcion',flat=True)
            dependencia=datos.dependencia.descripcion
            unidadreg=datos.dependencia.unidades_regionales.descripcion
            idprev=idprev
            #envio de datos al template updatehechos.html
            jerarqui_a=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=actuante).values('jerarquia_id'))
            jerarqui_p=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=preventor).values('jerarquia_id'))
            form1=Hechos.objects.filter(preventivo=idprev)
            today = datetime.datetime.now()
            info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
                     'caratula':caratula,'idhec':idhec,'involus':involus,'involuscra':involuscra,'datosper':datosper,
                     'actuante':actuante,'today':today,'datosgral':datosgral,'hechodeli':hechodeli,'i':i,'tieneelementos':tieneelementos,
                     'preventor':preventor,'jerarqui_a':jerarqui_a,'jerarqui_p':jerarqui_p,'depe':depe,'elementos':elementos,
                     'autoridades':autoridades,'personas':personas,'lugar':lugar,'lati':lati,'longi':longi,'tienehecho':tienehecho,
                     'errors': errors, 'grabo':grabo,'form':form, 'ciudad': ciudad,'condiciones':condiciones,'tienelugar':tienelugar,
                     'state':state, 'continua':continua,'delito':delito,'descripcion':descripcion,'idprev':idprev,'countinvolus':countinvolus,
                     'unidadreg':unidadreg,'dependencia':dependencia,'unireg':unireg,'fecha_cierre':fecha_cierre,
                     'destino': destino,'form1':form1,'ftiposdelitos':ftiposdelitos,'tamaño':5,}


            return render(request,'./preventivoi.html', info)



    #datos del preventivos
    datos=Preventivos.objects.get(id=idprev)
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    fecha_cierre=datos.fecha_cierre
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    autoridades= datos.autoridades.values_list('descripcion',flat=True)
    dependencia=datos.dependencia.descripcion
    unidadreg=datos.dependencia.unidades_regionales.descripcion
    idprev=idprev
    #envio de datos al template updatehechos.html
    jerarqui_a=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=actuante).values('jerarquia_id'))
    jerarqui_p=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=preventor).values('jerarquia_id'))
    form1=Hechos.objects.filter(preventivo=idprev)
    today = datetime.datetime.now()
    info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
                'caratula':caratula,'idprev':idprev,'tienehecho':tienehecho,'tienelugar':tienelugar,
                'actuante':actuante,'elementos':elementos,'fecha_cierre':fecha_cierre,
                'preventor':preventor,'jerarqui_a':jerarqui_a,'jerarqui_p':jerarqui_p,'depe':depe,
                'autoridades':autoridades,'countinvolus':countinvolus,'unireg':unireg,
                'errors': errors,'unidadreg':unidadreg,'dependencia':dependencia,
                'state':state,
                'destino': destino,}
    #return render(request,'./preventivoi.html', info, context_instance=RequestContext(request))

    return render(request,'./preventivoi.html', info)

def generar_pdf(html):
        # Función para generar el archivo PDF y devolverlo mediante HttpResponse
        result = StringIO.StringIO()
        pdf = pisa.pisaDocument(StringIO.StringIO(html.encode("UTF-8")), result,encoding='UTF-8',
                                                                                        link_callback=fetch_resources)

        if not pdf.err:
                                converted_pdf= HttpResponse(result.getvalue(), content_type='application/pdf')
                                pdf = result.getvalue()
                                return converted_pdf


        return HttpResponse('Error al generar el PDF: %s' % cgi.escape(html))

def fetch_resources(uri, rel):
        import os.path
        from django.conf import settings
        path = os.path.join(
                        settings.STATIC_ROOT,
                        uri.replace(settings.STATIC_URL, ""))

        return path

@login_required
@group_required(["policia","investigaciones","radio"])
def updatehechos(request,idprev):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors =[]
    continua=''
    grabo='fin'
    ids = Preventivos.objects.get(id = idprev)
    depe= ids.dependencia
    #en caso de agregar delitos
    form=HechosForm(request.POST, request.FILES)
    boton='no'
    hecho = Hechos.objects.get(preventivo=idprev)
    idhec=hecho.id
    ftiposdelitos=DelitoForm()
    motivo=request.POST.get('motivo')
    modos=RefModosHechoForm(instance=hecho)
    delito =HechosDelito.objects.filter(hechos = hecho,borrado__isnull=True)
    descripcion=hecho.descripcion
    idhec=hecho.id
    #datos del preventivos
    datos=Preventivos.objects.get(id=idprev)
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    fecha_autorizacion=datos.fecha_autorizacion
    autoridades= datos.autoridades.values_list('descripcion',flat=True)
    dependencia=datos.dependencia.descripcion
    unidadreg=datos.dependencia.unidades_regionales.descripcion
    descripcionhecho=''
    idprev=idprev
    #envio de datos al template updatehechos.html



    if request.POST.get('continua')=="Agregar":
        if request.POST.get('delito'):
                    continua="si"
                    hechoDelito = HechosDelito()
                    hechoDelito.hechos = hecho
                    hechoDelito.refdelito = RefDelito.objects.get(id = request.POST.get('delito'))
                    if request.POST.get('modos'):
                     hechoDelito.refmodoshecho = RefModosHecho.objects.get(id = request.POST.get('modos'))
                    if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or 'RADIO' in request.user.userprofile.depe.descripcion:
                         try:
                            hechoDelito.save()
                            descripcion=request.POST.get('descripcion')
                            continua='no'
                            grabo='si'
                            if request.user.userprofile.depe==depe or 'MUJER' in request.user.userprofile.depe.descripcion:
                               tipodelito=RefDelito.objects.get(id = request.POST.get('delito'))
                               if 'VIOLENCIA FAMILIAR' in tipodelito.descripcion or 'Violencia Familiar' in tipodelito.descripcion:
                                   boton='si'
                         except IntegrityError:
                            errors.append('Delito que intenta agregar ya fue cargado. Si considera que es un Error, comuniquese con el Administrador.-')
                            #return render(request, './updatehechos.html', {'errors': errors})
                         #hechoDelito.save()
                         delitos =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)
                    else:
                         errors.append('No se puede modificar preventivos de otras dependencias.')

    else:
     if request.POST.get("borrar"):

                 hechoDelito = HechosDelito()
                 hechoDelito = HechosDelito.objects.filter(id = request.POST.get("borrar"))
                 hechoDelito.delete()
                 descripcion=request.POST.get('descripcion')
                 #.update(borrado='S')
                 delitos =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)

     else:
             if request.POST.get('grabar')=="Modificar":

                 delito =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)


                 if form.is_valid():

                     hecho.descripcion=form.cleaned_data['descripcion']
                     hecho.motivo=form.cleaned_data['motivo']
                     hecho.fecha_desde=form.cleaned_data['fecha_desde']
                     hecho.fecha_hasta=form.cleaned_data['fecha_hasta']

                     if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or 'RADIO' in request.user.userprofile.depe.descripcion:
                            if hecho.fecha_desde > ids.fecha_denuncia or hecho.fecha_hasta > ids.fecha_denuncia:
                                errors.append('La Fecha de Denuncia nunca puede ser menor a la Fecha y Hora del Hecho sucedido')
                            else:
                                hecho.save()
                                form=HechosForm(instance=hecho)
                                descripcion = hecho.descripcion
                     else:
                             errors.append('No se puede modificar preventivos de otras dependencias.')
                 else:

                         errors.append('No se realizo ninguna modificación. Verifique los datos ingresados.')
                         info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
                        'caratula':caratula,'idhec':idhec,
                        'actuante':actuante,'unidadreg':unidadreg,'dependencia':dependencia,
                        'preventor':preventor,'descripcionhecho':descripcionhecho,
                        'autoridades':autoridades,
                        'errors':errors,'grabo':grabo,'fecha_autorizacion':fecha_autorizacion,
                        'state':state, 'continua':continua,'delito':delito,'descripcion':descripcion,
                        'destino': destino,'form':form,'ftiposdelitos':ftiposdelitos,'idprev':idprev,}
                         return render(request,'./updatehechos.html',info)

             else:

                        si = Hechos()
                        form=HechosForm(instance=hecho)
                        #mostrar datos del hecho cargado para modificar o imprimir
                        existe= Hechos.objects.filter(preventivo=idprev)

                        if existe:
                            hecho = Hechos.objects.get(preventivo=idprev)
                        else:
                            return HttpResponseRedirect('../')


    info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
    'caratula':caratula,'idhec':idhec,'boton':boton,
    'actuante':actuante,
    'preventor':preventor,'depe':depe,
    'autoridades':autoridades,'unidadreg':unidadreg,'dependencia':dependencia,
    'errors':errors,'grabo':grabo,'fecha_autorizacion':fecha_autorizacion,
    'state':state, 'continua':continua,'delito':delito,'descripcion':descripcion,
    'destino': destino,'form':form,'ftiposdelitos':ftiposdelitos,'idprev':idprev,}

    return render(request,'./updatehechos.html',info)


@login_required
@group_required(["policia","investigaciones","visita","radio"])
def selectPrev(request,prev):
     state= request.session.get('state')
     destino= request.session.get('destino')
     preventivo = Preventivos.objects.get(id = prev)
     fecha_autorizacion=preventivo.fecha_autorizacion

     idprev=preventivo.id
     depe= preventivo.dependencia
     unireg=depe.unidades_regionales.descripcion
     form = PreventivosForm()
     errors=[]
     idhec=0
     idper=0
     tieneHecho = False
     tienePersonas=False
     tienelugar=False
     tieneelemento = False
     datosinvo=[]
     lista=''
     lugarhecho=''
     descri=''
     hechodeli=''
     fecha_desde=''
     fecha_hasta=''
     modosref=''
     delito=''
     boton='no'
     if Preventivos.objects.get(id=prev).has_hecho():
        tieneHecho = True
        if len(Elementos.objects.filter(hechos = Hechos.objects.get(preventivo = prev).id)) > 0:
            hecho=Hechos.objects.get(preventivo=prev)
            lista = Elementos.objects.filter(hechos = hecho.id,borrado__isnull=True)
            tieneelemento = True

        idhec=Hechos.objects.get(preventivo=idprev)

        datas=Hechos.objects.filter(preventivo=idprev).values()
        datoshecho=[]
        for data in datas:
             datoshecho.append(data)
        for j in datoshecho:
                delito =HechosDelito.objects.filter(hechos = j['id'],borrado__isnull=True)
                idmodo=j['motivo_id']
                descri=j['descripcion'].encode('utf8')
                modosref=RefMotivosHecho.objects.get(id=idmodo)
                fecha_desde=j['fecha_desde']
                fecha_hasta=j['fecha_hasta']
        cometidos=[]
        hechodeli=""
        for d in delito:
            cometidos.append(d)

        for i in cometidos:
            if i.refmodoshecho:
                 hechodeli=hechodeli+' '+str(i)+'  Modalidad :'+str(i.refmodoshecho)
            else:
                 hechodeli=hechodeli+' '+str(i)+'  Sin Modalidad'

        idhec=idhec.id
        ######### MODIFICACION ##########
        if len(Lugar.objects.filter(hecho = Hechos.objects.get(preventivo = prev).id)) > 0:
            hecho=Hechos.objects.get(preventivo=idprev)
            idlugar = Hechos.objects.get(id=hecho.id).lugar_hecho
            lugar=Hechos.objects.get(id=hecho.id).lugar_hecho
            lugarhecho='LUGAR : '+idlugar.tipo_lugar.descripcion+' --- '+'  ZONA :'+str(idlugar.zona)+' ---  '+'  UBICACION : '+idlugar.calle.descripcion+' NRO.: '+str(idlugar.altura)
            tienelugar = True

        if  Hechos.objects.get(id=Preventivos.objects.get(id=prev).hecho.id).involu.all():

                datosinvo=Hechos.objects.get(id=Preventivos.objects.get(id=prev).hecho.id).involu.all()

                tienePersonas= True


     if request.POST.get('button') == 'Modificar':
        form = PreventivosForm(request.POST, request.FILES)
        if form.is_valid():
            preventivo.fecha_denuncia = form.cleaned_data['fecha_denuncia']
            preventivo.caratula = form.cleaned_data['caratula']
            preventivo.actuante = form.cleaned_data['actuante']
            preventivo.preventor = form.cleaned_data['preventor']
            ureg=Dependencias.objects.get(descripcion__contains=request.user.userprofile.depe.descripcion)


            if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or  'RADIO' in request.user.userprofile.depe.descripcion and depe.ciudad_id==ureg.ciudad_id:
                 preventivo.save()
                 preventivo.autoridades.clear()
                 for grabauto in form.cleaned_data['autoridades']:
                        preventivo.autoridades.add(int(RefAutoridad.objects.get(descripcion=grabauto).id))
            else:

                 errors.append('No se puede modificar preventivos de dependencias que no pertenezcan a la Jurisdiccion.-')


        else:
            id_depe=Dependencias.objects.filter(descripcion__exact=depe).values('id')
            form.fields['actuante'].queryset = Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=1) | Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=3)
            form.fields['preventor'].queryset= Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=2) | Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=3)
            form.fields['actuante'].initial= preventivo.actuante.id
            form.fields['preventor'].initial = preventivo.preventor.id
            id_ciudad=Dependencias.objects.filter(descripcion__exact=depe).values('ciudad')
            form.fields['autoridades'].queryset=RefCiudades.objects.get(id=id_ciudad).ciu_autori.all()
            autoridades= preventivo.autoridades.all()
            autoridad=[]
            for seleccion in autoridades:
                    ids=int(RefAutoridad.objects.get(descripcion=seleccion).id)
                    autoridad.append(ids)
            form.fields['autoridades'].initial=autoridad
            form.fields['autoridades'].widget.attrs["onclick"] = False
            return render(request,'./updateprev.html',{'boton':boton,'fecha_desde':fecha_desde,'fecha_hasta':fecha_hasta,'idmodo':modosref,'delito':delito,'preventivo':preventivo,'fecha_autorizacion':fecha_autorizacion,'unireg':unireg,'idprev':idprev,'form':form,'state':state, 'destino': destino,'depe':depe,'tieneHecho':tieneHecho,'tienelugar':tienelugar,'tienePersonas':tienePersonas,'idhec':idhec,'idper':idper,})

     form = PreventivosForm(instance = preventivo)
     id_depe=Dependencias.objects.filter(descripcion__exact=depe).values('id')
     form.fields['actuante'].queryset = Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=1) | Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=3)
     form.fields['preventor'].queryset= Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=2) | Actuantes.objects.filter(dependencia_id__exact=id_depe,funcion__exact=3)
     form.fields['actuante'].initial= preventivo.actuante.id
     form.fields['preventor'].initial = preventivo.preventor.id
     id_ciudad=Dependencias.objects.filter(descripcion__exact=depe).values('ciudad')
     form.fields['autoridades'].queryset=RefCiudades.objects.get(id=id_ciudad).ciu_autori.all()
     autoridades= preventivo.autoridades.all()
     autoridad=[]
     for seleccion in autoridades:
            ids=int(RefAutoridad.objects.get(descripcion=seleccion).id)
            autoridad.append(ids)
     form.fields['autoridades'].initial=autoridad
     form.fields['autoridades'].widget.attrs["onclick"] = False

     return render(request,'./updateprev.html',{'boton':boton,'errors':errors,'fecha_desde':fecha_desde,'fecha_hasta':fecha_hasta,'idmodo':modosref,'delito':delito,'preventivo':preventivo,'unireg':unireg,'fecha_autorizacion':fecha_autorizacion,'lista':lista,'lugarhecho':lugarhecho,'datosinvo':datosinvo,'descripcion':descri,'hechodeli':hechodeli,'idprev':idprev,'form':form,'state':state, 'destino': destino,'depe':depe,'tieneHecho':tieneHecho,'tienelugar':tienelugar,'tienePersonas':tienePersonas,'idhec':idhec,'idper':idper,'tieneelemento':tieneelemento,})

@login_required
@transaction.atomic
@group_required(["administrador","policia","investigaciones","radio"])
def persinvo(request,idhec,idper):

    state= request.session.get('state')
    destino= request.session.get('destino')
    hechos = Hechos.objects.get(id = idhec)
    idciu = hechos.preventivo.dependencia.ciudad_id
    depe = hechos.preventivo.dependencia
    ids = Preventivos.objects.get(id = hechos.preventivo_id)
    text=''
    errors=[]
    mostrar="0"
    todos=[]
    comb=""
    roles=""
    formro=""
    domicilios=""
    estadete="no"
    datosinvo=[]
    tieneHecho = False
    tienePersonas=False
    tienelugar=False
    siexistepoli=False
    formp = PersonasForm(request.POST)
    domicilios = Domicilios(request.POST)
    dom = DomiciliosForm(request.POST)
    formr = PersInvolucradasForm(request.POST)
    formpa = PadresForm(request.POST)
    formc = CiudadesForm()
    formd = DepartamentosForm()
    formpr = ProvinciasForm()
    personas=''

    if request.POST.get('grabarciu')=="Grabar":
         formc = CiudadesForm(request.POST, request.FILES)
         descripcion = request.POST.get('descripcion')
         pais = request.POST.get('pais')

         if not descripcion or not pais:
                 errors.append('Ingrese una referencia(Pais) a la que pertenece la ciudad')
         else:
                         if not(len(descripcion)>=4 and len(descripcion)< 80):
                                         errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
                         else:
                                                 formc = CiudadesForm(request.POST, request.FILES)

                                                 if formc.is_valid():
                                                         formc.save()

                                                 else:
                                                            errors.append('La ciudad que UD. desea Guardar ya Existe. Verifique')
    #gaba domicilio
    if request.POST.get('grabadomi')=='Guardar':
         formp = PersonasForm(request.POST, request.FILES)
         dom = DomiciliosForm(request.POST)
         personas = Personas.objects.get(id = idper)

         reside=RefCiudades.objects.get(id=request.POST.get('ciudad_res'))
         if dom.is_valid():
                try:
                 domicilios                      = Domicilios()
                 domicilios.barrio_codigo        = dom.cleaned_data['barrio_codigo']
                 domicilios.calle                = dom.cleaned_data['calle']
                 domicilios.altura               = dom.cleaned_data['altura']
                 domicilios.entre                = dom.cleaned_data['entre']
                 domicilios.fecha_desde          = date.today()
                 domicilios.fecha_hasta          = dom.cleaned_data['fecha_hasta']
                 domicilios.fecha_actualizacion  = dom.cleaned_data['fecha_actualizacion']
                 domicilios.tipos_domicilio      = dom.cleaned_data['tipos_domicilio']
                 domicilios.ref_zona             = dom.cleaned_data['ref_zona']
                 domicilios.departamento         = dom.cleaned_data['departamento']
                 domicilios.piso                 = dom.cleaned_data['piso']
                 domicilios.lote                 = dom.cleaned_data['lote']
                 domicilios.sector               = dom.cleaned_data['sector']
                 domicilios.manzana              = dom.cleaned_data['manzana']
                 domicilios.ref_ciudades         = reside
                 domicilios.personas             = personas
                 domicilios.save()
                except IntegrityError:
                    errors.append('Datos duplicados.-')
         else:
                mostrar="no"

    #graba barrios segun la ciudad
    if request.POST.get('grabab')=='Grabar':
        formp=PersonasForm(request.POST, request.FILES)
        formbarrios = BarriadasForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        ciudad = request.POST.get('ciudad')
        mostrar="no"
        if not descripcion:
            errors.append('Ingrese un Barrio')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<100):
                    errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
             else:
                    if ciudad == 'Seleccione ciudad':
                        errors.append('Debe seleccionar una ciudad')
                    else:


                            if formbarrios.is_valid():
                             formbarrios.save()



    ##graba calles segun idciu
    if request.POST.get('grabac')=='Grabar':
        formcalles = AddressForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        ciudad = request.POST.get('ciudad')

        mostrar="no"
        if not descripcion:
            errors.append('Ingrese el nombre de la calle')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<150):
                    errors.append('El dato ingresado debe tener entre 4 y 150 caracteres')
             else:
                    if not ciudad:
                        errors.append('Debe seleccionar una ciudad')
                    else:
                         if formcalles.is_valid():
                                formcalles.save()






    if Hechos.objects.get(id = idhec):
     tieneHecho = True
     idhec=Hechos.objects.get(preventivo=ids).id
     ######### MODIFICACION ##########
    if len(Lugar.objects.filter(hecho = hechos.id)) > 0:
        tienelugar = True
    if  hechos.involu.all():
                tienePersonas= True


    if request.POST.get('nuevo')=="Nuevo" :
         formp=PersonasForm()
         domicilios = Domicilios()
         dom = DomiciliosForm()
         formr = PersInvolucradasForm()
         formpa = PadresForm()
         idper=0
         mostrar="no"


    else:
        if request.POST.get('search')=="Buscar":
         texto=request.POST.get('texto')

         if texto:
            query_string=texto
            entry_query = get_query(query_string, ['nro_doc', 'apellidos','nombres',])
         else:
            query_string='%'
            entry_query = get_query(query_string, ['nro_doc', 'apellidos','nombres',])



         filtro=Personas.objects.filter(entry_query)

         if filtro:
                if filtro not in todos:
                        todos.append(filtro)
                        mostrar="es"
         else:
                errors.append('No existen Personas con la referencia de búsqueda ingresada')
                mostrar="0"

        if idper != '0' and mostrar!='es':
            mostrar="si"
            filtros=Personas.objects.filter(id = idper)
            if filtros not in todos:
                 todos.append(filtros)

            personas = Personas.objects.get(id = idper)

            formp = PersonasForm(instance=personas)
            domicilios = Domicilios()
            dom = DomiciliosForm()

            formr = PersInvolucradasForm()
            filt= Padres.objects.filter(persona = personas.id)
            if filt:
                 idpapis= Padres.objects.get(persona = personas.id)
                 formpa = PadresForm(instance=idpapis)
            else:
                 formpa=PadresForm()

            if len(Domicilios.objects.filter(personas = idper)) > 0:
             domicilios = Domicilios.objects.filter(personas = idper)[0]
             dom = DomiciliosForm(instance = domicilios)
             dom.fields['barrio_codigo'].queryset = RefBarrios.objects.filter(ciudad=personas.ciudad_res)
             dom.fields['calle'].queryset = dom.fields['entre'].queryset= RefCalles.objects.filter(ciudad=personas.ciudad_res)
            else:

             dom = DomiciliosForm()

             dom.fields['barrio_codigo'].queryset = RefBarrios.objects.filter(ciudad=personas.ciudad_res)
             dom.fields['calle'].queryset = dom.fields['entre'].queryset= RefCalles.objects.filter(ciudad=personas.ciudad_res)

        else:
            if (request.POST.get('grabab') is None and request.POST.get('grabac') is None):
                 formp = PersonasForm()
                 domicilios = Domicilios()
                 dom = DomiciliosForm()
                 formr = PersInvolucradasForm()
                 formpa = PadresForm()



        if request.POST.get('dele'):
            persoinvoluc=PersInvolucradas.objects.filter(id=request.POST.get('dele'))

            if persoinvoluc:
                persoinvolu=PersInvolucradas.objects.get(id=request.POST.get('dele'))
                if 'si' in persoinvolu.detenido:
                    if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or 'RADIO' in request.user.userprofile.depe.descripcion:
                        Detenidos.objects.filter(persona_id = persoinvolu.persona_id).update(borrado='S',observaciones=request.user.username)
                        PersInvolucradas.objects.get(id=request.POST.get('dele')).delete()
                    else:
                        errors.append('No se puede modificar preventivos de otras dependencias.')
                else:
                    PersInvolucradas.objects.get(id=request.POST.get('dele')).delete()

            else:
                        errors.append('No se existe persona involucrada.')
            mostrar="0"
        if request.POST.get('grabar')=="Guardar":

                 formp = PersonasForm(request.POST, request.FILES)          #obtiene los datos de la persona en un formulario persona
                 dom = DomiciliosForm(request.POST,request.FILES)           #obtiene los datos del domicilio en un formulario domicilio
                 formr = PersInvolucradasForm(request.POST,request.FILES)   #obtiene los datos de persona involucrada en un formulario persona involucrada
                 formpa = PadresForm(request.POST,request.FILES)


                 anionac=datetime.datetime.strptime(request.POST.get('fecha_nac'),'%d/%m/%Y').strftime('%Y')
                 anioactual=datetime.datetime.now()
                 aniohoy=anioactual.today().year
                 if anionac>=1900:
                  dife=aniohoy-int(anionac)
                  #if (dife>=18 and menoris=='no') or  (dife<=17 and menoris=='si'):
                  #if (dife>=18) or  (dife<=17):
                  if request.POST.get('fechahoradetencion'):
                            fechadete=datetime.datetime.strptime(request.POST.get('fechahoradetencion'), '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y')

                            fecha_denuncia=datetime.datetime.strptime(request.POST.get('fecha_denuncia'), '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y %H:%M:%S')
                            fd = time.strptime(fecha_denuncia, "%d/%m/%Y %H:%M:%S")
                            fdet = time.strptime(fechadete, "%d/%m/%Y")
                            if fdet<fd:
                                    errors.append('La Fecha y hora de Detencion nunca debe ser menor a la de Denuncia del Hecho sucedido')
                                    mostrar="no"
                                    estadete="no"
                  if idper!='0':

                         perso=Personas.objects.get(id=idper)
                         fil=Padres.objects.filter(persona=perso.id)
                         if fil:
                                papis= Padres.objects.get(persona = perso.id)
                                formpa = PadresForm(instance=papis)
                         else:
                                formpa=PadresForm()
                                papis=Padres()
                         #buscar en personal
                         findpoli=Personal.objects.filter(persona_id=perso.id)
                         if findpoli:
                                siexistepoli=True
                  else:

                         perso=Personas()
                         papis=Padres()

                         iddom='1'

                  if len(Domicilios.objects.filter(personas = idper)) > 0:
                         domicilios = Domicilios.objects.filter(personas = idper)[0]
                         iddom=domicilios.id
                  else:
                         iddom='1'
                         domicilios=Domicilios()



                  if formp.is_valid() or idper!='0':

                            if formp.is_valid():

                             perso.apellidos  = formp.cleaned_data['apellidos']
                             perso.nombres    = formp.cleaned_data['nombres']
                             perso.tipo_doc   = formp.cleaned_data['tipo_doc']
                             perso.nro_doc    = formp.cleaned_data['nro_doc']
                             perso.fecha_nac  = formp.cleaned_data['fecha_nac']
                             perso.sexo_id    = formp.cleaned_data['sexo_id']
                             perso.pais_nac   = formp.cleaned_data['pais_nac']
                             perso.ciudad_nac = formp.cleaned_data['ciudad_nac']
                             perso.pais_res   = formp.cleaned_data['pais_res']
                             perso.ciudad_res = formp.cleaned_data['ciudad_res']
                             perso.ocupacion  = formp.cleaned_data['ocupacion']
                             perso.alias      = formp.cleaned_data['alias']
                             perso.estado_civil = formp.cleaned_data['estado_civil']
                             idpoli=formp.cleaned_data['ocupacion']
                             refpoli=RefOcupacion()
                             if idpoli:
                                refpoli=RefOcupacion.objects.get(descripcion=idpoli)


                            else:

                                 if request.POST.get('ocupacion')=='None' or request.POST.get('ocupacion')=='':
                                     refpoli=RefOcupacion.objects.get(descripcion='EMPLEADO')
                                     texto='EMPLEADO'
                                 else:
                                     idpoli=request.POST.get('ocupacion')
                                     refpoli=RefOcupacion.objects.get(id=idpoli)
                                     texto=refpoli.descripcion

                            if dom.is_valid():


                                     domicilios.barrio_codigo        = dom.cleaned_data['barrio_codigo']
                                     domicilios.calle                = dom.cleaned_data['calle']
                                     domicilios.altura               = dom.cleaned_data['altura']
                                     domicilios.entre                = dom.cleaned_data['entre']
                                     domicilios.fecha_desde          = dom.cleaned_data['fecha_desde']
                                     domicilios.fecha_hasta          = dom.cleaned_data['fecha_hasta']
                                     domicilios.fecha_actualizacion  = dom.cleaned_data['fecha_actualizacion']
                                     domicilios.tipos_domicilio      = dom.cleaned_data['tipos_domicilio']
                                     domicilios.ref_zona             = dom.cleaned_data['ref_zona']
                                     domicilios.departamento         = dom.cleaned_data['departamento']
                                     domicilios.piso                 = dom.cleaned_data['piso']
                                     domicilios.lote                 = dom.cleaned_data['lote']
                                     domicilios.sector               = dom.cleaned_data['sector']
                                     domicilios.manzana              = dom.cleaned_data['manzana']


                                     if idper!='0':
                                            persom=Personas.objects.get(id=idper)


                                            if formp.is_valid() or idper!='0':
                                                persom.ciudad_res = formp.cleaned_data['ciudad_res']
                                                persom.ocupacion  = formp.cleaned_data['ocupacion']
                                                persom.alias      = formp.cleaned_data['alias']
                                                persom.estado_civil = formp.cleaned_data['estado_civil']
                                                try:

                                                    persom.save()

                                                except IntegrityError:
                                                    errors.append('')
                                            else:

                                                 mostrar='si'
                                                 errors.append(formp.errors.as_text)

                                     else:

                                             perso.save()




                                     if idper!='0':

                                                     idpersu=Personas.objects.get(id=idper)

                                                     personas=idpersu
                                     else:
                                                     idpersu=Personas.objects.get(id=perso.id)

                                                     personas=idpersu

                                     domicilios.personas             = personas
                                     domicilios.ref_ciudades         = formp.cleaned_data['ciudad_res']

                                     domicilios.save()

                                     if refpoli:
                                        if refpoli.descripcion.find('POLICI')>=0:
                                                    policia=Personal()
                                                    policia.persona_id = personas
                                                    policia.credencial=0
                                                    try:
                                                        policia.save()
                                                    except IntegrityError:
                                                        errors.append('')
                                        else:
                                                 if siexistepoli:
                                                        #borro esa persona en personal
                                                        borrar=Personal.objects.get(persona_id=personas).delete()

                                     papis.persona=personas
                                     papis.padre_apellidos=request.POST.get('padre_apellidos')
                                     papis.padre_nombres=request.POST.get('padre_nombres')
                                     papis.madre_apellidos=request.POST.get('madre_apellidos')
                                     papis.madre_nombres=request.POST.get('madre_nombres')
                                     try:
                                            papis.save()
                                     except IntegrityError:
                                         errors.append('Datos existente en Padres')

                                     if formr.is_valid():

                                         persoin=PersInvolucradas()
                                         detenidos = Detenidos()



                                         persoin.persona=personas
                                         detenidos.persona = personas
                                         persoin.hechos=hechos

                                         persoin.roles = formr.cleaned_data['roles']
                                         if dife>=18:
                                            persoin.menor = 'no'
                                         else:
                                            persoin.menor = 'si'
                                            #formr.cleaned_data['menor']
                                         if 'APREHENDIDO' in persoin.roles.descripcion or  'APRENDIDO' in persoin.roles.descripcion or 'DETENIDO' in persoin.roles.descripcion:
                                             persoin.detenido = formr.cleaned_data['detenido']
                                         else:
                                             persoin.detenido='no'

                                         persoin.cargado_prev=True
                                         if persoin.detenido=='si':
                                            #
                                            # if 'no' in estadete:
                                            detenidos.hechos  = hechos
                                            detenidos.fechahoradetencion = formr.cleaned_data['fechahoradetencion']
                                            persoin.fechahoradetencion = formr.cleaned_data['fechahoradetencion']

                                            try:
                                             detenidos.save()
                                            except IntegrityError:
                                             errors.append('Datos existente en Detenidos')


                                         persoin.infraganti = formr.cleaned_data['infraganti']
                                         persoin.juridica = formr.cleaned_data['juridica']
                                         if persoin.juridica=='si':
                                                persoin.razon_social = formr.cleaned_data['razon_social']
                                                persoin.cuit = formr.cleaned_data['cuit']
                                                persoin.nrocuit = formr.cleaned_data['nrocuit']
                                         else:
                                                persoin.razon_social = 'SIN DESCRIPCION'

                                         try:
                                            mostrar='0'
                                            persoin.save()

                                         except IntegrityError:
                                            errors.append('Datos inexistente en Carateristicas segun el Rol')



                                     else:
                                         errors.append('Error faltan datos en seccion de Rol de la Persona')
                                         filtros=Personas.objects.filter(id = personas.id)
                                         if filtros not in todos:
                                             todos.append(filtros)

                            else:
                                         mostrar='no'


                  else:
                    mostrar='no'

                 else:

                    mostrar='no'
                    errors.append('Año de Nacimiento debe ser mayor a 1900')

    try:
        noposee = Personas.objects.get(id=idper).tipo_doc.descripcion
    except Exception as e:
        noposee=""

    datosinvo=hechos.involu.all()
    hec = Hechos.objects.get(id = idhec)
    datos = Preventivos.objects.get(id = ids.id)
    idprev=ids.id
    form=HechosForm(instance=hec)
    ftiposdelitos=DelitoForm()
    modos=RefModosHechoForm(instance=hec)
    delito =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)
    descripcion=hec.descripcion
    idhec=hec.id
    motivo=hec.motivo
    fecha_desde=hec.fecha_desde
    fecha_hasta=hec.fecha_hasta
    #datos del preventivos
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    fecha_autorizacion=datos.fecha_autorizacion
    autoridades= datos.autoridades.values_list('descripcion',flat=True)
    dependencia=datos.dependencia.descripcion
    unidadreg=datos.dependencia.unidades_regionales.descripcion

    #envio de datos al template updatehechos.html
    formcalles= AddressForm()
    formbarrios = BarriadasForm()
    formciu=RefCiudades.objects.all()

    info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
    'caratula':caratula,'idhec':idhec,'domicilios':domicilios,'formro':formro,'formcalles':formcalles,
    'actuante':actuante,'fecha_desde':fecha_desde,'fecha_hasta':fecha_hasta,'formbarrios':formbarrios,
    'preventor':preventor,'mostrar':mostrar,'formr':formr,'datosinvo':datosinvo,'fecha_autorizacion':fecha_autorizacion,
    'autoridades':autoridades,'formp':formp,'dom':dom,'roles':roles,'formciu':formciu,'idper':idper,'personas':personas,
    'errors': errors,'motivo':motivo,'todos':todos,'comb':comb,'idciu':idciu, 'tieneHecho':tieneHecho,
    'tienePersonas':tienePersonas,'tienelugar':tienelugar,'formd':formd,'formpr':formpr,'formc':formc,
    'state':state,'delito':delito,'descripcion':descripcion,'formpa':formpa,'depe':depe,'unidadreg':unidadreg,'dependencia':dependencia,
    'destino': destino,'form':form,'ftiposdelitos':ftiposdelitos,'idprev':idprev,'preventivo':datos,'noposee':noposee,}
    return render(request,'./personasin.html',info)

@login_required
@transaction.atomic
@group_required(["administrador","policia","investigaciones","radio"])
def persinvom(request,idhec,idper):
    state= request.session.get('state')
    destino= request.session.get('destino')
    hechos = Hechos.objects.get(id = idhec)
    idciu = hechos.preventivo.dependencia.ciudad_id
    depe = hechos.preventivo.dependencia
    text=''
    errors=[]
    mostrar="0"
    todos=[]
    comb=""
    roles=""
    formro=""
    domicilios=""
    datosinvo=[]
    detenido=''
    estadetenido=False
    personas=''
    razon=False
    otros=False
    #gaba domicilio

    if request.POST.get('grabadomi')=='Guardar':
         formp = PersonasForm(request.POST, request.FILES)
         dom = DomiciliosForm(request.POST)
         personas = Personas.objects.get(id = request.POST.get('idper'))
         reside=RefCiudades.objects.get(id=request.POST.get('ciudad_res'))
         if dom.is_valid():
                try:
                 domicilios                      = Domicilios()
                 domicilios.barrio_codigo        = dom.cleaned_data['barrio_codigo']
                 domicilios.calle                = dom.cleaned_data['calle']
                 domicilios.altura               = dom.cleaned_data['altura']
                 domicilios.entre                = dom.cleaned_data['entre']
                 domicilios.fecha_desde          = date.today()
                 domicilios.fecha_hasta          = dom.cleaned_data['fecha_hasta']
                 domicilios.fecha_actualizacion  = dom.cleaned_data['fecha_actualizacion']
                 domicilios.tipos_domicilio      = dom.cleaned_data['tipos_domicilio']
                 domicilios.ref_zona             = dom.cleaned_data['ref_zona']
                 domicilios.departamento         = dom.cleaned_data['departamento']
                 domicilios.piso                 = dom.cleaned_data['piso']
                 domicilios.lote                 = dom.cleaned_data['lote']
                 domicilios.sector               = dom.cleaned_data['sector']
                 domicilios.manzana              = dom.cleaned_data['manzana']
                 domicilios.ref_ciudades         = reside
                 domicilios.personas             = personas
                 domicilios.save()
                except IntegrityError:
                    errors.append('Datos duplicados.-')
         else:
                mostrar="no"

    if request.POST.get('ver'):

        idpin=request.POST.get('ver')
        idpersi=PersInvolucradas.objects.get(id=request.POST.get('ver'))

        idper=idpersi.persona_id

        mostrar="si"
        filtros=Personas.objects.filter(id = idper)
        if filtros not in todos:
             todos.append(filtros)

        personas = Personas.objects.get(id = idper)
        formp = PersonasForm(instance=personas)
        idpapis= Padres.objects.get(persona = personas.id)
        formpa = PadresForm(instance=idpapis)
        tdni=personas.tipo_doc
        domicilios = Domicilios()
        dom = DomiciliosForm()

        detenido=PersInvolucradas.objects.get(id=idpin, hechos_id=idhec)

        if 'APRE' in detenido.roles.descripcion or 'DETE' in detenido.roles.descripcion:
             estadetenido=True
        else:
            if 'DENUNCIADO' in detenido.roles.descripcion or 'VICTIMA' in detenido.roles.descripcion:
                razon=True
            else:
                if 'SOS' not in detenido.roles.descripcion:
                    otros=True

        formr = PersInvolucradasForm(instance=idpersi)
        formr.fields['roles'].initial=idpersi.roles
        if personas.ciudad_res:
             formp.fields['pais_res'].initial=personas.ciudad_res.pais
        if len(Domicilios.objects.filter(personas = idper)) > 0:
         domicilios = Domicilios.objects.filter(personas = idper)[0]
         dom = DomiciliosForm(instance = domicilios)
        else:
         dom.fields['barrio_codigo'].queryset = RefBarrios.objects.filter(ciudad=personas.ciudad_res)
         dom.fields['calle'].queryset = dom.fields['entre'].queryset= RefCalles.objects.filter(ciudad=personas.ciudad_res)
    else:

        if request.POST.get('grabar')=="Modificar":
                 formr = PersInvolucradasForm(request.POST, request.FILES)
                 formp = PersonasForm()
                 dom = DomiciliosForm()
                 formpa = PadresForm()
                 perso=Personas.objects.get(id=request.POST.get('idper'))
                 fecha_nac=request.POST.get('fecha_nac')

                 if formr.is_valid():
                        persoin=PersInvolucradas.objects.get(persona=perso.id,hechos_id=idhec)
                        persoin.roles = formr.cleaned_data['roles']
                        anionac=datetime.datetime.strptime(fecha_nac,'%d/%m/%Y').strftime('%Y')
                        anioactual=datetime.datetime.now()
                        aniohoy=anioactual.today().year

                        if anionac>=1900:
                            dife=aniohoy-int(anionac)
                            if 'APREHENDIDO' in persoin.roles.descripcion or  'APRENDIDO' in persoin.roles.descripcion or 'DETENIDO' in persoin.roles.descripcion:

                                     if formr.cleaned_data['fechahoradetencion']:
                                         fechadete=formr.cleaned_data['fechahoradetencion'].strftime('%d/%m/%Y %H:%M:%S')
                                         fecha_denuncia=datetime.datetime.strptime(request.POST.get('fecha_denuncia'),'%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y')
                                         fecha_dete=datetime.datetime.strptime(fechadete, '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y %H:%M:%S')

                                         fd = time.strptime(fecha_denuncia, "%d/%m/%Y")
                                         fde = time.strptime(fecha_dete, "%d/%m/%Y %H:%M:%S")
                                         if fde<fd:
                                                    errors.append('La Fecha y hora de Detencion nunca debe ser menor a la de Denuncia del Hecho sucedido')
                                         else:
                                            detenidos = Detenidos()
                                            dete=Detenidos.objects.filter(persona = perso)
                                            if dete:
                                               Detenidos.objects.filter(persona = perso).update(fechahoradetencion=formr.cleaned_data['fechahoradetencion'],libertad='N',borrado='',observaciones=request.user.username+'se equivoco a asignarle el rol')
                                            else:
                                               detenidos.persona=perso
                                               detenidos.hechos  = hechos
                                               detenidos.fechahoradetencion = formr.cleaned_data['fechahoradetencion']
                                               detenidos.libertad=''

                                            persoin.juridica='no'
                                            persoin.razon_social='SIN DESCRIPCION'

                                            persoin.detenido = 'si'
                                            persoin.tentativa = formr.cleaned_data['tentativa']
                                            persoin.infraganti = formr.cleaned_data['infraganti']
                                            persoin.fechahoradetencion = formr.cleaned_data['fechahoradetencion']

                                            try:
                                             if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or 'RADIO' in request.user.userprofile.depe.descripcion:
                                                 if fde>=fd:
                                                        detenidos.save()
                                             else:
                                                 errors.append('No se puede modificar preventivos de otras dependencias.')
                                            except IntegrityError:
                                             errors.append('La Persona Detenida ya registra antecedentes')
                                     else:
                                        mostrar='no'
                                        errors.append('Faltan Datos necesarios en Persona Detenida y/o Aprehendida. Verifique.- ')
                            else:
                                     persoin.detenido ='no'
                                     persoin.tentativa='no'
                                     persoin.infraganti='no'
                                     nulo='01/01/1900 00:00:00'
                                     persoin.fechahoradetencion=datetime.datetime.strptime(nulo, '%d/%m/%Y %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                                     Detenidos.objects.filter(persona = perso).update(libertad='S',borrado='S',observaciones=request.user.username+'se equivoco a asignarle el rol')
                                     PersInvolucradas.objects.filter(persona = perso).update(detenido=persoin.detenido)
                                     persoin.juridica = formr.cleaned_data['juridica']
                                     if persoin.juridica=='si':

                                        persoin.razon_social = formr.cleaned_data['razon_social']
                                        persoin.cuit = formr.cleaned_data['cuit']
                                        persoin.nrocuit = formr.cleaned_data['nrocuit']
                                     else:
                                        persoin.juridica='no'
                                        persoin.razon_social='SIN DESCRIPCION'
                                        persoin.cuit=formr.cleaned_data['cuit']
                                        persoin.nrocuit=formr.cleaned_data['nrocuit']

                            if dife>=18:
                                persoin.menor = 'no'
                            else:
                                persoin.menor = 'si'

                            try:
                                persoin.save()
                            except IntegrityError:
                                errors.append('Datos existente en Personas Involucradas')

                 else:

                        mostrar="no"
        else:

         formp = PersonasForm()
         domicilios = Domicilios()
         dom = DomiciliosForm()
         formr = PersInvolucradasForm()
         formpa= PadresForm()


    if request.POST.get('dele'):
         persoinvoluc=PersInvolucradas.objects.filter(id=request.POST.get('dele'))

         if persoinvoluc:
             persoinvolu=PersInvolucradas.objects.get(id=request.POST.get('dele'))
             if 'si' in persoinvolu.detenido:
                 if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or 'RADIO' in request.user.userprofile.depe.descripcion:
                     Detenidos.objects.filter(persona_id = persoinvolu.persona_id).update(borrado='S',observaciones=request.user.username)
                     PersInvolucradas.objects.get(id=request.POST.get('dele')).delete()
                 else:
                     errors.append('No se puede modificar preventivos de otras dependencias.')
             else:
                 PersInvolucradas.objects.get(id=request.POST.get('dele')).delete()

         else:
                errors.append('No se existe persona involucrada.')
         mostrar="0"


    datosinvo=Hechos.objects.get(id=idhec).involu.all()

    hec = Hechos.objects.get(id = idhec)
    filtroprev = Hechos.objects.filter(id= idhec).values('preventivo_id')
    datos = Preventivos.objects.get(id= filtroprev)
    idprev=datos.id
    form=HechosForm(instance=hec)
    ftiposdelitos=DelitoForm()
    modos=RefModosHechoForm(instance=hec)
    delito =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)
    descripcion=hec.descripcion
    idhec=hec.id
    motivo=hec.motivo
    fecha_desde=hec.fecha_desde
    fecha_hasta=hec.fecha_hasta
    #datos del preventivos
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    autoridades= datos.autoridades.values_list('descripcion',flat=True)
    dependencia=datos.dependencia.descripcion
    unidadreg=datos.dependencia.unidades_regionales.descripcion

    #envio de datos al template updatehechos.html
    fecha_autorizacion=datos.fecha_autorizacion
    formcalles= AddressForm()
    formbarrios = BarriadasForm()
    formciu =RefCiudades.objects.filter(id=idciu)

    info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
    'caratula':caratula,'idhec':idhec,'domicilios':domicilios,'formro':formro,'formcalles':formcalles,
    'actuante':actuante,'fecha_desde':fecha_desde,'fecha_hasta':fecha_hasta,'formbarrios':formbarrios,
    'preventor':preventor,'mostrar':mostrar,'formr':formr,'datosinvo':datosinvo,'fecha_autorizacion':fecha_autorizacion,
    'autoridades':autoridades,'formp':formp,'dom':dom,'roles':roles,'formciu':formciu,'estadetenido':estadetenido,'otros':otros,
    'errors': errors,'motivo':motivo,'todos':todos,'comb':comb,'idciu':idciu,'formpa':formpa,'unidadreg':unidadreg,'dependencia':dependencia,
    'state':state,'delito':delito,'descripcion':descripcion,'idper':idper,'detenido':detenido,'depe':depe,'personas':personas,
    'destino': destino,'form':form,'ftiposdelitos':ftiposdelitos,'idprev':idprev,'razon':razon,}
    return render(request,'./editpersoin.html',info)

@login_required
@transaction.atomic
@group_required(["policia","investigaciones","radio"])
def lugar_hecho(request,idhecho,idprev):
    state= request.session.get('state')
    destino= request.session.get('destino')
    preventivo = Preventivos.objects.get(id = idprev)
    idciu = preventivo.dependencia.ciudad_id
    depe = preventivo.dependencia.id
    depes=Dependencias.objects.get(id=depe)
    lugarhecho=''
    numero=''
    notienePer=False
    errors=[]
    if  Hechos.objects.get(id=Preventivos.objects.get(id=idprev).hecho.id).involu.all():

                datosinvo=Hechos.objects.get(id=Preventivos.objects.get(id=idprev).hecho.id).involu.all()

                notienePer= True
    if not notienePer:
            state= request.session.get('state')
            destino= request.session.get('destino')
            return render(request,'./errorsinper.html',{'state':state, 'destino': destino})

    if not Hechos.objects.get(id=idhecho).has_lugar():

        form = LugarForm()
        lugar = Lugar()
        form.fields['calle'].queryset = RefCalles.objects.filter(ciudad = preventivo.dependencia.ciudad)
        form.fields['barrio'].queryset = RefBarrios.objects.filter(ciudad = preventivo.dependencia.ciudad)
        form.fields['dependencia'].queryset = Dependencias.objects.filter(unidades_regionales = preventivo.dependencia.unidades_regionales_id)

    else:
        idlugar = Hechos.objects.get(id=idhecho).lugar_hecho
        lugar=Hechos.objects.get(id=idhecho).lugar_hecho



        form = LugarForm(instance=lugar)
        if request.POST.get('grabar') == 'Modificar':
         lugar=Lugar.objects.get(id=idlugar.id)
         form.fields['calle'].queryset = RefCalles.objects.filter(ciudad = preventivo.dependencia.ciudad)
         form.fields['barrio'].queryset = RefBarrios.objects.filter(ciudad = preventivo.dependencia.ciudad)
         form.fields['dependencia'].queryset = Dependencias.objects.filter(unidades_regionales = preventivo.dependencia.unidades_regionales_id)



    if request.POST.get('button') == 'Guardar' or request.POST.get('grabar') == 'Modificar':
            form = LugarForm(request.POST)

            if form.is_valid():
                lugar.tipo_lugar          = form.cleaned_data['tipo_lugar']
                lugar.zona                = form.cleaned_data['zona']
                nbarrio                   = form.cleaned_data['nuevo_barrio']
                lugar.barrio              = form.cleaned_data['barrio']
                lugar.nro_casa            = form.cleaned_data['nro_casa']
                lugar.manzana             = form.cleaned_data['manzana']
                lugar.lote                = form.cleaned_data['lote']
                lugar.sector              = form.cleaned_data['sector']
                lugar.edificio            = form.cleaned_data['edificio']
                lugar.escalera            = form.cleaned_data['escalera']
                lugar.piso                = form.cleaned_data['piso']
                lugar.departamento        = form.cleaned_data['departamento']



                ncalle,numero = street_name(form.cleaned_data['callen'])
                try:
                    calle = RefCalles.objects.get(descripcion=ncalle.upper(),ciudad=preventivo.dependencia.ciudad.id)
                except Exception as e:
                    calle=RefCalles()
                    calle.ciudad = RefCiudades.objects.get(id=preventivo.dependencia.ciudad_id)
                    calle.descripcion = ncalle
                    calle.save()

                lugar.calle =calle




                if not form.cleaned_data['altura']:

                     if numero:
                         numero2 = int(numero)
                     else:
                         numero2 = None
                     lugar.altura = numero2
                else:
                     if numero or numero==form.cleaned_data['altura'] :
                         if numero!=form.cleaned_data['altura']:
                                numero2=form.cleaned_data['altura']
                         else:
                                numero2 = form.cleaned_data['altura']
                         lugar.altura=numero2

                     else:
                         numero2 = None
                         lugar.altura              = form.cleaned_data['altura']

                lugar.latitud             = form.cleaned_data['latitud']
                lugar.longitud            = form.cleaned_data['longitud']
                lugar.dependencia         = form.cleaned_data['dependencia']
                lugar.hecho               = Hechos.objects.get(id=idhecho)
                if nbarrio:
                    barrio = RefBarrios()
                    barrio.ciudad = RefCiudades.objects.get(id=preventivo.dependencia.ciudad_id)
                    barrio.descripcion = nbarrio
                    barrio.save()
                    lugar.barrio = barrio
                if request.user.userprofile.depe_id==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or 'RADIO' in request.user.userprofile.depe.descripcion:

                     lugar.save()
                     if request.POST.get('grabar') == 'Modificar':
                         existen=lugar.cond_climaticas.through.objects.filter(lugar_id=idlugar).delete()

                     if form.cleaned_data['cond_climaticas']:
                            for clima in form.cleaned_data['cond_climaticas']:
                                 lugar.cond_climaticas.add(clima)
                     else:

                         if request.POST.get('grabar') == 'Modificar' :
                            if request.POST.get('cond_climaticas') is not None:
                                for clima in request.POST.get('cond_climaticas'):
                                        lugar.cond_climaticas.add(clima)


                else:
                     errors.append('No se puede guardar ubicacion del lugar.')



                form = LugarForm(instance=lugar)
                form.fields['latitud'].initial = lugar.latitud
                form.fields['longitud'].initial = lugar.longitud

    hechos = Hechos.objects.get(id = idhecho)
    idciu = hechos.preventivo.dependencia.ciudad_id
    depe = hechos.preventivo.dependencia
    ids = Preventivos.objects.get(id = hechos.preventivo_id)
    ftiposdelitos=DelitoForm()
    modos=RefModosHechoForm()
    datos=Preventivos.objects.get(id=ids.id)
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    fecha_autorizacion=datos.fecha_autorizacion
    autoridades= datos.autoridades.values_list('descripcion',flat=True)
    dependencia=datos.dependencia.descripcion
    unidadreg=datos.dependencia.unidades_regionales.descripcion

    values={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
            'caratula':caratula,'idhec':idhecho,
            'actuante':actuante,'idprev':idprev,
            'preventor':preventor,
            'autoridades':autoridades,'unidadreg':unidadreg,'dependencia':dependencia,
            'state':state, 'fecha_autorizacion':fecha_autorizacion,
            'destino': destino,
            'form': form,'depe':depe,
            'idhecho':idhecho,
            'idprev':idprev,
            'preventivo':preventivo,
            'lugar':lugar,'notienePer':notienePer,
            'ciudad': preventivo.dependencia.ciudad_id,


    }
    return render(request,'./crime_scene.html',values)

def street_name(string):

    nstring =''
    indice = 0
    number = ''
    numbers=''

    while string[indice] != '-' and string[indice] != ',' and indice < len(string):

        nstring = nstring + string[indice]
        indice = indice +1

    cant = 0


    while nstring[len(nstring)-1].isdigit():
        number=nstring[len(nstring)-1]+number

        nstring = nstring[:-1]


    return (nstring,number)

def autorizar(request,idprev):
    preventivo = Preventivos.objects.get(id=idprev)
    msg =""
    try:
        preventivo.fecha_autorizacion = datetime.datetime.now()
        preventivo.autoriza = request.user.username
        preventivo.save()
        msg = "El preventivo fue autorizado."
    except Exception as e:
        msg = "no se pudo autorizar el preventivo."
    return HttpResponse(msg)

#informar a autoridades por email
@login_required
@transaction.atomic
@group_required(["policia","investigaciones","radio"])
def informe(request,idhec,idprev,aforo):
        info_enviado = False
        name=""
        docu=""
        jerarca=""
        destino=""
        errors=""
        grabo=""
        continua=""
        mail=""
        comment=""
        formulario=""
        info_enviado= True
        tienehecho=False
        tienelugar=False
        tieneper=False
        tieneelementos=False
        state= request.session.get('state')
        destino= request.session.get('destino')
        preventivo = Preventivos.objects.get(id = idprev)
        #grabo la fecha de autorizacion solo la primera vez que se informa el preventivo
        ciudad= preventivo.dependencia.ciudad
        depe=preventivo.dependencia
        #Datos del Hecho delicitivo atraves del nro de preventivo
        hecho = Hechos.objects.get(preventivo=idprev)

        form=HechosForm(instance=hecho)
        ftiposdelitos=DelitoForm()
        motivo=request.POST.get('motivo')
        modos=RefModosHechoForm(instance=hecho)
        descripcion=hecho.descripcion
        idhec=hecho.id
        delito =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)
        modalidad=''

        eleinvo=Elementos.objects.filter(hechos=hecho.id,ampliacion_id__isnull=True,borrado__isnull=True).all()
        cometidos=[]
        hechodeli=""
        for d in delito:
            tienehecho=True
            cometidos.append(d)

        for i in cometidos:
            if i.refmodoshecho:
                 hechodeli=hechodeli+'*'+str(i)+'  Modalidad :'+str(i.refmodoshecho)+'<br>'
            else:
                 hechodeli=hechodeli+'*'+str(i)+'  Sin Modalidad <br>'

        #Datos de las Personas involucradas en el hecho
        involuscra=[]
        datosper=""
        titempo=""
        involus=Hechos.objects.get(id=hecho.id).involu.all()
        datosgral=""
        eleminvo=[]
        #Datos del lugar del hecho
        lugar = Hechos.objects.get(id=idhec).lugar_hecho
        condiciones= lugar.cond_climaticas.values_list('descripcion',flat=True)
        for tiempo in condiciones:
             titempo=titempo+' * '+str(tiempo)+'<br>'

        laticiudad = RefCiudades.objects.get(id=preventivo.dependencia.ciudad_id)
        lati=laticiudad.lat
        longi=laticiudad.longi
        tienelugar=True
        for p in Hechos.objects.get(id=hecho.id).involu.all():
            #aqui comprobar que datos son null de personas
            bandera,personai=funverifica(p.persona.id)
            if p.menor=='':
                p.menor="NO"
            if p.juridica=='si':
                if p.razon_social!=None:
                  perjuridica=str(p.razon_social)
                if RefTipoDocumento.objects.get(id=p.cuit_id)!='Null':
                  perjuridica=perjuridica+'-'+str(RefTipoDocumento.objects.get(id=p.cuit_id))
                if p.nrocuit!=0:
                  perjuridica=perjuridica+'-'+str(p.nrocuit)
            domi=Personas.objects.get(id=p.persona.id).persodom.all()
            if domi:
                for l in Personas.objects.get(id=p.persona.id).persodom.all():
                     dad=Personas.objects.get(id=p.persona.id).padre.all()
                     tieneper=True
                     if dad:

                            for la in Personas.objects.get(id=p.persona.id).padre.all():
                                    roles='<u>'+p.roles.descripcion+'</u><br><br>'
                                    if bandera:
                                        if p.juridica=='si':
                                            persona='<dd>'+p.persona.tipo_doc+': '+p.persona.nro_doc+personai+'</dd><dd>Personeria Juridica :'+perjuridica+'</dd>'
                                        else:
                                            persona='<dd>'+p.persona.tipo_doc+': '+p.persona.nro_doc+personai+'</dd>'
                                    else:
                                        if p.juridica=='si':
                                            persona='<dd>'+p.persona.tipo_doc+': '+p.persona.nro_doc+'</dd><dd>Personeria Juridica :'+perjuridica+'</dd><dd>Ocupacion :'+p.persona.ocupacion+', Estado Civil :'+' '+p.persona.estado_civil+', Menor de Edad : '+p.menor.upper()+'<dd>Nacido en: '+p.persona.pais_nac+', '+p.persona.ciudad_nac+', Fecha Nac: '+p.persona.fecha_nac.strftime("%d/%m/%Y")
                                        else:
                                            persona='<dd>'+p.persona.tipo_doc.descripcion+': '+p.persona.nro_doc+', Ocupacion :'+p.persona.ocupacion.descripcion+', Estado Civil :'+' '+p.persona.estado_civil.descripcion+', Menor de Edad : '+p.menor.upper()+'<dd>Nacido en: '+p.persona.pais_nac.descripcion+', '+p.persona.ciudad_nac.descripcion+', Fecha Nac: '+p.persona.fecha_nac.strftime("%d/%m/%Y")+'</dd>'
                                    domi='<dd>Reside en : '+p.persona.ciudad_res.descripcion +',  Domicilio : '+ l.calle.descripcion if l.calle else ' ' +'  Nro.: '+l.altura if l.altura else ' '  +'</dd>'
                                    if la.padre_apellidos or la.padre_nombres or la.madre_apellidos or la.madre_nombres:
                                            padys='<dd>Hijo de : '+la.padre_apellidos+', '+la.padre_nombres+'  y de : '+la.madre_apellidos+', '+la.madre_nombres+'<br><br></dd>'
                                    else:
                                            padys='<dd>no registra datos de los padres'+'<br><br></dd>'
                                    
                                    datosgral=roles+persona+domi+padys
                     else:

                         roles='<u>'+p.roles+'</u><br><br>'
                         if bandera:
                            if p.juridica=='si':
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')+str('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')
                            else:
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')
                         else:
                            if p.juridica=='si':
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+'</dd>')+str('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')+'<dd>Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))
                            else:
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')

                         domi='<dd>Reside en : '+str(p.persona.ciudad_res)+',  Domicilio : '+str(l.calle)+'  Nro.: '+str(l.altura)+'<br></dd>'
                         padys='<dd>no registra datos de los padres'+'<br><br></dd>'
                         datosgral=roles+persona+domi+padys
                involuscra.append(datosgral)
            else:
                roles='<u>'+p.roles+'</u><br><br>'
                if bandera:
                    if p.juridica=='si':
                        persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')+str('</dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')
                    else:
                        persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')
                else:
                    if p.juridica=='si':
                        persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+'</dd>')+str('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')+'<dd>Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))
                    else:
                        persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')

                domi='<dd>no registra domicilio'+'<br></dd>'
                padys='<dd>no registra datos de los padres'+'<br><br></dd>'
                datosgral=roles+persona+domi+padys

                involuscra.append(datosgral)
        for i in involuscra:
             datosper=datosper+' * '+i+'<br>'

        datosgral=''

        deta=''
        detav=''
        eleme=''
        elementos=''
        i=1
        for eles in eleinvo:
            obdata=[]
            obdatav=[]
            deta=''
            detav=''
            tieneelementos=True
            if len(Elementosarmas.objects.filter(idelemento=eles.id))>0:

                 idar = Elementosarmas.objects.filter(idelemento=eles.id).values('idarma')
                 obdata=Armas.objects.filter(id=idar)

                 for extra in obdata:
                    titu='<ul><li><u> Carateristicas Generales : </u></li></ul>'
                    tabla='<dd> '+str(extra.subtipos)+'  ---  Tipo/s : '+str(extra.tipos)+'  ---  Sistema de Disparo : '+str(extra.sistema_disparo)+'   --- Marcas : '+str(extra.marcas)+'</dd><br>'
                    tipos='<dd> Calibre : '+str(extra.calibre)+'  --- Modelo : '+str(extra.modelo)+'  --- Nro Serie : '+str(extra.nro_arma)+'   ---  Propietario : '+str(extra.nro_doc)+' - '+str(extra.propietario)+'</dd><br>'
                    deta=titu+tabla+tipos

            if len(Elementoscars.objects.filter(idelemento=eles.id))>0:

                 idarv = Elementoscars.objects.filter(idelemento=eles.id).values('idvehiculo')

                 obdatav=Vehiculos.objects.filter(id=idarv)
                 for extrav in obdatav:
                    tituv='<ul><li><u> Carateristicas Generales : </u></li></ul>'
                    tablav='<dd> Marca/s : '+str(extrav.idmarca)+'  ---   Modelo : '+str(extrav.modelo)+'  ---  Dominio : '+str(extrav.dominio)+'   ---  Año : '+str(extrav.anio)+'</dd><br>'
                    tiposv='<dd> Tipo/s : '+str(extrav.tipov)+' ---  Nro Chasis : '+str(extrav.nchasis)+' ---  Nro. Motor : '+str(extrav.nmotor)+'</dd><br>'+'<dd> Propietario : '+str(extrav.nro_doc)+' - '+str(extrav.propietario)+'</dd><br>'
                    detav=tituv+tablav+tiposv



            tipo='<br><dd><u>'+str(eles.tipo)+'</u></dd><br><br>'
            ampli=''
            rubro=' Elemento/s '+str(eles.tipo)
            rubros='Rubro y Categoria :'+str(eles.rubro)+' --- '+str(eles.categoria)
            canti=' Cantidad : '+str(eles.cantidad)+' --- '+str(eles.unidadmed)
            obse=' Observaciones : '+str(eles.descripcion.encode("utf8"))
            if deta:
                 if detav:
                        eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+detav+'<br>'
                 else:
                        eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+deta+'<br>'

            else:
                 if detav:
                        eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+detav+'<br>'
                 else:
                        eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+deta+'<br>'
            eleminvo.append(eleme)
            i=i+1

        for j in eleminvo:
            elementos=elementos+j

        infor=''
        autoridad=''
        #datos del preventivos
        datos=Preventivos.objects.get(id=idprev)
        nro=datos.nro
        anio=datos.anio
        fecha_denuncia=str(timezone.localtime(datos.fecha_denuncia).strftime("%d/%m/%Y %H:%M:%S"))
        fecha_carga=str(timezone.localtime(datos.fecha_carga).strftime("%d/%m/%Y %H:%M:%S"))
        caratula=datos.caratula
        actuante=datos.actuante
        preventor=datos.preventor
        today = datetime.datetime.now()
        autoridades=datos.autoridades.values_list('descripcion',flat=True)
        dependencia=datos.dependencia.descripcion
        unidadreg=datos.dependencia.unidades_regionales.descripcion

        for a in autoridades:
                autoridad=autoridad+'*'+str(a)+'<br>'
        informa=datos.autoridades.values_list('email',flat=True)
        jerarqui_a=RefJerarquias.objects.get(id=Actuantes.objects.filter(id=actuante.id).values('jerarquia_id'))
        jerarqui_p=RefJerarquias.objects.get(id=Actuantes.objects.filter(id=preventor.id).values('jerarquia_id'))
        form1=Hechos.objects.filter(preventivo=idprev)

        for value in form1:

         hecho='Motivo : '+str(value.motivo)+'<br>'+'Fecha y Hora Inicial :'+str(timezone.localtime(value.fecha_desde).strftime("%d/%m/%Y %H:%M:%S"))+' --- Fecha y Hora Final :'+str(timezone.localtime(value.fecha_hasta).strftime("%d/%m/%Y %H:%M:%S"))+'<br>'+'Fecha de Carga: '+str(value.fecha_carga.strftime("%d/%m/%Y"))+'<br>'
         hecho1='Delitos cometidos : <br>'+'<dd>'+str(hechodeli)+'</dd><br>'+'Descripcion Breve del Hecho :'+descripcion+'<br>'

        hechodelitos=hecho+hecho1


        #envio de datos al template updatehechos.html

        subject ='Preventivo Nro : '+str(nro)+'/'+str(anio)+'--Dependencia : '+str(depe)+' Ciudad de : '+str(ciudad)
        from_email =  'divsistemasjp@policia.chubut.gov.ar'
        to=str(infor)

        titulo =str('<u>'+'Preventivo Nro : '+str(nro)+'/'+str(anio)+'</u>')
        tresto='--Dependencia : '+str(depe)+' Ciudad de : '+str(ciudad)+'</u><br>'
        titulo1='Fecha de Denuncia : '+fecha_denuncia+'<br>'+'Fecha de Carga: '+fecha_carga+'<br>'+'Caratula :'+caratula+'<br>'+'Actuante : '+jerarqui_a.descripcion+' '+actuante.apeynombres+' --- '+' Preventor :'+jerarqui_p.descripcion+' '+preventor.apeynombres+'<br>'+'Autoridades a informar :'+'<br><dd>'+autoridad+'<br><dd><hr>'
        titulo2=hechodelitos+'<hr>'+'Personas Involucradas'+'<hr>'+datosper+'<hr>'

        #ubicacion
        titulo3='Ubicacion Geografica y caracteristicas del Lugar del Hecho <br><hr>'+'Ciudad :'+str(ciudad)+' --- Ubicacion : '+lugar.calle.descripcion+' --- Nro :'+str(lugar.altura)+'<br>Tipo de Lugar : '+str(lugar.tipo_lugar)+' --  Zona: '+str(lugar.zona)+'<br>'
        titulo4='Barrio : '+lugar.barrio.descripcion if lugar.barrio else '' +' ---  Nro casa: '+str(lugar.nro_casa)+'<br>'+'Lote : '+lugar.lote+' --- Manzana: '+lugar.manzana+' --- Sector: '+lugar.sector+'<br>'
        titulo5='Edificio : '+str(lugar.edificio)+' --- Escalera: '+str(lugar.escalera)+' --- Departamento: '+str(lugar.departamento)+' ---  Piso: '+str(lugar.piso)+'<br><hr>'
        titulo6='Condiciones Climaticas del lugar <br><hr>'+'<dd>'+str(titempo)+'</dd><br><hr>'
        titulo7='Elementos del Hecho <br><hr>'+str(elementos)+'</dd><br><hr>'
        text_content=titulo+tresto+titulo1+titulo2+titulo3+titulo4+titulo5+titulo6+titulo7

        if request.user.userprofile.depe.descripcion != 'INVESTIGACIONES':

            informa=datos.autoridades.values_list('email',flat=True)
            #agregar email 2jefeacei para que reciba los preventivos
            envio=1
            mailsrch = re.compile(r'[\w\-][\w\-\.]+@[\w\-][\w\-\.]+[a-zA-Z]{1,4}')
            informa = mailsrch.findall(str(informa))
            informa = eliminarduplicados(informa)
            try:
                envio,nstring,subject,text_content,from_email=envioemail(envio,informa,subject,text_content,from_email,request)
                if not preventivo.fecha_envio:
                    fecha_envio=preventivo.fecha_envio
                    grabarfa = Preventivos.objects.filter(id = idprev).update(fecha_envio=datetime.datetime.now())
                if not preventivo.aforo:
                    grabarAforo = Preventivos.objects.filter(id = idprev).update(aforo=aforo)
            except smtplib.SMTPException as e:
                error = Errores()
                error.usuario = request.user
                error.descripcion = e
                error.save()
                return HttpResponseBadRequest()
            except Exception as e:
                error = Errores()
                error.usuario = request.user
                error.descripcion = e
                error.save()
                return HttpResponseBadRequest()



        info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,'tieneelementos':tieneelementos,
             'caratula':caratula,'idhec':idhec,'involus':involus,'involuscra':involuscra,'datosper':datosper,
             'actuante':actuante,'today':today,'datosgral':datosgral,'hechodeli':hechodeli,'elementos':elementos,
             'preventor':preventor,'jerarqui_a':jerarqui_a,'jerarqui_p':jerarqui_p,'depe':depe,'tienehecho':tienehecho,
             'autoridades':autoridades,'personas':personas,'lugar':lugar,'lati':lati,'longi':longi,'tienelugar':tienelugar,
             'errors': errors, 'grabo':grabo,'form':form, 'ciudad': ciudad,'condiciones':condiciones,'tieneper':tieneper,
             'state':state, 'continua':continua,'delito':delito,'descripcion':descripcion,'idprev':idprev,
             'destino': destino,'form1':form1,'ftiposdelitos':ftiposdelitos,'tamaño':5,}


        try:
            if request.session['reenvio']:
                request.session['msg'] = 'El preventivo se reenvio con exito.'
                preventivo.reenviado = True
                preventivo.save()
                return HttpResponseRedirect(reverse('reenvio'))
        except Exception as e:
            print( e)

        return render(request,'./informado.html', info)


@login_required
@transaction.atomic
@group_required(["policia","investigaciones","radio"])
def elementos(request,idhecho):
    state= request.session.get('state')
    destino= request.session.get('destino')
    hecho = Hechos.objects.get(id=idhecho)
    depe = hecho.preventivo.dependencia
    tiene=False
    tieneob=False
    idele=''
    errors = []
    notienePer=False

    if  Hechos.objects.get(id=Preventivos.objects.get(id=hecho.preventivo_id).hecho.id).involu.all():

                datosinvo=Hechos.objects.get(id=Preventivos.objects.get(id=hecho.preventivo_id).hecho.id).involu.all()

                notienePer= True
    if not notienePer:
            state= request.session.get('state')
            destino= request.session.get('destino')
            return render(request,'./errorsinper.html',{'state':state, 'destino': destino})

    if request.POST.get('dele'):
            elementosin=Elementos.objects.filter(id=request.POST.get('dele'))

            if elementosin:
                if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion.find('INVESTIGACIONES')>=0 or request.user.userprofile.depe.descripcion.find('RADIO')>=0:
                        obs="elemento borrado por usuario : "+request.user.username

                        Elementos.objects.filter(id = request.POST.get('dele')).update(borrado='S',observaciones=obs)

                else:
                        errors='No se puede borrar elementos de un Hecho de otras dependencias.'
            else:
                        errors='No se existe elemento involucrado.'

    if request.POST.get('button') == 'Guardar':
        form = ElementosForm(request.POST)
        formar=ArmasForm(request.POST)
        formv=VehiculosForm(request.POST)
        if form.is_valid():
            elemento = Elementos()
            elemento.tipo           = form.cleaned_data['tipo']
            elemento.rubro          = form.cleaned_data['rubro']
            elemento.categoria      = form.cleaned_data['categoria']
            elemento.cantidad       = form.cleaned_data['cantidad']
            elemento.unidadmed      = form.cleaned_data['unidadmed']
            elemento.descripcion    = form.cleaned_data['descripcion']
            elemento.hechos         = hecho
            elemento.cargado_prev=True
            if elemento.rubro.descripcion=='VEHICULOS' or elemento.rubro.descripcion== 'AUTOMOTORES':

                    if formv.is_valid():
                     elemento.save()
                     vehicle = Vehiculos()
                     vehicle.idmarca         = formv.cleaned_data['idmarca']
                     vehicle.modelo          = formv.cleaned_data['modelo']
                     vehicle.anio            = formv.cleaned_data['anio']
                     vehicle.tipov           = formv.cleaned_data['tipov']
                     vehicle.dominio         = formv.cleaned_data['dominio']
                     vehicle.nmotor          = formv.cleaned_data['nmotor']
                     vehicle.nchasis         = formv.cleaned_data['nchasis']
                     vehicle.nro_doc         = formv.cleaned_data['nro_doc']
                     vehicle.propietario     = formv.cleaned_data['propietario']
                     nueva_marcav            = formv.cleaned_data['nueva_marcav']
                     if nueva_marcav:
                        marca = RefTrademark()
                        marca.descripcion = nueva_marcav
                        marca.save()
                        vehicle.idmarca       = marca
                     vehicle.save()
                     elevehi = Elementoscars()
                     elevehi.idelemento        = elemento
                     elevehi.idvehiculo        = vehicle
                     elevehi.save()
                     formv = VehiculosForm()
                     errors.append('Elementos guardados')
            else:
                 if elemento.categoria.descripcion =='FUEGO' or elemento.categoria.descripcion=='DE FUEGO':
                         if formar.is_valid():
                             elemento.save()
                             armas=Armas()
                             armas.tipos    = formar.cleaned_data['tipos']
                             armas.subtipos = formar.cleaned_data['subtipos']
                             armas.sistema_disparo = formar.cleaned_data['sistema_disparo']
                             armas.marcas = formar.cleaned_data['marcas']
                             armas.calibre = formar.cleaned_data['calibre']
                             armas.modelo = formar.cleaned_data['modelo']
                             armas.nro_arma=formar.cleaned_data['nro_arma']
                             armas.nro_doc = formar.cleaned_data['nro_doc']
                             armas.propietario = formar.cleaned_data['propietario']
                             nueva_marca       = formar.cleaned_data['nueva_marca']

                             if nueva_marca:
                                 marca = RefTrademark()
                                 marca.descripcion = nueva_marca
                                 marca.save()
                                 armas.marcas = marca

                             armas.save()

                             elearmas=Elementosarmas()
                             elearmas.idelemento=elemento
                             elearmas.idarma=armas
                             elearmas.save()
                             form = ElementosForm()
                             formar=ArmasForm()
                             errors.append('Elementos guardados')
                 else:
                             elemento.save()
                             errors.append('Elementos guardados')
                             form=ElementosForm()
        else:
             errors.append('Faltan Datos verifique')

    if request.POST.get('guardarubro') == 'Guardar':
        formrub = ItemForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        mostrar="no"
        if not descripcion:
            errors.append('Ingrese un Rubro')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<100):
                    errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
             else:
                 if formrub.is_valid():
                        formrub.save()
    if request.POST.get('guardacategoria') == 'Guardar':
        formcat = CategoryForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        rubro = request.POST.get('rubro')
        mostrar="no"
        if not descripcion:
            errors.append('Ingrese una categoria')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<100):
                    errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
             else:
                if not rubro:
                    errors.append('Seleccione Rubro')
                else:
                    if formcat.is_valid():
                        formcat.save()
    if request.POST.get('guardamedida') == 'Guardar':
        formumed = UnidadMedidasForm(request.POST, request.FILES)

        descripcion = request.POST.get('descripcion')
        mostrar="no"
        if not descripcion:
            errors.append('Ingrese unidad de medida')
        else:
             if formumed.is_valid():
                formumed.save()


    lista = Elementos.objects.filter(hechos = hecho.id,borrado__isnull=True)
    formv=VehiculosForm()
    form = ElementosForm()
    elemento = Elementos()
    formrub = ItemForm()
    formcat = CategoryForm()
    formumed = UnidadMedidasForm()
    formar=ArmasForm()
    hechos = Hechos.objects.get(id = idhecho)
    idciu = hechos.preventivo.dependencia.ciudad_id
    depe = hechos.preventivo.dependencia
    ids = Preventivos.objects.get(id = hechos.preventivo_id)
    formprev = PreventivosForm(instance = ids)
    idprev=ids.id
    ftiposdelitos=DelitoForm()
    modos=RefModosHechoForm()
    datos=Preventivos.objects.get(id=hechos.preventivo_id)
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    fecha_autorizacion=datos.fecha_autorizacion
    autoridades= datos.autoridades.values_list('descripcion',flat=True)




    values={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
            'caratula':caratula,'formar':formar,'tiene':tiene,'tieneob':tieneob,
            'actuante':actuante,'idprev':idprev,'idele':idele,
            'preventor':preventor,'fecha_autorizacion':fecha_autorizacion,
            'autoridades':autoridades,'formprev':formprev,
            'state':state,'depe':depe,
            'destino': destino,
            'preventivo':hecho.preventivo,
            'hecho':hecho,
            'form':form,'errors':errors,
            'lista':lista,'vehiculo':formv,
            'elemento':elemento,
            'formrub':formrub,
            'formcat':formcat,
            'formumed':formumed,
    }
    return render(request,'./objects.html',values)


@login_required
@transaction.atomic
@group_required(["policia","investigaciones","radio"])
def elemento(request,idhecho,elemento):
    state= request.session.get('state')
    destino= request.session.get('destino')
    hecho = Hechos.objects.get(id=idhecho)
    elementox = Elementos.objects.get(id=elemento)
    form = ElementosForm(instance=elementox)
    formar = ArmasForm()
    obdata=''
    tieneob=False
    tiene=False
    tienecar=False
    formar=[]
    formv=[]
    if len(Elementosarmas.objects.filter(idelemento=elementox.id))>0:

                 idar = Elementosarmas.objects.filter(idelemento=elementox.id).values('idarma')
                 obdata=Armas.objects.get(id=idar)
                 formar=ArmasForm(instance=obdata)
                 tiene=True
    if len(Elementoscars.objects.filter(idelemento=elementox.id))>0:

                 idar = Elementoscars.objects.filter(idelemento=elementox.id).values('idvehiculo')
                 obdata=Vehiculos.objects.get(id=idar)
                 formv=VehiculosForm(instance=obdata)
                 tienecar=True

    lista = Elementos.objects.filter(hechos = hecho.id,borrado__isnull=True)
    if len(Elementos.objects.filter(hechos = hecho.id,borrado__isnull=True))>0:
        tieneob = True

    hechos = Hechos.objects.get(id = idhecho)
    idciu = hechos.preventivo.dependencia.ciudad_id
    depe = hechos.preventivo.dependencia
    ids = Preventivos.objects.get(id = hechos.preventivo_id)
    idprev=ids
    ftiposdelitos=DelitoForm()
    modos=RefModosHechoForm()
    datos=Preventivos.objects.get(id=ids.id)
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    fecha_autorizacion=datos.fecha_autorizacion
    autoridades= datos.autoridades.values_list('descripcion',flat=True)


    values={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
            'caratula':caratula,'formar':formar,'tiene':tiene,'idele':elemento,'vehiculo':formv,
            'actuante':actuante,'idprev':idprev,'tieneob':tieneob,'tienecar':tienecar,
            'preventor':preventor,'fecha_autorizacion':fecha_autorizacion,
            'autoridades':autoridades,'depe':depe,
            'destino': destino,'state':state,
            'preventivo':hecho.preventivo,
            'hecho':hecho,
            'form':form,
            'lista':lista,
            'elemento':elementox,
    }
    return render(request,'./objects.html',values)

@login_required
def get_categories(request,rubro):
                data = request.POST
                categorias = RefCategory.objects.filter(rubro= rubro)
                data = serializers.serialize("json", categorias)

                return HttpResponse(data, content_type='application/json')

@login_required
@group_required(["policia","investigaciones","radio","visita","judiciales"])
def seemaps(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    delito=''
    preven=[]
    prevs = []
    lugar=[]
    injusticia=''
    haydatos=False
    refdelitosdes=[]
    depe=request.user.userprofile.depe_id
    ureg=request.user.userprofile.ureg_id
    preventivo=Dependencias.objects.get(id=depe)
    depes=depe
    ureg=ureg
    manydelis=[]
    lugar=Lugar()
    form = MapaForm()
    ayerfue=''
    buscarciu=''
    buscar=''
    hoyes=''
    today = datetime.datetime.now()
    if request.POST.get('ver')=="Visualizar":
            datos=request.POST

            ciudad = datos.__getitem__('ciudades')
            ureg = datos.__getitem__('ureg')
            depes= datos.__getitem__('depe')
            fechadesde=datos.__getitem__('fecha_desde')
            fechahasta=datos.__getitem__('fecha_hasta')
            delitos=request.POST.getlist('delito')
            delitoslen=len(request.POST.getlist('delito'))
            if depes:
                depese=Dependencias.objects.get(id=depes)
                buscar=depese.unidades_regionales.descripcion+' - '+depese.descripcion
            else:
                if ciudad:
                     buscarciu=RefCiudades.objects.get(id=ciudad).descripcion
                     buscar=buscarciu
                else:
                     form.errors['__all__'] = form.error_class(["debe seleccionar una Ciudad y/o Unidad y dependencia en cada caso"])

            i=0
            if (fechadesde and fechahasta and delitos and ureg and depes) or (fechadesde and fechahasta and delitos and ciudad) or (fechadesde and fechahasta and ureg and depes) or (fechadesde and fechahasta and ciudad):
                    hoy=str(datetime.datetime.strptime(fechadesde,"%d/%m/%Y").date())+' 00:00:01'
                    ayer=str(datetime.datetime.strptime(fechahasta,"%d/%m/%Y").date())+' 23:59:59'
                    hoy=datetime.datetime.strptime(hoy,'%Y-%m-%d %H:%M:%S')
                    ayer=datetime.datetime.strptime(ayer,'%Y-%m-%d %H:%M:%S')
                    ayerfue=ayer.strftime('%d/%m/%Y %H:%M:%S')
                    hoyes=hoy.strftime('%d/%m/%Y %H:%M:%S')

                    if not ciudad and not depes:
                         ciudad=preventivo.ciudad.id

                    if depes:

                        preventivo=Dependencias.objects.get(id=depes)

                        if delitoslen>0:
                         for delis in delitos:

                            filtrorefd=HechosDelito.objects.filter(refdelito_id=delis,borrado__isnull=True).values('hechos_id')
                            refdelitosdes.append(RefDelito.objects.get(id=delis).descripcion)
                            if filtrorefd:

                                preventivos=Preventivos.objects.filter(dependencia=depes,fecha_autorizacion__range=[hoy,ayer]).values()


                                for desde in preventivos:
                                     idp=desde['id']

                                     idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                     if idhecho:
                                            for ids in idhecho:
                                                 idhec=ids['id']

                                                 for cfiltro in filtrorefd:
                                                     if idhec==cfiltro['hechos_id']:
                                                            filtrode=HechosDelito.objects.filter(hechos_id=idhec,borrado__isnull=True).values('refdelito_id')
                                                            for correr in filtrode:
                                                                    injusticia=injusticia+RefDelito.objects.get(id=correr['refdelito_id']).descripcion+', '
                                                                    delitosin=injusticia
                                                            datalugar=Lugar.objects.filter(hecho_id=idhec).values('id')
                                                            if datalugar:


                                                             reglugar=Lugar.objects.get(id=datalugar)
                                                             prevs={'id'+str(i):str(desde['id']),'prev'+str(i):str(desde['nro'])+'/'+str(desde['anio']),'denuncia'+str(i):str(datetime.datetime.strftime(desde['fecha_denuncia'], '%d/%m/%Y')),'latitud'+str(i):str(reglugar.latitud),'longitud'+str(i):str(reglugar.longitud),'delitos'+str(i):str(injusticia.encode("utf8"))}
                                                             preven.append(prevs)
                                                             haydatos=True
                                                             i=i+1
                                                             injusticia=''
                                    #muestra delitos
                        else:
                            filtrorefd=HechosDelito.objects.all().filter(borrado__isnull=True).values('hechos_id','refdelito_id')
                            for delis in filtrorefd:
                                refdelitosdes.append(RefDelito.objects.get(id=delis['refdelito_id']).descripcion)


                            if filtrorefd:
                                if ayer==hoy:
                                    preventivos=Preventivos.objects.filter(dependencia=depes,fecha_autorizacion__startswith=hoy).values()
                                else:
                                    preventivos=Preventivos.objects.filter(dependencia=depes,fecha_autorizacion__range=(hoy,ayer)).values()

                                for desde in preventivos:
                                     idp=desde['id']

                                     idhecho=Hechos.objects.filter(preventivo_id=idp).values()
                                     if idhecho:
                                            for ids in idhecho:
                                                 idhec=ids['id']

                                                 for cfiltro in filtrorefd:
                                                     if idhec==cfiltro['hechos_id']:
                                                            filtrode=HechosDelito.objects.filter(hechos_id=idhec,borrado__isnull=True).values('refdelito_id')
                                                            for correr in filtrode:
                                                                    injusticia=injusticia+RefDelito.objects.get(id=correr['refdelito_id']).descripcion+', '
                                                                    delitosin=injusticia
                                                            datalugar=Lugar.objects.filter(hecho_id=idhec).values('id')
                                                            if datalugar:



                                                             reglugar=Lugar.objects.get(id=datalugar)
                                                             prevs={'id'+str(i):str(desde['id']),'prev'+str(i):str(desde['nro'])+'/'+str(desde['anio']),'denuncia'+str(i):str(datetime.datetime.strftime(desde['fecha_denuncia'], '%d/%m/%Y')),'latitud'+str(i):str(reglugar.latitud),'longitud'+str(i):str(reglugar.longitud),'delitos'+str(i):str(injusticia.encode("utf8"))}

                                                             preven.append(prevs)
                                                             haydatos=True
                                                             i=i+1
                                                            injusticia=''
                    if ciudad:

                     depes=Dependencias.objects.filter(ciudad_id=ciudad).values('id')
                     if depes:
                        for dp in depes:
                             depes=dp['id']
                        preventivo=Dependencias.objects.get(id=depes)

                        if delitoslen>0:
                         for delis in delitos:

                            filtrorefd=HechosDelito.objects.filter(refdelito_id=delis,borrado__isnull=True).values('hechos_id')
                            refdelitosdes.append(RefDelito.objects.get(id=delis).descripcion)

                            if filtrorefd:
                                depes=Dependencias.objects.filter(ciudad_id=ciudad).values('id')
                                if depes:
                                    for crias in depes:
                                            dp=crias['id']
                                            preventivos=Preventivos.objects.filter(dependencia=dp,fecha_autorizacion__range=[hoy,ayer]).values()
                                            for desde in preventivos:
                                                    idp=desde['id']
                                                    idhecho=Hechos.objects.filter(preventivo_id=idp).values()
                                                    if idhecho:
                                                            for ids in idhecho:
                                                                 idhec=ids['id']


                                                                 for cfiltro in filtrorefd:
                                                                        if idhec==cfiltro['hechos_id']:
                                                                             filtrode=HechosDelito.objects.filter(hechos_id=idhec,borrado__isnull=True).values('refdelito_id')

                                                                             for correr in filtrode:
                                                                                     injusticia=injusticia+RefDelito.objects.get(id=correr['refdelito_id']).descripcion+', '
                                                                                     delitosin=injusticia
                                                                             datalugar=Lugar.objects.filter(hecho_id=idhec).values('id')

                                                                             if datalugar:
                                                                                    reglugar=Lugar.objects.get(id=datalugar)




                                                                                    prevs={'id'+str(i):str(desde['id']),'prev'+str(i):str(desde['nro'])+'/'+str(desde['anio']),'denuncia'+str(i):str(datetime.datetime.strftime(desde['fecha_denuncia'], '%d/%m/%Y')),'latitud'+str(i):str(reglugar.latitud),'longitud'+str(i):str(reglugar.longitud),'delitos'+str(i):str(injusticia.encode("utf8"))}
                                                                                    preven.append(prevs)
                                                                                    haydatos=True

                                                                                    i=i+1
                                                                             injusticia=''

                        else:
                            filtrorefd=HechosDelito.objects.all().filter(borrado__isnull=True).values('hechos_id','refdelito_id')
                            for delis in filtrorefd:
                                refdelitosdes.append(RefDelito.objects.get(id=delis['refdelito_id']).descripcion)



                            if filtrorefd:
                                depes=Dependencias.objects.filter(ciudad_id=ciudad).values('id')
                                if depes:
                                    for crias in depes:
                                            dp=crias['id']

                                            if ayer==hoy:
                                                preventivos=Preventivos.objects.filter(dependencia=dp,fecha_autorizacion__startswith=hoy).values()
                                            else:
                                                preventivos=Preventivos.objects.filter(dependencia=dp,fecha_autorizacion__range=[hoy,ayer]).values()


                                            for desde in preventivos:
                                                    idp=desde['id']
                                                    idhecho=Hechos.objects.filter(preventivo_id=idp).values()
                                                    if idhecho:
                                                            for ids in idhecho:
                                                                 idhec=ids['id']


                                                                 for cfiltro in filtrorefd:
                                                                        if idhec==cfiltro['hechos_id']:
                                                                             filtrode=HechosDelito.objects.filter(hechos_id=idhec,borrado__isnull=True).values('refdelito_id')

                                                                             for correr in filtrode:
                                                                                     injusticia=injusticia+RefDelito.objects.get(id=correr['refdelito_id']).descripcion+', '
                                                                                     delitosin=injusticia
                                                                             datalugar=Lugar.objects.filter(hecho_id=idhec).values('id')

                                                                             if datalugar:
                                                                                    reglugar=Lugar.objects.get(id=datalugar)

                                                                                    prevs={'id'+str(i):str(desde['id']),'prev'+str(i):str(desde['nro'])+'/'+str(desde['anio']),'denuncia'+str(i):str(datetime.datetime.strftime(desde['fecha_denuncia'], '%d/%m/%Y')),'latitud'+str(i):str(reglugar.latitud),'longitud'+str(i):str(reglugar.longitud),'delitos'+str(i):str(injusticia.encode('ISO-8859-1'))}
                                                                                    preven.append(prevs)
                                                                                    haydatos=True
                                                                                    i=i+1
                                                                                    injusticia=''


    values={'buscar':buscar,'buscarciu':buscarciu,'ayerfue':ayerfue,'hoyes':hoyes,'destino': destino,'state':state,'form':form,'preventivo':preventivo,'haydatos':haydatos,'preven':preven,'depes':depes,'ureg':ureg,'refdelitosdes':refdelitosdes,}
    return render(request,'./mapsanality.html',values)

#definicion para ver personas involuvradas en hechos
@login_required
def verperin(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors=[]
    todos=[]
    filtro=''
    mostrar=''
    idhecho=''
    if request.POST.get('search')=='Buscar':
         texto=request.POST.get('texto')

         if texto:
            query_string=texto
            entry_query = get_query(query_string, ['nro_doc', 'apellidos','nombres',])
         else:
            query_string='%'
            entry_query = get_query(query_string, ['nro_doc', 'apellidos','nombres',])



         filtro=Personas.objects.filter(entry_query)


         if filtro:
                if filtro not in todos:
                        todos.append(filtro)
                        mostrar="es"
         else:
                errors.append('No existen Personas con la referencia de búsqueda ingresada')
                mostrar="0"


    info={'errors': errors,'state':state,'destino': destino,'mostrar':mostrar,'filtro':filtro,'todos':todos,'idhecho':idhecho,}
    return render(request,'./peoplesin.html',info)

@login_required
@group_required(["policia","investigaciones","radio"])
def verpersin(request,idper):
    state= request.session.get('state')
    destino= request.session.get('destino')
    personas = Personas.objects.get(id = idper)
    lista =PersInvolucradas.objects.all()
    todosdata=[]
    datosinvo=''
    alldata=[]
    idhec=''
    text=''
    errors=[]
    mostrar="1"
    todos=[]
    comb=""
    roles=""
    formro=""
    filtro=''
    domicilios=""
    datosinvo=[]
    detenido=''
    persinv=[]
    delitos=''
    perhechos= PersInvolucradas.objects.filter(persona=personas.id).values()
    perhechos1= PersInvolucradas.objects.filter(persona=personas.id)

    if perhechos1:
         for idhecho in perhechos:

                idhec=idhecho['hechos_id']
                idperso=idhecho['id']


         if perhechos1 not in todosdata:
                todosdata.append(perhechos1)


         hechos = Hechos.objects.get(id = idhec)

         idciu = hechos.preventivo.dependencia.ciudad_id
         depe = hechos.preventivo.dependencia

         if request.POST.get('vermas'):
                filtro='si'
                datosinvo=Hechos.objects.filter(id=request.POST.get('vermas'))
                relacion=PersInvolucradas.objects.all().filter(persona_id=personas,hechos_id=request.POST.get('vermas'))

                delitos =HechosDelito.objects.all().filter(hechos = request.POST.get('vermas'),borrado__isnull=True)
         if datosinvo:
                if datosinvo not in alldata:
                     alldata.append(datosinvo)
                     persinv.append(relacion)

         dataper=PersonasForm(instance=personas)

         dataper.fields['tipo_doc'].initial=dataper['tipo_doc']
         dataper.fields['ocupacion'].initial=dataper['ocupacion']
         hec = Hechos.objects.get(id = idhec)
         filtroprev = Hechos.objects.filter(id= idhec).values('preventivo_id')
         datos = Preventivos.objects.get(id= filtroprev)
         idprev=datos
         formprev=PreventivosForm(instance=idprev)
         form=HechosForm(instance=hec)
         ftiposdelitos=DelitoForm()
         modos=RefModosHechoForm(instance=hec)
         delito =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)
         descripcion=hec.descripcion
         idhec=hec.id
         motivo=hec.motivo
         fecha_desde=hec.fecha_desde
         fecha_hasta=hec.fecha_hasta
         #datos del preventivos
         nro=datos.nro
         anio=datos.anio
         fecha_denuncia=datos.fecha_denuncia
         fecha_carga=datos.fecha_carga
         caratula=datos.caratula
         actuante=datos.actuante
         preventor=datos.preventor
         autoridades= datos.autoridades.values_list('descripcion',flat=True)
         #envio de datos al template updatehechos.html
         formcalles= AddressForm()
         formbarrios = BarriadasForm()
         formciu =RefCiudades.objects.filter(id=idciu)
         info={'delitos':delitos,'persinv':persinv,'alldata':alldata,'filtro':filtro,'todosdata':todosdata,'perhechos':perhechos,'dataper':dataper,'datos':datos,'datosinvo':datosinvo,'errors': errors,'state':state,'destino': destino,'mostrar':mostrar,'filtro':filtro,'todos':todos,}
    else:
         errors.append("La persona seleccionada no registra datos que los involucre en algun Hecho")
         info={'errors': errors,'state':state,'destino': destino,'mostrar':mostrar,'filtro':filtro,'todos':todos,}
    return render(request,'./peoplesin.html',info)

#definicion para ver elementos involuvrados en hechos
@login_required
def verobjin(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors=[]
    todos=[]
    filtro=''
    mostrar=''
    mostrarsi=''
    idhecho=''
    alldata=[]
    filtro=''
    texto=0
    todosa=[]
    idprev=0
    form=ElementosForm()
    if request.POST.get('search')=='Buscar':
         texto=request.POST.get('categorias')
         elementos=Elementos.objects.filter(categoria=texto,borrado__isnull=True)
         if elementos:
                if elementos not in todos:
                        todos.append(elementos)
                mostrar="es"
         else:
                errors.append('No existen Elementos con la referencia de búsqueda ingresada')
                mostrar="0"

    if request.POST.get('vermas'):
         texto=request.POST.get('idcate')

         elementos=Elementos.objects.all().filter(hechos_id=request.POST.get('vermas'),borrado__isnull=True)

         if elementos:
                 if elementos not in todos:
                         todos.append(elementos)
                         mostrarsi="es"
         else:
                errors.append('No existen Elementos con la referencia de búsqueda ingresada')
                mostrar="0"
         filtro='si'

         datosinvo=Hechos.objects.filter(id=request.POST.get('vermas'))

         if datosinvo:
                if datosinvo not in alldata:
                     alldata.append(datosinvo)

         hec = Hechos.objects.get(id = request.POST.get('vermas'))
         filtroprev = Hechos.objects.filter(id= request.POST.get('vermas')).values('preventivo_id')
         datos = Preventivos.objects.get(id= filtroprev)
         idprev=datos
         formprev=PreventivosForm(instance=idprev)
         formh=HechosForm(instance=hec)
         ftiposdelitos=DelitoForm()
         modos=RefModosHechoForm(instance=hec)
         delito =HechosDelito.objects.filter(hechos = request.POST.get('vermas'),borrado__isnull=True)
         descripcion=hec.descripcion
         idhec=hec.id
         motivo=hec.motivo
         fecha_desde=hec.fecha_desde
         fecha_hasta=hec.fecha_hasta
         #datos del preventivos
         nro=datos.nro
         idprev=datos.id
         anio=datos.anio
         fecha_denuncia=datos.fecha_denuncia
         fecha_carga=datos.fecha_carga
         caratula=datos.caratula
         actuante=datos.actuante
         preventor=datos.preventor
         autoridades= datos.autoridades.values_list('descripcion',flat=True)

    info={'idprev':idprev,'todosa':todosa,'mostrarsi':mostrarsi,'texto':texto,'filtro':filtro,'alldata':alldata,'errors': errors,'state':state,'destino': destino,'mostrar':mostrar,'filtro':filtro,'todos':todos,'idhecho':idhecho,'form':form,}
    return render(request,'./objectsin.html',info)

#definicion de los reposrtes estadisticos
@login_required
@group_required(["policia","investigaciones","radio"])
def repestadis(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    delito=''
    preven=[]
    prevs = []
    lugar=[]
    injusticia=''
    haydatos=False
    refdelitosdes=[]
    depe=request.user.userprofile.depe_id
    ureg=request.user.userprofile.ureg_id
    preventivo=Dependencias.objects.get(id=depe)
    depes=depe
    ureg=ureg
    manydelis=[]
    lugar=Lugar()
    form = MapaForm()
    ayerfue=''
    hoyes=''
    today = datetime.datetime.now()
    dparti=0
    inpoli=0
    ordenj=0
    otros=0
    months_choices = []
    anios=[]
    son=0
    tipodel=''
    for i in range(1,13):
         months_choices.append(ugettext(datetime.date(date.today().year, i, 1).strftime('%b')))

    for j in range(date.today().year-6, date.today().year+1):
         anios.append(j)
    anios=sorted(anios, reverse = True)

    tiposdelitos=DelitoForm()

    if request.POST.get('ver')=="Visualizar":
            datos=request.POST

            ciudad = datos.__getitem__('ciudades')
            ureg = datos.__getitem__('ureg')
            depes= datos.__getitem__('depe')
            mes=datos.__getitem__('mes')
            anio=datos.__getitem__('anios')
            tipo=datos.__getitem__('tipodel')
            delitos=datos.__getitem__('delitoe')
            texto=datos.__getitem__('buscartex')
            mesnro=strptime(mes,'%b').tm_mon
            filtrohecho=Hechos.objects.filter(descripcion__icontains=texto)
            if not ciudad and not depes:
                 ciudad=preventivo.ciudad.id


            if depes:
                preventivo=Dependencias.objects.get(id=depes)
                preventivos=Preventivos.objects.filter(dependencia=depes,fecha_denuncia__month=mesnro,fecha_denuncia__year=anio).values()



    values={'tiposdelitos':tiposdelitos,'anios':anios,'month_choices':months_choices,'ayerfue':ayerfue,'hoyes':hoyes,'destino': destino,'state':state,'form':form,'preventivo':preventivo,'haydatos':haydatos,'preven':preven,'depes':depes,'ureg':ureg,'refdelitosdes':refdelitosdes,}
    return render(request,'./reportes.html',values)

#definicion de los reposrtes estadisticos
@login_required
@group_required(["policia","investigaciones","radio"])
def repforages(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors=[]
    #hoy = datetime.datetime.now()
    rango=[]
    rango1=[]
    rango2=[]
    rango3=[]
    rango4=[]
    rango5=[]
    rango6=[]
    rango7=[]
    rangos=[]
    rangos1=[]
    rangos2=[]
    rangos3=[]
    rangos4=[]
    rangos5=[]
    rangos6=[]
    rangos7=[]
    rangoms=[]
    rangosr=[]
    rangosr1=[]
    rangosr2=[]
    rangosr3=[]
    rangosr4=[]
    rangosr5=[]
    rangosr6=[]
    rangosr7=[]
    rangomsr=[]
    robos=0
    homi=0
    hurtos=0
    otros=0
    injusticia=''
    delitosin=''
    rangom=[]
    trobos=0
    thomi=0
    thurtos=0
    totros=0
    fechadesde=0
    fechahasta=0
    mesnro=0
    sumo=0
    saludo=''
    anios=0
    rango1a17=[]
    tro=0
    tho=0
    thu=0
    reporte=False
    tot=0
    ciudad=''
    depes=''
    ureg=''
    lugar=''
    feme={}
    masc={}
    sumof=0
    sumom=0
    sumofr=0
    sumomr=0
    sumosinr=0
    sinsexo={}
    sumorf=0
    sumorm=0
    sinsexor={}
    sumosin=0
    rephomisexo=False
    reprobosexo=False

    if request.POST.get('ver')=="Visualizar" or request.POST.get('expo')=="Exportar" or request.POST.get('expo1')=="Exportar" or request.POST.get('expor')=="Exportar":
            datosform=request.POST

            if request.POST.get('expo')=="Exportar" or request.POST.get('expo1')=="Exportar" or request.POST.get('expor')=="Exportar":

             ciudad=request.POST.get('ciu')
             ureg=request.POST.get('ure')
             depes=request.POST.get('dep')
             fechadesde=request.POST.get('mesi')
             fechahasta=request.POST.get('aniosi')

            else:


             ciudad = datosform.__getitem__('ciudades')
             ureg = datosform.__getitem__('ureg')
             depes= datos__getitem__('depe')
             fechadesde=request.POST.get('mes')
             fechahasta=request.POST.get('anios')
            filtrorefd=Hechos.objects.all()


            if ciudad:

                lugar=RefCiudades.objects.get(id=ciudad)
                depes=Dependencias.objects.filter(ciudad_id=ciudad).values('id')
                if depes:
                     for dp in depes:
                             depes=dp['id']
                     preventivo=Dependencias.objects.get(id=depes)
                     depes=Dependencias.objects.filter(ciudad_id=ciudad).values('id')

                     if depes:
                                    for crias in depes:
                                            dp=crias['id']
                                            preventivos=Preventivos.objects.all().filter(dependencia=dp,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()

                                            for desde in preventivos:
                                                    idp=desde['id']

                                                    idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                    if idhecho:
                                                         for ids in idhecho:
                                                                 idprevh=ids['preventivo_id']
                                                                 if idprevh==idp:
                                                                            idhecho=ids['id']
                                                                            filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')
                                                                            torobos=[]
                                                                            sexo=''
                                                                            sexor=''
                                                                            for indelis in filtrodelih:
                                                                                    descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])

                                                                                    if descripcion.descripcion.find('HOMI')>=0 or descripcion.descripcion.find('ROBO')>=0 or descripcion.descripcion.find('HURTO')>=0 or (descripcion.descripcion.find('HOMI')<0 or descripcion.descripcion.find('ROBO')<0 or descripcion.descripcion.find('HURTO')<0):
                                                                                            filtroperin=PersInvolucradas.objects.all().filter(hechos_id=idhecho)
                                                                                            sumo=sumo+len(filtroperin)

                                                                                            for datos in filtroperin:

                                                                                                            if datos.persona.fecha_nac:
                                                                                                                if descripcion.descripcion.find('ROBO')>=0:
                                                                                                                     robos=robos+1
                                                                                                                     sexor=datos.persona.sexo_id

                                                                                                                     if sexor:
                                                                                                                            sexr=sexor.descripcion
                                                                                                                            if sexor.descripcion.find('FEME')>=0:
                                                                                                                                 sumorf+=1
                                                                                                                            else:
                                                                                                                                 if sexor.descripcion.find('MASCU')>=0:
                                                                                                                                        sumorm+=1
                                                                                                                     else:
                                                                                                                                 sumosinr+=1

                                                                                                                else:
                                                                                                                    if descripcion.descripcion.find('HOMI')>=0:
                                                                                                                         homi=homi+1
                                                                                                                         sexo=datos.persona.sexo_id

                                                                                                                         if sexo:
                                                                                                                                 sex=sexo.descripcion
                                                                                                                                 if sexo.descripcion.find('FEME')>=0:
                                                                                                                                        sumof+=1
                                                                                                                                 else:
                                                                                                                                     if sexo.descripcion.find('MASCU')>=0:
                                                                                                                                         sumom+=1
                                                                                                                         else:
                                                                                                                                         sumosin+=1

                                                                                                                    else:
                                                                                                                        if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                             hurtos=hurtos+1
                                                                                                                        else:
                                                                                                                             otros=otros+1

                                                                                                                years=int(datetime.date.today().strftime('%Y'))-int(datos.persona.fecha_nac.strftime('%Y'))
                                                                                                                if years>=5:
                                                                                                                    if years>=5 and years<=15:
                                                                                                                         rangom.append(str(descripcion))
                                                                                                                         if sexo:
                                                                                                                             rangoms.append(str(sexo))
                                                                                                                         if sexor:
                                                                                                                             rangomsr.append(str(sexor))
                                                                                                                    else:
                                                                                                                         if years>=16 and years<=25:
                                                                                                                             rango.append(str(descripcion))
                                                                                                                             if sexo:
                                                                                                                                 rangos.append(str(sexo))
                                                                                                                             if sexor:
                                                                                                                                 rangosr.append(str(sexor))
                                                                                                                         else:
                                                                                                                             if years>=26 and years<=35:
                                                                                                                                 rango1.append(str(descripcion))
                                                                                                                                 if sexo:
                                                                                                                                     rangos1.append(str(sexo))
                                                                                                                                 if sexor:
                                                                                                                                     rangosr1.append(str(sexor))
                                                                                                                             else:
                                                                                                                                 if years>=36 and years<=45:
                                                                                                                                    rango2.append(str(descripcion))
                                                                                                                                    if sexo:
                                                                                                                                        rangos2.append(str(sexo))
                                                                                                                                    if sexor:
                                                                                                                                        rangosr2.append(str(sexor))
                                                                                                                                 else:
                                                                                                                                        if years>=46 and years<=55:
                                                                                                                                            rango3.append(str(descripcion))
                                                                                                                                            if sexo:
                                                                                                                                                 rangos3.append(str(sexo))
                                                                                                                                            if sexor:
                                                                                                                                                rangosr3.append(str(sexor))
                                                                                                                                        else:
                                                                                                                                             if years>=56 and years<=65:
                                                                                                                                                 rango4.append(str(descripcion))
                                                                                                                                                 if sexo:
                                                                                                                                                    rangos4.append(str(sexo))
                                                                                                                                                 if sexor:
                                                                                                                                                    rangosr4.append(str(sexor))
                                                                                                                                             else:
                                                                                                                                                 if years>=66 and years<=75:
                                                                                                                                                        rango5.append(str(descripcion))
                                                                                                                                                        if sexo:
                                                                                                                                                            rangos5.append(str(sexo))
                                                                                                                                                        if sexor:
                                                                                                                                                            rangosr5.append(str(sexor))
                                                                                                                                                 else:
                                                                                                                                                        rango6.append(str(descripcion))
                                                                                                                                                        if sexo:
                                                                                                                                                            rangos6.append(str(sexo))
                                                                                                                                                        if sexor:
                                                                                                                                                            rangosr6.append(str(sexor))



                                                                                                                else:
                                                                                                                     rango7.append(str(descripcion))
                                                                                                                     if sexo:
                                                                                                                         rangos7.append(str(sexo))
                                                                                                                     if sexor:
                                                                                                                         rangosr7.append(str(sexor))

            else:
                if depes:

                                            preventivo=Dependencias.objects.get(id=depes)
                                            ciudad=""
                                            depe=preventivo.descripcion
                                            unireg=preventivo.unidades_regionales.descripcion
                                            lugar=ciudad+' -- '+unireg+' -- '+depe

                                            preventivos=Preventivos.objects.all().filter(dependencia=depes,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()
                                            for desde in preventivos:
                                                    idp=desde['id']

                                                    idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                    if idhecho:
                                                         for ids in idhecho:
                                                                 idprevh=ids['preventivo_id']
                                                                 if idprevh==idp:
                                                                            idhecho=ids['id']
                                                                            filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')
                                                                            torobos=[]
                                                                            sexo=''
                                                                            sexor=''
                                                                            for indelis in filtrodelih:
                                                                                    descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])
                                                                                    if descripcion.descripcion.find('HOMI')>=0 or descripcion.descripcion.find('ROBO')>=0 or descripcion.descripcion.find('HURTO')>=0 or (descripcion.descripcion.find('HOMI')<0 or descripcion.descripcion.find('ROBO')<0 or descripcion.descripcion.find('HURTO')<0):
                                                                                         filtroperin=PersInvolucradas.objects.all().filter(hechos_id=idhecho)
                                                                                         sumo=sumo+len(filtroperin)
                                                                                         for datos in filtroperin:

                                                                                                            if datos.persona.fecha_nac:
                                                                                                                if descripcion.descripcion.find('ROBO')>=0:
                                                                                                                     robos=robos+1
                                                                                                                     sexor=datos.persona.sexo_id

                                                                                                                     if sexor:
                                                                                                                            sexr=sexor.descripcion
                                                                                                                            if sexor.descripcion.find('FEME')>=0:
                                                                                                                                 sumorf+=1
                                                                                                                            else:
                                                                                                                                 if sexor.descripcion.find('MASCU')>=0:
                                                                                                                                        sumorm+=1
                                                                                                                     else:
                                                                                                                                 sumosinr+=1

                                                                                                                else:
                                                                                                                    if descripcion.descripcion.find('HOMI')>=0:
                                                                                                                         homi=homi+1
                                                                                                                         sexo=datos.persona.sexo_id

                                                                                                                         if sexo:
                                                                                                                                 sex=sexo.descripcion
                                                                                                                                 if sexo.descripcion.find('FEME')>=0:
                                                                                                                                        sumof+=1
                                                                                                                                 else:
                                                                                                                                     if sexo.descripcion.find('MASCU')>=0:
                                                                                                                                         sumom+=1
                                                                                                                         else:
                                                                                                                                         sumosin+=1

                                                                                                                    else:
                                                                                                                        if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                             hurtos=hurtos+1
                                                                                                                        else:
                                                                                                                             otros=otros+1

                                                                                                                years=int(datetime.date.today().strftime('%Y'))-int(datos.persona.fecha_nac.strftime('%Y'))
                                                                                                                if years>=5:
                                                                                                                    if years>=5 and years<=15:
                                                                                                                         rangom.append(str(descripcion))
                                                                                                                         if sexo:
                                                                                                                             rangoms.append(str(sexo))
                                                                                                                         if sexor:
                                                                                                                             rangomsr.append(str(sexor))
                                                                                                                    else:
                                                                                                                         if years>=16 and years<=25:
                                                                                                                             rango.append(str(descripcion))
                                                                                                                             if sexo:
                                                                                                                                 rangos.append(str(sexo))
                                                                                                                             if sexor:
                                                                                                                                 rangosr.append(str(sexor))
                                                                                                                         else:
                                                                                                                             if years>=26 and years<=35:
                                                                                                                                 rango1.append(str(descripcion))
                                                                                                                                 if sexo:
                                                                                                                                     rangos1.append(str(sexo))
                                                                                                                                 if sexor:
                                                                                                                                     rangosr1.append(str(sexor))
                                                                                                                             else:
                                                                                                                                 if years>=36 and years<=45:
                                                                                                                                    rango2.append(str(descripcion))
                                                                                                                                    if sexo:
                                                                                                                                        rangos2.append(str(sexo))
                                                                                                                                    if sexor:
                                                                                                                                        rangosr2.append(str(sexor))
                                                                                                                                 else:
                                                                                                                                        if years>=46 and years<=55:
                                                                                                                                            rango3.append(str(descripcion))
                                                                                                                                            if sexo:
                                                                                                                                                 rangos3.append(str(sexo))
                                                                                                                                            if sexor:
                                                                                                                                                rangosr3.append(str(sexor))
                                                                                                                                        else:
                                                                                                                                             if years>=56 and years<=65:
                                                                                                                                                 rango4.append(str(descripcion))
                                                                                                                                                 if sexo:
                                                                                                                                                    rangos4.append(str(sexo))
                                                                                                                                                 if sexor:
                                                                                                                                                    rangosr4.append(str(sexor))
                                                                                                                                             else:
                                                                                                                                                 if years>=66 and years<=75:
                                                                                                                                                        rango5.append(str(descripcion))
                                                                                                                                                        if sexo:
                                                                                                                                                            rangos5.append(str(sexo))
                                                                                                                                                        if sexor:
                                                                                                                                                            rangosr5.append(str(sexor))
                                                                                                                                                 else:
                                                                                                                                                        rango6.append(str(descripcion))
                                                                                                                                                        if sexo:
                                                                                                                                                            rangos6.append(str(sexo))
                                                                                                                                                        if sexor:
                                                                                                                                                            rangosr6.append(str(sexor))



                                                                                                                else:
                                                                                                                     rango7.append(str(descripcion))
                                                                                                                     if sexo:
                                                                                                                         rangos7.append(str(sexo))
                                                                                                                     if sexor:
                                                                                                                         rangosr7.append(str(sexor))
                else:

                    lugar="LA PROVINCIA DE CHUBUT"
                    ciudad=''
                    idprovincia=RefProvincia.objects.get(descripcion__icontains='CHUBUT')
                    provis=RefCiudades.objects.all().filter(provincia_id=idprovincia.id).values('id')

                    for ciuda in provis:
                                de=Dependencias.objects.filter(ciudad_id=ciuda['id']).values('id')
                                if de:
                                 for dp in de:
                                         depe=dp['id']
                                 preventivo=Dependencias.objects.get(id=depe)
                                 dep=Dependencias.objects.filter(ciudad_id=ciuda['id']).values('id')

                                 if dep:
                                                for crias in dep:
                                                        dp=crias['id']

                                                        preventivos=Preventivos.objects.all().filter(dependencia=dp,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()

                                                        for desde in preventivos:
                                                                idp=desde['id']

                                                                idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                                if idhecho:
                                                                     for ids in idhecho:
                                                                             idprevh=ids['preventivo_id']
                                                                             if idprevh==idp:
                                                                                        idhecho=ids['id']
                                                                                        filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')
                                                                                        torobos=[]
                                                                                        sexo=''
                                                                                        sexor=''
                                                                                        for indelis in filtrodelih:
                                                                                                 descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])
                                                                                                 if descripcion.descripcion.find('HOMI')>=0 or descripcion.descripcion.find('ROBO')>=0 or descripcion.descripcion.find('HURTO')>=0 or (descripcion.descripcion.find('HOMI')<0 or descripcion.descripcion.find('ROBO')<0 or descripcion.descripcion.find('HURTO')<0):
                                                                                                     filtroperin=PersInvolucradas.objects.all().filter(hechos_id=idhecho)
                                                                                                     sumo=sumo+len(filtroperin)

                                                                                                     for datos in filtroperin:

                                                                                                            if datos.persona.fecha_nac:
                                                                                                                if descripcion.descripcion.find('ROBO')>=0:
                                                                                                                     robos=robos+1
                                                                                                                     sexor=datos.persona.sexo_id

                                                                                                                     if sexor:
                                                                                                                            sexr=sexor.descripcion
                                                                                                                            if sexor.descripcion.find('FEME')>=0:
                                                                                                                                 sumorf+=1
                                                                                                                            else:
                                                                                                                                 if sexor.descripcion.find('MASCU')>=0:
                                                                                                                                        sumorm+=1
                                                                                                                     else:
                                                                                                                                 sumosinr+=1

                                                                                                                else:
                                                                                                                    if descripcion.descripcion.find('HOMI')>=0:
                                                                                                                         homi=homi+1
                                                                                                                         sexo=datos.persona.sexo_id

                                                                                                                         if sexo:
                                                                                                                                 sex=sexo.descripcion
                                                                                                                                 if sexo.descripcion.find('FEME')>=0:
                                                                                                                                        sumof+=1
                                                                                                                                 else:
                                                                                                                                     if sexo.descripcion.find('MASCU')>=0:
                                                                                                                                         sumom+=1
                                                                                                                         else:
                                                                                                                                         sumosin+=1

                                                                                                                    else:
                                                                                                                        if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                             hurtos=hurtos+1
                                                                                                                        else:
                                                                                                                             otros=otros+1

                                                                                                                years=int(datetime.date.today().strftime('%Y'))-int(datos.persona.fecha_nac.strftime('%Y'))
                                                                                                                if years>=5:
                                                                                                                    if years>=5 and years<=15:
                                                                                                                         rangom.append(str(descripcion))
                                                                                                                         if sexo:
                                                                                                                             rangoms.append(str(sexo))
                                                                                                                         if sexor:
                                                                                                                             rangomsr.append(str(sexor))
                                                                                                                    else:
                                                                                                                         if years>=16 and years<=25:
                                                                                                                             rango.append(str(descripcion))
                                                                                                                             if sexo:
                                                                                                                                 rangos.append(str(sexo))
                                                                                                                             if sexor:
                                                                                                                                 rangosr.append(str(sexor))
                                                                                                                         else:
                                                                                                                             if years>=26 and years<=35:
                                                                                                                                 rango1.append(str(descripcion))
                                                                                                                                 if sexo:
                                                                                                                                     rangos1.append(str(sexo))
                                                                                                                                 if sexor:
                                                                                                                                     rangosr1.append(str(sexor))
                                                                                                                             else:
                                                                                                                                 if years>=36 and years<=45:
                                                                                                                                    rango2.append(str(descripcion))
                                                                                                                                    if sexo:
                                                                                                                                        rangos2.append(str(sexo))
                                                                                                                                    if sexor:
                                                                                                                                        rangosr2.append(str(sexor))
                                                                                                                                 else:
                                                                                                                                        if years>=46 and years<=55:
                                                                                                                                            rango3.append(str(descripcion))
                                                                                                                                            if sexo:
                                                                                                                                                 rangos3.append(str(sexo))
                                                                                                                                            if sexor:
                                                                                                                                                rangosr3.append(str(sexor))
                                                                                                                                        else:
                                                                                                                                             if years>=56 and years<=65:
                                                                                                                                                 rango4.append(str(descripcion))
                                                                                                                                                 if sexo:
                                                                                                                                                    rangos4.append(str(sexo))
                                                                                                                                                 if sexor:
                                                                                                                                                    rangosr4.append(str(sexor))
                                                                                                                                             else:
                                                                                                                                                 if years>=66 and years<=75:
                                                                                                                                                        rango5.append(str(descripcion))
                                                                                                                                                        if sexo:
                                                                                                                                                            rangos5.append(str(sexo))
                                                                                                                                                        if sexor:
                                                                                                                                                            rangosr5.append(str(sexor))
                                                                                                                                                 else:
                                                                                                                                                        rango6.append(str(descripcion))
                                                                                                                                                        if sexo:
                                                                                                                                                            rangos6.append(str(sexo))
                                                                                                                                                        if sexor:
                                                                                                                                                            rangosr6.append(str(sexor))



                                                                                                                else:
                                                                                                                     rango7.append(str(descripcion))
                                                                                                                     if sexo:
                                                                                                                         rangos7.append(str(sexo))
                                                                                                                     if sexor:
                                                                                                                         rangosr7.append(str(sexor))

    matriz=''
    dictms={}
    valores={}
    dato={}
    dada={}
    datas={}
    rob={}
    robar={}
    tsexofr=0
    tsexomr=0
    tsexosr=0
    tsexof=0
    tsexom=0
    tsexos=0


    if rango7:
        for i in rango7:
             if i.find('HOMI')>=0:

                    if 'MASCULINO' in rangos7:
                            tsexom+=1
                    else:
                            if 'FEMENINO' in rangos7:
                                 tsexof+=1
                            else:
                                 tsexos+=1
        dada['00-05']=[tsexom,tsexof,tsexos]
        datas['00-05']=['00-05',tsexom,tsexof,tsexos]

        for i in rango7:
             if i.find('ROBO')>=0:

                    if 'MASCULINO' in rangosr7:
                            tsexomr+=1
                    else:
                            if 'FEMENINO' in rangosr7:
                                 tsexofr+=1
                            else:
                                 tsexosr+=1
        rob['00-05']=[tsexomr,tsexofr,tsexosr]
        robar['00-05']=['00-05',tsexomr,tsexofr,tsexosr]

        for i in rango7:
            if i.find('ROBO')>=0:
                 trobos=trobos+1
            else:
                if i.find('HOMI')>=0:
                     thomi=thomi+1

                else:
                    if i.find('HURTO')>=0:
                         thurtos=thurtos+1
                    else:
                         if i.find('ROBO')<0 or i.find('HOMI')<0 or  i.find('HURTO')<0:

                                totros=totros+1

        a='  '+' Robos : '+str(trobos)+' -- '+' Homicidios : '+str(thomi)+' -- '+' Hurtos :'+str(thurtos)+' -- '+' Otros : '+str(totros)
        dictms.update({' 00 - 05 ':str(a)})

        valores['00-05']=['00-05',trobos,thomi,thurtos,totros]
        dato['00-05']=[trobos,thomi,thurtos,totros]



    trobos=0
    thomi=0
    thurtos=0
    totros=0
    tsexof=0
    tsexom=0
    tsexos=0
    tsexofr=0
    tsexomr=0
    tsexosr=0
    if rangom:
        for i in rangom:
             if i.find('HOMI')>=0:

                    if 'MASCULINO' in rangoms:
                            tsexom+=1
                    else:
                            if 'FEMENINO' in rangoms:
                                 tsexof+=1
                            else:
                                 tsexos+=1

        dada['05-15']=[tsexom,tsexof,tsexos]
        datas['05-15']=['05-15',tsexom,tsexof,tsexos]

        for i in rangom:
             if i.find('ROBO')>=0:

                    if 'MASCULINO' in rangomsr:
                            tsexomr+=1
                    else:
                            if 'FEMENINO' in rangomsr:
                                 tsexofr+=1
                            else:
                                 tsexosr+=1
        rob['05-15']=[tsexomr,tsexofr,tsexosr]
        robar['05-15']=['00-05',tsexomr,tsexofr,tsexosr]

        for i in rangom:

            if i.find('ROBO')>=0:
                 trobos=trobos+1

            else:
                if i.find('HOMI')>=0:
                     thomi=thomi+1

                else:
                    if i.find('HURTO')>=0:
                         thurtos=thurtos+1

                    else:
                         totros=totros+1


        a='  '+' Robos : '+str(trobos)+' -- '+' Homicidios : '+str(thomi)+' -- '+' Hurtos :'+str(thurtos)+' -- '+' Otros : '+str(totros)
        dictms.update({'05 - 15 ':str(a)})

        valores['05-15']=['05-15',trobos,thomi,thurtos,totros]
        dato['05-15']=[trobos,thomi,thurtos,totros]

    trobos=0
    thomi=0
    thurtos=0
    totros=0
    tsexof=0
    tsexom=0
    tsexos=0
    tsexofr=0
    tsexomr=0
    tsexosr=0
    if rango:
        for i in rango:
             if i.find('HOMI')>=0:

                    if 'MASCULINO' in rangos:
                            tsexom+=1
                    else:
                            if 'FEMENINO' in rangos:
                                 tsexof+=1
                            else:
                                 tsexos+=1
        dada['16-25']=[tsexom,tsexof,tsexos]
        datas['16-25']=['16-25',tsexom,tsexof,tsexos]

        for i in rango:
             if i.find('ROBO')>=0:

                    if 'MASCULINO' in rangosr:
                            tsexomr+=1
                    else:
                            if 'FEMENINO' in rangosr:
                                 tsexofr+=1
                            else:
                                 tsexosr+=1
        rob['16-25']=[tsexomr,tsexofr,tsexosr]
        robar['16-25']=['16-25',tsexomr,tsexofr,tsexosr]

        for i in rango:
            if i.find('ROBO')>=0:
                 trobos=trobos+1
            else:
                if i.find('HOMI')>=0:
                     thomi=thomi+1

                else:
                    if i.find('HURTO')>=0:
                         thurtos=thurtos+1
                    else:
                         if i.find('ROBO')<0 or i.find('HOMI')<0 or  i.find('HURTO')<0:
                                totros=totros+1

        a='  '+' Robos : '+str(trobos)+' -- '+' Homicidios : '+str(thomi)+' -- '+' Hurtos :'+str(thurtos)+' -- '+' Otros : '+str(totros)
        dictms.update({'16 - 25 ':str(a)})
        valores['16-25']=['16-25',trobos,thomi,thurtos,totros]
        dato['16-25']=[trobos,thomi,thurtos,totros]



    trobos=0
    thomi=0
    thurtos=0
    totros=0
    tsexof=0
    tsexom=0
    tsexos=0
    tsexofr=0
    tsexomr=0
    tsexosr=0
    if rango1:

        for i in rango1:
             if i.find('HOMI')>=0:

                    if 'MASCULINO' in rangos1:
                            tsexom+=1
                    else:
                            if 'FEMENINO' in rangos1:
                                 tsexof+=1
                            else:
                                 tsexos+=1
        dada['26-35']=[tsexom,tsexof,tsexos]
        datas['26-35']=['26-35',tsexom,tsexof,tsexos]

        for i in rango1:
             if i.find('ROBO')>=0:

                    if 'MASCULINO' in rangosr1:
                            tsexomr+=1
                    else:
                            if 'FEMENINO' in rangosr1:
                                 tsexofr+=1
                            else:
                                 tsexosr+=1
        rob['26-35']=[tsexomr,tsexofr,tsexosr]
        robar['26-35']=['26-35',tsexomr,tsexofr,tsexosr]

        for i in rango1:
            if i.find('ROBO')>=0:
                 trobos=trobos+1
            else:
                if i.find('HOMI')>=0:
                     thomi=thomi+1

                else:
                    if i.find('HURTO')>=0:
                         thurtos=thurtos+1
                    else:
                         if i.find('ROBO')<0 or i.find('HOMI')<0 or  i.find('HURTO')<0:

                                totros=totros+1
        a='  '+' Robos : '+str(trobos)+' -- '+' Homicidios : '+str(thomi)+' -- '+' Hurtos :'+str(thurtos)+' -- '+' Otros : '+str(totros)
        dictms.update({'26 - 35 ':str(a)})
        valores['26-35']=['26-35',trobos,thomi,thurtos,totros]
        dato['26-35']=[trobos,thomi,thurtos,totros]

    trobos=0
    thomi=0
    thurtos=0
    totros=0
    tsexof=0
    tsexom=0
    tsexos=0
    tsexofr=0
    tsexomr=0
    tsexosr=0
    if rango2:
        for i in rango2:
             if i.find('HOMI')>=0:

                    if 'MASCULINO' in rangos2:
                            tsexom+=1
                    else:
                            if 'FEMENINO' in rangos2:
                                 tsexof+=1
                            else:
                                 tsexos+=1
        dada['36-45']=[tsexom,tsexof,tsexos]
        datas['36-45']=['36-45',tsexom,tsexof,tsexos]

        for i in rango2:
             if i.find('ROBO')>=0:

                    if 'MASCULINO' in rangosr2:
                            tsexomr+=1
                    else:
                            if 'FEMENINO' in rangosr2:
                                 tsexofr+=1
                            else:
                                 tsexosr+=1
        rob['36-45']=[tsexomr,tsexofr,tsexosr]
        robar['36-45']=['36-45',tsexomr,tsexofr,tsexosr]

        for i in rango2:
            if i.find('ROBO')>=0:
                 trobos=trobos+1
            else:
                if i.find('HOMI')>=0:
                     thomi=thomi+1

                else:
                    if i.find('HURTO')>=0:
                         thurtos=thurtos+1
                    else:
                         if i.find('ROBO')<0 or i.find('HOMI')<0 or  i.find('HURTO')<0:

                                totros=totros+1
        a='  '+' Robos : '+str(trobos)+' -- '+' Homicidios : '+str(thomi)+' -- '+' Hurtos :'+str(thurtos)+' -- '+' Otros : '+str(totros)
        dictms.update({'36 - 45 ':str(a)})
        valores['36-45']=['36-45',trobos,thomi,thurtos,totros]
        dato['36-45']=[trobos,thomi,thurtos,totros]
    trobos=0
    thomi=0
    thurtos=0
    totros=0
    tsexof=0
    tsexom=0
    tsexos=0
    tsexofr=0
    tsexomr=0
    tsexosr=0
    if rango3:
        for i in rango3:
             if i.find('HOMI')>=0:

                    if 'MASCULINO' in rangos3:
                            tsexom+=1
                    else:
                            if 'FEMENINO' in rangos3:
                                 tsexof+=1
                            else:
                                 tsexos+=1
        dada['46-55']=[tsexom,tsexof,tsexos]
        datas['46-55']=['46-55',tsexom,tsexof,tsexos]

        for i in rango3:
             if i.find('ROBO')>=0:

                    if 'MASCULINO' in rangosr3:
                            tsexomr+=1
                    else:
                            if 'FEMENINO' in rangosr3:
                                 tsexofr+=1
                            else:
                                 tsexosr+=1
        rob['46-55']=[tsexomr,tsexofr,tsexosr]
        robar['46-55']=['46-55',tsexomr,tsexofr,tsexosr]

        for i in rango3:
            if i.find('ROBO')>=0:
                 trobos=trobos+1
            else:
                if i.find('HOMI')>=0:
                     thomi=thomi+1
                else:
                    if i.find('HURTO')>=0:
                         thurtos=thurtos+1
                    else:
                         if i.find('ROBO')<0 or i.find('HOMI')<0 or  i.find('HURTO')<0:

                                totros=totros+1
        a='  '+' Robos : '+str(trobos)+' -- '+' Homicidios : '+str(thomi)+' -- '+' Hurtos :'+str(thurtos)+' -- '+' Otros : '+str(totros)
        dictms.update({'46 - 55 ':str(a)})
        valores['46-55']=['46-55',trobos,thomi,thurtos,totros]
        dato['46-55']=[trobos,thomi,thurtos,totros]
    trobos=0
    thomi=0
    thurtos=0
    totros=0
    tsexof=0
    tsexom=0
    tsexos=0
    tsexofr=0
    tsexomr=0
    tsexosr=0
    if rango4:
        for i in rango4:
             if i.find('HOMI')>=0:

                    if 'MASCULINO' in rangos4:
                            tsexom+=1
                    else:
                            if 'FEMENINO' in rangos4:
                                 tsexof+=1
                            else:
                                 tsexos+=1
        dada['56-66']=[tsexom,tsexof,tsexos]
        datas['56-65']=['56-65',tsexom,tsexof,tsexos]

        for i in rango4:
             if i.find('ROBO')>=0:

                    if 'MASCULINO' in rangosr4:
                            tsexomr+=1
                    else:
                            if 'FEMENINO' in rangosr4:
                                 tsexofr+=1
                            else:
                                 tsexosr+=1
        rob['56-65']=[tsexomr,tsexofr,tsexosr]
        robar['56-65']=['56-65',tsexomr,tsexofr,tsexosr]

        for i in rango4:
            if i.find('ROBO')>=0:
                 trobos=trobos+1
            else:
                if i.find('HOMI')>=0:
                     thomi=thomi+1

                else:
                    if i.find('HURTO')>=0:
                         thurtos=thurtos+1
                    else:
                         if i.find('ROBO')<0 or i.find('HOMI')<0 or  i.find('HURTO')<0:

                                totros=totros+1
        a='  '+' Robos : '+str(trobos)+' -- '+' Homicidios : '+str(thomi)+' -- '+' Hurtos :'+str(thurtos)+' -- '+' Otros : '+str(totros)
        dictms.update({'56 - 66 ':str(a)})
        valores['56-66']=['56-66',trobos,thomi,thurtos,totros]
        dato['56-65']=[trobos,thomi,thurtos,totros]

    trobos=0
    thomi=0
    thurtos=0
    totros=0
    tsexof=0
    tsexom=0
    tsexos=0
    tsexofr=0
    tsexomr=0
    tsexosr=0
    if rango5:
        for i in rango5:
             if i.find('HOMI')>=0:

                    if 'MASCULINO' in rangos5:
                            tsexom+=1
                    else:
                            if 'FEMENINO' in rangos5:
                                 tsexof+=1
                            else:
                                 tsexos+=1
        dada['66-75']=[tsexom,tsexof,tsexos]
        datas['66-75']=['66-75',tsexom,tsexof,tsexos]

        for i in rango5:
             if i.find('ROBO')>=0:

                    if 'MASCULINO' in rangosr5:
                            tsexomr+=1
                    else:
                            if 'FEMENINO' in rangosr5:
                                 tsexofr+=1
                            else:
                                 tsexosr+=1
        rob['66-75']=[tsexomr,tsexofr,tsexosr]
        robar['66-75']=['66-75',tsexomr,tsexofr,tsexosr]

        for i  in rango5:
            if i.find('ROBO')>=0:
                 trobos=trobos+1
            else:
                if i.find('HOMI')>=0:
                     thomi=thomi+1
                else:
                    if i.find('HURTO')>=0:
                         thurtos=thurtos+1
                    else:
                         if i.find('ROBO')<0 or i.find('HOMI')<0 or  i.find('HURTO')<0:

                                totros=totros+1

        a='  '+' Robos : '+str(trobos)+' -- '+' Homicidios : '+str(thomi)+' -- '+' Hurtos :'+str(thurtos)+' -- '+' Otros : '+str(totros)
        dictms.update({'66 - 75 ':str(a)})
        valores['66-75']=['66-75',trobos,thomi,thurtos,totros]
        dato['66-75']=[trobos,thomi,thurtos,totros]
    trobos=0
    thomi=0
    thurtos=0
    totros=0
    tsexof=0
    tsexom=0
    tsexos=0
    tsexofr=0
    tsexomr=0
    tsexosr=0
    if rango6:
        for i in rango6:
             if i.find('HOMI')>=0:

                    if 'MASCULINO' in rangos6:
                            tsexom+=1
                    else:
                            if 'FEMENINO' in rangos6:
                                 tsexof+=1
                            else:
                                 tsexos+=1
        dada['76-99']=[tsexom,tsexof,tsexos]
        datas['76-99']=['76-99',tsexom,tsexof,tsexos]

        for i in rango6:
             if i.find('ROBO')>=0:

                    if 'MASCULINO' in rangosr6:
                            tsexomr+=1
                    else:
                            if 'FEMENINO' in rangosr6:
                                 tsexofr+=1
                            else:
                                 tsexosr+=1
        rob['76-99']=[tsexomr,tsexofr,tsexosr]
        robar['76-99']=['76-99',tsexomr,tsexofr,tsexosr]
        for i in rango6:
            if i.find('ROBO')>=0:
                 trobos=trobos+1
            else:
                if i.find('HOMI')>=0:
                     thomi=thomi+1

                else:
                    if i.find('HURTO')>=0:
                         thurtos=thurtos+1
                    else:
                         if i.find('ROBO')<0 or i.find('HOMI')<0 or  i.find('HURTO')<0:

                                totros=totros+1

        a='  '+' Robos : '+str(trobos)+' -- '+' Homicidios : '+str(thomi)+' -- '+' Hurtos :'+str(thurtos)+' -- '+' Otros : '+str(totros)
        dictms.update({' 76 - 99':str(a)})
        valores['76-99']=['76-99',trobos,thomi,thurtos,totros]
        dato['76-99']=[trobos,thomi,thurtos,totros]

    matriz=' Robos : '+str(robos)+' -'+' Homicidios : '+str(homi)+' -'+' Hurtos : '+str(hurtos)+' -'+' Otros Delitos : '+str(otros)+'  >>> '+str(sumo)+' personas involucradas'
    matrix=[]

    for key,value in sorted(dictms.iteritems()):
        reporte=True
        rephomisexo=True
        reprobosexo=True
        matrix.append(key+':'+value)


    key=valores.keys()

    arreglo=[]



    if request.POST.get('expo')=="Exportar":


        # se crea el libro y se obtiene la hoja


        libro = Workbook()
        hoja = libro.get_active_sheet()
        hoja.title = "Personas Involucradas"

        # Ahora, se obtiene las celdas en la cuál se colocará el nombre
        # del campo. como son 8 campos, se necesita 8 celdas
        celda = hoja.cell("B1")
        celda.value=" Elementos recopilados desde la Base de Datos para Estadísticas en el ámbito de "+str(lugar)
        celda = hoja.cell("B2")
        celda.value=" Datos del Periodo de Fecha : MES / AÑO  "+str(fechadesde)+'/'+str(fechahasta)
        celda = hoja.cell("B3")
        celda.value=" Tabla con datos obtenidos de Personas involucradas discriminadas por Delitos y Rango de Edades "



        rango_celdas = hoja.range("B4:F4")

        # se crea una tupla con los nombres de los campos

        nombre_campos = "Rango de Edades", "ROBOS", "HOMICIDIOS", "HURTOS", "OTROS DELITOS"
        # ahora, se asigna cada nombre de campo a cada celda

        for campo in rango_celdas:
                indice = 0  # se crea un contador para acceder a la tupla
                for celda in campo:
                        celda.value = nombre_campos[indice]
                        indice += 1

        # ya se tiene los nombres de los campos
        # ahora se obtiene el rango de celdas en donde irán los datos

        longitud = 4 + len(valores)


        celdas_datos = hoja.range("B5:F{0}".format(longitud))
        # ahora vamos a dar los valores a las celdas con los datos

        fila=0

        for key,value in sorted(valores.iteritems()):

         datos=(valores[key])
         indice=0
         for celda in celdas_datos[fila]:
                 celda.value = datos[indice]
                 indice+=1
         fila += 1

         # se crea un objeto httpresponse y se pasa como parámetro el content_type
         # diciendo que es excel
        response = HttpResponse(content_type="application/ms-excel")  # HttpResponse viene del modulo django.http
        nombre_archivo = "persoinvo.xlsx"
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        libro.save(response)
        return response

    if request.POST.get('expo1')=="Exportar":


        # se crea el libro y se obtiene la hoja


        libro = Workbook()
        hoja = libro.get_active_sheet()
        hoja.title = "Pers_Inv_Homicidios"

        # Ahora, se obtiene las celdas en la cuál se colocará el nombre
        # del campo. como son 8 campos, se necesita 8 celdas
        celda = hoja.cell("B1")
        celda.value=" Elementos recopilados desde la Base de Datos para Estadísticas en el ámbito de "+str(lugar)
        celda = hoja.cell("B2")
        celda.value=" Datos del Periodo de Fecha : MES / AÑO  "+str(fechadesde)+'/'+str(fechahasta)
        celda = hoja.cell("B3")
        celda.value=" Tabla con datos obtenidos de Personas involucradas en Homicidios por Rango de Edad y Sexo "



        rango_celdas = hoja.range("B4:E4")

        # se crea una tupla con los nombres de los campos

        nombre_campos = "Rango de Edades", "MASCULINOS", "FEMENINOS", "SIN DATOS"
        # ahora, se asigna cada nombre de campo a cada celda

        for campo in rango_celdas:
                indice = 0  # se crea un contador para acceder a la tupla
                for celda in campo:
                        celda.value = nombre_campos[indice]
                        indice += 1

        # ya se tiene los nombres de los campos
        # ahora se obtiene el rango de celdas en donde irán los datos

        longitud1 = 4 + len(datas)


        celdas_datos1 = hoja.range("B5:E{0}".format(longitud1))
        # ahora vamos a dar los valores a las celdas con los datos

        fila1=0

        for key,value in sorted(datas.iteritems()):

         datos1=(datas[key])

         indice1=0

         for celda in celdas_datos1[fila1]:
                 celda.value = datos1[indice1]
                 indice1+=1
         fila1 += 1


         # se crea un objeto httpresponse y se pasa como parámetro el content_type
         # diciendo que es excel
        response = HttpResponse(content_type="application/ms-excel")  # HttpResponse viene del modulo django.http
        nombre_archivo = "persinvhom.xlsx"
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        libro.save(response)
        return response

    if request.POST.get('expor')=="Exportar":


        # se crea el libro y se obtiene la hoja


        libro = Workbook()
        hoja = libro.get_active_sheet()
        hoja.title = "Pers_Inv_Robos"

        # Ahora, se obtiene las celdas en la cuál se colocará el nombre
        # del campo. como son 8 campos, se necesita 8 celdas
        celda = hoja.cell("B1")
        celda.value=" Elementos recopilados desde la Base de Datos para Estadísticas en el ámbito de "+str(lugar)
        celda = hoja.cell("B2")
        celda.value=" Datos del Periodo de Fecha : MES / AÑO  "+str(fechadesde)+'/'+str(fechahasta)
        celda = hoja.cell("B3")
        celda.value=" Tabla con datos obtenidos de Personas involucradas en Robos por Rango de Edad y Sexo "



        rango_celdas = hoja.range("B4:E4")

        # se crea una tupla con los nombres de los campos

        nombre_campos = "Rango de Edades", "MASCULINOS", "FEMENINOS", "SIN DATOS"
        # ahora, se asigna cada nombre de campo a cada celda

        for campo in rango_celdas:
                indice = 0  # se crea un contador para acceder a la tupla
                for celda in campo:
                        celda.value = nombre_campos[indice]
                        indice += 1

        # ya se tiene los nombres de los campos
        # ahora se obtiene el rango de celdas en donde irán los datos

        longitud1 = 4 + len(robar)


        celdas_datos1 = hoja.range("B5:E{0}".format(longitud1))
        # ahora vamos a dar los valores a las celdas con los datos

        fila1=0

        for key,value in sorted(robar.iteritems()):

         datos1=(robar[key])

         indice1=0

         for celda in celdas_datos1[fila1]:
                 celda.value = datos1[indice1]
                 indice1+=1
         fila1 += 1


         # se crea un objeto httpresponse y se pasa como parámetro el content_type
         # diciendo que es excel
        response = HttpResponse(content_type="application/ms-excel")  # HttpResponse viene del modulo django.http
        nombre_archivo = "persinvrobos.xlsx"
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        libro.save(response)
        return response

    form = MapaForm()
    anios =[]
    meses={}
    anios={}

    ayer =datetime.date.today().year-4
    hoy = datetime.date.today().year
    rango=(hoy-ayer)+1
    for i in range(rango):
         if ayer<=hoy:
             a=ayer+i
             anios[i]=a



    for i in range(12):

        i+=1

        meses[i]=calendar.month_name[i]

    values={'reprobosexo':reprobosexo,'rob':rob,'dada':dada,'rephomisexo':rephomisexo,'anios':anios,'meses':meses,'fechadesde':fechadesde,'fechahasta':fechahasta,'dato':dato,'ciudad':ciudad,'depes':depes,'ureg':ureg,'destino': destino,'state':state,'form':form,'reporte':reporte,'dictms':dictms,'matriz':matriz,'matrix':matrix,'lugar':lugar,'valores':valores,'arreglo':arreglo,}
    return render(request,'./repoestadis.html',values)

#definicion de los reposrtes estadisticos
@login_required
@group_required(["policia","investigaciones","radio"])
def rephechos(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors=[]
    rango=[]
    rangot={}
    tothechos=0
    robos=[]
    homi=0
    hurtos=0
    otros=0
    injusticia=''
    delitosin=''
    rangoip=[]
    rangooj=[]
    rangoot=[]

    indice=0
    sumo=0
    saludo=''
    anios=0
    rangof=[]

    rephec=False
    tot=0
    veces=0
    veces1=0
    veces2=0
    veces3=0
    indice=0
    lista={}
    ciudad=''
    ureg=''
    depes=''
    lugar=''
    fechadesde=0
    fechahasta=0
    mesnro=0
    delitos={}
    vecesd=0
    dictdp={}
    dictip={}
    dictoj={}
    dictot={}

    if request.POST.get('ver')=="Visualizar" or request.POST.get('expo')=="Exportar":
            datosform=request.POST

            if request.POST.get('expo')=="Exportar":

             ciudad=request.POST.get('ciu')
             ureg=request.POST.get('ure')
             depes=request.POST.get('dep')
             fechadesde=request.POST.get('mesi')
             fechahasta=request.POST.get('aniosi')

            else:


             ciudad = datosform.__getitem__('ciudades')
             ureg = datosform.__getitem__('ureg')
             depes= datosform.__getitem__('depe')
             fechadesde=request.POST.get('mes')
             fechahasta=request.POST.get('anios')


            filtrorefd=Hechos.objects.all()



            if ciudad:

                lugar=RefCiudades.objects.get(id=ciudad)
                depes=Dependencias.objects.filter(ciudad_id=ciudad).values('id')
                if depes:
                     for dp in depes:
                             depes=dp['id']
                     preventivo=Dependencias.objects.get(id=depes)
                     depes=Dependencias.objects.filter(ciudad_id=ciudad).values('id')

                     if depes:
                                    for crias in depes:
                                            dp=crias['id']

                                            preventivos=Preventivos.objects.all().filter(dependencia=dp,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()

                                            for desde in preventivos:
                                                    idp=desde['id']

                                                    idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                    if idhecho:

                                                         for ids in idhecho:
                                                                 idprevh=ids['preventivo_id']
                                                                 if idprevh==idp:
                                                                        idhecho=ids['id']
                                                                        motivo=ids['motivo_id']
                                                                        tipohe=RefMotivosHecho.objects.get(id=motivo)
                                                                        tipohecho=tipohe.descripcion
                                                                        tothechos=tothechos+1
                                                                        filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')

                                                                        for indelis in filtrodelih:
                                                                                descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])
                                                                                if tipohecho.find('DENUNCIA')>=0:
                                                                                     veces=veces+1
                                                                                     rango.append(descripcion.descripcion)


                                                                                else:
                                                                                    if tipohecho.find('INTER')>=0:
                                                                                         veces1=veces1+1
                                                                                         rangoip.append(descripcion.descripcion)

                                                                                    else:
                                                                                         if tipohecho.find('ORDEN')>=0:
                                                                                                veces2=veces2+1
                                                                                                rangooj.append(descripcion.descripcion)

                                                                                         else:
                                                                                                veces3=veces3+1
                                                                                                rangoot.append(descripcion.descripcion)


                                                                                lista[descripcion.descripcion]='delitos'




            else:
                if depes:

                                            preventivo=Dependencias.objects.get(id=depes)
                                            ciudad=""
                                            #preventivo.ciudad.descripcion
                                            depe=preventivo.descripcion
                                            unireg=preventivo.unidades_regionales.descripcion
                                            lugar=ciudad+' -- '+unireg+' -- '+depe

                                            preventivos=Preventivos.objects.all().filter(dependencia=depes,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()
                                            for desde in preventivos:
                                                    idp=desde['id']

                                                    idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                    if idhecho:
                                                         for ids in idhecho:
                                                                 idprevh=ids['preventivo_id']
                                                                 if idprevh==idp:
                                                                            idhecho=ids['id']
                                                                            motivo=ids['motivo_id']
                                                                            tipohe=RefMotivosHecho.objects.get(id=motivo)
                                                                            tipohecho=tipohe.descripcion
                                                                            tothechos=tothechos+1
                                                                            filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')
                                                                            for indelis in filtrodelih:
                                                                                descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])
                                                                                if tipohecho.find('DENUNCIA')>=0:
                                                                                     veces=veces+1
                                                                                     rango.append(descripcion.descripcion)


                                                                                else:
                                                                                    if tipohecho.find('INTER')>=0:
                                                                                         veces1=veces1+1
                                                                                         rangoip.append(descripcion.descripcion)

                                                                                    else:
                                                                                         if tipohecho.find('ORDEN')>=0:
                                                                                                veces2=veces2+1
                                                                                                rangooj.append(descripcion.descripcion)

                                                                                         else:
                                                                                                veces3=veces3+1
                                                                                                rangoot.append(descripcion.descripcion)

                                                                                lista[descripcion.descripcion]='delitos'



                else:

                    lugar="LA PROVINCIA DE CHUBUT"
                    ciudad=''
                    idprovincia=RefProvincia.objects.get(descripcion__icontains='CHUBUT')
                    provis=RefCiudades.objects.all().filter(provincia_id=idprovincia.id).values('id')
                    for ciuda in provis:
                                de=Dependencias.objects.filter(ciudad_id=ciuda['id']).values('id')
                                if de:
                                 for dp in de:
                                         depe=dp['id']
                                 preventivo=Dependencias.objects.get(id=depe)
                                 dep=Dependencias.objects.filter(ciudad_id=ciuda['id']).values('id')

                                 if dep:
                                                for crias in dep:
                                                        dp=crias['id']

                                                        preventivos=Preventivos.objects.all().filter(dependencia=dp,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()

                                                        for desde in preventivos:
                                                                idp=desde['id']

                                                                idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                                if idhecho:
                                                                     for ids in idhecho:
                                                                             idprevh=ids['preventivo_id']
                                                                             if idprevh==idp:
                                                                                        idhecho=ids['id']
                                                                                        motivo=ids['motivo_id']
                                                                                        tipohe=RefMotivosHecho.objects.get(id=motivo)
                                                                                        tipohecho=tipohe.descripcion
                                                                                        tothechos=tothechos+1
                                                                                        filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')
                                                                                        for indelis in filtrodelih:
                                                                                                descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])
                                                                                                if tipohecho.find('DENUNCIA')>=0:
                                                                                                     veces=veces+1
                                                                                                     rango.append(descripcion.descripcion)


                                                                                                else:
                                                                                                    if tipohecho.find('INTER')>=0:
                                                                                                         veces1=veces1+1
                                                                                                         rangoip.append(descripcion.descripcion)

                                                                                                    else:
                                                                                                         if tipohecho.find('ORDEN')>=0:
                                                                                                                veces2=veces2+1
                                                                                                                rangooj.append(descripcion.descripcion)

                                                                                                         else:
                                                                                                                veces3=veces3+1
                                                                                                                rangoot.append(descripcion.descripcion)

                                                                                                lista[descripcion.descripcion]='delitos'
    if rango:

         lstr=repetidos(rango)
         lstc=contar(rango,lstr)
         for key,value in sorted(lstc.iteritems()):
            dictdp[key]=value

    if rangoip:
         lstr=repetidos(rangoip)
         lstc=contar(rangoip,lstr)
         for key,value in sorted(lstc.iteritems()):
            dictip[key]=value

    if rangooj:
         lstr=repetidos(rangooj)
         lstc=contar(rangooj,lstr)
         for key,value in sorted(lstc.iteritems()):
            dictoj[key]=value

    if rangoot:
         lstr=repetidos(rangoot)
         lstc=contar(rangoot,lstr)
         for key,value in sorted(lstc.iteritems()):
            dictot[key]=value

    if lista:

         rephec=True
         for key,value in sorted(lista.iteritems()):
                 lstr=repetidos(lista)
                 delitos[key]=value










    if request.POST.get('expo')=="Exportar":


        # se crea el libro y se obtiene la hoja


        libro = Workbook()
        hoja = libro.get_active_sheet()
        hoja.title = "Cantidad de Hechos"

        # Ahora, se obtiene las celdas en la cuál se colocará el nombre
        # del campo. como son 8 campos, se necesita 8 celdas
        celda = hoja.cell("B1")
        celda.value=" Elementos recopilados desde la Base de Datos para Estadísticas en el ámbito de "+str(lugar)
        celda = hoja.cell("B2")
        celda.value=" Datos del Periodo de Fecha : MES / AÑO  "+str(fechadesde)+'/'+str(fechahasta)
        celda = hoja.cell("B3")
        celda.value=" Tabla con datos obtenidos de Hechos Delictivos según Tipos de Hechos y Delitos "



        rango_celdas = hoja.range("B4:F4")

        # se crea una tupla con los nombres de los campos

        nombre_campos = "Delitos", "DENUNCIA POLICIAL", "INTERVENCION POLICIAL","ORDEN JUDICIAL", "OTROS TIPOS"
        # ahora, se asigna cada nombre de campo a cada celda

        for campo in rango_celdas:
                indice = 0  # se crea un contador para acceder a la tupla
                for celda in campo:
                        celda.value = nombre_campos[indice]
                        indice += 1

        # ya se tiene los nombres de los campos
        # ahora se obtiene el rango de celdas en donde irán los datos

        longitud = 4 + len(delitos)


        celdas_datos = hoja.range("B5:B{0}".format(longitud))
        # ahora vamos a dar los valores a las celdas con los datos

        fila=0
        filad=0
        paso=0
        for key,value in delitos.items():

         datos=key

         indice=0
         for celda in celdas_datos[fila]:
                 celda.value = datos


         fila += 1


        celdas_datosd = hoja.range("C5:C{0}".format(longitud))

        for key,value in delitos.items():

                if key in dictdp.keys():


                     for celda in celdas_datosd[filad]:
                            celda.value = dictdp[key]


                else:

                        for celda in celdas_datosd[filad]:
                            celda.value = 0
                filad+=1

        celdas_datosi = hoja.range("D5:D{0}".format(longitud))
        filai=0
        for key,value in delitos.items():

                if key in dictip.keys():


                     for celda in celdas_datosi[filai]:
                            celda.value = dictip[key]


                else:

                        for celda in celdas_datosi[filai]:
                            celda.value = 0
                filai+=1

        celdas_datoso = hoja.range("E5:E{0}".format(longitud))
        filao=0
        for key,value in delitos.items():

                if key in dictoj.keys():


                     for celda in celdas_datoso[filao]:
                            celda.value = dictoj[key]


                else:

                        for celda in celdas_datoso[filao]:
                            celda.value = 0
                filao+=1

        celdas_datost = hoja.range("F5:F{0}".format(longitud))
        filat=0
        for key,value in delitos.items():

                if key in dictot.keys():


                     for celda in celdas_datost[filat]:
                            celda.value = dictoj[key]


                else:

                        for celda in celdas_datost[filat]:
                            celda.value = 0
                filat+=1



         # se crea un objeto httpresponse y se pasa como parámetro el content_type
         # diciendo que es excel
        response = HttpResponse(content_type="application/ms-excel")  # HttpResponse viene del modulo django.http
        nombre_archivo = "canthechos.xlsx"
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        libro.save(response)
        return response

    form = MapaForm()
    anios =[]
    meses={}
    anios={}

    ayer =datetime.date.today().year-4
    hoy = datetime.date.today().year
    rango=(hoy-ayer)+1
    for i in range(rango):
         if ayer<=hoy:
             a=ayer+i
             anios[i]=a



    for i in range(12):

        i+=1

        meses[i]=calendar.month_name[i]

    values={'fechadesde':fechadesde,'fechahasta':fechahasta,'anios':anios,'meses':meses,'delitos':delitos,'dictdp':dictdp,'dictip':dictip,'dictoj':dictoj,'dictot':dictot,'ciudad':ciudad,'ureg':ureg,'depes':depes,'destino': destino,'state':state,'form':form,'matriz':tothechos,'rephec':rephec,'lugar':lugar,'rangot':rangot,}
    return render(request,'./rephechos.html',values)

#homicidios segun tipo de lugar
#definicion de los reposrtes estadisticos
@login_required
@group_required(["policia","investigaciones","radio"])
def killings(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors=[]
    rango=[]
    rangot={}
    tothechos=0
    robos=[]
    homi=0
    hurtos=0
    otros=0
    injusticia=''
    delitosin=''
    rangoip=[]
    rangooj=[]
    rangoot=[]

    indice=0
    sumo=0
    saludo=''
    anios=0
    rangof=[]

    canhec=False
    tot=0
    veces=0
    veces1=0
    veces2=0
    veces3=0
    indice=0
    lista={}
    ciudad=''
    ureg=''
    depes=''
    lugar=''
    fechadesde=0
    fechahasta=0
    mesnro=0
    vecesd=0
    dictch={}
    dictvi=[]
    dictim=[]
    dictot={}
    dictci={}
    sumovi=0
    sumoim=0
    sumolt=0
    sumoar=0
    sumosar=0
    dictar={}
    dictlt={}
    if request.POST.get('ver')=="Visualizar" or request.POST.get('expo')=="Exportar":
            datosform=request.POST

            if request.POST.get('expo')=="Exportar":

             ciudad=request.POST.get('ciu')
             ureg=request.POST.get('ure')
             depes=request.POST.get('dep')
             fechadesde=request.POST.get('mesi')
             fechahasta=request.POST.get('aniosi')

            else:


             ciudad = datosform.__getitem__('ciudades')
             ureg = datosform.__getitem__('ureg')
             depes= datosform.__getitem__('depe')
             fechadesde=request.POST.get('mes')
             fechahasta=request.POST.get('anios')


            filtrorefd=Hechos.objects.all()



            if ciudad:

                lugar=RefCiudades.objects.get(id=ciudad)
                depes=Dependencias.objects.filter(ciudad_id=ciudad).values('id')
                if depes:
                     for dp in depes:
                             depes=dp['id']
                     preventivo=Dependencias.objects.get(id=depes)
                     depes=Dependencias.objects.filter(ciudad_id=ciudad).values('id')

                     if depes:
                                    for crias in depes:
                                            dp=crias['id']

                                            preventivos=Preventivos.objects.all().filter(dependencia=dp,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()

                                            for desde in preventivos:
                                                    idp=desde['id']

                                                    idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                    if idhecho:

                                                         for ids in idhecho:
                                                                 idprevh=ids['preventivo_id']
                                                                 if idprevh==idp:
                                                                        idhecho=ids['id']

                                                                        filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')
                                                                        for indelis in filtrodelih:
                                                                            descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])
                                                                            if descripcion.descripcion.find('HOMI')>=0:
                                                                                #total de hechos donde hubo homicidios
                                                                                tothechos=tothechos+1
                                                                                #personas involucradas en homicidios rol de victima / imputado
                                                                                filtroperin=PersInvolucradas.objects.all().filter(hechos_id=idhecho).all()
                                                                                if filtroperin:
                                                                                     for datos in filtroperin:
                                                                                             if datos.roles.descripcion.find('VICTI')>=0:
                                                                                                    sumovi+=1
                                                                                                    dictvi.append(sumovi)
                                                                                             else:
                                                                                                    if datos.roles.descripcion.find('IMPUT')>=0:
                                                                                                         sumoim+=1
                                                                                                         dictim.append(sumoim)
                                                                                #tipo del lugar del homicidio.
                                                                                lugars=Hechos.objects.get(id=idhecho).lugar_hecho.values('tipo_lugar')
                                                                                sumolt=0
                                                                                if lugars:
                                                                                     descripcion=RefLugares.objects.get(id=lugars)
                                                                                     sumolt+=1
                                                                                     if descripcion in dictlt:
                                                                                            dictlt[descripcion]+=sumolt
                                                                                     else:
                                                                                            dictlt[descripcion]=sumolt
                                                                                #armas en el hecho
                                                                                elem=Hechos.objects.get(id=idhecho).eleinvolu.values('categoria')
                                                                                if elem:
                                                                                     for filas in elem:
                                                                                            categoria=RefCategory.objects.get(id=filas['categoria'])
                                                                                            if categoria.descripcion.find('DE FUEGO')>=0 or categoria.descripcion.find('BLANCA')>=0 or categoria.descripcion.find('IMPROPIA')>=0:
                                                                                                 sumoar+=1
                                                                                                 if categoria in dictar:
                                                                                                        dictar[categoria]+=sumoar
                                                                                                 else:
                                                                                                        dictar[categoria]=sumoar
                                                                                                 sumoar=0
                                                                                            break

                                                                                else:
                                                                                        sumosar+=1
                                                                                        dictar['Sin Armas']=sumosar






            else:
                if depes:

                                            preventivo=Dependencias.objects.get(id=depes)
                                            ciudad=""
                                            depe=preventivo.descripcion
                                            unireg=preventivo.unidades_regionales.descripcion
                                            lugar=ciudad+' -- '+unireg+' -- '+depe

                                            preventivos=Preventivos.objects.all().filter(dependencia=depes,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()
                                            for desde in preventivos:
                                                    idp=desde['id']

                                                    idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                    if idhecho:
                                                         for ids in idhecho:
                                                                 idprevh=ids['preventivo_id']
                                                                 if idprevh==idp:
                                                                            idhecho=ids['id']
                                                                            motivo=ids['motivo_id']

                                                                            filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')
                                                                            for indelis in filtrodelih:
                                                                                descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])
                                                                                if descripcion.descripcion.find('HOMI')>=0:
                                                                                 #total de hechos donde hubo homicidios
                                                                                 tothechos=tothechos+1
                                                                                 #personas involucradas en homicidios rol de victima / imputado
                                                                                 filtroperin=PersInvolucradas.objects.all().filter(hechos_id=idhecho).all()
                                                                                 if filtroperin:
                                                                                     for datos in filtroperin:
                                                                                             if datos.roles.descripcion.find('VICTI')>=0:
                                                                                                    sumovi+=1
                                                                                                    dictvi.append(sumovi)
                                                                                             else:
                                                                                                    if datos.roles.descripcion.find('IMPUT')>=0:
                                                                                                         sumoim+=1
                                                                                                         dictim.append(sumoim)
                                                                                 #tipo del lugar del homicidio.
                                                                                 lugars=Hechos.objects.get(id=idhecho).lugar_hecho.values('tipo_lugar')
                                                                                 sumolt=0
                                                                                 if lugars:
                                                                                     descripcion=RefLugares.objects.get(id=lugars)
                                                                                     sumolt+=1
                                                                                     if descripcion in dictlt:
                                                                                            dictlt[descripcion]+=sumolt
                                                                                     else:
                                                                                            dictlt[descripcion]=sumolt
                                                                                 #armas en el hecho
                                                                                 elem=Hechos.objects.get(id=idhecho).eleinvolu.values('categoria')
                                                                                 if elem:
                                                                                     for filas in elem:
                                                                                            categoria=RefCategory.objects.get(id=filas['categoria'])
                                                                                            if categoria.descripcion.find('DE FUEGO')>=0 or categoria.descripcion.find('BLANCA')>=0 or categoria.descripcion.find('IMPROPIA')>=0:
                                                                                                    sumoar+=1
                                                                                                    if categoria in dictar:
                                                                                                         dictar[categoria]+=sumoar
                                                                                                    else:
                                                                                                         dictar[categoria]=sumoar
                                                                                                    sumoar=0
                                                                                            break

                                                                                 else:
                                                                                        sumosar+=1
                                                                                        dictar['Sin Armas']=sumosar




                else:

                    lugar="LA PROVINCIA DE CHUBUT"
                    ciudad=''
                    idprovincia=RefProvincia.objects.get(descripcion__icontains='CHUBUT')
                    provis=RefCiudades.objects.all().filter(provincia_id=idprovincia.id).values('id')

                    for ciuda in provis:
                                de=Dependencias.objects.filter(ciudad_id=ciuda['id']).values('id')
                                if de:
                                 for dp in de:
                                         depe=dp['id']
                                 preventivo=Dependencias.objects.get(id=depe)
                                 dep=Dependencias.objects.filter(ciudad_id=ciuda['id']).values('id')

                                 if dep:
                                                for crias in dep:
                                                        dp=crias['id']

                                                        preventivos=Preventivos.objects.all().filter(dependencia=dp,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()

                                                        for desde in preventivos:
                                                                idp=desde['id']

                                                                idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                                if idhecho:
                                                                     for ids in idhecho:
                                                                             idprevh=ids['preventivo_id']
                                                                             if idprevh==idp:
                                                                                        idhecho=ids['id']

                                                                                        filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')
                                                                                        for indelis in filtrodelih:
                                                                                                descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])
                                                                                                if descripcion.descripcion.find('HOMI')>=0:
                                                                                                     #total de hechos donde hubo homicidios
                                                                                                     tothechos=tothechos+1
                                                                                                     #personas involucradas en homicidios rol de victima / imputado
                                                                                                     filtroperin=PersInvolucradas.objects.all().filter(hechos_id=idhecho).all()
                                                                                                     if filtroperin:
                                                                                                            for datos in filtroperin:
                                                                                                                if datos.roles.descripcion.find('VICTI')>=0:
                                                                                                                     sumovi+=1
                                                                                                                     dictvi.append(sumovi)
                                                                                                                else:
                                                                                                                    if datos.roles.descripcion.find('IMPUT')>=0:
                                                                                                                         sumoim+=1
                                                                                                                         dictim.append(sumoim)
                                                                                                     #tipo del lugar del homicidio.
                                                                                                     lugars=Hechos.objects.get(id=idhecho).lugar_hecho.values('tipo_lugar')
                                                                                                     sumolt=0

                                                                                                     if lugars:
                                                                                                            descripcion=RefLugares.objects.get(id=lugars)
                                                                                                            sumolt+=1
                                                                                                            if descripcion in dictlt:
                                                                                                                 dictlt[descripcion]+=sumolt
                                                                                                            else:
                                                                                                                 dictlt[descripcion]=sumolt
                                                                                                     #armas en el hecho
                                                                                                     elem=Hechos.objects.get(id=idhecho).eleinvolu.values('categoria')

                                                                                                     if elem:

                                                                                                            for filas in elem:

                                                                                                                categoria=RefCategory.objects.get(id=filas['categoria'])
                                                                                                                if categoria.descripcion.find('DE FUEGO')>=0 or categoria.descripcion.find('BLANCA')>=0 or categoria.descripcion.find('IMPROPIA')>=0:
                                                                                                                     sumoar+=1
                                                                                                                     if categoria in dictar:
                                                                                                                            dictar[categoria]+=sumoar
                                                                                                                     else:
                                                                                                                            dictar[categoria]=sumoar
                                                                                                                     sumoar=0
                                                                                                                break
                                                                                                     else:
                                                                                                            sumosar+=1
                                                                                                            dictar['Sin Armas']=sumosar


    dictvic={}
    dictimp={}
    if tothechos:
        canhec=True
        dictot={'HOMICIDIOS':0}
        dictch['total']=tothechos
        if dictvi:
            for vi in dictvi:
                    dictvic['total']=int(vi)
        else:
             dictvic['total']=0

        if dictim:
            for im in dictim:
                 dictimp['total']=int(im)
        else:
            dictimp['total']=0


    if request.POST.get('expo')=="Exportar":


        # se crea el libro y se obtiene la hoja


        libro = Workbook()
        hoja = libro.get_active_sheet()
        hoja.title = "Resumen de Hechos"

        # Ahora, se obtiene las celdas en la cuál se colocará el nombre
        # del campo. como son 8 campos, se necesita 8 celdas
        celda = hoja.cell("B1")
        celda.value=" Elementos recopilados desde la Base de Datos para Estadísticas en el ámbito de "+str(lugar)
        celda = hoja.cell("B2")
        celda.value=" Datos del Periodo de Fecha : MES / AÑO  "+str(fechadesde)+'/'+str(fechahasta)
        celda = hoja.cell("B3")
        celda.value=" Resumen de Homicidios - Total de Hechos, de Victimas y de Imputados "



        rango_celdas = hoja.range("B5:E5")

        # se crea una tupla con los nombres de los campos

        nombre_campos = "Delitos", "TOTAL DE HECHOS", "TOTAL DE VICTIMAS","TOTAL DE IMPUTADOS"
        # ahora, se asigna cada nombre de campo a cada celda

        for campo in rango_celdas:
                indice = 0  # se crea un contador para acceder a la tupla
                for celda in campo:
                        celda.value = nombre_campos[indice]
                        indice += 1

        # ya se tiene los nombres de los campos
        # ahora se obtiene el rango de celdas en donde irán los datos

        longitud = 5 + len(dictot)


        celdas_datos = hoja.range("B6:B{0}".format(longitud))
        # ahora vamos a dar los valores a las celdas con los datos

        fila=0

        for key,value in dictot.items():

         datos=key

         for celda in celdas_datos[fila]:
                 celda.value = datos



        celdas_datos = hoja.range("C6:C{0}".format(longitud))
        for key,value in dictch.items():
         datos=value
         for celda in celdas_datos[fila]:
                     celda.value = datos

        celdas_datos = hoja.range("D6:D{0}".format(longitud))
        for key,value in dictvic.items():
         datos=value
         for celda in celdas_datos[fila]:
                     celda.value = datos

        celdas_datos = hoja.range("E6:E{0}".format(longitud))
        for key,value in dictimp.items():
         datos=value
         for celda in celdas_datos[fila]:
                     celda.value = datos

        # lugar
        celda = hoja.cell("B8")
        celda.value=" Distribución de frecuencias de Homicidios por Tipo de Lugar de ocurrencia "




        rango_celdas1 = hoja.range("B9:C9")


        nombre_campos1 = "Tipo de Lugar", "TOTAL DE HECHOS"
        # ahora, se asigna cada nombre de campo a cada celda

        for campo in rango_celdas1:
                indice = 0  # se crea un contador para acceder a la tupla
                for celda in campo:
                        celda.value = nombre_campos1[indice]
                        indice += 1

        # ya se tiene los nombres de los campos
        # ahora se obtiene el rango de celdas en donde irán los datos

        longitud1 = 9 + len(dictlt)


        celdas_datos1 = hoja.range("B10:B{0}".format(longitud1))
        # ahora vamos a dar los valores a las celdas con los datos

        fila1=0

        for key,value in dictlt.items():


         for celda in celdas_datos1[fila1]:
                 celda.value = str(key)
         fila1+=1

        filac1=0
        celdas_datosc1 = hoja.range("C10:C{0}".format(longitud1))
        for key,value in dictlt.items():
            for celda in celdas_datosc1[filac1]:
                 celda.value = value
            filac1+=1

        #armas

        celda = hoja.cell("H8")
        celda.value=" Distribución de frecuencias de Homicidios por Clase de Armas "

        rango_celdas1 = hoja.range("H9:I9")


        nombre_campos1 = "Clase de Arma", "TOTAL DE HECHOS"
        # ahora, se asigna cada nombre de campo a cada celda

        for campo in rango_celdas1:
                indice = 0  # se crea un contador para acceder a la tupla
                for celda in campo:
                        celda.value = nombre_campos1[indice]
                        indice += 1

        # ya se tiene los nombres de los campos
        # ahora se obtiene el rango de celdas en donde irán los datos

        longitud1 = 9 + len(dictlt)


        celdas_datos1 = hoja.range("H10:H{0}".format(longitud1))
        # ahora vamos a dar los valores a las celdas con los datos

        fila1=0

        for key,value in dictar.items():


         for celda in celdas_datos1[fila1]:
                 celda.value = str(key)
         fila1+=1
        filaa1=0
        celdas_datos1 = hoja.range("I10:I{0}".format(longitud1))
        # ahora vamos a dar los valores a las celdas con los datos
        for key,value in dictar.items():
            for celda in celdas_datos1[filaa1]:
                 celda.value = value
            filaa1+=1

         # se crea un objeto httpresponse y se pasa como parámetro el content_type
         # diciendo que es excel
        response = HttpResponse(content_type="application/ms-excel")  # HttpResponse viene del modulo django.http
        nombre_archivo = "tothechos.xlsx"
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        libro.save(response)
        return response

    form = MapaForm()
    anios =[]
    meses={}
    anios={}

    ayer =datetime.date.today().year-4
    hoy = datetime.date.today().year
    rango=(hoy-ayer)+1
    for i in range(rango):
         if ayer<=hoy:
             a=ayer+i
             anios[i]=a



    for i in range(12):

        i+=1

        meses[i]=calendar.month_name[i]

    values={'fechadesde':fechadesde,'fechahasta':fechahasta,'anios':anios,'meses':meses,
    'delitos':delitos,'dictch':dictch,'dictvic':dictvic,'dictimp':dictimp,'dictot':dictot,'dictar':dictar,
    'ciudad':ciudad,'ureg':ureg,'depes':depes,'destino': destino,'state':state,'dictlt':dictlt,
    'form':form,'matriz':tothechos,'canhec':canhec,'lugar':lugar,'rangot':rangot,}
    return render(request,'./homicidios.html',values)

#cantidad de automotores segun delitos
@login_required
@group_required(["policia","investigaciones","radio"])
def repautos(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors=[]
    rangose=[]
    tothechos=0
    robos=[]
    injusticia=''
    delitosin=''
    rangod=[]
    rangou=[]
    rangoot=[]
    indice=0
    sumo=0
    anios=0
    rephec=False
    tot=0
    veces=0
    veces1=0
    veces2=0
    veces3=0
    indice=0
    ciudad=''
    ureg=''
    depes=''
    lugar=''
    fechadesde=0
    fechahasta=0
    mesnro=0
    delitos={}
    vecesd=0
    dictse={}
    robose=0
    homise=0
    hurtose=0
    otrose=0
    dictde={}
    robode=0
    homide=0
    hurtode=0
    otrode=0
    dictut={}
    robosu=0
    homisu=0
    hurtosu=0
    otrosu=0
    dictot={}
    roboso=0
    homiso=0
    hurtoso=0
    otroso=0
    lista={}
    listvehs=[]
    listvehd=[]
    listvehu=[]
    listvehot=[]
    if request.POST.get('ver')=="Visualizar" or request.POST.get('expo')=="Exportar":
            datosform=request.POST

            if request.POST.get('expo')=="Exportar":

             ciudad=request.POST.get('ciu')
             ureg=request.POST.get('ure')
             depes=request.POST.get('dep')
             fechadesde=request.POST.get('mesi')
             fechahasta=request.POST.get('aniosi')

            else:


             ciudad = datosform.__getitem__('ciudades')
             ureg = datosform.__getitem__('ureg')
             depes= datosform.__getitem__('depe')
             fechadesde=request.POST.get('mes')
             fechahasta=request.POST.get('anios')


            filtrorefd=Hechos.objects.all()



            if ciudad:

                lugar=RefCiudades.objects.get(id=ciudad)
                depes=Dependencias.objects.filter(ciudad_id=ciudad).values('id')
                if depes:
                     for dp in depes:
                             depes=dp['id']
                     preventivo=Dependencias.objects.get(id=depes)
                     depes=Dependencias.objects.filter(ciudad_id=ciudad).values('id')

                     if depes:
                                    for crias in depes:
                                            dp=crias['id']

                                            preventivos=Preventivos.objects.all().filter(dependencia=dp,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()

                                            for desde in preventivos:
                                                    idp=desde['id']

                                                    idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                    if idhecho:

                                                         for ids in idhecho:
                                                                 idprevh=ids['preventivo_id']
                                                                 if idprevh==idp:
                                                                        idhecho=ids['id']

                                                                        tothechos=tothechos+1

                                                                        filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')
                                                                        for indelis in filtrodelih:

                                                                                descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])
                                                                                elementos=Elementos.objects.filter(hechos_id=idhecho,borrado__isnull=True).values()
                                                                                for indelis in elementos:

                                                                                        descri=RefTipoelementos .objects.get(id=indelis['tipo_id'])
                                                                                        crubro=RefCategory.objects.get(id=indelis['categoria_id'])
                                                                                        if RefItems.objects.get(id=indelis['rubro_id']).descripcion.find('VEHI')>=0:
                                                                                            if descri.descripcion.find('SECUESTRA')>=0:
                                                                                                 veces=veces+1
                                                                                                 if descripcion.descripcion.find('ROBO')>=0:
                                                                                                        robose=+1
                                                                                                 else:
                                                                                                        if descripcion.descripcion.find('HOMI')>=0:
                                                                                                            homise+=1
                                                                                                        else:
                                                                                                            if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                 hurtose+=1
                                                                                                            else:
                                                                                                                 otrose+=1

                                                                                                 rangose.append(descri.descripcion)
                                                                                            else:
                                                                                             if descri.descripcion.find('DENUNCIA')>=0:
                                                                                                 veces1=veces1+1
                                                                                                 if descripcion.descripcion.find('ROBO')>=0:
                                                                                                        robode=+1
                                                                                                 else:
                                                                                                        if descripcion.descripcion.find('HOMI')>=0:
                                                                                                            homide+=1
                                                                                                        else:
                                                                                                            if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                 hurtode+=1
                                                                                                            else:
                                                                                                                 otrode+=1

                                                                                                 rangod.append(descri.descripcion)
                                                                                             else:
                                                                                                 if descri.descripcion.find('UTILIZADO')>=0:
                                                                                                        veces2=veces2+1
                                                                                                        if descripcion.descripcion.find('ROBO')>=0:
                                                                                                                robosu=+1
                                                                                                        else:
                                                                                                                if descripcion.descripcion.find('HOMI')>=0:
                                                                                                                    homisu+=1
                                                                                                                else:
                                                                                                                    if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                         hurtosu+=1
                                                                                                                    else:
                                                                                                                         otrosu+=1

                                                                                                        rangou.append(descri.descripcion)
                                                                                                 else:
                                                                                                         veces3=veces3+1
                                                                                                         if descripcion.descripcion.find('ROBO')>=0:
                                                                                                                roboso=+1
                                                                                                         else:
                                                                                                                if descripcion.descripcion.find('HOMI')>=0:
                                                                                                                    homiso+=1
                                                                                                                else:
                                                                                                                    if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                         hurtoso+=1
                                                                                                                    else:
                                                                                                                         otroso+=1

                                                                                                         rangoot.append(descri.descripcion)
                                                                                            lista[crubro.descripcion]='vehiculos'




            else:
                if depes:

                                            preventivo=Dependencias.objects.get(id=depes)
                                            ciudad=""
                                            #preventivo.ciudad.descripcion
                                            depe=preventivo.descripcion
                                            unireg=preventivo.unidades_regionales.descripcion
                                            lugar=ciudad+' -- '+unireg+' -- '+depe

                                            preventivos=Preventivos.objects.all().filter(dependencia=depes,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()
                                            for desde in preventivos:
                                                    idp=desde['id']

                                                    idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                    if idhecho:
                                                         for ids in idhecho:
                                                                 idprevh=ids['preventivo_id']
                                                                 if idprevh==idp:
                                                                            idhecho=ids['id']
                                                                            motivo=ids['motivo_id']
                                                                            tipohe=RefMotivosHecho.objects.get(id=motivo)
                                                                            tipohecho=tipohe.descripcion
                                                                            tothechos=tothechos+1
                                                                            filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')
                                                                            for indelis in filtrodelih:
                                                                                descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])
                                                                                elementos=Elementos.objects.filter(hechos_id=idhecho,borrado__isnull=True).values()
                                                                                for indelis in elementos:

                                                                                        descri=RefTipoelementos .objects.get(id=indelis['tipo_id'])
                                                                                        crubro=RefCategory.objects.get(id=indelis['categoria_id'])
                                                                                        if RefItems.objects.get(id=indelis['rubro_id']).descripcion.find('VEHI')>=0:
                                                                                            if descri.descripcion.find('SECUESTRA')>=0:
                                                                                                 veces=veces+1
                                                                                                 if descripcion.descripcion.find('ROBO')>=0:
                                                                                                        robose=+1
                                                                                                 else:
                                                                                                        if descripcion.descripcion.find('HOMI')>=0:
                                                                                                            homise+=1
                                                                                                        else:
                                                                                                            if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                 hurtose+=1
                                                                                                            else:
                                                                                                                 otrose+=1

                                                                                                 rangose.append(descri.descripcion)
                                                                                            else:
                                                                                             if descri.descripcion.find('DENUNCIA')>=0:
                                                                                                 veces1=veces1+1
                                                                                                 if descripcion.descripcion.find('ROBO')>=0:
                                                                                                        robode=+1
                                                                                                 else:
                                                                                                        if descripcion.descripcion.find('HOMI')>=0:
                                                                                                            homide+=1
                                                                                                        else:
                                                                                                            if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                 hurtode+=1
                                                                                                            else:
                                                                                                                 otrode+=1

                                                                                                 rangod.append(descri.descripcion)
                                                                                             else:
                                                                                                 if descri.descripcion.find('UTILIZADO')>=0:
                                                                                                        veces2=veces2+1
                                                                                                        if descripcion.descripcion.find('ROBO')>=0:
                                                                                                                robosu=+1
                                                                                                        else:
                                                                                                                if descripcion.descripcion.find('HOMI')>=0:
                                                                                                                    homisu+=1
                                                                                                                else:
                                                                                                                    if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                         hurtosu+=1
                                                                                                                    else:
                                                                                                                         otrosu+=1

                                                                                                        rangou.append(descri.descripcion)
                                                                                                 else:
                                                                                                         veces3=veces3+1
                                                                                                         if descripcion.descripcion.find('ROBO')>=0:
                                                                                                                roboso=+1
                                                                                                         else:
                                                                                                                if descripcion.descripcion.find('HOMI')>=0:
                                                                                                                    homiso+=1
                                                                                                                else:
                                                                                                                    if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                         hurtoso+=1
                                                                                                                    else:
                                                                                                                         otroso+=1

                                                                                                         rangoot.append(descri.descripcion)
                                                                                            lista[crubro.descripcion]='vehiculos'



                else:

                    lugar="LA PROVINCIA DE CHUBUT"
                    ciudad=''
                    idprovincia=RefProvincia.objects.get(descripcion__icontains='CHUBUT')
                    provis=RefCiudades.objects.all().filter(provincia_id=idprovincia.id).values('id')

                    for ciuda in provis:
                                de=Dependencias.objects.filter(ciudad_id=ciuda['id']).values('id')
                                if de:
                                 for dp in de:
                                         depe=dp['id']
                                 preventivo=Dependencias.objects.get(id=depe)
                                 dep=Dependencias.objects.filter(ciudad_id=ciuda['id']).values('id')

                                 if dep:
                                                for crias in dep:
                                                        dp=crias['id']

                                                        preventivos=Preventivos.objects.all().filter(dependencia=dp,fecha_autorizacion__isnull=False,fecha_denuncia__month=fechadesde,fecha_denuncia__year=fechahasta).values()

                                                        for desde in preventivos:
                                                                idp=desde['id']

                                                                idhecho=Hechos.objects.filter(preventivo_id=idp).values()

                                                                if idhecho:
                                                                     for ids in idhecho:
                                                                             idprevh=ids['preventivo_id']
                                                                             if idprevh==idp:
                                                                                        idhecho=ids['id']
                                                                                        motivo=ids['motivo_id']
                                                                                        tipohe=RefMotivosHecho.objects.get(id=motivo)
                                                                                        tipohecho=tipohe.descripcion
                                                                                        tothechos=tothechos+1
                                                                                        filtrodelih=HechosDelito.objects.all().filter(hechos_id=idhecho,borrado__isnull=True).values('id','refdelito_id','hechos_id')
                                                                                        for indelis in filtrodelih:
                                                                                                descripcion=RefDelito.objects.get(id=indelis['refdelito_id'])
                                                                                                elementos=Elementos.objects.filter(hechos_id=idhecho,borrado__isnull=True).values()
                                                                                                for indelis in elementos:

                                                                                                        descri=RefTipoelementos .objects.get(id=indelis['tipo_id'])
                                                                                                        crubro=RefCategory.objects.get(id=indelis['categoria_id'])
                                                                                                        if RefItems.objects.get(id=indelis['rubro_id']).descripcion.find('VEHI')>=0:
                                                                                                            if descri.descripcion.find('SECUESTRA')>=0:
                                                                                                                 veces=veces+1
                                                                                                                 if descripcion.descripcion.find('ROBO')>=0:
                                                                                                                        robose=+1
                                                                                                                 else:
                                                                                                                        if descripcion.descripcion.find('HOMI')>=0:
                                                                                                                            homise+=1
                                                                                                                        else:
                                                                                                                            if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                                 hurtose+=1
                                                                                                                            else:
                                                                                                                                 otrose+=1

                                                                                                                 rangose.append(descri.descripcion)
                                                                                                            else:
                                                                                                             if descri.descripcion.find('DENUNCIA')>=0:
                                                                                                                 veces1=veces1+1
                                                                                                                 if descripcion.descripcion.find('ROBO')>=0:
                                                                                                                        robode=+1
                                                                                                                 else:
                                                                                                                        if descripcion.descripcion.find('HOMI')>=0:
                                                                                                                            homide+=1
                                                                                                                        else:
                                                                                                                            if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                                 hurtode+=1
                                                                                                                            else:
                                                                                                                                 otrode+=1

                                                                                                                 rangod.append(descri.descripcion)
                                                                                                             else:
                                                                                                                 if descri.descripcion.find('UTILIZADO')>=0:
                                                                                                                        veces2=veces2+1
                                                                                                                        if descripcion.descripcion.find('ROBO')>=0:
                                                                                                                                robosu=+1
                                                                                                                        else:
                                                                                                                                if descripcion.descripcion.find('HOMI')>=0:
                                                                                                                                    homisu+=1
                                                                                                                                else:
                                                                                                                                    if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                                         hurtosu+=1
                                                                                                                                    else:
                                                                                                                                         otrosu+=1

                                                                                                                        rangou.append(descri.descripcion)
                                                                                                                 else:
                                                                                                                         veces3=veces3+1
                                                                                                                         if descripcion.descripcion.find('ROBO')>=0:
                                                                                                                                roboso=+1
                                                                                                                         else:
                                                                                                                                if descripcion.descripcion.find('HOMI')>=0:
                                                                                                                                    homiso+=1
                                                                                                                                else:
                                                                                                                                    if descripcion.descripcion.find('HURTO')>=0:
                                                                                                                                         hurtoso+=1
                                                                                                                                    else:
                                                                                                                                         otroso+=1

                                                                                                                         rangoot.append(descri.descripcion)
                                                                                                            lista[crubro.descripcion]='vehiculos'

    if rangose:
         dictse={'robos':robose,'homicidios':homise,'hurtos':hurtose,'otros':otrose}
    if rangod:
         dictde={'robos':robode,'homicidios':homide,'hurtos':hurtode,'otros':otrode}
    if rangou:
         dictut={'robos':robode,'homicidios':homide,'hurtos':hurtode,'otros':otrode}
    if rangoot:
         dictot={'robos':robode,'homicidios':homide,'hurtos':hurtode,'otros':otrode}

    if lista:

         rephec=True
         for key,value in sorted(lista.iteritems()):
                 lstr=repetidos(lista)
                 delitos[key]=value

    if request.POST.get('expo')=="Exportar":


        # se crea el libro y se obtiene la hoja


        libro = Workbook()
        hoja = libro.get_active_sheet()
        hoja.title = "Vehiculos_Delitos"

        # Ahora, se obtiene las celdas en la cuál se colocará el nombre
        # del campo. como son 8 campos, se necesita 8 celdas
        celda = hoja.cell("B1")
        celda.value=" Elementos recopilados desde la Base de Datos para Estadísticas en el ámbito de "+str(lugar)
        celda = hoja.cell("B2")
        celda.value=" Datos del Periodo de Fecha : MES / AÑO  "+str(fechadesde)+'/'+str(fechahasta)

        if len(dictde)!=0:

            celda = hoja.cell("B3")
            celda.value=" Tabla con datos obtenidos de Vehiculos DENUNCIADOS segun Delitos"



            rango_celdas = hoja.range("B4:F4")

            # se crea una tupla con los nombres de los campos

            nombre_campos = "Vehiculos", "ROBOS", "HOMICIDIOS","HURTOS", "OTROS"
            # ahora, se asigna cada nombre de campo a cada celda

            for campo in rango_celdas:
                    indice = 0  # se crea un contador para acceder a la tupla
                    for celda in campo:
                            celda.value = nombre_campos[indice]
                            indice += 1

            # ya se tiene los nombres de los campos
            # ahora se obtiene el rango de celdas en donde irán los datos

            longitud = 4 + len(delitos)


            celdas_datos = hoja.range("B5:B{0}".format(longitud))
            # ahora vamos a dar los valores a las celdas con los datos

            fila=0
            filad=0
            paso=0
            for key,value in delitos.items():

             datos=key
             for celda in celdas_datos[fila]:
                     celda.value = datos


             fila += 1
             if len(dictde)==0:
                    longitud=8
             else:
                    longitud=4+len(dictde)

             celdas_datost = hoja.range("C5:F{0}".format(longitud))
             filat=0

             for celdas in celdas_datost[filat]:

                                 if 'robos' in  dictde.keys():
                                        celdas.value = dictde.values()[filat]
                                 else:
                                        if 'homicidios' in dictde.keys():
                                            celdas.value =dictde.values()[filat]
                                        else:
                                            if 'hurtos' in dictde.keys():
                                                 celdas.value = dictde.values()[filat]
                                            else:
                                                 if 'otros' in dictde.keys():
                                                        celdas.value = dictde.values()[filat]

                                 filat+=1
        else:
            celda = hoja.cell("B{0}".format(len(delitos)+6))
            celda.value=" No se encontraron datos de Vehiculos DENUNCIADOS segun Delitos"

        if len(dictse)!=0:
                celda = hoja.cell("B30")
                celda.value=" Tabla con datos obtenidos de Vehiculos SECUESTRADOS segun Delitos"



                rango_celdas = hoja.range("B31:F31")

                # se crea una tupla con los nombres de los campos

                nombre_campos = "Vehiculos", "ROBOS", "HOMICIDIOS","HURTOS", "OTROS"
                # ahora, se asigna cada nombre de campo a cada celda

                for campo in rango_celdas:
                        indice = 0  # se crea un contador para acceder a la tupla
                        for celda in campo:
                                celda.value = nombre_campos[indice]
                                indice += 1

                # ya se tiene los nombres de los campos
                # ahora se obtiene el rango de celdas en donde irán los datos
                longitudse = 32 + len(delitos)


                celdas_datos = hoja.range("B32:B{0}".format(longitudse))
                # ahora vamos a dar los valores a las celdas con los datos

                filase=0
                for key,value in delitos.items():

                 datose=key
                 for celda in celdas_datos[filase]:
                         celda.value = datose


                 filase += 1

                 if len(dictse)==0:
                        longitudse=36
                 else:
                        longitudse=32+len(dictse)

                 celdas_datost = hoja.range("C33:F{0}".format(longitudse))
                 filat=0
                 for celdas in celdas_datost[filat]:

                                     if 'robos' in  dictse.keys():
                                            celdas.value = dictse.values()[filat]
                                     else:
                                            if 'homicidios' in dictse.keys():
                                                celdas.value =dictse.values()[filat]
                                            else:
                                                if 'hurtos' in dictse.keys():
                                                     celdas.value = dictse.values()[filat]
                                                else:
                                                     if 'otros' in dictse.keys():
                                                            celdas.value = dictse.values()[filat]

                                     filat+=1
        else:

            celda = hoja.cell("B{0}".format(len(delitos)+7))
            celda.value=" No se encontraron datos de Vehiculos SECUESTRADOS segun Delitos"

        if len(dictut)!=0:
                celda = hoja.cell("B60")
                celda.value=" Tabla con datos obtenidos de Vehiculos UTILIZADOS segun Delitos"



                rango_celdas = hoja.range("B61:F61")

                # se crea una tupla con los nombres de los campos

                nombre_campos = "Vehiculos", "ROBOS", "HOMICIDIOS","HURTOS", "OTROS"
                # ahora, se asigna cada nombre de campo a cada celda

                for campo in rango_celdas:
                        indice = 0  # se crea un contador para acceder a la tupla
                        for celda in campo:
                                celda.value = nombre_campos[indice]
                                indice += 1

                # ya se tiene los nombres de los campos
                # ahora se obtiene el rango de celdas en donde irán los datos
                longitud = 61 + len(delitos)


                celdas_datos = hoja.range("B62:B{0}".format(longitud))
                # ahora vamos a dar los valores a las celdas con los datos

                fila=0
                filad=0
                paso=0
                for key,value in delitos.items():

                 datos=key
                 for celda in celdas_datos[fila]:
                         celda.value = datos


                 fila += 1
                 if len(dictse)==0:
                        longitud=8
                 else:
                        longitud=4+len(dictut)
                 celdas_datost = hoja.range("C5:F{0}".format(longitud))
                 filat=0

                 for celdas in celdas_datost[filat]:

                                     if 'robos' in  dictut.keys():
                                            celdas.value = dictut.values()[filat]
                                     else:
                                            if 'homicidios' in dictut.keys():
                                                celdas.value =dictut.values()[filat]
                                            else:
                                                if 'hurtos' in dictut.keys():
                                                     celdas.value = dictut.values()[filat]
                                                else:
                                                     if 'otros' in dictut.keys():
                                                            celdas.value = dictut.values()[filat]

                                     filat+=1
        else:
                celda = hoja.cell("B{0}".format(len(delitos)+8))
                celda.value=" No se encontraron datos de Vehiculos UTILIZADOS segun Delitos"

        if len(dictot)!=0:
                celda = hoja.cell("B90")
                celda.value=" Tabla con datos obtenidos de Vehiculos UTILIZADOS y/o SECUESTRADOS segun Delitos"



                rango_celdas = hoja.range("B91:F91")

                # se crea una tupla con los nombres de los campos

                nombre_campos = "Vehiculos", "ROBOS", "HOMICIDIOS","HURTOS", "OTROS"
                # ahora, se asigna cada nombre de campo a cada celda

                for campo in rango_celdas:
                        indice = 0  # se crea un contador para acceder a la tupla
                        for celda in campo:
                                celda.value = nombre_campos[indice]
                                indice += 1

                # ya se tiene los nombres de los campos
                # ahora se obtiene el rango de celdas en donde irán los datos
                longitud = 91 + len(delitos)


                celdas_datos = hoja.range("B92:B{0}".format(longitud))
                # ahora vamos a dar los valores a las celdas con los datos

                fila=0
                filad=0
                paso=0
                for key,value in delitos.items():

                 datos=key
                 for celda in celdas_datos[fila]:
                         celda.value = datos


                 fila += 1
                 if len(dictot)==0:
                        longitud=8
                 else:
                        longitud=4+len(dictot)
                 celdas_datost = hoja.range("C5:F{0}".format(longitud))
                 filat=0

                 for celdas in celdas_datost[filat]:

                                     if 'robos' in  dictot.keys():
                                            celdas.value = dictot.values()[filat]
                                     else:
                                            if 'homicidios' in dictot.keys():
                                                celdas.value =dictot.values()[filat]
                                            else:
                                                if 'hurtos' in dictot.keys():
                                                     celdas.value = dictot.values()[filat]
                                                else:
                                                     if 'otros' in dictot.keys():
                                                            celdas.value = dictot.values()[filat]

                                     filat+=1

        else:
                celda = hoja.cell("B{0}".format(len(delitos)+9))
                celda.value=" No se encontraron datos de Vehiculos UTILIZADOS y/o SECUESTRADOS segun Delitos"



         # se crea un objeto httpresponse y se pasa como parámetro el content_type
         # diciendo que es excel
        response = HttpResponse(content_type="application/ms-excel")  # HttpResponse viene del modulo django.http
        nombre_archivo = "cantcars.xlsx"
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        libro.save(response)
        return response

    form = MapaForm()
    anios =[]
    meses={}
    anios={}

    ayer =datetime.date.today().year-4
    hoy = datetime.date.today().year
    rango=(hoy-ayer)+1
    for i in range(rango):
         if ayer<=hoy:
             a=ayer+i
             anios[i]=a



    for i in range(12):

        i+=1

        meses[i]=calendar.month_name[i]

    values={'rangose':rangose,'rangod':rangod,'rangou':rangou,'rangoot':rangoot,'fechadesde':fechadesde,'fechahasta':fechahasta,'anios':anios,'meses':meses,'delitos':delitos,'dictse':dictse,'dictde':dictde,'dictut':dictut,'dictot':dictot,'ciudad':ciudad,'ureg':ureg,'depes':depes,'destino': destino,'state':state,'form':form,'matriz':tothechos,'rephec':rephec,'lugar':lugar,}
    return render(request,'./repautos.html',values)


def repetidos(rango):
     lstr=[]

     for key in rango:
        if key not in lstr:
            lstr.append(key)

     return lstr
def contar(rango,lstr):
     dictm={}

     for eti in lstr:
         dictm[eti]=rango.count(eti)
     return dictm

def funverifica(idper):
    persona=''
    bandera=False
    esci=''
    ocu=''
    pnac=''
    datos=Personas.objects.get(id=idper)
    if datos.estado_civil is None:
         esci=', Estado Civil :'+str('No posee')
    if datos.ocupacion is None:
         ocu=', Ocupacion :'+str('No posee')
    if datos.pais_nac is None and datos.ciudad_nac is None:
         pnac=', Nacido en : '+str('sin datos')


    persona=str(ocu)+str(esci)+str(pnac)
    if persona:
         bandera=True

    return bandera,persona


@login_required
@group_required(["policia","investigaciones","radio"])
def ampliacion(request,idprev):
    state= request.session.get('state')
    destino= request.session.get('destino')
    preventivo = Preventivos.objects.get(id = idprev)
    ampliaciones = preventivo.ampli.all()
    depe=preventivo.dependencia
    ampliacion = AmpliacionForm()
    if request.POST.get('grabar') == 'Grabar':
        ampliacion = AmpliacionForm(request.POST, request.FILES)
        amp        = Ampliacion()

        if ampliacion.is_valid():
            amp.fecha         = datetime.datetime.now()
            amp.titulo        = ampliacion.cleaned_data['titulo']
            amp.descripcion   = ampliacion.cleaned_data['descripcion']
            amp.preventivo    = preventivo
            amp.cierre_causa  = ampliacion.cleaned_data['cierre_causa']
            if amp.cierre_causa:
                preventivo.fecha_cierre = amp.fecha_cierre  = datetime.datetime.now()

            amp.save()
            for aut in ampliacion.cleaned_data['autoridades']:
                amp.autoridades.add(aut)
            preventivo.save()

            #ampliacion = AmpliacionForm()
            id_ciudad=Dependencias.objects.filter(descripcion__exact=depe).values('ciudad')
            ampliacion.fields['autoridades'].queryset=RefCiudades.objects.get(id=id_ciudad).ciu_autori.all()
            autoridades= preventivo.autoridades.all()
            autoridad=[]
            for seleccion in autoridades:
              ids=int(RefAutoridad.objects.get(descripcion=seleccion).id)
              autoridad.append(ids)
            ampliacion.fields['autoridades'].initial=autoridad
            ampliacion.fields['autoridades'].widget.attrs["onclick"] = False
            #instance=amp
            return HttpResponseRedirect(reverse('ver_ampliacion',args=[int(preventivo.id),int(amp.id)]))
    else:
          id_ciudad=Dependencias.objects.filter(descripcion__exact=depe).values('ciudad')
          ampliacion.fields['autoridades'].queryset=RefCiudades.objects.get(id=id_ciudad).ciu_autori.all()
          autoridades= preventivo.autoridades.all()
          autoridad=[]
          for seleccion in autoridades:
            ids=int(RefAutoridad.objects.get(descripcion=seleccion).id)
            autoridad.append(ids)
          ampliacion.fields['autoridades'].initial=autoridad
          ampliacion.fields['autoridades'].widget.attrs["onclick"] = False
          #ampliacion.fields['autoridades'].initial = preventivo.autoridades.all()


    values={'destino': destino,'state':state,'preventivo':preventivo,'ampliaciones':ampliaciones,'ampliacion':ampliacion,'depe':depe,}

    return render(request,'./ampliaciones.html',values)

@login_required
@group_required(["policia","investigaciones","radio"])
def ver_ampliacion(request,idprev,idamp):
    state= request.session.get('state')
    destino= request.session.get('destino')
    preventivo = Preventivos.objects.get(id = idprev)
    hecho = Hechos.objects.get(preventivo_id=preventivo.id)
    depe=preventivo.dependencia
    ampliaciones = preventivo.ampli.all()
    ampli = Ampliacion.objects.get(id=idamp)
    ampliacion = AmpliacionForm(instance=ampli)
    tieneelemento=False
    tienepersonas=False
    verelemento=False
    verpers=False
    if ampli.fin_edicion:
        finaliza=False
    else:
        finaliza=True
    tiene=[]
    tienep=[]
    personas= PersInvolucradas.objects.all().filter(ampliacion_id=ampli.id).order_by('id')

    veriele= Elementos.objects.all().filter(ampliacion_id=ampli.id,borrado__isnull=True).order_by('id')
    if personas:
         tienepersonas=True
    if veriele:
        tieneelemento=True
        ampliaciones = Ampliacion.objects.all().filter(id=ampli.id)
        for v in veriele:
             tiene.append(v)
        if request.POST.get('vere'):
                verelemento=True
    if personas:
        tienepersonas=True
        ampliaciones = Ampliacion.objects.all().filter(id=ampli.id)
        for v in personas:
             tienep.append(v)
        if request.POST.get('verper'):
                verpers=True

    id_ciudad=Dependencias.objects.filter(descripcion__exact=depe).values('ciudad')
    ampliacion.fields['autoridades'].queryset=RefCiudades.objects.get(id=id_ciudad).ciu_autori.all()
    autoridades= preventivo.autoridades.all()
    autoridad=[]
    for seleccion in autoridades:
        ids=int(RefAutoridad.objects.get(descripcion=seleccion).id)
        autoridad.append(ids)

    ampliacion.fields['autoridades'].initial=autoridad
    ampliacion.fields['autoridades'].widget.attrs["onclick"] = False
    values={'depe':depe,'tienep':tienep,'personas':personas,'verpers':verpers,'finaliza':finaliza,'tiene':tiene,'veriele':veriele,'verelemento':verelemento,'tieneelemento':tieneelemento,'id':idamp,'destino': destino,'state':state,'preventivo':preventivo,'ampliaciones':ampliaciones,'ampliacion':ampliacion,'tienepersonas':tienepersonas,}

    return render(request,'./ampliaciones.html',values)

@login_required
@group_required(["policia","investigaciones","radio"])
def amplia_ele(request,idprev,idamp):
    state= request.session.get('state')
    destino= request.session.get('destino')
    preventivo = Preventivos.objects.get(id = idprev)
    hechos = Hechos.objects.get(preventivo_id=preventivo.id)
    depe = preventivo.dependencia
    ampliaciones = preventivo.ampli.all()
    ampli = Ampliacion.objects.get(id=idamp)
    fechaenvio=ampli.fecha_autorizacion
    ampliacion = AmpliacionForm(instance=ampli)
    idprev=preventivo.id
    nro=preventivo.nro
    anio=preventivo.anio
    fecha_denuncia=preventivo.fecha_denuncia
    fecha_carga=preventivo.fecha_carga
    caratula=preventivo.caratula
    actuante=preventivo.actuante
    preventor=preventivo.preventor
    fecha_autorizacion=preventivo.fecha_autorizacion
    autoridades= preventivo.autoridades.values_list('descripcion',flat=True)

    formv=VehiculosForm()
    form = ElementosForm()
    elemento = Elementos()
    formrub = ItemForm()
    formcat = CategoryForm()
    formumed = UnidadMedidasForm()
    formar=ArmasForm()
    errors = []

    if request.POST.get('dele'):
            elementosin=Elementos.objects.filter(id=request.POST.get('dele'))

            if elementosin:
                if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion.find('INVESTIGACIONES')>=0 or request.user.userprofile.depe.descripcion.find('RADIO')>=0:
                        obs="elemento borrado por usuario : "+request.user.username

                        Elementos.objects.filter(id = request.POST.get('dele')).update(borrado='S',observaciones=obs)

                else:
                        errors='No se puede borrar elementos de un Hecho de otras dependencias.'
            else:
                        errors='No se existe elemento involucrado.'

    if request.POST.get('button') == 'Guardar':
        form = ElementosForm(request.POST)
        formar=ArmasForm(request.POST)
        formv=VehiculosForm(request.POST)


        if form.is_valid():
            elemento = Elementos()
            elemento.tipo           = form.cleaned_data['tipo']
            elemento.rubro          = form.cleaned_data['rubro']
            elemento.categoria      = form.cleaned_data['categoria']
            elemento.cantidad       = form.cleaned_data['cantidad']
            elemento.unidadmed      = form.cleaned_data['unidadmed']
            elemento.descripcion    = form.cleaned_data['descripcion']
            elemento.hechos         = hechos
            elemento.ampliacion     = ampli

            if elemento.rubro.descripcion=='VEHICULOS' or elemento.rubro.descripcion== 'AUTOMOTORES':

                    if formv.is_valid():
                     elemento.save()
                     vehicle = Vehiculos()
                     vehicle.idmarca         = formv.cleaned_data['idmarca']
                     vehicle.modelo          = formv.cleaned_data['modelo']
                     vehicle.anio            = formv.cleaned_data['anio']
                     vehicle.tipov           = formv.cleaned_data['tipov']
                     vehicle.dominio         = formv.cleaned_data['dominio']
                     vehicle.nmotor          = formv.cleaned_data['nmotor']
                     vehicle.nchasis         = formv.cleaned_data['nchasis']
                     vehicle.nro_doc         = formv.cleaned_data['nro_doc']
                     vehicle.propietario     = formv.cleaned_data['propietario']
                     nueva_marcav            = formv.cleaned_data['nueva_marcav']
                     if nueva_marcav:
                        marca = RefTrademark()
                        marca.descripcion = nueva_marcav
                        marca.save()
                        vehicle.idmarca       = marca
                     vehicle.save()
                     elevehi = Elementoscars()
                     elevehi.idelemento        = elemento
                     elevehi.idvehiculo        = vehicle
                     elevehi.save()
                     formv = VehiculosForm()
                     errors.append('Elementos guardados')
            else:
                 if elemento.categoria.descripcion =='FUEGO' or elemento.categoria.descripcion=='DE FUEGO':
                         if formar.is_valid():
                             elemento.save()
                             armas=Armas()
                             armas.tipos    = formar.cleaned_data['tipos']
                             armas.subtipos = formar.cleaned_data['subtipos']
                             armas.sistema_disparo = formar.cleaned_data['sistema_disparo']
                             armas.marcas = formar.cleaned_data['marcas']
                             armas.calibre = formar.cleaned_data['calibre']
                             armas.modelo = formar.cleaned_data['modelo']
                             armas.nro_arma=formar.cleaned_data['nro_arma']
                             armas.nro_doc = formar.cleaned_data['nro_doc']
                             armas.propietario = formar.cleaned_data['propietario']
                             nueva_marca       = formar.cleaned_data['nueva_marca']

                             if nueva_marca:
                                 marca = RefTrademark()
                                 marca.descripcion = nueva_marca
                                 marca.save()
                                 armas.marcas = marca

                             armas.save()

                             elearmas=Elementosarmas()
                             elearmas.idelemento=elemento
                             elearmas.idarma=armas
                             elearmas.save()
                             form = ElementosForm()
                             formar=ArmasForm()
                             errors.append('Elementos guardados')
                 else:
                             elemento.save()
                             errors.append('Elementos guardados')
                             form=ElementosForm()
        else:
             errors.append('Faltan Datos verifique')

    if request.POST.get('guardarubro') == 'Guardar':
        formrub = ItemForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        mostrar="no"
        if not descripcion:
            errors.append('Ingrese un Rubro')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<100):
                    errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
             else:
                 if formrub.is_valid():
                        formrub.save()
    if request.POST.get('guardacategoria') == 'Guardar':
        formcat = CategoryForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        rubro = request.POST.get('rubro')
        mostrar="no"
        if not descripcion:
            errors.append('Ingrese una categoria')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<100):
                    errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
             else:
                if not rubro:
                    errors.append('Seleccione Rubro')
                else:
                    if formcat.is_valid():
                        formcat.save()
    if request.POST.get('guardamedida') == 'Guardar':
        formumed = UnidadMedidasForm(request.POST, request.FILES)

        descripcion = request.POST.get('descripcion')
        mostrar="no"
        if not descripcion:
            errors.append('Ingrese unidad de medida')
        else:
             if formumed.is_valid():
                formumed.save()


    lista = Elementos.objects.filter(hechos = hechos.id,cargado_prev=True,ampliacion__isnull=True,borrado__isnull=True)
    listam= Elementos.objects.filter(hechos = hechos.id,ampliacion_id=idamp,borrado__isnull=True)


    values={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
            'caratula':caratula,'actuante':actuante,'idprev':idprev,'preventor':preventor,'fecha_autorizacion':fecha_autorizacion,
            'autoridades':autoridades,'idprev':idprev,'hecho':hechos,'destino': destino,'state':state,'preventivo':preventivo,
            'ampliaciones':ampliaciones,'ampliacion':ampliacion,'ampli':ampli,
            'vehiculo':formv,'formar':formar,'form':form,'idamp':idamp,'listam':listam,
            'elemento':elemento,'lista':lista,'fechaenvio':fechaenvio,
            'formrub':formrub,
            'formcat':formcat,
            'formumed':formumed,}

    return render(request,'./objampli.html',values)


@group_required(["policia","investigaciones","radio"])
@login_required
def reporampli(request,idprev,idamp):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors=[]
    continua=''
    tienepersonas=False
    tieneelementos=False
    i=0
    grabo='fin'
    preventivo = Preventivos.objects.get(id = idprev)
    ciudad= preventivo.dependencia.ciudad
    depe=preventivo.dependencia
    unireg=depe.unidades_regionales.descripcion
    image=""
    countinvolus=0
    #Datos del Hecho delicitivo atraves del nro de preventivo
    if len(Ampliacion.objects.filter(id=idamp))>0:
            amplia = Ampliacion.objects.get(id=idamp)
            tieneampli=True
            form=AmpliacionForm(instance=amplia)
            involuscra=[]
            eleminvo=[]
            datosper=""
            elementos=""

            involus=PersInvolucradas.objects.filter(ampliacion=amplia.id).all()
            eleinvo=Elementos.objects.filter(ampliacion=amplia.id,borrado__isnull=True).all()

            datosgral=""
            lugar=''
            lati=''
            longi=''
            condiciones=''
            #Datos del lugar del hecho



            if len(PersInvolucradas.objects.filter(ampliacion=amplia.id).all())>0:
                 tienepersonas=True
                 countinvolus=PersInvolucradas.objects.filter(ampliacion=amplia.id).count()
                 for p in PersInvolucradas.objects.filter(ampliacion=amplia.id).all():

                     bandera,personai = funverifica(p.persona.id)

                     domi=Personas.objects.get(id=p.persona.id).persodom.all()
                     if domi:
                        for l in Personas.objects.get(id=p.persona.id).persodom.all():
                         #datosgral=str(p.roles)+'-'+str(p)+' '+str(p.persona.tipo_doc)+' :'+str(p.persona.nro_doc)
                         dad=Personas.objects.get(id=p.persona.id).padre.all()

                         if dad:

                                for la in Personas.objects.get(id=p.persona.id).padre.all():
                                    if p.menor=='':
                                         p.menor="NO"
                                    roles='<u>'+str(p.roles)+'</u>'+' : '
                                    if bandera:
                                         persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')
                                    else:
                                        persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')

                                    domi='<dd>Reside en : '+str(p.persona.ciudad_res)+',  Domicilio : '+str(l.calle)+'  Nro.: '+str(l.altura)+'</dd>'
                                    if la.padre_apellidos or la.padre_nombres or la.madre_apellidos or la.madre_nombres:
                                         padys='<dd>Hijo de : '+str(la.padre_apellidos.encode("utf8"))+', '+str(la.padre_nombres.encode("utf8"))+' y de : '+str(la.madre_apellidos.encode("utf8"))+', '+str(la.madre_nombres.encode("utf8"))+'<br><br></dd>'
                                    else:
                                         padys='<dd>no registra datos de los padres'+'<br></dd>'
                                    datosgral=roles+persona+domi+padys
                         else:
                                 if p.menor=='':
                                         p.menor="NO"
                                 roles='<u>'+str(p.roles)+'</u>'+' : '
                                 if bandera:
                                         persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')
                                 else:
                                        persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')

                                 #persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+',  Estado Civil : '+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')
                                 domi='<dd>Reside en : '+str(p.persona.ciudad_res)+',  Domicilio : '+str(l.calle)+'  Nro.: '+str(l.altura)+'</dd>'
                                 padys='<dd>no registra datos de los padres'+'<br></dd>'
                                 datosgral=roles+persona+domi+padys
                         involuscra.append(datosgral)
                     else:
                        if p.menor=='':
                                         p.menor="NO"
                        roles='<u>'+str(p.roles)+'</u>'+' : '
                        if bandera:
                             persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')
                        else:
                             persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')

                        #persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+',  Estado Civil : '+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')
                        domi='<dd>no registra domicilio'+'</dd>'
                        padys='<dd>no registra datos de los padres'+'<br></dd>'
                        datosgral=roles+persona+domi+padys

                        involuscra.append(datosgral)
                 for i in involuscra:
                     datosper=datosper+i

                    #datosper.append(persona)
                 datosgral=''
                 obdata=[]
                 obdatav=[]
                 deta=''
                 detav=''
                 eleme=''
                 hay=[]


            if len(Elementos.objects.filter(ampliacion=amplia.id).all())>0:
                    countinvolus=Elementos.objects.filter(ampliacion=amplia.id).count()
                    i=1
                    for eles in eleinvo:
                        tieneelementos=True
                        obdata=[]
                        obdatav=[]
                        deta=''
                        detav=''

                        if len(Elementosarmas.objects.filter(idelemento=eles.id))>0:

                             idar = Elementosarmas.objects.filter(idelemento=eles.id).values('idarma')
                             tieneelementos=True
                             obdata=Armas.objects.filter(id=idar)
                             for extra in obdata:
                                titu='<ul><li><u> Carateristicas Generales : </u></li></ul>'
                                tabla='<dd> '+str(extra.subtipos)+'  ---  Tipo/s : '+str(extra.tipos)+'  ---  Sistema de Disparo : '+str(extra.sistema_disparo)+'   --- Marcas : '+str(extra.marcas)+'</dd>'
                                tipos='<dd> Calibre : '+str(extra.calibre)+'  --- Modelo : '+str(extra.modelo)+'  --- Nro Serie : '+str(extra.nro_arma)+'   ---  Propietario : '+str(extra.nro_doc)+' - '+str(extra.propietario)+'</dd>'
                                deta=titu+tabla+tipos

                        if len(Elementoscars.objects.filter(idelemento=eles.id))>0:
                             tieneelementos=True
                             idarv = Elementoscars.objects.filter(idelemento=eles.id).values('idvehiculo')

                             obdatav=Vehiculos.objects.filter(id=idarv)
                             for extrav in obdatav:
                                tituv='<ul><li><u> Carateristicas Generales : </u></li></ul>'
                                tablav='<dd> Marca/s : '+str(extrav.idmarca)+'  ---   Modelo : '+str(extrav.modelo)+'  ---  Dominio : '+str(extrav.dominio)+'   ---  Año : '+str(extrav.anio)+'</dd>'
                                tiposv='<dd> Tipo/s : '+str(extrav.tipov)+' ---  Nro Chasis : '+str(extrav.nchasis)+' ---  Nro. Motor : '+str(extrav.nmotor)+'</dd>'+'<dd> Propietario : '+str(extrav.nro_doc)+' - '+str(extrav.propietario)+'</dd>'
                                detav=tituv+tablav+tiposv



                        tipo='<br><dd><u>'+str(eles.tipo)+'</u></dd>'

                        rubro=' Elemento/s '+str(eles.tipo)
                        rubros='Rubro y Categoria :'+str(eles.rubro)+' --- '+str(eles.categoria)
                        canti=' Cantidad : '+str(eles.cantidad)+' --- '+str(eles.unidadmed)
                        obse=' Observaciones : '+str(eles.descripcion.encode("utf8"))


                        if deta:
                             if detav:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+detav+'<br>'
                             else:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+deta+'<br>'

                        else:
                             if detav:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+detav+'<br>'
                             else:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'


                        eleminvo.append(eleme)
                        i=i+1

                    for ja in eleminvo:
                        elementos=elementos+ja


            #datos del preventivos
            datos=Preventivos.objects.get(id=idprev)
            nro=datos.nro
            anio=datos.anio
            fecha_denuncia=datos.fecha_denuncia
            fecha_carga=datos.fecha_carga
            fecha_cierre=datos.fecha_cierre
            caratula=datos.caratula
            actuante=datos.actuante
            preventor=datos.preventor
            autoridades= datos.autoridades.values_list('descripcion',flat=True)
            autoridada= amplia.autoridades.values_list('descripcion',flat=True)

            dependencia=datos.dependencia.descripcion
            unidadreg=datos.dependencia.unidades_regionales.descripcion
            idprev=idprev
            #envio de datos al template updatehechos.html
            jerarqui_a=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=actuante).values('jerarquia_id'))
            jerarqui_p=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=preventor).values('jerarquia_id'))
            form1=Hechos.objects.filter(preventivo=idprev)
            today = datetime.datetime.now()
            info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
                     'caratula':caratula,'involus':involus,'involuscra':involuscra,'datosper':datosper,'tienepersonas':tienepersonas,
                     'actuante':actuante,'today':today,'datosgral':datosgral,'i':i,'tieneelementos':tieneelementos,
                     'preventor':preventor,'jerarqui_a':jerarqui_a,'jerarqui_p':jerarqui_p,'depe':depe,'elementos':elementos,
                     'autoridades':autoridades,'personas':personas,'lugar':lugar,'lati':lati,'longi':longi,'autoridada':autoridada,
                     'errors': errors, 'grabo':grabo,'form':form, 'ciudad': ciudad,'condiciones':condiciones,
                     'state':state, 'continua':continua,'idprev':idprev,'countinvolus':countinvolus,'fecha_cierre':fecha_cierre,
                     'unidadreg':unidadreg,'dependencia':dependencia,'unireg':unireg,'amplia':amplia,
                     'destino': destino,'form1':form1,'tamaño':5,}

            #return render(request,'./preventivoi.html', info, context_instance=RequestContext(request))


            return render(request,'./ampliacioni.html', info)



    #datos del preventivos
    datos=Preventivos.objects.get(id=idprev)
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    autoridades= datos.autoridades.values_list('descripcion',flat=True)
    autoridada= amplia.autoridades.values_list('descripcion',flat=True)
    fecha_cierre=datos.fecha_cierre
    dependencia=datos.dependencia.descripcion
    unidadreg=datos.dependencia.unidades_regionales.descripcion
    idprev=idprev
    #envio de datos al template updatehechos.html
    jerarqui_a=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=actuante).values('jerarquia_id'))
    jerarqui_p=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=preventor).values('jerarquia_id'))
    form1=Hechos.objects.filter(preventivo=idprev)
    today = datetime.datetime.now()
    info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
                'caratula':caratula,'idprev':idprev,'tienehecho':tienehecho,'tienelugar':tienelugar,
                'actuante':actuante,'elementos':elementos,'autoridada':autoridada,'fecha_cierre':fecha_cierre,
                'preventor':preventor,'jerarqui_a':jerarqui_a,'jerarqui_p':jerarqui_p,'depe':depe,
                'autoridades':autoridades,'countinvolus':countinvolus,'unireg':unireg,
                'errors': errors,'unidadreg':unidadreg,'dependencia':dependencia,'tienepersonas':tienepersonas,
                'state':state,
                'destino': destino,}
    #return render(request,'./preventivoi.html', info, context_instance=RequestContext(request))

    return render(request,'./preventivoi.html', info)


@login_required
@group_required(["policia","investigaciones","visita","radio"])
def verampli(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    form=SearchPreveForm()
    anio=''
    nro=''
    fecha_carga=''
    fecha_cargah=''
    titulo=""
    todos=[]
    total='no'
    errors=[]
    jurisdi=''
    if request.POST.get('search')=="Buscar":
         form=SearchPreveForm(request.POST, request.FILES)
         titulo=request.POST.get('titulo')
         fecha_carga=request.POST.get('fecha_cargas')
         fecha_cargah=request.POST.get('fecha_cargah')
         ureg=request.POST.get('ureg')
         depe=request.POST.get('depe')
         unidadregi=Dependencias.objects.get(descripcion__contains=request.user.userprofile.depe.descripcion)
         jurisdi=unidadregi.ciudad.descripcion

         if titulo and not ureg and not depe:
                query_string=titulo
                entry_query = get_query(query_string, ['titulo',])

                filtro=Ampliacion.objects.filter(entry_query).order_by('fecha','titulo')
                if filtro not in todos:
                        todos=[]
                        todos.append(filtro)




         #En el caso que solamente consulte ureg
         #si solo tengo ureg primero filtro en todas las dependencias de esa unidad
         #depes=Dependencias.objects.filter(unidades_regionales=ureg)
         #genero un arreglo ej. prev[] hago un for depende in depes agrego a prev.append(preventivos.objects.filter(dependencia=depende))
         if ureg and not depe and not fecha_carga:
            depes=Dependencias.objects.filter(unidades_regionales=ureg)
            if titulo:
                 for son in depes:
                         #preventivos q estan en las dependencias
                         estan=Preventivos.objects.filter(dependencia=son).order_by('anio','nro','dependencia')
                         for c in estan:
                                 #buscar por cada preventivo en ampliaciones
                                 am=Ampliacion.objects.all().filter(preventivo=c.id,titulo__icontains=titulo).order_by('fecha',)
                                 todos.append(am)
            else:
                            for son in depes:
                                 #preventivos q estan en las dependencias
                                 estan=Preventivos.objects.filter(dependencia=son).order_by('id')
                                 for c in estan:

                                     am=Ampliacion.objects.all().filter(preventivo=c.id).order_by('fecha',)

                                     if am not in todos:
                                         todos=[]
                                         todos.append(am)

         else:
            if fecha_carga and fecha_cargah and ureg and not depe:
                depes=Dependencias.objects.filter(unidades_regionales=ureg)
                if titulo:
                     for son in depes:
                         #preventivos q estan en las dependencias
                         estan=Preventivos.objects.filter(dependencia=son).order_by('anio','nro','dependencia')
                         for c in estan:
                                 #buscar por cada preventivo en ampliaciones
                                 am=Ampliacion.objects.all().filter(preventivo=c.id,titulo__icontains=titulo).order_by('fecha',)
                                 todos.append(am)
                else:
                    for son in depes:
                        #preventivos q estan en las dependencias
                        estan=Preventivos.objects.filter(dependencia=son).order_by('anio','nro','dependencia')
                        for c in estan:
                             #buscar por cada preventivo en ampliaciones
                             am=Ampliacion.objects.all().filter(preventivo=c.id,fecha__range =(datetime.datetime.strptime(fecha_carga,"%d/%m/%Y").date(),datetime.datetime.strptime(fecha_cargah,"%d/%m/%Y").date())).order_by('fecha')
                             todos.append(am)
            else:
                if fecha_carga  and ureg and not depe:
                     depes=Dependencias.objects.filter(unidades_regionales=ureg)
                     for son in depes:
                            #preventivos q estan en las dependencias
                            estan=Preventivos.objects.filter(dependencia=son).order_by('id')
                            for c in estan:

                                 am=Ampliacion.objects.all().filter(preventivo=c.id,fecha = datetime.datetime.strptime(fecha_carga,"%d/%m/%Y").date()).order_by('fecha')
                                 todos.append(am)

                else:
                    if fecha_carga and fecha_cargah:
                            #preventivos q estan en las dependencias
                            am=Ampliacion.objects.all().filter(fecha__range =(datetime.datetime.strptime(fecha_carga,"%d/%m/%Y").date(), datetime.datetime.strptime(fecha_cargah,"%d/%m/%Y").date())).order_by('fecha')
                            todos.append(am)



         if ureg and depe:

            if fecha_carga and fecha_cargah:
                     depes=Dependencias.objects.filter(unidades_regionales=ureg)
                     for son in depes:
                            #preventivos q estan en las dependencias
                            estan=Preventivos.objects.filter(dependencia=son).order_by('id')
                            for c in estan:

                                 am=Ampliacion.objects.all().filter(preventivo=c.id,fecha__range =(datetime.datetime.strptime(fecha_carga,"%d/%m/%Y").date(), datetime.datetime.strptime(fecha_cargah,"%d/%m/%Y").date())).order_by('fecha')
                                 todos.append(am)

            else:
             if fecha_carga:
                        depes=Dependencias.objects.filter(unidades_regionales=ureg)
                        for son in depes:
                            #preventivos q estan en las dependencias
                            estan=Preventivos.objects.filter(dependencia=son).order_by('id')
                            for c in estan:

                                 am=Ampliacion.objects.all().filter(preventivo=c.id,fecha =datetime.datetime.strptime(fecha_carga,"%d/%m/%Y").date()).order_by('fecha')
                                 todos.append(am)
             else:
                    if titulo:
                         estan=Preventivos.objects.filter(dependencia_id=depe)
                         for c in estan:
                                todos.append(Ampliacion.objects.filter(preventivo=c.id,titulo__icontains=titulo).order_by('fecha'))
                    else:
                         #preventivos q estan en las dependencias
                         estan=Preventivos.objects.filter(dependencia_id=depe)
                         for c in estan:

                                 am=Ampliacion.objects.all().filter(preventivo=c.id).order_by('fecha')
                                 todos.append(am)


         else:
            if ureg:
             depes=Dependencias.objects.filter(unidades_regionales=ureg)

             if fecha_carga and fecha_cargah:

                     for son in depes:
                            fil=Preventivos.objects.filter(dependencia=son, fecha_carga__range =(datetime.datetime.strptime(fecha_carga,"%d/%m/%Y").date(),datetime.datetime.strptime(fecha_cargah,"%d/%m/%Y").date())).order_by('anio','nro','dependencia')
             else:
                if fecha_carga:

                     for son in depes:
                            fil=Preventivos.objects.filter(dependencia=son, fecha_carga=datetime.datetime.strptime(fecha_carga,"%d/%m/%Y").date()).order_by('anio','nro','dependencia')
                else:
                    if titulo:
                     todos=[]
                     for son in depes:
                         #preventivos q estan en las dependencias
                         estan=Preventivos.objects.filter(dependencia=son).order_by('anio','nro','dependencia')
                         for c in estan:
                                 #buscar por cada preventivo en ampliaciones
                                 am=Ampliacion.objects.all().filter(preventivo=c.id,titulo__icontains=titulo).order_by('fecha',)
                                 todos.append(am)
                    else:
                        for son in depes:
                                 #preventivos q estan en las dependencias
                                 estan=Preventivos.objects.filter(dependencia=son).order_by('anio','nro','dependencia')
                                 for c in estan:
                                     prev=c.id
                                     am=Ampliacion.objects.all().filter(preventivo=prev).order_by('fecha',)
                                     todos.append(am)


         #hago filtro si viene ureg y depe

         #aqui hago filtro si viene fecha de carga con cualquier otro valor concatenar arreglos con append

         if fecha_carga and fecha_cargah and not ureg and not depe:
                fecha_cargas=datetime.datetime.strptime(fecha_carga,"%d/%m/%Y").date()
                fecha_cargah=datetime.datetime.strptime(fecha_cargah,"%d/%m/%Y").date()
                if titulo:
                     am=Ampliacion.objects.filter(fecha__range=(fecha_cargas,fecha_cargah),titulo__icontains=titulo).order_by('fecha')
                     todos.append(am)
                else:
                     todos=[]
                     todos.append(Ampliacion.objects.all().filter(fecha__range=(fecha_cargas,fecha_cargah)).order_by('fecha'))
                                #'id','nro','anio','caratula','fecha_carga'))
         else:
             if fecha_carga and not ureg and not depe:
                fecha_cargas=datetime.datetime.strptime(fecha_carga,"%d/%m/%Y").date()
                #fecha_cargah=datetime.datetime.strptime(fecha_cargah,"%d/%m/%Y").date()
                if titulo:
                     #preventivos q estan en las dependencias
                     estan=Preventivos.objects.filter(dependencia=son).order_by('anio','nro','dependencia')
                     for c in estan:
                            #buscar por cada preventivo en ampliaciones
                            am=Ampliacion.objects.get(preventivo==c.id)
                            if filti not in todos:
                                 todos=[]
                                 todos.append(filti)
                else:

                     am=Ampliacion.objects.all().filter(fecha = fecha_cargas).order_by('fecha',)
                     todos.append(am)




    else:
        if request.POST.get('reset')=="Limpiar":
             todos=''
             total='si'

        else:

                 todos.append(Preventivos.objects.all())
                 total=Preventivos.objects.all().count


    info={'fecha_carga':fecha_carga,'fecha_cargah':fecha_cargah,'jurisdi':jurisdi,
    'titulo':titulo,'todos':todos,'total':total,'errors':errors,
    'state':state,'destino': destino,'form':form}

    return render(request,'./seeampli.html',info)


@login_required
@transaction.atomic
@group_required(["policia","investigaciones","radio"])
def eleampli(request,idhecho,elemento,idamp):
    state= request.session.get('state')
    destino= request.session.get('destino')

    hecho = Hechos.objects.get(id=idhecho)
    preventivo=Preventivos.objects.get(id=hecho.preventivo_id)
    depe= preventivo.dependencia
    elementox = Elementos.objects.get(id=elemento)
    amplia = Ampliacion.objects.get(id=idamp)
    form = ElementosForm(instance=elementox)
    formar = ArmasForm()
    obdata=''
    tieneob=False
    tiene=False
    tienecar=False
    formar=[]
    formv=[]
    errors=[]



    if request.POST.get('dele'):
            elementosin=Elementos.objects.filter(id=request.POST.get('dele'))

            if elementosin:
                if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion.find('INVESTIGACIONES')>=0  or request.user.userprofile.depe.descripcion.find('RADIO')>=0:
                        obs="elemento borrado por usuario : "+request.user.username

                        Elementos.objects.filter(id = request.POST.get('dele')).update(borrado='S',observaciones=obs)

                else:
                        errors='No se puede borrar elementos de un Hecho de otras dependencias.'
            else:
                        errors='No se existe elemento involucrado.'
            form = ElementosForm()
            formar = ArmasForm()

    if request.POST.get('button') == 'Guardar':
        form = ElementosForm(request.POST)
        formar=ArmasForm(request.POST)
        formv=VehiculosForm(request.POST)

        if form.is_valid():
            elemento = Elementos()
            elemento.tipo           = form.cleaned_data['tipo']
            elemento.rubro          = form.cleaned_data['rubro']
            elemento.categoria      = form.cleaned_data['categoria']
            elemento.cantidad       = form.cleaned_data['cantidad']
            elemento.unidadmed      = form.cleaned_data['unidadmed']
            elemento.descripcion    = form.cleaned_data['descripcion']
            elemento.hechos         = hecho
            elemento.ampliacion_id     = idamp

            if elemento.rubro.descripcion=='VEHICULOS' or elemento.rubro.descripcion== 'AUTOMOTORES':

                    if formv.is_valid():
                     elemento.save()
                     vehicle = Vehiculos()
                     vehicle.idmarca         = formv.cleaned_data['idmarca']
                     vehicle.modelo          = formv.cleaned_data['modelo']
                     vehicle.anio            = formv.cleaned_data['anio']
                     vehicle.tipov           = formv.cleaned_data['tipov']
                     vehicle.dominio         = formv.cleaned_data['dominio']
                     vehicle.nmotor          = formv.cleaned_data['nmotor']
                     vehicle.nchasis         = formv.cleaned_data['nchasis']
                     vehicle.nro_doc         = formv.cleaned_data['nro_doc']
                     vehicle.propietario     = formv.cleaned_data['propietario']
                     nueva_marcav            = formv.cleaned_data['nueva_marcav']
                     if nueva_marcav:
                        marca = RefTrademark()
                        marca.descripcion = nueva_marcav
                        marca.save()
                        vehicle.idmarca       = marca
                     vehicle.save()
                     elevehi = Elementoscars()
                     elevehi.idelemento        = elemento
                     elevehi.idvehiculo        = vehicle
                     elevehi.save()
                     formv = VehiculosForm()
                     errors.append('Elementos guardados')
            else:
                 if elemento.categoria.descripcion =='FUEGO' or elemento.categoria.descripcion=='DE FUEGO':
                         if formar.is_valid():
                             elemento.save()
                             armas=Armas()
                             armas.tipos    = formar.cleaned_data['tipos']
                             armas.subtipos = formar.cleaned_data['subtipos']
                             armas.sistema_disparo = formar.cleaned_data['sistema_disparo']
                             armas.marcas = formar.cleaned_data['marcas']
                             armas.calibre = formar.cleaned_data['calibre']
                             armas.modelo = formar.cleaned_data['modelo']
                             armas.nro_arma=formar.cleaned_data['nro_arma']
                             armas.nro_doc = formar.cleaned_data['nro_doc']
                             armas.propietario = formar.cleaned_data['propietario']
                             nueva_marca       = formar.cleaned_data['nueva_marca']

                             if nueva_marca:
                                 marca = RefTrademark()
                                 marca.descripcion = nueva_marca
                                 marca.save()
                                 armas.marcas = marca

                             armas.save()

                             elearmas=Elementosarmas()
                             elearmas.idelemento=elemento
                             elearmas.idarma=armas
                             elearmas.save()
                             form = ElementosForm()
                             formar=ArmasForm()
                             errors.append('Elementos guardados')
                 else:
                             elemento.save()
                             errors.append('Elementos guardados')
                             form=ElementosForm()
        else:
             errors.append('Faltan Datos verifique')

    if request.POST.get('guardarubro') == 'Guardar':
        formrub = ItemForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        mostrar="no"
        if not descripcion:
            errors.append('Ingrese un Rubro')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<100):
                    errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
             else:
                 if formrub.is_valid():
                        formrub.save()
    if request.POST.get('guardacategoria') == 'Guardar':
        formcat = CategoryForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        rubro = request.POST.get('rubro')
        mostrar="no"
        if not descripcion:
            errors.append('Ingrese una categoria')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<100):
                    errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
             else:
                if not rubro:
                    errors.append('Seleccione Rubro')
                else:
                    if formcat.is_valid():
                        formcat.save()
    if request.POST.get('guardamedida') == 'Guardar':
        formumed = UnidadMedidasForm(request.POST, request.FILES)

        descripcion = request.POST.get('descripcion')
        mostrar="no"
        if not descripcion:
            errors.append('Ingrese unidad de medida')
        else:
             if formumed.is_valid():
                formumed.save()


    if len(Elementosarmas.objects.filter(idelemento_id=elementox.id))>0:

                 idar = Elementosarmas.objects.filter(idelemento_id=elementox.id).values('idarma')
                 obdata=Armas.objects.get(id=idar)
                 formar=ArmasForm(instance=obdata)
                 tiene=True

    if len(Elementoscars.objects.filter(idelemento_id=elementox.id))>0:

                 idar = Elementoscars.objects.filter(idelemento_id=elementox.id).values('idvehiculo')
                 obdata=Vehiculos.objects.get(id=idar)
                 formv=VehiculosForm(instance=obdata)
                 tienecar=True

    lista = Elementos.objects.filter(hechos = hecho.id,cargado_prev=True,ampliacion__isnull=True,borrado__isnull=True)
    listam= Elementos.objects.filter(hechos = hecho.id,ampliacion_id=idamp,borrado__isnull=True)

    if len(Elementos.objects.filter(hechos = hecho.id,borrado__isnull=True))>0:
        tieneob = True

    hechos = Hechos.objects.get(id = idhecho)
    idciu = hechos.preventivo.dependencia.ciudad_id
    depe = hechos.preventivo.dependencia
    ids = Preventivos.objects.get(id = hechos.preventivo_id)
    idprev=ids
    ftiposdelitos=DelitoForm()
    modos=RefModosHechoForm()
    datos=Preventivos.objects.get(id=ids.id)
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    fecha_autorizacion=datos.fecha_autorizacion
    autoridades= datos.autoridades.values_list('descripcion',flat=True)


    values={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,
            'caratula':caratula,'formar':formar,'tiene':tiene,'idele':elemento,'vehiculo':formv,
            'actuante':actuante,'idprev':idprev,'tieneob':tieneob,'tienecar':tienecar,
            'preventor':preventor,'fecha_autorizacion':fecha_autorizacion,'amplia':amplia,
            'autoridades':autoridades,'idamp':idamp,'listam':listam,
            'destino': destino,'state':state,
            'preventivo':hecho.preventivo,
            'hecho':hecho,
            'form':form,
            'lista':lista,
            'elemento':elementox,
    }
    return render(request,'./objampli.html',values)


@login_required
def helpassword(request):
    state= request.session.get('state')
    destino= request.session.get('destino')

    values={'destino': destino,'state':state,}

    return render(request,'./helpassw.html',values)


@login_required
@group_required(["policia","investigaciones","radio"])
def amplia_pers(request,idprev,idamp):

    state= request.session.get('state')
    destino= request.session.get('destino')
    preventivo = Preventivos.objects.get(id=idprev)
    involucrados = preventivo.hecho.involu.all().order_by('roles')
    enprev = involucrados.filter(cargado_prev=True,ampliacion__isnull=True)
    modif_amp = involucrados.filter(cargado_prev=True,ampliacion__isnull=False)
    enamp = involucrados.filter(cargado_prev=False)
    errors = []
    filtro = 0;
    persoinvform=personainv = persona = None
    if request.POST.get('dele'):
        personainv = PersInvolucradas.objects.get(id=request.POST.get('dele'))
        try:
            personainv.delete()
        except IntegrityError:
                                    errors.append('')
    if request.POST.get('ver'):
        personainv = PersInvolucradas.objects.get(id=request.POST.get('ver'))
        persona = Personas.objects.get(id=personainv.persona.id)
        persoinvform = PersInvolucradasForm(instance=personainv)
    if request.POST.get('modificar') == 'Modificar':
        persinv = PersInvolucradasForm(request.POST, request.FILES)
        if persinv.is_valid():
            personainv                  = PersInvolucradas.objects.get(id=request.POST.get('idinv'))
            per                         = PersInvolucradas()


            if personainv.ampliacion_id and not Ampliacion.objects.get(id=idamp).fin_edicion:

                personainv.roles                   = persinv.cleaned_data['roles']


                if not personainv.fechahoradetencion:
                    personainv.fechahoradetencion    = persinv.cleaned_data['fechahoradetencion']
                if persinv.cleaned_data['fechahoralibertad'] :
                    personainv.fechahoralibertad     = persinv.cleaned_data['fechahoralibertad']
                    personainv.detenido              = 'no'
                    dete=Detenidos.objects.filter(persona = personainv.persona.id)
                    if dete:
                       Detenidos.objects.filter(persona = personainv.persona.id).update(libertad='S',borrado='',observaciones=request.user.username+'asigna fecha de libertad')
                    else:
                       detenidos.persona=persona
                       detenidos.hechos  = hechos
                       detenidos.fechahoradetencion = formr.cleaned_data['fechahoradetencion']
                       detenidos.libertad=''
                       detenidos.save()
                personainv.save()
            elif not personainv.ampliacion_id and not Ampliacion.objects.get(id=idamp).fin_edicion:
                per.hechos                  = Hechos.objects.get(id=personainv.hechos.id)
                per.roles                   = persinv.cleaned_data['roles']
                per.persona                 = personainv.persona
                per.juridica                = personainv.juridica
                per.razon_social            = personainv.razon_social
                per.cuit                    = personainv.cuit
                per.nrocuit                 = personainv.nrocuit
                per.menor                   = personainv.menor
                per.infraganti              =  personainv.infraganti



                per.fechahoradetencion    = persinv.cleaned_data['fechahoradetencion']

                if persinv.cleaned_data['fechahoralibertad'] :
                    per.fechahoralibertad     = persinv.cleaned_data['fechahoralibertad']
                    per.detenido              = 'si'
                else:
                    per.fechahoralibertad     = personainv.fechahoralibertad
                    per.detenido              = personainv.detenido
                per.cargado_prev            = personainv.cargado_prev
                per.ampliacion              = Ampliacion.objects.get(id=idamp)
                dete=Detenidos.objects.filter(persona = personainv.persona.id)

                if dete:
                   Detenidos.objects.filter(persona = personainv.persona.id).update(fechahoradetencion=persinv.cleaned_data['fechahoradetencion'],libertad='N',borrado='',observaciones=request.user.username+'se cambio el rol en ampliacion')
                else:
                   detenidos.persona=persi.persona
                   detenidos.hechos  = hechos
                   detenidos.fechahoradetencion = formr.cleaned_data['fechahoradetencion']
                   detenidos.libertad=''
                   detenidos.save()

                per.save()
    if request.POST.get('search')=="Buscar":
        texto=request.POST.get('texto')

        if texto:
            query_string=texto
            entry_query = get_query(query_string, ['nro_doc', 'apellidos','nombres',])
        else:
            query_string='%'
            entry_query = get_query(query_string, ['nro_doc', 'apellidos','nombres',])



        filtro=Personas.objects.filter(entry_query)


    values={'destino'       :      destino,
                    'state'         :      state,
                    'preventivo'    :      preventivo,
                    'involucrados'  :      involucrados,
                    'persona'       :      persona,
                    'persoinv'      :      personainv,
                    'persoinvform'  :      persoinvform,
                    'enprev'        :      enprev,
                    'modif_amp'     :      modif_amp,
                    'enamp'         :      enamp,
                    'idamp'         :      idamp,
                    'filtro'        :      filtro
                    }

    return render(request,'./amplipers.html',values)

@login_required
@group_required(["policia","investigaciones","radio"])
def amplia_per(request,idprev,idamp,idper):

    state= request.session.get('state')
    destino= request.session.get('destino')
    preventivo = Preventivos.objects.get(id=idprev)
    involucrados = preventivo.hecho.involu.all().order_by('roles')
    enprev = involucrados.filter(cargado_prev=True,ampliacion__isnull=True)
    modif_amp = involucrados.filter(cargado_prev=True,ampliacion__isnull=False)
    enamp = involucrados.filter(cargado_prev=False)
    formp = PersonasForm()
    dom = DomiciliosForm()
    formr = PersInvolucradasForm()
    formpa=PadresForm()
    hechos = preventivo.hecho
    errors=[]
    siexistepoli=""
    estadetes=True
    if request.POST.get('dele'):
        personainv = PersInvolucradas.objects.get(id=request.POST.get('dele'))
        try:
            personainv.delete()
        except IntegrityError:
                                    errors.append('')
    if request.POST.get('grabar') == 'Guardar':
     formp = PersonasForm(request.POST, request.FILES)          #obtiene los datos de la persona en un formulario persona
     dom = DomiciliosForm(request.POST,request.FILES)           #obtiene los datos del domicilio en un formulario domicilio
     formr = PersInvolucradasForm(request.POST,request.FILES)   #obtiene los datos de persona involucrada en un formulario persona involucrada
     formpa = PadresForm(request.POST,request.FILES)

     #fecha_detencion=datetime.datetime.strptime(request.POST.get('fechahoradetencion'), '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y')
     anionac=datetime.datetime.strptime(request.POST.get('fecha_nac'),'%d/%m/%Y').strftime('%Y')
     anioactual=datetime.datetime.now()
     aniohoy=anioactual.today().year
     if anionac>=1900:
      dife=aniohoy-int(anionac)
      if request.POST.get('fechahoradetencion'):
            fechadete=datetime.datetime.strptime(request.POST.get('fechahoradetencion'), '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y')
            fecha_denuncia=preventivo.fecha_denuncia.strftime('%d/%m/%Y')

            fd = time.strptime(fecha_denuncia, "%d/%m/%Y")
            fdet = time.strptime(fechadete, "%d/%m/%Y")

            if fechadete<fecha_denuncia:
                    errors.append('La Fecha y hora de Detencion nunca debe ser menor a la de Denuncia del Hecho sucedido')
                    mostrar="no"
                    estadete="no"
                    personas = Personas.objects.get(id=idper)
                    formp = PersonasForm(instance=personas)
                    formr = PersInvolucradasForm()
                    valuesi={'destino'       :      destino,
                    'state'         :      state,
                    'preventivo'    :      preventivo,
                    'involucrados'  :      involucrados,
                    'persona'       :      persona,
                    'enprev'        :      enprev,
                    'modif_amp'     :      modif_amp,
                    'enamp'         :      enamp,
                    'idamp'         :      idamp,
                    'formp'         :      formp,
                    'dom'           :      dom,
                    'formr'         :      formr,
                    'formpa'        :      formpa,
                    'errors'        : errors,
                    'idper'         : idper
                    }
                    return render(request,'./amplipers.html',valuesi)

      if idper!='0':
         perso=Personas.objects.get(id=idper)
         fil=Padres.objects.filter(persona=perso.id)
         if fil:
                papis= Padres.objects.get(persona = perso.id)
                formpa = PadresForm(instance=papis)
         else:
                formpa=PadresForm()
                papis=Padres()
         #buscar en personal
         findpoli=Personal.objects.filter(persona_id=perso.id)
         if findpoli:
                siexistepoli=True
      else:

         perso=Personas()
         papis=Padres()

         iddom='1'
      if len(Domicilios.objects.filter(personas = idper)) > 0:
         domicilios = Domicilios.objects.filter(personas = idper)[0]
         iddom=domicilios.id
      else:
         iddom='1'
         domicilios=Domicilios()



      if formp.is_valid() or idper!='0':
            if formp.is_valid():

             perso.apellidos  = formp.cleaned_data['apellidos']
             perso.nombres    = formp.cleaned_data['nombres']
             perso.tipo_doc   = formp.cleaned_data['tipo_doc']
             perso.nro_doc    = formp.cleaned_data['nro_doc']
             perso.fecha_nac  = formp.cleaned_data['fecha_nac']
             perso.sexo_id    = formp.cleaned_data['sexo_id']
             perso.pais_nac   = formp.cleaned_data['pais_nac']
             perso.ciudad_nac = formp.cleaned_data['ciudad_nac']
             perso.pais_res   = formp.cleaned_data['pais_res']
             perso.ciudad_res = formp.cleaned_data['ciudad_res']
             perso.ocupacion  = formp.cleaned_data['ocupacion']
             perso.alias      = formp.cleaned_data['alias']


             perso.estado_civil = formp.cleaned_data['estado_civil']
             idpoli=formp.cleaned_data['ocupacion']
             refpoli=RefOcupacion()
             if idpoli:
                refpoli=RefOcupacion.objects.get(descripcion=idpoli)


            else:

                 if request.POST.get('ocupacion')=='None' or request.POST.get('ocupacion')=='':
                     refpoli=RefOcupacion.objects.get(descripcion='EMPLEADO')
                     texto='EMPLEADO'
                 else:
                     idpoli=request.POST.get('ocupacion')
                     refpoli=RefOcupacion.objects.get(id=idpoli)
                     texto=refpoli.descripcion

            if dom.is_valid():


                     domicilios.barrio_codigo        = dom.cleaned_data['barrio_codigo']
                     domicilios.calle                = dom.cleaned_data['calle']
                     domicilios.altura               = dom.cleaned_data['altura']
                     domicilios.entre                = dom.cleaned_data['entre']
                     domicilios.fecha_desde          = dom.cleaned_data['fecha_desde']
                     domicilios.fecha_hasta          = dom.cleaned_data['fecha_hasta']
                     domicilios.fecha_actualizacion  = dom.cleaned_data['fecha_actualizacion']
                     domicilios.tipos_domicilio      = dom.cleaned_data['tipos_domicilio']
                     domicilios.ref_zona             = dom.cleaned_data['ref_zona']
                     domicilios.departamento         = dom.cleaned_data['departamento']
                     domicilios.piso                 = dom.cleaned_data['piso']
                     domicilios.lote                 = dom.cleaned_data['lote']
                     domicilios.sector               = dom.cleaned_data['sector']
                     domicilios.manzana              = dom.cleaned_data['manzana']


                     if idper!='0':
                            persom=Personas.objects.get(id=idper)


                            if formp.is_valid() or idper!='0':
                                persom.ciudad_res = formp.cleaned_data['ciudad_res']
                                persom.ocupacion  = formp.cleaned_data['ocupacion']
                                persom.alias      = formp.cleaned_data['alias']
                                persom.estado_civil = formp.cleaned_data['estado_civil']
                                try:


                                    persom.save()

                                except IntegrityError:
                                    errors.append('')
                            else:

                                 mostrar='si'
                                 errors.append(formp.errors.as_text)

                     else:

                             perso.save()




                     if idper!='0':

                                     idpersu=Personas.objects.get(id=idper)

                                     personas=idpersu
                     else:
                                     idpersu=Personas.objects.get(id=perso.id)

                                     personas=idpersu

                     domicilios.personas             = personas
                     domicilios.ref_ciudades         = formp.cleaned_data['ciudad_res']

                     domicilios.save()

                     if refpoli:
                        if refpoli.descripcion.find('POLICI')>=0:
                                    policia=Personal()
                                    policia.persona_id = personas
                                    policia.credencial=0
                                    try:
                                        policia.save()
                                    except IntegrityError:
                                        errors.append('')
                        else:
                                 if siexistepoli:
                                        #borro esa persona en personal
                                        borrar=Personal.objects.get(persona_id=personas).delete()

                     papis.persona=personas
                     papis.padre_apellidos=request.POST.get('padre_apellidos')
                     papis.padre_nombres=request.POST.get('padre_nombres')
                     papis.madre_apellidos=request.POST.get('madre_apellidos')
                     papis.madre_nombres=request.POST.get('madre_nombres')
                     try:
                            papis.save()
                     except IntegrityError:
                         errors.append('Datos existente en Padres')

                     if formr.is_valid():

                         persoin=PersInvolucradas()
                         detenidos = Detenidos()


                         persoin.persona=personas
                         detenidos.persona = personas
                         persoin.hechos=hechos

                         persoin.roles = formr.cleaned_data['roles']

                         if dife>=18:
                            persoin.menor='no'
                         else:
                            persoin.menor='si'

                         if 'APREHENDIDO' in persoin.roles.descripcion or  'APRENDIDO' in persoin.roles.descripcion or 'DETENIDO' in persoin.roles.descripcion:
                             persoin.detenido = formr.cleaned_data['detenido']
                         else:
                             persoin.detenido='no'
                         persoin.cargado_prev = False
                         persoin.ampliacion = Ampliacion.objects.get(id=idamp)

                         if persoin.detenido=='si':


                                 detenidos.hechos  = hechos
                                 detenidos.fechahoradetencion = formr.cleaned_data['fechahoradetencion']
                                 persoin.fechahoradetencion = formr.cleaned_data['fechahoradetencion']
                                 detenidos.save()

                         persoin.infraganti = formr.cleaned_data['infraganti']
                         persoin.juridica = formr.cleaned_data['juridica']
                         if persoin.juridica=='si':
                                persoin.razon_social = formr.cleaned_data['razon_social']
                                persoin.cuit = formr.cleaned_data['cuit']
                                persoin.nrocuit = formr.cleaned_data['nrocuit']
                         else:
                                persoin.razon_social = 'SIN DESCRIPCION'


                         try:


                                    mostrar='0'
                                    persoin.save()

                         except IntegrityError:
                             errors.append('Datos inexistente en Carateristicas segun el Rol')




                     else:
                         errors.append('Error faltan datos en seccion de Rol de la Persona')
                         filtros=Personas.objects.filter(id = personas.id)
                         if filtros not in todos:
                             todos.append(filtros)

            else:
                         mostrar='no'


      else:
        mostrar='no'


     else:
        mostrar='no'
     return HttpResponseRedirect('../')
    if idper != '0':
        personas = Personas.objects.get(id=idper)
        formp = PersonasForm(instance=personas)
        domicilios = Domicilios()
        dom = DomiciliosForm()

        formr = PersInvolucradasForm()
        filt= Padres.objects.filter(persona = personas.id)
        if filt:
             idpapis= Padres.objects.get(persona = personas.id)
             formpa = PadresForm(instance=idpapis)

        if len(Domicilios.objects.filter(personas = idper)) > 0:
         domicilios = Domicilios.objects.filter(personas = idper)[0]
         dom = DomiciliosForm(instance = domicilios)
         dom.fields['barrio_codigo'].queryset = RefBarrios.objects.filter(ciudad=personas.ciudad_res)
         dom.fields['calle'].queryset = dom.fields['entre'].queryset= RefCalles.objects.filter(ciudad=personas.ciudad_res)
        else:

         dom = DomiciliosForm()

         dom.fields['barrio_codigo'].queryset = RefBarrios.objects.filter(ciudad=personas.ciudad_res)
         dom.fields['calle'].queryset = dom.fields['entre'].queryset= RefCalles.objects.filter(ciudad=personas.ciudad_res)
        values={'destino'       :      destino,
                    'state'         :      state,
                    'preventivo'    :      preventivo,
                    'involucrados'  :      involucrados,
                    'persona'       :      persona,
                    #'persoinv'      :      personainv,
                    #'persoinvform'  :      persoinvform,
                    'enprev'        :      enprev,
                    'modif_amp'     :      modif_amp,
                    'enamp'         :      enamp,
                    'idamp'         :      idamp,
                    #'filtro'        :      filtro,
                    'formp'         :      formp,
                    'dom'           :      dom,
                    'formr'         :      formr,
                    'formpa'        :      formpa
                    }


    return render(request,'./amplipers.html',values)

@login_required
@group_required(["policia","investigaciones","radio"])
def finalizar(request,idprev,idamp):
    state= request.session.get('state')
    destino= request.session.get('destino')
    preventivo = Preventivos.objects.get(id = idprev)
    depe = preventivo.dependencia
    hecho = Hechos.objects.get(preventivo_id=preventivo.id)
    ampliaciones = preventivo.ampli.all()
    ampli = Ampliacion.objects.get(id=idamp)
    ampliacion = AmpliacionForm(instance=ampli)
    id_ciudad=Dependencias.objects.filter(descripcion__exact=depe).values('ciudad')
    ampliacion.fields['autoridades'].queryset=RefCiudades.objects.get(id=id_ciudad).ciu_autori.all()
    autoridades= preventivo.autoridades.all()
    autoridad=[]
    for seleccion in autoridades:
        ids=int(RefAutoridad.objects.get(descripcion=seleccion).id)
        autoridad.append(ids)

    ampliacion.fields['autoridades'].initial=autoridad
    ampli.fin_edicion=True
    ampli.save()
    finaliza=False

    values={'finaliza':finaliza,'id':idamp,'destino': destino,'state':state,'preventivo':preventivo,'ampliaciones':ampliaciones,'ampliacion':ampliacion}

    return render(request,'./ampliaciones.html',values)

@login_required
@group_required(["policia","investigaciones","radio"])
def enviar(request,idprev,idamp):
    state= request.session.get('state')
    destino= request.session.get('destino')
    preventivo = Preventivos.objects.get(id = idprev)
    depe = preventivo.dependencia
    hecho = Hechos.objects.get(preventivo_id=preventivo.id)
    ampliaciones = preventivo.ampli.all()
    ampli = Ampliacion.objects.get(id=idamp)
    ampliacion = AmpliacionForm(instance=ampli)
    id_ciudad=Dependencias.objects.filter(descripcion__exact=depe).values('ciudad')
    fecha_autorizacion=ampli.fecha_autorizacion
    grabarfa = Ampliacion.objects.filter(id = idamp).update(fecha_autorizacion=date.today())
    if len(Ampliacion.objects.filter(id=idamp))>0:
            amplia = Ampliacion.objects.get(id=idamp)
            tieneampli=True
            form=AmpliacionForm(instance=amplia)
            involuscra=[]
            eleminvo=[]
            datosper=""
            elementos=""
            involus=PersInvolucradas.objects.filter(ampliacion=amplia.id).all()
            eleinvo=Elementos.objects.filter(ampliacion=amplia.id,borrado__isnull=True).all()

            datosgral=""
            lugar=''
            lati=''
            longi=''
            condiciones=''
            perjuridica=''
            #Datos del lugar del hecho



            if len(PersInvolucradas.objects.filter(ampliacion=amplia.id).all())>0:
                 tienepersonas=True
                 countinvolus=PersInvolucradas.objects.filter(ampliacion=amplia.id).count()
                 for p in PersInvolucradas.objects.filter(ampliacion=amplia.id).all():

                     bandera,personai = funverifica(p.persona.id)
                     if p.menor=='':
                        p.menor="NO"
                     if p.juridica=='si':
                        if p.razon_social!=None:
                          perjuridica=str(p.razon_social)

                        if RefTipoDocumento.objects.get(id=p.cuit_id)!='Null':
                          perjuridica=perjuridica+'-'+str(RefTipoDocumento.objects.get(id=p.cuit_id))
                        if p.nrocuit!=0:
                          perjuridica=perjuridica+'-'+str(p.nrocuit)
                     domi=Personas.objects.get(id=p.persona.id).persodom.all()
                     if domi:
                        for l in Personas.objects.get(id=p.persona.id).persodom.all():
                         #datosgral=str(p.roles)+'-'+str(p)+' '+str(p.persona.tipo_doc)+' :'+str(p.persona.nro_doc)
                         dad=Personas.objects.get(id=p.persona.id).padre.all()

                         if dad:

                                for la in Personas.objects.get(id=p.persona.id).padre.all():
                                    


                                    roles='<u>'+str(p.roles)+'</u>'+' : '
                                    if bandera:
                                        if p.juridica=='si':
                                            persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')+str('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')
                                        else:
                                            persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')
                                    else:
                                        if p.juridica=='si':
                                           persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+'</dd>')+str('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')+'<dd>Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))
                                        else:
                                           persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+' , Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')

                                    domi='<dd>Reside en : '+str(p.persona.ciudad_res)+',  Domicilio : '+str(l.calle)+'  Nro.: '+str(l.altura)+'</dd>'
                                    if la.padre_apellidos or la.padre_nombres or la.madre_apellidos or la.madre_nombres:
                                         padys='<dd> Hijo de : '+str(la.padre_apellidos.encode("utf8"))+', '+str(la.padre_nombres.encode("utf8"))+' y de : '+str(la.madre_apellidos.encode("utf8"))+', '+str(la.madre_nombres.encode("utf8"))+'<br><br></dd>'
                                    else:
                                         padys='<dd>no registra datos de los padres'+'<br></dd>'
                                    datosgral=roles+persona+domi+padys
                         else:

                                 roles='<u>'+str(p.roles)+'</u>'+' : '
                                 if bandera:
                                    if p.juridica=='si':
                                        persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')+str('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')
                                    else:
                                        persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')
                                 else:
                                    if p.juridica=='si':
                                        persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+'</dd>')+str('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')+'<dd>Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))
                                    else:
                                        persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')

                                 #persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+',  Estado Civil : '+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')
                                 domi='<dd>Reside en : '+str(p.persona.ciudad_res)+',  Domicilio : '+str(l.calle)+'  Nro.: '+str(l.altura)+'</dd>'
                                 padys='<dd>no registra datos de los padres'+'<br></dd>'
                                 datosgral=roles+persona+domi+padys
                         involuscra.append(datosgral)
                     else:

                        roles='<u>'+str(p.roles)+'</u>'+' : '
                        if bandera:
                            if p.juridica=='si':
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')+str('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')
                            else:
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+str(personai)+'</dd>')
                        else:
                            if p.juridica=='si':
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+'</dd>')+str('<dd>'+'Personeria Juridica :'+str(perjuridica)+'</dd>')+'<dd>Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))
                            else:
                                persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+', Estado Civil :'+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')

                        #persona=str(p)+str('<dd>'+str(p.persona.tipo_doc)+': '+str(p.persona.nro_doc)+', Ocupacion :'+str(p.persona.ocupacion)+',  Estado Civil : '+' '+str(p.persona.estado_civil)+', Menor de Edad : '+str(p.menor.upper())+'<dd>Nacido en: '+str(p.persona.pais_nac)+', '+str(p.persona.ciudad_nac)+', Fecha Nac: '+str(p.persona.fecha_nac.strftime("%d/%m/%Y"))+'</dd>')
                        domi='<dd>no registra domicilio'+'</dd>'
                        padys='<dd>no registra datos de los padres'+'<br></dd>'
                        datosgral=roles+persona+domi+padys

                        involuscra.append(datosgral)
                 for i in involuscra:
                     datosper=datosper+i

                    #datosper.append(persona)
                 datosgral=''
                 obdata=[]
                 obdatav=[]
                 deta=''
                 detav=''
                 eleme=''
                 hay=[]


            if len(Elementos.objects.filter(ampliacion=amplia.id).all())>0:
                    countinvolus=Elementos.objects.filter(ampliacion=amplia.id).count()
                    i=1
                    for eles in eleinvo:
                        tieneelementos=True
                        obdata=[]
                        obdatav=[]
                        deta=''
                        detav=''

                        if len(Elementosarmas.objects.filter(idelemento=eles.id))>0:

                             idar = Elementosarmas.objects.filter(idelemento=eles.id).values('idarma')
                             tieneelementos=True
                             obdata=Armas.objects.filter(id=idar)
                             for extra in obdata:
                                titu='<ul><li><u> Carateristicas Generales : </u></li></ul>'
                                tabla='<dd> '+str(extra.subtipos)+'  ---  Tipo/s : '+str(extra.tipos)+'  ---  Sistema de Disparo : '+str(extra.sistema_disparo)+'   --- Marcas : '+str(extra.marcas)+'</dd>'
                                tipos='<dd> Calibre : '+str(extra.calibre)+'  --- Modelo : '+str(extra.modelo)+'  --- Nro Serie : '+str(extra.nro_arma)+'   ---  Propietario : '+str(extra.nro_doc)+' - '+str(extra.propietario)+'</dd>'
                                deta=titu+tabla+tipos

                        if len(Elementoscars.objects.filter(idelemento=eles.id))>0:
                             tieneelementos=True
                             idarv = Elementoscars.objects.filter(idelemento=eles.id).values('idvehiculo')

                             obdatav=Vehiculos.objects.filter(id=idarv)
                             for extrav in obdatav:
                                tituv='<ul><li><u> Carateristicas Generales : </u></li></ul>'
                                tablav='<dd> Marca/s : '+str(extrav.idmarca)+'  ---   Modelo : '+str(extrav.modelo)+'  ---  Dominio : '+str(extrav.dominio)+'   ---  Año : '+str(extrav.anio)+'</dd>'
                                tiposv='<dd> Tipo/s : '+str(extrav.tipov)+' ---  Nro Chasis : '+str(extrav.nchasis)+' ---  Nro. Motor : '+str(extrav.nmotor)+'</dd>'+'<dd> Propietario : '+str(extrav.nro_doc)+' - '+str(extrav.propietario)+'</dd>'
                                detav=tituv+tablav+tiposv



                        tipo='<br><dd><u>'+str(eles.tipo)+'</u></dd>'

                        rubro=' Elemento/s '+str(eles.tipo)
                        rubros='Rubro y Categoria :'+str(eles.rubro)+' --- '+str(eles.categoria)
                        canti=' Cantidad : '+str(eles.cantidad)+' --- '+str(eles.unidadmed)
                        obse=' Observaciones : '+str(eles.descripcion.encode("utf8"))


                        if deta:
                             if detav:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+detav+'<br>'
                             else:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+deta+'<br>'

                        else:
                             if detav:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'+detav+'<br>'
                             else:
                                 eleme='<br>'+str(i)+' --'+rubro+'<br>'+rubros+'<br>'+canti+'<br>'+obse+'<br>'


                        eleminvo.append(eleme)
                        i=i+1

                    for ja in eleminvo:
                        elementos=elementos+ja
            #datos del preventivos
            datos=Preventivos.objects.get(id=idprev)
            nro=datos.nro
            anio=datos.anio
            fecha_denuncia=datos.fecha_denuncia
            fecha_carga=datos.fecha_carga
            caratula=datos.caratula
            actuante=datos.actuante
            preventor=datos.preventor
            autoridades= datos.autoridades.values_list('descripcion',flat=True)
            autoridada= amplia.autoridades.values_list('descripcion',flat=True)
            fecha_cierre=datos.fecha_cierre
            dependencia=datos.dependencia.descripcion
            ciudad=datos.dependencia.ciudad
            unidadreg=datos.dependencia.unidades_regionales.descripcion
            idprev=idprev
            #envio de datos al template updatehechos.html
            jerarqui_a=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=actuante).values('jerarquia_id'))
            jerarqui_p=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=preventor).values('jerarquia_id'))
            autoridad=''
            for a in autoridada:
                 autoridad=autoridad+'*'+str(a)+'<br>'

            subject ='Ampliacion de Preventivo Nro : '+str(nro)+'/'+str(anio)+'--Dependencia : '+str(dependencia)+' Ciudad de : '+str(ciudad)
            from_email =  'divsistemasjp@policia.chubut.gov.ar'
            #to=str(infor)
            titulo2=str('Fecha Ampliacion : '+str(amplia.fecha.strftime("%d/%m/%Y"))+'<br>'+'Titulo : '+str(amplia.titulo)+'<br>'+' Autoridades a Informar : '+'<br>'+str(autoridad)+'<br>'+' Texto Ampliatorio : '+str(amplia.descripcion)+'<br>')

            titulo ='Datos del Preventivos '+'<br><hr>'+str('<u>'+'Preventivo Nro : '+str(nro)+'/'+str(anio)+'</u>')
            tresto='--Dependencia : '+str(dependencia)+' Ciudad de : '+str(ciudad)+'</u><br>'
            titulo1=str('Fecha de Denuncia : '+str(fecha_denuncia.strftime("%d/%m/%Y"))+'<br>'+'Fecha de Carga: '+str(fecha_carga.strftime("%d/%m/%Y"))+'<br>'+'Caratula :'+str(caratula.encode("utf8"))+'<br>'+'Actuante : '+str(jerarqui_a)+' '+str(actuante)+' --- '+' Preventor :'+str(jerarqui_p)+' '+str(preventor)+'<br>'+'Autoridades a informar :'+'<br><dd>'+str(autoridad)+'</dd><hr>')

            #ubicacion
            if datosper:
                    titulo3='Personas Involucradas: '+str(datosper)+'<br><br><hr>'
            else:
                    titulo3='Sin Personas involucradas'+'<br><br><hr>'
            if elementos:
                 titulo4='Elementos : '+str(elementos)+'<br><br><hr>'
            else:
                 titulo4="Sin Elementos"+'<br><br><hr>'
            text_content=titulo2+titulo3+titulo4+titulo+tresto+titulo1

            if request.user.userprofile.depe.descripcion != 'INVESTIGACIONES':
                informa=amplia.autoridades.values_list('email',flat=True)
                #agregar email 2jefeacei para que reciba los preventivos
                
                direcciones=[]
                indice=0
                nstring=''
                acumula=''
                envio=1
                for dire in informa:
                        direcciones.append(dire)
                        #direcciones.append('fydsoftware@gmail.com')
                        indice=0
                        nstring=''
                        if dire.find(',')>=0 or dire.find(';')>=0:

                             while indice < len(dire):

                                    if dire[indice] != ',' and dire[indice] != ';' and dire[indice]!='':
                                         nstring = nstring + dire[indice]
                                         indice = indice +1

                                    else:

                                         envio,nstring,subject,text_content,from_email=envioemail(envio,nstring,subject,text_content,from_email,request)
                                         indice=indice+1
                                         nstring=''

                             if nstring:

                                    envio,nstring,subject,text_content,from_email=envioemail(envio,nstring,subject,text_content,from_email,request)
                                    nstring=''

                        else:
                             nstring=dire
                             envio,nstring,subject,text_content,from_email=envioemail(envio,nstring,subject,text_content,from_email,request)
                             nstring=''



    finaliza=False
    ampliacion.fields['autoridades'].queryset=RefCiudades.objects.get(id=id_ciudad).ciu_autori.all()
    autoridades= preventivo.autoridades.all()
    autoridad=[]
    for seleccion in autoridades:
        ids=int(RefAutoridad.objects.get(descripcion=seleccion).id)
        autoridad.append(ids)

    ampliacion.fields['autoridades'].initial=autoridad
    values={'finaliza':finaliza,'id':idamp,'destino': destino,'state':state,'preventivo':preventivo,'ampliaciones':ampliaciones,'ampliacion':ampliacion}
    return render(request,'./ampliaciones.html',values)



def verificardni(request,tdni,dni):
    data = request.POST
    persona = Personas.objects.filter(nro_doc = dni,tipo_doc=tdni)
    data = serializers.serialize("json", persona)
    return HttpResponse(data, content_type='application/json')

def envioemail(envio,nstring,subject,text_content,from_email,request):
    msg = EmailMultiAlternatives(subject,text_content,from_email, nstring)
    msg.attach_alternative(text_content,'text/html')
    msg.send(fail_silently=False)

    return(envio,nstring,subject,text_content,from_email)

@login_required
@group_required(["administrador"])
def sendfechas(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    fecha_carga = datetime.datetime.now()
    return render(request,'./enviowebservice.html',{'state':state, 'destino': destino,'fecha_carga':fecha_carga})


@login_required
def enviadop(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    fecha_carga = datetime.datetime.now()
    totenviados=0
    errors=''


    if request.POST.get('search')=='Informar':
       fecha_cargad=request.POST.get('fecha_cargas')
       fecha_cargah=request.POST.get('fecha_cargah')

       if fecha_cargad and fecha_cargah:
        hoy=datetime.datetime.strptime(fecha_cargad,"%d/%m/%Y")
        ayer=(datetime.datetime.strptime(fecha_cargah,"%d/%m/%Y")+timedelta(days=1)).date()
        #hoy=datetime.datetime.strptime(fecha_cargad,"%d/%m/%Y").date()
        #ayer=datetime.datetime.strptime(fecha_cargah,"%d/%m/%Y").date()
        grabarfa = Preventivos.objects.filter(fecha_autorizacion__range=(hoy,ayer),fecha_autorizacion__isnull=False,sendwebservice=0)
        #fecha_autorizacion=datetime.datetime.strptime(,"%d/%m/%Y").strftime('%Y-%m-%d'),sendwebservice=0)
        #date.today())
        personas={}
        datosdict={}
        cantpersonas=''
        totenviados=0
        alturalugar=''
        for hay in grabarfa:

            preventivo = Preventivos.objects.get(id=hay.id)
            ciudad= preventivo.dependencia.ciudad
            depe=preventivo.dependencia
            #Datos del Hecho delicitivo atraves del nro de preventivo
            if Hechos.objects.filter(preventivo=preventivo.id):
                hecho = Hechos.objects.get(preventivo=preventivo.id)
                form=HechosForm(instance=hecho)
                ftiposdelitos=DelitoForm()
                motivo=request.POST.get('motivo')
                modos=RefModosHechoForm(instance=hecho)
                descripcion=hecho.descripcion
                idhec=hecho.id
                delito =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)
                modalidad=''
                eleinvo=Elementos.objects.filter(hechos=hecho.id,ampliacion_id__isnull=True,borrado__isnull=True).all()
                cometidos=[]
                hechodeli=""
                modus=""
                depes=preventivo.dependencia.ciudad_id
                depes=RefCiudades.objects.get(id=depes)


                for d in delito:
                    tienehecho=True
                    cometidos.append(d)

                for i in cometidos:

                    #if i.refmodoshecho:
                    hechodeli=hechodeli+unicode(str(i).strip(),'UTF-8')+'|'
                    #+' Modalidad :'+unicode(str(i.refmodoshecho),'UTF-8')+'|'
                    #else:
                    #hechodeli=hechodeli+unicode(str(i).strip(),'UTF-8')
                    #+' Sin Modalidad'+'|'
                    if i.refmodoshecho!=None:
                       modus=modus+unicode(str(i.refmodoshecho),'UTF-8')+'|'


                
                #Datos de las Personas involucradas en el hecho
                involuscra=[]
                datosper=""
                titempo=""
                involus=Hechos.objects.get(id=hecho.id).involu.all()
                datosgral=""
                eleminvo=[]
                #Datos del lugar del hecho
                lugar = Hechos.objects.get(id=idhec).lugar_hecho
                laticiudad = RefCiudades.objects.get(id=preventivo.dependencia.ciudad_id)
                localidad=laticiudad.descripcion
                lati=laticiudad.lat
                longi=laticiudad.longi
                tienelugar=True
                tienepersona=False
                involucrados=0
                #hacer el mapeo de comisarias,calles,barrios del lugar del hecho y de la personas involucradas

                localcria=Localidad.objects.get(descripcion__icontains=depes)
                localcria=int(localcria.idLocalidad)
                if lugar.calle:
                   callelugar=str(lugar.calle).strip()
                   if lugar.barrio:
                      barriolugar=str(lugar.barrio).replace('Bº','').strip()
                   else:
                      barriolugar=''
                else:
                   callelugar=''

                idComisaria=0
                idCalleHecho=0
                idBarrioHecho=0
                idBarrio=0
                idCalle=0
                idRolPersona=0
                idTipoOcupacion=0
                idEstadocivil=0
                naciona=0
                sector=''
                departamento=''
                piso=''
                escalera=''
                ocupacion=''
                lugarbarrio=''
                try:
                  for cal in  Calles.objects.filter(idLocalidad=localcria,descripcion__icontains=callelugar):
                     idCalleHecho=cal.idCalle
                except ObjectDoesNotExist:
                     idCalleHecho=0

                try:
                  for barr in Barrio.objects.filter(idLocalidad=localcria,descripcion__icontains=barriolugar):
                     idBarrioHecho=barr.idBarrio
                except ObjectDoesNotExist:
                     idBarrioHecho=0

                for crias in Comisarias.objects.filter(idLocalidad=localcria):
                    depen=str(depe)
                    criasdepe=crias.descripcion.upper()[-7:]

                    if criasdepe in depen:
                        idComisaria=crias.idorganismo
                
                cantper=0
                cantpersonas=''
                detenidos=0
                ciudad_res=''
                calle=''
                altura=0
                perjuridica=''
                for p in Hechos.objects.get(id=hecho.id).involu.all():
                    #aqui comprobar que datos son null de personas

                    bandera=True
                    tienepersona=True
                    involucrados=involucrados+1
                    rolin=str(p.roles)[2:].lower()
                    rolper=RolPersonas.objects.all()
                    for rolesper in rolper:
                        rp=rolesper.descripcion[2:]

                        if rolin==rp:
                           idRolPersona=int(str(rolesper))

                    if 'si' in p.detenido:
                       #detenidos=detenidos+1
                       detenidos=1
                    else:
                       detenidos=0

                    if 'si' in p.tentativa:
                         tentativa=1
                    else:
                         tentativa=0
                    if 'si' in p.infraganti:
                        infraganti=1
                    else:
                        infraganti=0
                    domi=Personas.objects.get(id=p.persona.id).persodom.all()
                    if p.juridica=='no':
                       #pf='FISICA'
                       pf=1
                    else:
                       #pf='JURIDICA'
                       pf=0
                       if p.razon_social!=None:
                          perjuridica=str(p.razon_social)

                       if RefTipoDocumento.objects.get(id=p.cuit_id)!='Null':
                          perjuridica=perjuridica+'-'+str(RefTipoDocumento.objects.get(id=p.cuit_id))
                       if p.nrocuit!=0:
                          perjuridica=perjuridica+'-'+str(p.nrocuit)

                    if p.menor=='':
                       p.menor="NO"



                    if p.persona.ocupacion!=None:
                       ocupacion=unicode(str(p.persona.ocupacion),'utf8')
                    else:
                       ocupacion=''

                    estadociv=str(p.persona.estado_civil).lower().capitalize()
                    a=Estadocivil.objects.all()
                    for civil in a:
                        civ=civil.descripcion.lower()
                        if civ==civ:
                            idEstadocivil=civil.idEstadoCivil

                    b=TipoOcupacion.objects.all()
                    ocupac=ocupacion.lower()
                    for ocu in b:
                        ocupa=ocu.descripcion.lower()
                        if ocupac==ocupa:
                            idTipoOcupacion=ocu.idtipoocupacion

                    if p.persona.pais_nac_id:
                        nacion=RefPaises.objects.get(id=p.persona.pais_nac_id)
                    else:
                        nacion='Descripcion'

                    naci=Nacionalidad.objects.all()
                    for nac in naci:
                        nacio=nac.descripcion
                        if nacion==nacio:
                             naciona=nac.id

                    tipo_doc=str(p.persona.tipo_doc)[:3]
                    doc=Tipodocumentos.objects.all()
                    tp_doc='INDOC'
                    for doc_tipo in doc:
                        tip=str(doc_tipo.idtipodocumento)[:3]

                        if tipo_doc==tip:
                           tp_doc=tip


                    if len(p.persona.nro_doc)<=8:
                       tipo_doc=p.persona.tipo_doc
                       nro_doc=p.persona.nro_doc

                       tipo_doc='N/T'
                       nro_doc='0'

                    if domi:
                        for l in Personas.objects.get(id=p.persona.id).persodom.all():
                             dad=Personas.objects.get(id=p.persona.id).padre.all()
                             tieneper=True
                             if p.persona.ciudad_res!=None:
                               ciudad_res=p.persona.ciudad_res
                               if l.calle!=None:
                                  calle=l.calle
                                  if str(l.altura)!='0':
                                     altura=l.altura
                                  else:
                                     altura=0
                               else:
                                  calle=''
                             else:
                                ciudad_res=''
                             descridomi=str(ciudad_res)+'-'+unicode(str(calle),'UTF-8').strip()+'-'+str(altura)
                             if l.calle:
                                calledom=str(l.calle).strip()
                                try:
                                    for cal in Calles.objects.filter(idLocalidad=localcria,descripcion__icontains=calledom):
                                         idCalle=cal.idCalle
                                except ObjectDoesNotExist:
                                         idCalle=0
                             if l.barrio_codigo:
                                barriodom=str(l.barrio_codigo).replace('Bº','').strip()
                                try:
                                    for barr in Barrio.objects.filter(idLocalidad=localcria,descripcion__icontains=barriodom):
                                         idBarrio=barr.idBarrio
                                except ObjectDoesNotExist:
                                         idBarrio=0
                             if dad:

                                    for la in Personas.objects.get(id=p.persona.id).padre.all():
                                            if p.juridica=='si':
                                                persona={'ApellidoyNombres':p.razon_social,'IdTipoDocumento':str(p.cuit_id),'DescripcionTipoDoc':str(RefTipoDocumento.objects.get(id=p.cuit_id)),'NroDocumento':p.nrocuit,'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'PersonaFisica':str(pf),'DescripcionPersonaJuridica':perjuridica,'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':ocupacion,'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),'IdNacionalidad':naciona}
                                            else:
                                                persona={'Apellido':p.persona.apellidos,'Nombre':p.persona.nombres,'IdTipoDocumento':tp_doc,'DescripcionTipoDoc':str(p.persona.tipo_doc),'NroDocumento':p.persona.nro_doc,'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'PersonaFisica':str(pf),'DescripcionPersonaJuridica':perjuridica,'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':ocupacion,'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),'IdNacionalidad':naciona}

                                            domi={'IdBarrio':idBarrio,'IdCalle':idCalle,'Nro':altura,'DescripcionDomicilio':descridomi}
                                            if la.padre_apellidos or la.padre_nombres or la.madre_apellidos or la.madre_nombres:
                                                    padys={'Hijode':la.padre_apellidos+','+la.padre_nombres+'-'+la.madre_apellidos+','+la.madre_nombres}
                                            else:
                                                    padys={'Hijode':'no registra datos de los padres'}

                                            dictpersona=persona
                                            dictpersona.update(domi)
                                            dictpersona.update(padys)
                             else:
                                 if p.juridica=='si':
                                     persona={'ApellidoyNombres':p.razon_social,'IdTipoDocumento':str(p.cuit_id),'DescripcionTipoDoc':str(RefTipoDocumento.objects.get(id=p.cuit_id)),'NroDocumento':p.nrocuit,'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'PersonaFisica':str(pf),'DescripcionPersonaJuridica':perjuridica,'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':ocupacion,'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),'IdNacionalidad':naciona}
                                 else:
                                    persona={'Apellido':p.persona.apellidos,'Nombre':p.persona.nombres,'IdTipoDocumento':tp_doc,'DescripcionTipoDoc':str(p.persona.tipo_doc),'NroDocumento':p.persona.nro_doc,'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'PersonaFisica':str(pf),'DescripcionPersonaJuridica':perjuridica,'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':str(p.persona.ocupacion),'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),'IdNacionalidad':naciona}

                                 domi={'IdBarrio':idBarrio,'IdCalle':idCalle,'Nro':altura,'DescripcionDomicilio':descridomi}
                                 padys={'Hijode':'no registra datos de los padres'}
                                 dictpersona=persona
                                 dictpersona.update(domi)
                                 dictpersona.update(padys)


                    else:

                        if p.juridica=='si':
                            persona={'ApellidoyNombres':p.razon_social,'IdTipoDocumento':str(p.cuit_id),'DescripcionTipoDoc':str(RefTipoDocumento.objects.get(id=p.cuit_id)),'NroDocumento':p.nrocuit,'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'PersonaFisica':str(pf),'DescripcionPersonaJuridica':perjuridica,'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':ocupacion,'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),'IdNacionalidad':naciona}
                        else:
                            persona={'Apellido':p.persona.apellidos,'Nombre':p.persona.nombres,'IdTipoDocumento':tp_doc,'DescripcionTipoDoc':str(p.persona.tipo_doc),'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'NroDocumento':p.persona.nro_doc,'PersonaFisica':str(pf),'DescripcionPersonaJuridica':perjuridica,'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':str(p.persona.ocupacion),'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),'IdNacionalidad':naciona}

                        domi={'DescripcionDomicilio':'no registra domicilio'}
                        padys={'Hijode':'no registra datos de los padres'}

                        dictpersona=persona
                        dictpersona.update(domi)
                        dictpersona.update(padys)
                    cantper=cantper+1
                    
                    personasxml=dicttoxml(dictpersona,attr_type=False,root='Personas')
                    xmlsper='<Persona>'+personasxml+'</Persona>'
                    cantpersonas=cantpersonas+xmlsper

                xmlsper='<PersonasPreventivo length='+'"'+str(cantper)+'"'+'>'+cantpersonas+'</PersonasPreventivo>'

                rotulo='CantidadElementos '+str(len(eleinvo))
                tituarmas=''
                tituvehiculos=''
                objectele={}
                elements=[]
                elementos=''
                obji=[]
                nrositems=0
                cantelemens=''
                for eles in eleinvo:
                    obdata=[]
                    obdatav=[]
                    if len(Elementosarmas.objects.filter(idelemento=eles.id))>0:

                         idar = Elementosarmas.objects.filter(idelemento=eles.id).values('idarma')
                         obdata=Armas.objects.filter(id=idar)

                         for extra in obdata:
                            tituarmas='Detalle Armas :'+str(extra.subtipos)+',Tipo/s : '+str(extra.tipos)+',Sistema de Disparo : '+str(extra.sistema_disparo)+',Marcas : '+str(extra.marcas)+',Calibre : '+str(extra.calibre)+',Modelo : '+str(extra.modelo)+',Nro Serie : '+str(extra.nro_arma)+',Propietario : '+str(extra.nro_doc)+'-'+str(extra.propietario)


                    if len(Elementoscars.objects.filter(idelemento=eles.id))>0:

                         idarv = Elementoscars.objects.filter(idelemento=eles.id).values('idvehiculo')

                         obdatav=Vehiculos.objects.filter(id=idarv)
                         for extrav in obdatav:
                            tituvehiculos='Detalle Vehiculo : '+str(extrav.idmarca)+',Modelo : '+str(extrav.modelo)+',Dominio : '+str(extrav.dominio)+',Año : '+str(extrav.anio)+' Tipos : '+str(extrav.tipov)+',Nro Chasis : '+str(extrav.nchasis)+',Nro. Motor : '+str(extrav.nmotor)+',Propietario : '+str(extrav.nro_doc)+'-'+str(extrav.propietario)

                    obse=html2text.html2text(eles.descripcion)
                    obse=strip_tags(eles.descripcion)
                    obse=obse.replace('&nbsp;','')
                    obser=obse.replace('"','')

                    rubros={'IdRelacionElemento':unicode(str(eles.tipo.id),'utf-8'),'RelacionElemento':unicode(str(eles.tipo.descripcion),'utf-8'),'DescripcionElemento':unicode(str(eles.rubro),'UTF-8')+'-'+unicode(str(eles.categoria),'UTF-8')+'-'+obser+'-'+str(eles.unidadmed)+'='+str(eles.cantidad),'TitularArmas':unicode(str(tituarmas),'utf-8'),'TitularVehiculo':unicode(str(tituvehiculos),'UTF-8')}
                    #por cada elemento hacer un tag para el xml
                    nrositems=nrositems+1
                    #ide={'IdItem_'+'nro_'+str(nrositems):str(eles.id)}
                    #rubrosi={'Descripcion':rubros}
                    #ide.update(rubrosi)
                    objetosxml=dicttoxml(rubros,attr_type=False,root='Elementos')
                    xmlsobj='<Elemento>'+objetosxml+'</Elemento>'
                    cantelemens=cantelemens+xmlsobj

                if nrositems>0:
                   xmlsobj='<Elementos length='+'"'+str(nrositems)+'"'+'>'+cantelemens+'</Elementos>'
                else:
                   xmlsobj='<Elementos length="0"></Elementos>'




                infor=''
                autoridad=''

                nro=preventivo.nro
                anio=preventivo.anio
                fecha_denuncia=preventivo.fecha_denuncia
                fecha_carga=preventivo.fecha_carga
                fecha_autorizacion=preventivo.fecha_autorizacion
                caratula=preventivo.caratula
                actuante=preventivo.actuante
                actuante=unicode(str(actuante), 'utf-8').strip()
                preventor=preventivo.preventor
                preventor=unicode(str(preventor), 'utf-8').strip()

                today = datetime.datetime.now()
                autoridades=preventivo.autoridades.values_list('descripcion',flat=True)
                dependencia=preventivo.dependencia.descripcion
                unidadreg=preventivo.dependencia.unidades_regionales.descripcion

                for a in autoridades:
                    autoridad=autoridad+str(a)+'|'

                jerarqui_a=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=actuante).values('jerarquia_id'))
                jerarqui_p=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=preventor).values('jerarquia_id'))
                form1=Hechos.objects.filter(preventivo=preventivo.id)

                for value in form1:


                    denuncia=html2text.html2text(descripcion,True)
                    denuncia=denuncia.encode('utf-8', 'xmlcharrefreplace')
                    denuncia=strip_tags(denuncia)
                    denuncia=denuncia.replace('&nbsp;','')
                    denuncia=denuncia.replace('"','')
                    motivo=str(value.motivo)
                    fecha_carga=fecha_carga.strftime("%d/%m/%Y %H:%m:%S")
                    fecha_autorizacion=fecha_autorizacion.strftime("%d/%m/%Y %H:%m:%S")
                    fechadesde=timezone.localtime(value.fecha_desde).strftime("%d/%m/%Y")
                    horadesde=timezone.localtime(value.fecha_desde).strftime("%H:%M")
                    fechahasta=timezone.localtime(value.fecha_hasta).strftime("%d/%m/%Y")
                    horahasta=timezone.localtime(value.fecha_hasta).strftime("%H:%M")
                    hora_denuncia= timezone.localtime(fecha_denuncia).strftime("%H:%M")
                    fecha_denuncia=timezone.localtime(fecha_denuncia).strftime("%d/%m/%Y")
                    idhecho=value.id
                    if value.fecha_esclarecido:
                         esclarecido=1
                    else:
                         esclarecido=0


                    if lugar.sector!=None:
                       sector=lugar.sector
                       domihecho=sector

                    if lugar.departamento!=None:
                       departamento=lugar.departamento
                       domihecho=domihecho+'-'+departamento

                    if lugar.piso!=0:
                       piso=lugar.piso
                       domihecho=domihecho+'-'+piso

                    if lugar.escalera!=None:
                       escalera=lugar.escalera
                       domihecho=domihecho+'-'+escalera


                    if lugar.altura==None:
                       alturalugar=0
                    else:
                       alturalugar=str(lugar.altura)

                    if lugar.barrio==None:
                       lugarbarrio=''
                    else:
                        lugarbarrio=unicode(str(lugar.barrio),'UTF-8')

                    hecho={'NroHecho':alturalugar,'Lat':lugar.latitud[0:10],'Lng':lugar.longitud[0:10],'DescripcionCalleHecho':unicode(str(lugar.calle),'UTF-8'),'IdCalleHecho':idCalleHecho,'IdBarrioHecho':idBarrioHecho,'DescripcionBarrioHecho':lugarbarrio,'DescripcionProvinciaHecho':'CHUBUT','DescripcionDomicilioHecho':domihecho,'MotivoDenuncia':motivo,'FechaHechoDesde':fechadesde,'FechaHechoHasta':fechahasta,'HoraDesde':horadesde,'HoraHasta':horahasta,'Esclarecido':esclarecido,'Tentativa':tentativa,'Detenidos':detenidos,'Flagrancia':infraganti}
                    denuncia={'Denuncia':denuncia}



                dichechos = hecho
                dichechos.update(denuncia)


                subject  ={'IdTipoPreventivo':1,'IdComisaria':int(idComisaria),'Numero':int(nro),'Anio':int(anio),'FechaCarga':fecha_carga,'FechaAutorizacion':fecha_autorizacion,'FechaEnvio':fecha_autorizacion,'FechaDenuncia':fecha_denuncia}
                subject1 ={'Caratula':caratula,'DelitosCometidos':hechodeli.strip(),'ModusOperandi':modus.strip(),'DescripcionActuante':str(jerarqui_a)+' - '+actuante,'DescripcionResponsable':str(jerarqui_p)+' - '+preventor,'Destinatarios':autoridad,'DescripcionLocalidadHecho':localidad,'LatLocalidad':float(lati),'LngLocalidad':float(longi),}

                datosp = subject
                datosp.update(subject1)


                datosp.update(dichechos)

                preventivosxml = dicttoxml(datosp,attr_type=False,root=False)



                coddestino='coironrw-test'
                fechahoy=datetime.datetime.now()

                xmlspre=preventivosxml


                xmls='<?xml version="1.0" encoding="utf-8"?>'+\
                '<soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">'+\
                '<soap:Header>'+\
                '<LoginInfo xmlns="http://sij.juschubut.gov.ar">'+\
                '<_usuario>policia-test</_usuario>'+\
                '<_password>policia-test</_password>'+\
                '<Usuario>policia-test</Usuario>'+\
                '<Password>policia-test</Password>'+\
                '</LoginInfo>'+\
                '</soap:Header>'+\
                '<soap:Body>'+\
                '<EnviarMensaje xmlns="http://sij.juschubut.gov.ar">'+\
                '<msg>'+\
                '<IdNodo>1</IdNodo>'+\
                '<Asunto>Info de Preventivos Spid</Asunto>'+\
                '<Cuerpo>'+\
                '<![CDATA[<?xml version="1.0" encoding="utf-8"?>'+\
                '<MensajePreventivo xmlns:xsd="http://www.w3.org/2001/XMLSchema" '+\
                'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" message-type="MensajeroPreventivo.Tranfer.MensajePreventivo, '+\
                'MensajeroPreventivo.Tranfer" message-version="1">'+xmlspre+xmlsper+xmlsobj+'</MensajePreventivo>'+']]>'+'</Cuerpo>'+\
                '<CodigoRemitente>policia-test</CodigoRemitente>'+\
                '<DescripcionRemitente>Prueba Policia</DescripcionRemitente>'+\
                '<CodigoDestino>coironrw-test</CodigoDestino>'+\
                '</msg>'+\
                '</EnviarMensaje>'+\
                '</soap:Body>'+\
                '</soap:Envelope>'


                user='policia-test'
                password='policia-test'
                params = { 'Authorization' : 'Basic %s' % base64.b64encode("user:password") }
                headers = {"Content-type": "application/x-www-form-urlencoded","Accept": "text/plain"}
                webservice = urllib2.Request('http://listas.juschubut.gov.ar/mensajero/mensajes.asmx')
                webservice = httplib.HTTPConnection('209.13.117.104',80)
                webservice.putrequest("POST", "http://listas.juschubut.gov.ar/mensajero/mensajes.asmx", params, headers)
                webservice.putheader("Host", "listas.juschubut.gov.ar")
                webservice.putheader("User-Agent", "Python Post")
                webservice.putheader("Content-type", "text/xml; charset=\"UTF-8\"")
                webservice.putheader("Content-length", "%d" % len(xmls))
                webservice.putheader("SOAPAction", "\"http://sij.juschubut.gov.ar/EnviarMensaje\"")
                webservice.endheaders()
                webservice.send(xmls)
                totenviados=totenviados+1
                ref=webservice.getresponse()
                refer=str(ref.status)+'-'+str(ref.reason)
                valorweb=0
                request.session['reason'] = refer
                if ref.status==200:
                   data = ref.read()
                   
                   #aqui actualizar el campo sendwebservice en preventivo a 1
                   user = User.objects.get(username='23140893')
                   prev = Preventivos.objects.get(id=hay.id)
                   judi=EnvioPreJudicial()
                   judi.preventivo=prev
                   judi.fecha_autorizacion=preventivo.fecha_autorizacion
                   judi.user=user
                   judi.enviado=1
                   judi.save()

                   valorweb=1
                   repla=Preventivos.objects.filter(id=hay.id).update(sendwebservice=valorweb)
                   lista=EnvioPreJudicial.objects.all()
                   webservice.close()

                else:
                   user = User.objects.get(username='23140893')
                   prev = Preventivos.objects.get(id=hay.id)
                   judi=EnvioPreJudicial()
                   judi.preventivo=prev
                   judi.fecha_autorizacion=preventivo.fecha_autorizacion
                   judi.user=user
                   judi.enviado=0
                   judi.save()
                   lista=EnvioPreJudicial.objects.all()


                datosdict={}
       else:
          errors="Ingrese Fecha Desde-Hasta"
    reason = 'None'
    if request.session.get('reason'):
        reason = request.session.get('reason')
    lista= Preventivos.objects.filter(sendwebservice=0) #EnvioPreJudicial.objects.all()
    fecha_carga=datetime.datetime.now()
    values={'errors':errors,'destino': destino,'state':state,'fecha_carga':fecha_carga,'lista':lista,'totenviados':totenviados,'reason':reason}
    return render(request,'./enviowebservice.html',values)


@login_required
def enviarp(request,idprev):
    state= request.session.get('state')
    destino= request.session.get('destino')
    preventivo = Preventivos.objects.get(id=idprev)                             # obtengo el preventivo a enviar
    ciudad= preventivo.dependencia.ciudad                                       # reservo la ciudad
    depe=preventivo.dependencia                                                 # reservo la dependencia
    enviado,sin_enviar = verificar_enviado(idprev)
    if enviado and len(sin_enviar) >= 1:
        marcar_enviados(sin_enviar)
        return HttpResponseRedirect('/spid/inicio/')
    
    #Datos del Hecho delicitivo atraves del nro de preventivo
    if Hechos.objects.filter(preventivo=preventivo.id):                         # si el preventivo tiene un hecho cargado
        hecho = Hechos.objects.get(preventivo=preventivo.id)                    # obtengo el hecho
        form=HechosForm(instance=hecho)                                         # instancio un formulario de hechos
        ftiposdelitos=DelitoForm()                                              # instancio un formulario de delitos
        motivo= hecho.motivo.descripcion                                        # obtengo el motivo de la denuncia
        modos=RefModosHechoForm(instance=hecho)                                 # instancio un formulario de modos
        descripcion=hecho.descripcion                                           # obtengo y reservo la descripcion del hecho
        idhec=hecho.id                                                          # obtengo y reservo el id del hecho
        delito =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True) # obtengo y reservo los delitos correspondientes al hecho
        modalidad=''                                                                # instancio la modalidad en vacio
        eleinvo=Elementos.objects.filter(hechos=hecho.id,ampliacion_id__isnull=True,borrado__isnull=True).all() # obtengo y reservo los elementos involucrados en el hecho
        cometidos=[]                                                            # creo un arreglo de hechos cometidos
        hechodeli=""                                                            # una cadena de delitos
        modus=""                                                                # una cadena de modus
        depes=preventivo.dependencia.ciudad_id                                  # obtengo el id de la ciudad
        depes=RefCiudades.objects.get(id=depes)                                 # instancio una ciudad


        for d in delito:                                                        # para cada delito en delito
            tienehecho=True                                                     # pongo en true tienehecho
            cometidos.append(d)                                                 # agrego el delito al arreglo de cometidos

        for i in cometidos:                                                     # por cada delito en cometidos

            hechodeli=hechodeli+str(i).strip()+'|'          # en la cadena hechodeli los concateno con formato utf-8
            if i.refmodoshecho!=None:                                           # si el hecho tiene modus operandi
               modus=modus+unicode(str(i.refmodoshecho),'UTF-8')+'|'            # lo concateno en modus en formato utf-8

        hechodeli = hechodeli.rstrip('|')
        modus = modus.rstrip('|')

        #Datos de las Personas involucradas en el hecho
        involuscra=[]                                                           # creo un arreglo de involucrados
        datosper=""                                                             # creo una cadena datosper
        titempo=""                                                              # creo una cadena titempo
        involus=Hechos.objects.get(id=hecho.id).involu.all()                    # obtengo a todas las personas involucradas en el hecho
        datosgral=""                                                            # creo una cadena datosgral
        eleminvo=[]                                                             # creo un arreglo eleminvo
        #Datos del lugar del hecho
        lugar = Hechos.objects.get(id=idhec).lugar_hecho                           # obtengo el lugar del hecho
        laticiudad = RefCiudades.objects.get(id=preventivo.dependencia.ciudad_id) # obtengo la ciudad de la dependencia donde se genero el preventivo
        localidad=laticiudad.descripcion                                        # Obtengo el nombre de la ciudad
        lati=laticiudad.lat                                                     # obtengo la latitud de la ciudad
        longi=laticiudad.longi                                                  # obtengo la longitud de la ciudad
        tienelugar=True                                                         # genero una bandera que indique que el hecho tiene lugar
        tienepersona=False                                                      # genero una bandera que indique que el hecho tiene persona
        involucrados=0                                                          # genero un contador de personas involucradas
        #hacer el mapeo de comisarias,calles,barrios del lugar del hecho y de la personas involucradas

        localcria=Localidad.objects.get(descripcion__icontains=depes)           # obtengo la localidad
        localcria=int(localcria.idLocalidad)                                    # obtengo el id en formato entero de la localidad
        if lugar.calle:                                                         # si el lugar tiene calle
           callelugar=str(lugar.calle).strip()                                  # obtengo el nombre de la calle
           if lugar.barrio:                                                     # si el lugar tiene barrio
              barriolugar=str(lugar.barrio).replace('Bº','').strip()            # obtengo el nombre del barrio
           else:
              barriolugar=''                                                    # sino tiene barrio envio el nombre vacio
        else:
           callelugar=''                                                        # sino tiene calle envio el nombre vacio

        idComisaria=0                                                           # |
        idCalleHecho=0                                                          #  \
        idBarrioHecho=0                                                         #   \
        idBarrio=0                                                              #    \
        idCalle=0                                                               #     \
        idRolPersona=0                                                          #      \
        idTipoOcupacion=0                                                       #       \
        idEstadocivil=0                                                         #        |==> Inicializo los id con 0 y otras variables con vacio
        naciona=0                                                               #       /
        sector=''                                                               #      /
        departamento=''                                                         #     /
        piso=''                                                                 #    /
        escalera=''                                                             #   /
        ocupacion=''                                                            #  /
        lugarbarrio=''                                                          # |

        idCalleHecho = buscar_calle(Localidad.objects.get(idLocalidad=localcria),callelugar) # Obtengo el id de calle segun el mapeo con las tablas del coiron

        """try:
          for barr in Barrio.objects.filter(idLocalidad=localcria,descripcion__icontains=barriolugar):  # para cada barrio que conicida con el del lugar
             idBarrioHecho=barr.idBarrio                                        # obtengo el id
        except ObjectDoesNotExist:                                              # si  no existe ninguno
             idBarrioHecho=0                                                    # lo dejo en cero"""
        idBarrioHecho = buscar_barrio(Localidad.objects.get(idLocalidad=localcria),barriolugar)

        for crias in Comisarias.objects.filter(idLocalidad=localcria):          # para cada comisaria que coincida con la indicada
            depen=str(depe)                                                     # obtengo su nombre
            criasdepe=crias.descripcion.upper()[-7:]                            # obtengo el nombre en mayusculas

            if criasdepe in depen:                                              # si la comisaria y la dependencia son iguales
                idComisaria=crias.idorganismo                                   # obtengo el numero de organismo asignado por el poder judicial
        
        cantper=0                                                               # cantidad de personas a cero
        cantpersonas=''                                                         # cantidad de personas vacio
        detenidos=0                                                             # detenidos a cero
        ciudad_res=''                                                           # ciudad de residencia vacio
        calle=''                                                                # calle vacio
        altura=0                                                                # altura a cero
        perjuridica=''                                                          # persona juridica vacio
        for p in Hechos.objects.get(id=hecho.id).involu.all():                  # para cada persona involucrada en el hecho
            #aqui comprobar que datos son null de personas

            bandera=True                                                        # bandera a True
            tienepersona=True                                                   # tiene persona a True
            involucrados=involucrados+1                                         # sumo un involucrado
            rolin=str(p.roles)[2:].lower()                                      # obtengo el rol de la persona
            rolper=RolPersonas.objects.all()                                    # obtengo un listado de roles
            for rolesper in rolper:                                             # para cada rol en rolper
                rp=rolesper.descripcion[2:]                                     # obtengo su descripcion segun el poder judicial

                if rolin==rp:                                                   # si coincide con el que policial
                   idRolPersona=int(str(rolesper))                              # obtengo su id judicial

            if 'si' in p.detenido:                                              # Si la persona esta detenida
               detenidos=1                                                      # incremento un detenido
            else:
               detenidos=0                                                      # sino lo dejo en cero

            if 'si' in p.tentativa:                                             # si el hecho fue en tentativa
                 tentativa=1                                                    # indico un uno en la bandera de tentativa
            else:
                 tentativa=0                                                    # sino lo dejo en cero
            if 'si' in p.infraganti:                                            # si el hecho fue en flagrancia
                infraganti=1                                                    # pongo la bandera en 1
            else:
                infraganti=0                                                    # sino la dejo en cero
            domi=Personas.objects.get(id=p.persona.id).persodom.all()           # obtengo los domicilios de la persona
            if p.juridica=='no':                                                # si no es persona juridica
               pf=1                                                             # pongo la bandera persona fisica en 1
            else:
               pf=0                                                             # sino la dejo en cero
               if p.razon_social!=None:                                         # y si la razon social tiene dato
                  perjuridica=str(p.razon_social)                               # obtengo el nombre de la persona juridica
               
               if RefTipoDocumento.objects.get(id=p.cuit_id)!='Null':           # si tiene cargado el cuit
                  perjuridica=perjuridica+'-'+str(RefTipoDocumento.objects.get(id=p.cuit_id))   # al nombre de la persona juridica le agrego su cuit
               if p.nrocuit!=0:                                                 # si el numero de cuit esta cargado
                  perjuridica=perjuridica+'-'+str(p.nrocuit)                    # lo agrego a la razon social

            if p.menor=='':                                                     # si la persona no es menor
               p.menor="NO"                                                     # marco que no lo es



            if p.persona.ocupacion!=None:                                       # si la ocupacion esta cargada
               ocupacion=unicode(str(p.persona.ocupacion),'utf8')               # obtengo su descripcion
            else:
               ocupacion=''                                                     # si no la dejo vacia

            estadociv=str(p.persona.estado_civil).lower().capitalize()          # obtengo el estado civil
            a=Estadocivil.objects.all()                                         # obtengo todos los estados civil del poder judicial
            for civil in a:                                                     # para cada uno de ellos
                civ=civil.descripcion.lower()                                   # obtengo su descripcion
                if civ==civ:
                    idEstadocivil=civil.idEstadoCivil                           # obtengo el id del estado civil segun el poder judicial

            b=TipoOcupacion.objects.all()                                       # obtengo todas las ocupaciones segun el poder judicial
            ocupac=ocupacion.lower()
            for ocu in b:                                                       # para cada ocupacion en la lista
                ocupa=ocu.descripcion.lower()                                   # obtengo su descripcion
                if ocupac==ocupa:                                               # si coincide con la de la persona
                    idTipoOcupacion=ocu.idtipoocupacion                         # obtengo el id segun el poder judicial

            if p.persona.pais_nac_id:                                           # si tiene cargada la nacionalidad
                nacion=RefPaises.objects.get(id=p.persona.pais_nac_id)          # obtengo el id del pais
            else:
                nacion='Descripcion'                                            # sino lo cargo con descripcion

            naci=Nacionalidad.objects.all()                                     # obtengo un listado de nacionalidades segun el poder judicial
            for nac in naci:                                                    # para cada una de ellas
                nacio=nac.descripcion                                           # obtengo su descripcion
                if nacion==nacio:                                               # si es igual a la nacionalidad en la persona
                     naciona=nac.id                                             # obtengo el id


            tp_doc = buscar_tipos_doc(p.persona.tipo_doc.descripcion)


            if len(p.persona.nro_doc)<=8:                                       # si el documento tiene hasta 8 digitos
               tipo_doc=p.persona.tipo_doc                                      # obtengo el tipo de documento
               nro_doc=p.persona.nro_doc                                        # obtengo en numero de documento

               tipo_doc='N/T'                                                   # tipo doc N/T
               nro_doc='0'                                                      # nro doc 0

            if domi:                                                            # si tiene domicilio
                for l in Personas.objects.get(id=p.persona.id).persodom.all():  # para cada domicilio en el listado
                     dad=Personas.objects.get(id=p.persona.id).padre.all()      # obtengo los padres de la persona
                     tieneper=True                                              # pongo la bandera de persona en verdadero
                     if p.persona.ciudad_res!=None:                             # si la ciudad de residencia esta cargada
                       ciudad_res=p.persona.ciudad_res                          # obtengo la ciudad de residencia
                       if l.calle!=None:                                        # si tiene calle
                          calle=l.calle                                         # obtengo la calle
                          if str(l.altura)!='0':                                # si tiene altura
                             altura=l.altura                                    # la obtengo
                          else:
                             altura=0                                           # sino la dejo en vacio
                       else:
                          calle=''                                              # sino tiene calle la dejo vacia
                     else:
                        ciudad_res=''                                           # sino tiene ciudad de residencia la dejo vacia
                     descridomi=str(ciudad_res)+'-'+unicode(str(calle),'UTF-8').strip()+'-'+str(altura) # genero la descripcion del domicilio
                     if l.calle:                                                # si el lugar tiene calle
                        calledom=str(l.calle).strip()                           # obtengo la calle

                        idCalle = buscar_calle(Localidad.objects.get(idLocalidad=localcria),calledom)
                     if l.barrio_codigo:                                        # si el lugar tiene codigo de barrio
                        barriodom=str(l.barrio_codigo).replace('Bº','').strip() # obtengo el barrio

                        idBarrio = buscar_barrio(Localidad.objects.get(idLocalidad=localcria),barriodom)
                     if dad:                                                    # si tiene datos de los padres

                            for la in Personas.objects.get(id=p.persona.id).padre.all(): # para cada uno
                                    if p.juridica=='si':                        # si es persona juridica
                                        persona={                               # genero el objeto persona
                                                'ApellidoyNombres':p.razon_social,  #agrego la razon social
                                                'IdTipoDocumento':str(p.cuit_id),   #agrego el cuit
                                                'DescripcionTipoDoc':str(RefTipoDocumento.objects.get(id=p.cuit_id)),   # agrego el tipo de documento
                                                'NroDocumento':p.nrocuit,           # agrego el numero de cuit
                                                'Alias':p.persona.alias,            # agrego el alias
                                                'IdTipoOcupacion':idTipoOcupacion,  # agrego la ocupacion
                                                'IdEstadoCivil':idEstadocivil,      # agrego el estado civil
                                                'PersonaFisica':str(pf),            # agrego la bandera persona fisica
                                                'DescripcionPersonaJuridica':perjuridica,   # agrego la bandera persona juridica
                                                'IdRolPersona':idRolPersona,        # agrego el id del rol de la persona
                                                'DescripcionRol':str(p.roles),      # agrego la descripcion del rol de la persona
                                                'Telefonos':p.persona.celular,      # agrego el telefono
                                                'Ocupacion':ocupacion,              # agrego la ocupacion de la persona
                                                'DescripcionEstadoCivil':str(p.persona.estado_civil),   # agrego la descripcion del estado civil
                                                'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%M:%S"),   # agrego la fecha de nacimiento
                                                'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'), # agrego el lugar de nacimiento
                                                'IdNacionalidad':naciona            # agrego la nacionalidad
                                                }
                                    else:                                       # si no es persona juridica
                                        persona={
                                                'Apellido':p.persona.apellidos,     # agrego el apellido
                                                'Nombre':p.persona.nombres,         # agrego el nombre
                                                'IdTipoDocumento':tp_doc,           # agrego el tipo de documento
                                                'DescripcionTipoDoc':str(p.persona.tipo_doc),   # agrego la descripcion del tipo de documento
                                                'NroDocumento':p.persona.nro_doc,   # agrego el numero de documento
                                                'Alias':p.persona.alias,            # agrego el alias
                                                'IdTipoOcupacion':idTipoOcupacion,  # agrego el id de la ocupacion
                                                'IdEstadoCivil':idEstadocivil,      # agrego el id del estado civil
                                                'PersonaFisica':str(pf),            # agrego la bandera persona fisica
                                                'DescripcionPersonaJuridica':perjuridica,   # agrego la descripcion de persona juridica
                                                'IdRolPersona':idRolPersona,        # agrego el rol de la persona
                                                'DescripcionRol':str(p.roles),      # agrego el rol de la persona
                                                'Telefonos':p.persona.celular,      # agrego el celular de la persona
                                                'Ocupacion':ocupacion,              # agrego la ocupacion
                                                'DescripcionEstadoCivil':str(p.persona.estado_civil),   # agrego la descripcion del estado civil
                                                'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%M:%S"), # agrego la fecha de nacimiento
                                                'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'), # agrego el lugar de nacimiento
                                                'IdNacionalidad':naciona            # agrego la nacionalidad
                                                }

                                    domi={                                          # genero el objeto domicilio
                                        'IdBarrio':idBarrio,                        # agrego el id de barrio
                                        'IdCalle':idCalle,                          # agrego el id de calle
                                        'Nro':altura,                               # agrego la altura del domicilio
                                        'DescripcionDomicilio':descridomi           # agrego la descripcion del domicilio
                                        }
                                    if la.padre_apellidos or la.padre_nombres or la.madre_apellidos or la.madre_nombres: # si tengo cargo el dato de alguno de los padres
                                            padys={                                 # genero el objeto de padres
                                                'Hijode':la.padre_apellidos+','     # agrego el apellido del padre
                                                +la.padre_nombres+'-'               # agrego el nombre del padre
                                                +la.madre_apellidos+','             # agrego el apellido de la madre
                                                +la.madre_nombres                   # agrego el nombre de la madre
                                                }
                                    else:
                                            padys={                                 # si no tiene datos de los padres
                                                'Hijode':'no registra datos de los padres' # indico que no los tiene
                                                }

                                    dictpersona=persona                             # creo un diccionario de persona
                                    dictpersona.update(domi)                        # le agrego el domicilio
                                    dictpersona.update(padys)                       # le agrego los padres
                     else:
                         if p.juridica=='si':
                            persona={
                                    'ApellidoyNombres':p.razon_social,
                                    'IdTipoDocumento':str(p.cuit_id),
                                    'DescripcionTipoDoc':str(RefTipoDocumento.objects.get(id=p.cuit_id)),
                                    'NroDocumento':p.nrocuit,
                                    'Alias':p.persona.alias,
                                    'IdTipoOcupacion':idTipoOcupacion,
                                    'IdEstadoCivil':idEstadocivil,
                                    'PersonaFisica':str(pf),
                                    'DescripcionPersonaJuridica':perjuridica,
                                    'IdRolPersona':idRolPersona,
                                    'DescripcionRol':str(p.roles),
                                    'Telefonos':p.persona.celular,
                                    'Ocupacion':ocupacion,
                                    'DescripcionEstadoCivil':str(p.persona.estado_civil),
                                    'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%M:%S"),
                                    'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),
                                    'IdNacionalidad':naciona
                                    }
                         else:
                            persona={
                                    'Apellido':p.persona.apellidos,
                                    'Nombre':p.persona.nombres,
                                    'IdTipoDocumento':tp_doc,
                                    'DescripcionTipoDoc':str(p.persona.tipo_doc),
                                    'NroDocumento':p.persona.nro_doc,
                                    'Alias':p.persona.alias,
                                    'IdTipoOcupacion':idTipoOcupacion,
                                    'IdEstadoCivil':idEstadocivil,
                                    'PersonaFisica':str(pf),
                                    'DescripcionPersonaJuridica':perjuridica,
                                    'IdRolPersona':idRolPersona,
                                    'DescripcionRol':str(p.roles),
                                    'Telefonos':p.persona.celular,
                                    'Ocupacion':str(p.persona.ocupacion),
                                    'DescripcionEstadoCivil':str(p.persona.estado_civil),
                                    'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%M:%S"),
                                    'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),
                                    'IdNacionalidad':naciona
                                    }

                         domi={
                                'IdBarrio':idBarrio,
                                'IdCalle':idCalle,
                                'Nro':altura,
                                'DescripcionDomicilio':descridomi
                                }
                         padys={
                                'Hijode':'no registra datos de los padres'
                                }
                         dictpersona=persona
                         dictpersona.update(domi)
                         dictpersona.update(padys)


            else:

                if p.juridica=='si':
                    persona={
                            'ApellidoyNombres':p.razon_social,
                            'IdTipoDocumento':str(p.cuit_id),
                            'DescripcionTipoDoc':str(RefTipoDocumento.objects.get(id=p.cuit_id)),
                            'NroDocumento':p.nrocuit,
                            'Alias':p.persona.alias,
                            'IdTipoOcupacion':idTipoOcupacion,
                            'IdEstadoCivil':idEstadocivil,
                            'PersonaFisica':str(pf),
                            'DescripcionPersonaJuridica':perjuridica,
                            'IdRolPersona':idRolPersona,
                            'DescripcionRol':str(p.roles),
                            'Telefonos':p.persona.celular,
                            'Ocupacion':ocupacion,
                            'DescripcionEstadoCivil':str(p.persona.estado_civil),
                            'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%M:%S"),
                            'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),
                            'IdNacionalidad':naciona
                            }
                else:
                    persona={
                            'Apellido':p.persona.apellidos,
                            'Nombre':p.persona.nombres,
                            'IdTipoDocumento':tp_doc,
                            'DescripcionTipoDoc':str(p.persona.tipo_doc),
                            'Alias':p.persona.alias,
                            'IdTipoOcupacion':idTipoOcupacion,
                            'IdEstadoCivil':idEstadocivil,
                            'NroDocumento':p.persona.nro_doc,
                            'PersonaFisica':str(pf),
                            'DescripcionPersonaJuridica':perjuridica,
                            'IdRolPersona':idRolPersona,
                            'DescripcionRol':str(p.roles),
                            'Telefonos':p.persona.celular,
                            'Ocupacion':str(p.persona.ocupacion),
                            'DescripcionEstadoCivil':str(p.persona.estado_civil),
                            'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%M:%S"),
                            'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),
                            'IdNacionalidad':naciona
                            }

                domi={
                        'DescripcionDomicilio':'no registra domicilio'
                    }
                padys={
                        'Hijode':'no registra datos de los padres'
                        }

                dictpersona=persona
                dictpersona.update(domi)
                dictpersona.update(padys)
            cantper=cantper+1                           # sumo uno a la cantidad de personas
            personasxml=dicttoxml(dictpersona,attr_type=False,root='Personas')  # genero un XML con el diccionario de persona
            xmlsper='<Persona>'+personasxml+'</Persona>'        # encierro el XML generado en los tags Persona
            cantpersonas=cantpersonas+xmlsper                   # agrego a cantersonas la persona generada

        xmlsper='<PersonasPreventivo length='+'"'+str(cantper)+'"'+'>'+cantpersonas+'</PersonasPreventivo>' # genero un tag PersonasPreventivo y le agrego las personas generadas y la cantidad

        # COMIENZO CON LOS ELEMENTOS

        rotulo='CantidadElementos '+str(len(eleinvo))                   # Obtengo la cantidad de elementos
        tituarmas=''                                                    # pongo tituarmas en vacion
        tituvehiculos=''                                                # pongo tituvehiculos en vacio
        objectele={}                                                    # genero un diccionario objectele
        elements=[]                                                     # genero un arrego de elementos
        elementos=''                                                    # genero una cadena elementos
        obji=[]
        nrositems=0                                                     # genero un indicador de numero de elemento
        cantelemens=''                                                  # genero una variable para guardar la cantidad de elementos
        for eles in eleinvo:                                            # para cada elemento involucrado
            obdata=[]                                                   # genero una arreglo de los datos del elemento
            obdatav=[]
            if len(Elementosarmas.objects.filter(idelemento=eles.id))>0: # si tiene armas

                 idar = Elementosarmas.objects.filter(idelemento=eles.id).values('idarma') # obtengo el id del arma
                 obdata=Armas.objects.filter(id=idar)                   # obtengo el arma

                 for extra in obdata:                                   # Genero todos los datos del arma
                    tituarmas='Detalle Armas :'+str(extra.subtipos)+',Tipo/s : '+str(extra.tipos)+',Sistema de Disparo : '+str(extra.sistema_disparo)+',Marcas : '+str(extra.marcas)+',Calibre : '+str(extra.calibre)+',Modelo : '+str(extra.modelo)+',Nro Serie : '+str(extra.nro_arma)+',Propietario : '+str(extra.nro_doc)+'-'+str(extra.propietario)


            if len(Elementoscars.objects.filter(idelemento=eles.id))>0: # si tiene vehiculos involucrados

                 idarv = Elementoscars.objects.filter(idelemento=eles.id).values('idvehiculo') # obtengo el id

                 obdatav=Vehiculos.objects.filter(id=idarv)         # obtengo el vehiculo
                 for extrav in obdatav:                             # genero todos los datos del vehiculo
                    tituvehiculos='Detalle Vehiculo : '+str(extrav.idmarca)+',Modelo : '+str(extrav.modelo)+',Dominio : '+str(extrav.dominio)+',Año : '+str(extrav.anio)+' Tipos : '+str(extrav.tipov)+',Nro Chasis : '+str(extrav.nchasis)+',Nro. Motor : '+str(extrav.nmotor)+',Propietario : '+str(extrav.nro_doc)+'-'+str(extrav.propietario)

            obse=html2text.html2text(eles.descripcion)                      # obtengo la descripcion de los elementos convertida en texto plano
            obse=strip_tags(eles.descripcion)                               # quito todos los tags html
            obse=obse.replace('&nbsp;','')                                  # genero los espacios
            obser=obse.replace('"','')                                      # reemplazo las comillas por vacio

            rubros={
                    'IdRelacionElemento':unicode(str(eles.tipo.id),'utf-8'),
                    'RelacionElemento':unicode(str(eles.tipo.descripcion),'utf-8'),
                    'DescripcionElemento':unicode(str(eles.rubro),'UTF-8')+'-'+unicode(str(eles.categoria),'UTF-8')+'-'+obser+'-'+str(eles.unidadmed)+'='+str(eles.cantidad),
                    'TitularArmas':unicode(str(tituarmas),'utf-8'),
                    'TitularVehiculo':unicode(str(tituvehiculos),'UTF-8')
                    }
            #por cada elemento hacer un tag para el xml
            nrositems=nrositems+1                                           # sumo uno a la cantidad de elementos
            objetosxml=dicttoxml(rubros,attr_type=False,root='Elementos')   # genero un objeto XML con el elemento generado
            xmlsobj='<Elemento>'+objetosxml+'</Elemento>'                   # agrego los tag elemento
            cantelemens=cantelemens+xmlsobj                                 # agrego el objeto a cantelemens

        if nrositems>0:                                                     # si hay mas de un elemento
           xmlsobj='<Elementos length='+'"'+str(nrositems)+'"'+'>'+cantelemens+'</Elementos>'   # genero el tag elementos con todos los elementos y la cantidad
        else:
           xmlsobj='<Elementos length="0"></Elementos>'                     # sino genero el tag vacio



        infor=''                                                            # genero una varible infor
        autoridad=''                                                        # genero una variable autoridad
        nro=preventivo.nro                                                  # obtengo el numero de preventivo
        anio=preventivo.anio                                                # obtengo el año del preventivo
        fecha_denuncia=preventivo.fecha_denuncia                            # obtengo la fecha de denuncia
        fecha_carga=preventivo.fecha_carga                                  # obtengo la fecha de carga
        fecha_autorizacion=preventivo.fecha_autorizacion                    # obtengo la fecha de autorizacion
        caratula=preventivo.caratula                                        # obtengo la caratula
        actuante=preventivo.actuante                                        # obtengo el actuante
        actuante=unicode(str(actuante), 'utf-8').strip()                    # convierto el formato del texto del actuante
        preventor=preventivo.preventor                                      # obtengo el preventor
        preventor=unicode(str(preventor), 'utf-8').strip()                  # convierto el formato del texto del preventor

        today = datetime.datetime.now()                                     # obtengo la fecha actual
        autoridades=preventivo.autoridades.values_list('descripcion',flat=True)     # obtengo el listado de autoridades a informar
        dependencia=preventivo.dependencia.descripcion                      # obtengo la dependencia
        unidadreg=preventivo.dependencia.unidades_regionales.descripcion    # obtengo la unidad regional

        for a in autoridades:                                               # para cada una de las autoridades
            autoridad=autoridad+str(a)+'|'                                  # las agrego a una cadena de autoridades
        autoridad = autoridad.rstrip('|')
        jerarqui_a=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=actuante).values('jerarquia_id')) # obtengo la jerarquia del actuante
        jerarqui_p=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=preventor).values('jerarquia_id')) # obtengo la jerarquia del preventor
        form1=Hechos.objects.filter(preventivo=preventivo.id)               # obtengo el hecho del preventivo

        for value in form1:                                                 # para cada hecho


            denuncia=html2text.html2text(descripcion,True)                  # obtengo la descripcion del hecho transformada en texto plano
            denuncia=denuncia.encode('utf8', 'xmlcharrefreplace')          # lo codifico en utf-8
            
            #denuncia=strip_tags(denuncia)                                   # le quito los tags html
            denuncia=denuncia.replace('&nbsp;','')                          # quito los espacios
            denuncia=denuncia.replace('"','')                               # reemplazo las comillas dobles por vacio
            motivo=str(value.motivo)                                        # obtengo el motivo de la denuncia
            fecha_carga=fecha_carga.strftime("%d/%m/%Y %H:%m:%S")           # obtengo la fecha de carga del preventivo
            fecha_autorizacion=fecha_autorizacion.strftime("%d/%m/%Y %H:%m:%S") # obtengo la fecha de autorizacion
            fechadesde=timezone.localtime(value.fecha_desde).strftime("%d/%m/%Y %H:%M:%S")  # obtengo la fecha desde
            horadesde=timezone.localtime(value.fecha_desde).strftime("%d/%m/%Y %H:%M:%S")       # obtengo la hora desde
            fechahasta=timezone.localtime(value.fecha_hasta).strftime("%d/%m/%Y %H:%M:%S")  # obtengo la fecha hasta
            horahasta=timezone.localtime(value.fecha_hasta).strftime("%d/%m/%Y %H:%M:%S")       # obtengo la hora hasta
            hora_denuncia= timezone.localtime(fecha_denuncia).strftime("%d/%m/%Y %H:%M:%S")     # obtengo la hora de denuncia
            fecha_denuncia=timezone.localtime(fecha_denuncia).strftime("%d/%m/%Y %H:%M:%S") # obtengo la fecha de denuncia
            idhecho=value.id                                                # obtengo el id del hecho

            if value.fecha_esclarecido:                                     # si el hecho esta esclarecido
                 esclarecido=1                                              # pongo la bandera de esclarecido en uno
            else:
                 esclarecido=0                                              # sino la pongo a cero

            if lugar.sector!=None:                                          # si el lugar tiene sector
               sector=lugar.sector                                          # obtengo el sector
               domihecho=sector                                             # agrego el sector al domicilio del hecho

            if lugar.departamento!=None:                                    # si el lugar tiene departamento
               departamento=lugar.departamento                              # obtengo el departamento
               domihecho=domihecho+'-'+departamento                         # lo agrego al domicilio del hecho

            if lugar.piso!=0:                                               # si el lugar tiene piso
               piso=lugar.piso                                              # lo obtengo
               domihecho=domihecho+'-'+piso                                 # lo agrego al domicilio del hecho

            if lugar.escalera!=None:                                        # si el lugar tiene escalera
               escalera=lugar.escalera                                      # obtengo la escalera
               domihecho=domihecho+'-'+escalera                             # la agrego al domicilio del echo


            if lugar.altura==None:                                          # si el lugar no tiene altura
               alturalugar=0                                                # pongo cero en la altura
            else:
               alturalugar=str(lugar.altura)                                # sino obtengo la altura

            if lugar.barrio==None:                                          # si el lugar no tiene barrio
               lugarbarrio=''                                               # dejo vacio el barrio
            else:
                lugarbarrio=unicode(str(lugar.barrio),'UTF-8')              # sino lo obtengo

            hecho={                                                         # genero el objeto de hecho
                    'NroHecho':alturalugar,                                 # agrego la altura del lugar
                    'Lat':lugar.latitud[0:10],                              # agrego la latitud
                    'Lng':lugar.longitud[0:10],                             # agrego la longitud
                    'DescripcionCalleHecho':unicode(str(lugar.calle),'UTF-8'),  # agrego la descripcion de la calle
                    'IdCalleHecho':idCalleHecho,                            # agrego el id de la calle
                    'IdBarrioHecho':idBarrioHecho,                          # agrego el id del barrio
                    'DescripcionBarrioHecho':lugarbarrio,                   # agrego la descripcion del barrio
                    'DescripcionProvinciaHecho':'CHUBUT',                   # agrego la descripcion de la provincia
                    'DescripcionDomicilioHecho':domihecho,                  # agrego la descripcion del domicilio del hecho
                    'MotivoDenuncia':motivo,                                # agrego el motivo de denuncia
                    'FechaHechoDesde':fechadesde,                           # agrego la fecha desde
                    'FechaHechoHasta':fechahasta,                           # agrego la fecha hasta
                    'HoraDesde':horadesde,                                  # agrego la hora desde
                    'HoraHasta':horahasta,                                  # agrego la hora hasta
                    'Esclarecido':esclarecido,                              # agrego el esclarecido
                    'Tentativa':tentativa,                                  # agrego la tentativa
                    'Detenidos':detenidos,                                  # agrego los detenidos
                    'Flagrancia':infraganti                                 # agrego la flagrancia
                    }
            denuncia={                                                      # genero el objeto de denuncia
                        'Denuncia':denuncia                                 # agrego la denuncia
                        }



        dichechos = hecho                                                   # genero el diccionario de hechos
        dichechos.update(denuncia)                                          # le agrego la denuncia

        subject  ={                                                         # genero el objeto asunto
                    'IdTipoPreventivo':1,                                   # agrego el tipo de preventivo
                    'IdComisaria':int(idComisaria),                         # agrego el id de la comisaria
                    'Numero':int(nro),                                      # agrego el numero de preventivo
                    'Anio':int(anio),                                       # agrego el año
                    'FechaCarga':fecha_carga,                               # agrego la fecha de carga
                    'FechaAutorizacion':fecha_autorizacion,                 # agrego la fecha de autorizacion
                    'FechaEnvio':fecha_autorizacion,                        # agrego la fecha de envio
                    'FechaDenuncia':fecha_denuncia                          # agrego la fecha de denuncia
                    }
        subject1 ={                                                         # genero un objeto asunto 1
                    'Caratula':caratula,                                    # agrego la caratula
                    'DelitosCometidos':unicode(str(hechodeli.strip()),'UTF-8'),                     # agrego los delitos
                    'ModusOperandi':modus.strip(),                          # agrego el modus
                    'DescripcionActuante':str(jerarqui_a)+' - '+actuante,   # agrego al actuante
                    'DescripcionResponsable':str(jerarqui_p)+' - '+preventor,   # agrego el preventor
                    'Destinatarios':autoridad,                              # agrego los destinatarios
                    'DescripcionLocalidadHecho':localidad,                  # agrego la localidad del hecho
                    'LatLocalidad':float(lati),                             # agrego la latitud de la localidad
                    'LngLocalidad':float(longi),                            # agrego la longitud de la localidad
                    }

        datosp = subject                                                    # genero un diccionario de datos del preventivo con el asunto
        datosp.update(subject1)                                             # agrego el asunto 1

        datosp.update(dichechos)                                            # le agrego el diccionario de hechos

        preventivosxml = dicttoxml(datosp,attr_type=False,root=False)       # Lo convierto en XML

        coddestino='coironrw-test'                                          # seteo el destino
        fechahoy=datetime.datetime.now()                                    # obtengo la fecha actual
        xmlspre=preventivosxml
        asunto = 'Preventivo '+str(preventivo.nro)+\
                 '/'+str(preventivo.anio)+' - '+\
                 str(subject['IdComisaria'])+' - '+\
                 preventivo.dependencia.descripcion +' - '+\
                 preventivo.dependencia.ciudad.descripcion
        # COMIENZO CON LA GENERACION DEL XML
        xmls='<?xml version="1.0" encoding="utf-8"?>'+\
                '<soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">'+\
                '<soap:Header>'+\
                '<LoginInfo xmlns="http://sij.juschubut.gov.ar">'+\
                '<_usuario>policia-test</_usuario>'+\
                '<_password>policia-test</_password>'+\
                '<Usuario>policia-test</Usuario>'+\
                '<Password>policia-test</Password>'+\
                '</LoginInfo>'+\
                '</soap:Header>'+\
                '<soap:Body>'+\
                '<EnviarMensaje xmlns="http://sij.juschubut.gov.ar">'+\
                '<msg>'+\
                '<IdNodo>1</IdNodo>'+\
                '<Asunto>'+ asunto.encode('utf8') +'</Asunto>'+\
                '<Cuerpo>'+\
                '<![CDATA[<?xml version="1.0" encoding="utf-8"?>'+\
                '<MensajePreventivo xmlns:xsd="http://www.w3.org/2001/XMLSchema" '+\
                'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" message-type="MensajeroPreventivo.Tranfer.MensajePreventivo, '+\
                'MensajeroPreventivo.Tranfer" message-version="1">'+xmlspre+xmlsper+xmlsobj+'</MensajePreventivo>'+']]>'+'</Cuerpo>'+\
                '<CodigoRemitente>policia-test</CodigoRemitente>'+\
                '<DescripcionRemitente>Prueba Policia</DescripcionRemitente>'+\
                '<CodigoDestino>coironrw-test</CodigoDestino>'+\
                '</msg>'+\
                '</EnviarMensaje>'+\
                '</soap:Body>'+\
                '</soap:Envelope>'
        user='policia-test'
        password='policia-test'
        params = { 'Authorization' : 'Basic %s' % base64.b64encode("user:password") }
        headers = {"Content-type": "application/x-www-form-urlencoded","Accept": "text/plain"}
        webservice = urllib2.Request('http://listas.juschubut.gov.ar/mensajero/mensajes.asmx')
        webservice = httplib.HTTPConnection('209.13.117.104',80)
        webservice.putrequest("POST", "http://listas.juschubut.gov.ar/mensajero/mensajes.asmx", params, headers)
        webservice.putheader("Host", "listas.juschubut.gov.ar")
        webservice.putheader("User-Agent", "Python Post")
        webservice.putheader("Content-type", "text/xml; charset=\"UTF-8\"")
        webservice.putheader("Content-length", "%d" % len(xmls))
        webservice.putheader("SOAPAction", "\"http://sij.juschubut.gov.ar/EnviarMensaje\"")
        webservice.endheaders()
        webservice.send(xmls)
        ref=webservice.getresponse()
        refer=str(ref.status)+'-'+str(ref.reason)
        valorweb=0
        
        if ref.status==200:
           data = ref.read()
           #aqui actualizar el campo sendwebservice en preventivo a 1
           user = request.user
           prev = Preventivos.objects.get(id=idprev)
           try:
               EnvioPreJudicial.objects.get(preventivo=prev)
               judi = EnvioPreJudicial.objects.get(preventivo = prev)
               judi.fecha_autorizacion = prev.fecha_autorizacion
               judi.user = user
               judi.enviado = 1
               judi.save()
           except ObjectDoesNotExist:
               judi=EnvioPreJudicial()
               judi.preventivo=prev
               judi.dependencia = prev.dependencia
               judi.fecha_autorizacion=preventivo.fecha_autorizacion
               judi.user=user
               judi.enviado=1
               judi.save()

           valorweb=1
           repla=Preventivos.objects.get(id=preventivo.id)
           repla.sendwebservice = valorweb
           repla.save()
           lista=EnvioPreJudicial.objects.all()
           webservice.close()
           #return judi.enviado)
        else:
           user = request.user
           prev = Preventivos.objects.get(id=idprev)
           try:
                EnvioPreJudicial.objects.get(preventivo = prev)
                judi = EnvioPreJudicial.objects.get(preventivo = prev)
                judi.fecha_autorizacion = prev.fecha_autorizacion
                judi.user = user
                judi.enviado = 0
                judi.save()
           except ObjectDoesNotExist:
                judi=EnvioPreJudicial()
                judi.preventivo=prev
                judi.dependencia = prev.dependencia
                judi.fecha_autorizacion=preventivo.fecha_autorizacion
                judi.user=user
                judi.enviado=0
                judi.save()
           lista=EnvioPreJudicial.objects.all()
           webservice.close()
           #return judi.enviado
        datosdict={}
    return HttpResponseRedirect('/spid/inicio/')


@login_required
@group_required(["administrador"])
def enviadoa(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    fecha_carga = datetime.datetime.now()
    totenviados=0
    errors=''
    if request.POST.get('search')=='Informar':
       fecha_cargad=request.POST.get('fecha_cargas')
       fecha_cargah=request.POST.get('fecha_cargah')
       if fecha_cargad and fecha_cargah:
        hoy=datetime.datetime.strptime(fecha_cargad,"%d/%m/%Y")
        ayer=(datetime.datetime.strptime(fecha_cargah,"%d/%m/%Y")+timedelta(days=1)).date()
        grabarfa = Ampliacion.objects.filter(fecha_autorizacion__range=(hoy,ayer),fecha_autorizacion__isnull=False,sendwebservice=0)
        #fecha_autorizacion=datetime.datetime.strptime(,"%d/%m/%Y").strftime('%Y-%m-%d'),sendwebservice=0)
        #date.today())
        personas={}
        datosdict={}
        cantpersonas=''
        totenviados=0
        tienepersonas=''
        tieneelementos=''

        grabo=''
        continua=''
        countinvolus=0
        unireg=0
        i=0
        for hay in grabarfam:
            datosgral=""
            lugar=''
            lati=''
            longi=''
            condiciones=''
            involucrados=0
            idComisaria=0
            idBarrio=0
            idCalle=0
            idRolPersona=0
            idTipoOcupacion=0
            idEstadocivil=0
            naciona=0
            sector=''
            departamento=''
            piso=''
            escalera=''
            ocupacion=''
            cantper=0
            preventivo = Preventivos.objects.get(id=hay.preventivo.id)
            ciudad= preventivo.dependencia.ciudad
            depe=preventivo.dependencia
            laticiudad = RefCiudades.objects.get(id=preventivo.dependencia.ciudad_id)
            localidad=laticiudad.descripcion
            lati=laticiudad.lat
            longi=laticiudad.longi
            tienelugar=True
            tienepersona=False
            involucrados=0
            depes=preventivo.dependencia.ciudad_id
            depes=RefCiudades.objects.get(id=depes)
            #hacer el mapeo de comisarias,calles,barrios del lugar del hecho y de la personas involucradas
            localcria=Localidad.objects.get(descripcion__icontains=depes)
            localcria=int(localcria.idLocalidad)

            amplia = Ampliacion.objects.get(id=hay.id)
            tieneampli=True
            form=AmpliacionForm(instance=amplia)
            involuscra=[]
            eleminvo=[]
            datosper=""
            elementos=""

            involus=PersInvolucradas.objects.filter(ampliacion=amplia.id).all()
            eleinvo=Elementos.objects.filter(ampliacion=amplia.id,borrado__isnull=True).all()

            for crias in Comisarias.objects.filter(idLocalidad=localcria):
                depen=str(depe)
                criasdepe=crias.descripcion.upper()[-7:]

                if criasdepe in depen:
                    idComisaria=crias.idorganismo

            #Datos del lugar del hecho

            if len(PersInvolucradas.objects.filter(ampliacion=amplia.id).all())>0:
                 tienepersonas=True
                 countinvolus=PersInvolucradas.objects.filter(ampliacion=amplia.id).count()

                 for p in PersInvolucradas.objects.filter(ampliacion=amplia.id).all():
                     bandera,personai = funverifica(p.persona.id)
                     domi=Personas.objects.get(id=p.persona.id).persodom.all()
                     involucrados=involucrados+1
                     rolin=str(p.roles)[2:].lower()
                     rolper=RolPersonas.objects.all()
                     for rolesper in rolper:
                        rp=rolesper.descripcion[2:]

                        if rolin==rp:
                           idRolPersona=int(str(rolesper))

                     if 'si' in p.detenido:
                       detenidos=detenidos+1

                     else:
                       detenidos=0

                     if 'si' in p.tentativa:
                         tentativa=1
                     else:
                         tentativa=0
                     if 'si' in p.infraganti:
                        infraganti=1
                     else:
                        infraganti=0

                     if p.juridica=='no':
                       pf='FISICA'
                       pf=1
                     else:
                       #pf='JURIDICA'
                       pf=0
                       if RefTipoDocumento.objects.get(id=p.cuit_id)!='Null':
                          perjuridica=str(RefTipoDocumento.objects.get(id=p.cuit_id))
                       if p.razon_social!=None:
                          perjuridica=perjuridica+'-'+str(p.razon_social)
                       if p.nrocuit!=0:
                          perjuridica=perjuridica+'-'+str(p.nrocuit)



                     if p.menor=='':
                       p.menor="NO"



                     if p.persona.ocupacion!=None:
                       ocupacion=unicode(str(p.persona.ocupacion),'utf8')
                     else:
                       ocupacion=''

                     estadociv=str(p.persona.estado_civil).lower().capitalize()
                     a=Estadocivil.objects.all()
                     for civil in a:
                        civ=civil.descripcion.lower()
                        if civ==civ:
                            idEstadocivil=civil.idEstadoCivil

                     b=TipoOcupacion.objects.all()
                     ocupac=ocupacion.lower()
                     for ocu in b:
                        ocupa=ocu.descripcion.lower()
                        if ocupac==ocupa:
                            idTipoOcupacion=ocu.idtipoocupacion

                     if p.persona.pais_nac_id:
                        nacion=RefPaises.objects.get(id=p.persona.pais_nac_id)
                     else:
                        nacion='Descripcion'

                     naci=Nacionalidad.objects.all()
                     for nac in naci:
                        nacio=nac.descripcion
                        if nacion==nacio:
                             naciona=nac.id

                     tipo_doc=str(p.persona.tipo_doc)[:3]
                     doc=Tipodocumentos.objects.all()
                     tp_doc='INDOC'
                     for doc_tipo in doc:
                        tip=str(doc_tipo.idtipodocumento)[:3]

                        if tipo_doc==tip:
                           tp_doc=tip


                     if len(p.persona.nro_doc)<=8:
                       tipo_doc=p.persona.tipo_doc
                       nro_doc=p.persona.nro_doc

                       tipo_doc='N/T'
                       nro_doc='0'

                     if domi:
                        for l in Personas.objects.get(id=p.persona.id).persodom.all():
                             dad=Personas.objects.get(id=p.persona.id).padre.all()
                             tieneper=True
                             if p.persona.ciudad_res!=None:
                               ciudad_res=p.persona.ciudad_res
                             if l.calle!=None:
                               calle=l.calle

                             if str(l.altura)!='0':
                               altura=l.altura

                             descridomi=str(ciudad_res)+str(calle)+str(altura)
                             if l.calle:
                                calledom=str(l.calle).strip()
                                try:
                                    for cal in Calles.objects.filter(idLocalidad=localcria,descripcion__icontains=calledom):
                                         idCalle=cal.idCalle
                                except ObjectDoesNotExist:
                                         idCalle=0
                             if l.barrio_codigo:
                                barriodom=str(l.barrio_codigo).replace('Bº','').strip()
                                try:
                                    for barr in Barrio.objects.filter(idLocalidad=localcria,descripcion__icontains=barriodom):
                                         idBarrio=barr.idBarrio
                                except ObjectDoesNotExist:
                                         idBarrio=0
                             if dad:

                                    for la in Personas.objects.get(id=p.persona.id).padre.all():
                                            if p.juridica=='si':
                                                persona={'ApellidoyNombres':p.razon_social,'IdTipoDocumento':str(p.cuit_id),'DescripcionTipoDoc':str(RefTipoDocumento.objects.get(id=p.cuit_id)),'NroDocumento':p.nrocuit,'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'PersonaFisica':str(pf),'DescripcionPersonaJuridica':perjuridica,'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':ocupacion,'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),'IdNacionalidad':naciona}
                                            else:
                                                persona={'Apellido':p.persona.apellidos,'Nombre':p.persona.nombres,'IdTipoDocumento':tp_doc,'DescripcionTipoDoc':str(p.persona.tipo_doc),'NroDocumento':p.persona.nro_doc,'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'PersonaFisica':str(pf),'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':ocupacion,'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+str(p.persona.ciudad_nac),'IdNacionalidad':naciona}

                                            domi={'IdBarrio':idBarrio,'IdCalle':idCalle,'Nro':altura,'DescripcionDomicilio':descridomi}
                                            if la.padre_apellidos or la.padre_nombres or la.madre_apellidos or la.madre_nombres:
                                                    padys={'Hijode':la.padre_apellidos+','+la.padre_nombres+'-'+la.madre_apellidos+','+la.madre_nombres}
                                            else:
                                                    padys={'Hijode':'no registra datos de los padres'}

                                            dictpersona=persona
                                            dictpersona.update(domi)
                                            dictpersona.update(padys)
                             else:
                                 if p.juridica=='si':
                                    persona={'ApellidoyNombres':p.razon_social,'IdTipoDocumento':str(p.cuit_id),'DescripcionTipoDoc':str(RefTipoDocumento.objects.get(id=p.cuit_id)),'NroDocumento':p.nrocuit,'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'PersonaFisica':str(pf),'DescripcionPersonaJuridica':perjuridica,'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':ocupacion,'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),'IdNacionalidad':naciona}
                                 else:
                                    persona={'Apellido':p.persona.apellidos,'Nombre':p.persona.nombres,'IdTipoDocumento':tp_doc,'DescripcionTipoDoc':str(p.persona.tipo_doc),'NroDocumento':p.persona.nro_doc,'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'PersonaFisica':str(pf),'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':str(p.persona.ocupacion),'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+str(p.persona.ciudad_nac),'IdNacionalidad':naciona}
                                 domi={'IdBarrio':idBarrio,'IdCalle':idCalle,'Nro':altura,'DescripcionDomicilio':descridomi}
                                 padys={'Hijode':'no registra datos de los padres'}
                                 dictpersona=persona
                                 dictpersona.update(domi)
                                 dictpersona.update(padys)


                     else:

                        if p.juridica=='si':
                            persona={'ApellidoyNombres':p.razon_social,'IdTipoDocumento':str(p.cuit_id),'DescripcionTipoDoc':str(RefTipoDocumento.objects.get(id=p.cuit_id)),'NroDocumento':p.nrocuit,'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'PersonaFisica':str(pf),'DescripcionPersonaJuridica':perjuridica,'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':ocupacion,'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+unicode(str(p.persona.ciudad_nac),'utf8'),'IdNacionalidad':naciona}
                        else:
                            persona={'Apellido':p.persona.apellidos,'Nombre':p.persona.nombres,'IdTipoDocumento':tp_doc,'DescripcionTipoDoc':str(p.persona.tipo_doc),'Alias':p.persona.alias,'IdTipoOcupacion':idTipoOcupacion,'IdEstadoCivil':idEstadocivil,'NroDocumento':p.persona.nro_doc,'PersonaFisica':str(pf),'IdRolPersona':idRolPersona,'DescripcionRol':str(p.roles),'Telefonos':p.persona.celular,'Ocupacion':str(p.persona.ocupacion),'DescripcionEstadoCivil':str(p.persona.estado_civil),'FechaNacimiento': p.persona.fecha_nac.strftime("%d/%m/%Y %H:%m:%S"),'LugarNacimiento':str(p.persona.pais_nac)+'-'+str(p.persona.ciudad_nac),'IdNacionalidad':naciona}
                        domi={'DescripcionDomicilio':'no registra domicilio'}
                        padys={'Hijode':'no registra datos de los padres'}

                        dictpersona=persona
                        dictpersona.update(domi)
                        dictpersona.update(padys)

                     cantper=cantper+1

                 personasxml=dicttoxml(dictpersona,attr_type=False,root='Personas')
                 xmlsper='<Persona>'+personasxml+'</Persona>'
                 cantpersonas=cantpersonas+xmlsper

            xmlsper='<PersonasAmliaciones length='+'"'+str(cantper)+'"'+'>'+cantpersonas+'</PersonasAmpliaciones>'

            rotulo='CantidadElementos '+str(len(eleinvo))
            tituarmas=''
            tituvehiculos=''
            objectele={}
            elements=[]
            elementos=''
            obji=[]
            nrositems=0
            cantelemens=''
            for eles in eleinvo:
                obdata=[]
                obdatav=[]
                if len(Elementosarmas.objects.filter(idelemento=eles.id))>0:
                     idar = Elementosarmas.objects.filter(idelemento=eles.id).values('idarma')
                     obdata=Armas.objects.filter(id=idar)
                     for extra in obdata:
                        tituarmas='Detalle Armas :'+str(extra.subtipos)+',Tipo/s : '+str(extra.tipos)+',Sistema de Disparo : '+str(extra.sistema_disparo)+',Marcas : '+str(extra.marcas)+',Calibre : '+str(extra.calibre)+',Modelo : '+str(extra.modelo)+',Nro Serie : '+str(extra.nro_arma)+',Propietario : '+str(extra.nro_doc)+' - '+str(extra.propietario)


                if len(Elementoscars.objects.filter(idelemento=eles.id))>0:
                     idarv = Elementoscars.objects.filter(idelemento=eles.id).values('idvehiculo')
                     obdatav=Vehiculos.objects.filter(id=idarv)
                     for extrav in obdatav:
                        tituvehiculos='Detalle Vehiculo : '+str(extrav.idmarca)+',Modelo : '+str(extrav.modelo)+',Dominio : '+str(extrav.dominio)+',Año : '+str(extrav.anio)+' Tipos : '+str(extrav.tipov)+',Nro Chasis : '+str(extrav.nchasis)+',Nro. Motor : '+str(extrav.nmotor)+',Propietario : '+str(extrav.nro_doc)+'-'+str(extrav.propietario)

                obse=html2text.html2text(eles.descripcion,True)
                obse=obse.encode('ascii', 'xmlcharrefreplace')
                obse=strip_tags(obse)
                obse=strip_tags(obse).replace('&nbsp;','')
                obser=obse.replace('"','')

                rubros={'IdRelacionElemento':unicode(str(eles.tipo.id),'utf-8'),'RelacionElemento':unicode(str(eles.tipo.descripcion),'utf-8'),'DescripcionElemento':unicode(str(eles.rubro),'UTF-8')+'-'+unicode(str(eles.categoria),'UTF-8')+'-'+obser+'-'+str(eles.unidadmed)+'='+str(eles.cantidad),'TitularArmas':unicode(str(tituarmas),'utf-8'),'TitularVehiculo':unicode(str(tituvehiculos),'UTF-8')}
                #por cada elemento hacer un tag para el xml
                nrositems=nrositems+1
                objetosxml=dicttoxml(rubros,attr_type=False,root='Elementos')
                xmlsobj='<Elemento>'+objetosxml+'</Elemento>'
                cantelemens=cantelemens+xmlsobj

            if nrositems>0:
               xmlsobj='<Elementos length='+'"'+str(nrositems)+'"'+'>'+cantelemens+'</Elementos>'
            else:
               xmlsobj='<Elementos length="0"></Elementos>'



            #datos del preventivos
            infor=''
            autoridad=''
            #datos=Preventivos.objects.get(id=hay.id)

            nro=preventivo.nro
            anio=preventivo.anio
            fecha_denuncia=preventivo.fecha_denuncia
            fecha_carga=preventivo.fecha_carga
            fecha_autorizacion=preventivo.fecha_autorizacion
            caratula=preventivo.caratula
            actuante=preventivo.actuante
            actuante=unicode(str(actuante), 'utf-8').strip()
            preventor=preventivo.preventor
            preventor=unicode(str(preventor), 'utf-8').strip()
            today = datetime.datetime.now()
            autoridades=preventivo.autoridades.values_list('descripcion',flat=True)
            dependencia=preventivo.dependencia.descripcion
            unidadreg=preventivo.dependencia.unidades_regionales.descripcion
            #timezone.localtime(value.fecha_desde).strftime("%d/%m/%Y %H:%M:%S")
            fecha=timezone.localtime(amplia.fecha).strftime("%d/%m/%Y %H:%M:%S")
            fecha_autoamplia=timezone.localtime(amplia.fecha_autorizacion).strftime("%d/%m/%Y %H:%M:%S")
            titulo=amplia.titulo
            #descripcion=html2text.html2text(str(amplia.descripcion),True)
            descripcion=strip_tags(amplia.descripcion).replace('&nbsp;','')
            descripcion=descripcion.replace('"','')
            cierre_causa=amplia.cierre_causa
            if cierre_causa==1:
               fecha_cierre=timezone.localtime(amplia.fecha_cierre).strftime("%d/%m/%Y %H:%m:%S")
            else:
               fecha_cierre=timezone.localtime(amplia.fecha_cierre).strftime("%d/%m/%Y %H:%m:%S")
               cierre_causa=0

            autoridada= amplia.autoridades.values_list('descripcion',flat=True)
            #for a in autoridades:
                #autoridad=autoridad+str(a)+' - '

            for a in autoridada:
                autoridad=autoridad+str(a)+'-'
            jerarqui_a=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=actuante).values('jerarquia_id'))
            jerarqui_p=RefJerarquias.objects.get(id=Actuantes.objects.filter(apeynombres=preventor).values('jerarquia_id'))

            subject  ={'TituloAmpliacion':titulo,'IdComisaria':int(idComisaria),'FechaCarga':fecha,'FechaAutorizacion':fecha_autoamplia,'FechaCierre':fecha_cierre,'DescripcionAmpliacion':descripcion,'Ampliacioncierre':cierre_causa}
            subject1 ={'Destinatarios':autoridad,'LatLocalidad':float(lati),'LngLocalidad':float(longi),}

            datosp = subject
            datosp.update(subject1)

            ampliacionesxml = dicttoxml(datosp,attr_type=False,root=False)
            coddestino='coironrw-test'
            fechahoy=datetime.datetime.now()
            xmlsamp=ampliacionesxml

            xmls='<?xml version="1.0" encoding="utf-8"?>'+\
            '<soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">'+\
            '<soap:Header>'+\
            '<LoginInfo xmlns="http://sij.juschubut.gov.ar">'+\
            '<_usuario>policia-test</_usuario>'+\
            '<_password>policia-test</_password>'+\
            '<Usuario>policia-test</Usuario>'+\
            '<Password>policia-test</Password>'+\
            '</LoginInfo>'+\
            '</soap:Header>'+\
            '<soap:Body>'+\
            '<EnviarMensaje xmlns="http://sij.juschubut.gov.ar">'+\
            '<msg>'+\
            '<IdNodo>1</IdNodo>'+\
            '<Asunto>Info de Ampliaciones Spid</Asunto>'+\
            '<Cuerpo>'+\
            '<![CDATA[<?xml version="1.0" encoding="utf-8"?><MensajeAmpliacion xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" message-type="MensajeroAmpliacion.Tranfer.MensajeAmpliacion, MensajeroAmpliacion.Tranfer" message-version="1">'+xmlsamp+xmlsper+xmlsobj+'</MensajeAmpliacion>'+']]>'+'</Cuerpo>'+\
            '<CodigoRemitente>policia-test</CodigoRemitente>'+\
            '<DescripcionRemitente>Prueba Policia</DescripcionRemitente>'+\
            '<CodigoDestino>coironrw-test</CodigoDestino>'+\
            '</msg>'+\
            '</EnviarMensaje>'+\
            '</soap:Body>'+\
            '</soap:Envelope>'


            user='policia-test'
            password='policia-test'
            params = { 'Authorization' : 'Basic %s' % base64.b64encode("user:password") }
            headers = {"Content-type": "application/x-www-form-urlencoded","Accept": "text/plain"}
            webservice = urllib2.Request('http://listas.juschubut.gov.ar/mensajero/mensajes.asmx')
            webservice = httplib.HTTPConnection('209.13.117.104',80)
            webservice.putrequest("POST", "http://listas.juschubut.gov.ar/mensajero/mensajes.asmx", params, headers)
            webservice.putheader("Host", "listas.juschubut.gov.ar")
            webservice.putheader("User-Agent", "Python Post")
            webservice.putheader("Content-type", "text/xml; charset=\"UTF-8\"")
            webservice.putheader("Content-length", "%d" % len(xmls))
            webservice.putheader("SOAPAction", "\"http://sij.juschubut.gov.ar/EnviarMensaje\"")
            webservice.endheaders()
            webservice.send(xmls)
            totenviados=totenviados+1
            ref=webservice.getresponse()
            refer=str(ref.status)+'-'+str(ref.reason)
            valorweb=0
            if ref.status==200:
               data = ref.read()
               #aqui actualizar el campo sendwebservice en preventivo a 1
               user = User.objects.get(username='23140893')
               ampli = Ampliacion.objects.get(id=hay.id)
               judi=EnvioAmpJudicial()
               judi.ampliacion=ampli.id
               judi.fecha_autorizacion=amplia.fecha_autorizacion
               judi.user=user
               judi.enviado=1
               judi.save()
               valorweb=1
               repla=Ampliacion.objects.filter(id=hay.id).update(sendwebservice=valorweb)
               lista=EnvioAmpJudicial.objects.all()
               webservice.close()
               #return render(request, './enviowebservice.html',{'refer':refer,})
            else:
               user = User.objects.get(username='23140893')
               ampli = Ampliacion.objects.get(id=hay.id)
               judi=EnvioAmpJudicial()
               judi.ampliacion=ampli
               judi.fecha_autorizacion=amplia.fecha_autorizacion
               judi.user=user
               judi.enviado=0
               judi.save()
               lista=EnvioAmpJudicial.objects.all()
               #return render(request, './errorHTTP.html',{'refer':refer,})

            datosdict={}
       else:
          errors="Ingrese Fecha Desde-Hasta"
    lista=EnvioAmpJudicial.objects.all()
    fecha_carga=datetime.datetime.now()
    values={'errors':errors,'destino': destino,'state':state,'fecha_carga':fecha_carga,'lista':lista,'totenviados':totenviados,}
    return render(request,'./enviowebamp.html',values)

@login_required
@permission_required('user.is_staff')
def violencia(request,idhec):
    state= request.session.get('state')
    destino= request.session.get('destino')
    errors=[]
    listacontrol=[]
    control=False
    boton='no'
    datosinteres=''
    grabo=False
    formvif = ViolenciaFliarForm()
    ftiposdelitos=DelitoForm()
    modos=RefModosHechoForm()
    hec=Hechos.objects.get(id=idhec)
    form=HechosForm(instance=hec)
    ftiposdelitos=DelitoForm()
    modos=RefModosHechoForm(instance=hec)
    delito =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)
    descripcion=hec.descripcion
    idhec=hec.id
    motivo=hec.motivo
    fecha_desde=hec.fecha_desde
    fecha_hasta=hec.fecha_hasta
    datos=Preventivos.objects.get(id=hec.preventivo_id)
    idprev=datos.id
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    autoridades= datos.autoridades.values_list('descripcion',flat=True)
    dependencia=datos.dependencia.descripcion
    unidadreg=datos.dependencia.unidades_regionales.descripcion
    fecha_autorizacion=datos.fecha_autorizacion
    if request.POST.get('grabar')=='Guardar Datos':
       formvif=ViolenciaFliarForm(request.POST, request.FILES)
       fecha=fecha_denuncia
       idhecho=request.POST.get('ids')

       if formvif.is_valid():
          formvio=ViolenciaFliar()
          formvio.fecha=fecha
          formvio.fecha_carga=fecha_carga
          formvio.intervencionsavd=formvif.cleaned_data['intervencionsavd']
          formvio.intervencionotro=formvif.cleaned_data['intervencionotro']
          formvio.intervencioncual=formvif.cleaned_data['intervencioncual']
          formvio.otrosdatosinteres=formvif.cleaned_data['otrosdatosinteres']
          formvio.hechos_id=idhecho
          try:
            #if not buscar:
               formvio.save()
            #else:
               #actualiza=ViolenciaFliar.objects.filter(hechos_id=idhecho).update(intervencionsavd=formvif.cleaned_data['intervencionsavd'],intervencionotro=formvif.cleaned_data['intervencionotro'],intervencioncual=formvif.cleaned_data['intervencioncual'],otrosdatosinteres=formvif.cleaned_data['otrosdatosinteres'])
          except IntegrityError:
            #if not buscar:
            errors='Hecho de Violencia Familiar fue actualizado'

            actualiza=ViolenciaFliar.objects.filter(hechos_id=idhecho).update(intervencionsavd=formvif.cleaned_data['intervencionsavd'],intervencionotro=formvif.cleaned_data['intervencionotro'],intervencioncual=formvif.cleaned_data['intervencioncual'],otrosdatosinteres=formvif.cleaned_data['otrosdatosinteres'])

          buscar=ViolenciaFliar.objects.get(hechos_id=idhecho)
          formvif=ViolenciaFliarForm(instance=buscar)
          datosinteres=buscar.otrosdatosinteres
          grabo=True

    info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,'datosinteres':datosinteres,
    'caratula':caratula,'boton':boton,'descripcion':descripcion,'fecha_desde':fecha_desde,'fecha_hasta':fecha_hasta,'motivo':motivo,
    'actuante':actuante,'fecha_autorizacion':fecha_autorizacion,'errors':errors,'grabo':grabo,
    'preventor':preventor,'idprev':idprev,'delito':delito,'formvif':formvif,
    'autoridades':autoridades,'errors': errors, 'dependencia':dependencia,'unidadreg':unidadreg,
    'state':state,'delitos':delitos,'destino': destino,'idhec':idhec}
    return render(request,'./formvif.html',info)

@login_required
@transaction.atomic
@group_required(["administrador","policia","investigaciones","radio"])
def persinvovif(request,idhec,idper):
    state= request.session.get('state')
    destino= request.session.get('destino')
    hechos = Hechos.objects.get(id = idhec)
    buscar=ViolenciaFliar.objects.get(hechos_id=idhec)
    datosinteres=html2text.html2text(buscar.otrosdatosinteres)

    idciu = hechos.preventivo.dependencia.ciudad_id
    depe = hechos.preventivo.dependencia
    ids = Preventivos.objects.get(id = hechos.preventivo_id)
    text=''
    errors=[]
    mostrar="0"
    todos=[]
    comb=""
    roles=""
    formro=""
    domicilios=""
    estadete="no"
    datosinvo=[]
    tieneHecho = False
    tienePersonas=False
    tienelugar=False
    siexistepoli=False
    formp = PersonasForm(request.POST)
    domicilios = Domicilios(request.POST)
    dom = DomiciliosForm(request.POST)
    formr = PerInvolViolenfliarForm(request.POST)
    formpa = PadresForm(request.POST)
    formc = CiudadesForm()
    formd = DepartamentosForm()
    formpr = ProvinciasForm()
    personas=''
    boton='no'
    grabo=False
    formvif = ViolenciaFliarForm(instance=buscar)
    ftiposdelitos=DelitoForm()
    modos=RefModosHechoForm()
    hec=Hechos.objects.get(id=idhec)
    form=HechosForm(instance=hec)
    ftiposdelitos=DelitoForm()
    modos=RefModosHechoForm(instance=hec)
    delito =HechosDelito.objects.filter(hechos = idhec,borrado__isnull=True)
    descripcion=hec.descripcion
    idhec=hec.id
    motivo=hec.motivo
    fecha_desde=hec.fecha_desde
    fecha_hasta=hec.fecha_hasta
    datos=Preventivos.objects.get(id=hec.preventivo_id)
    idprev=datos.id
    nro=datos.nro
    anio=datos.anio
    fecha_denuncia=datos.fecha_denuncia
    fecha_carga=datos.fecha_carga
    caratula=datos.caratula
    actuante=datos.actuante
    preventor=datos.preventor
    autoridades= datos.autoridades.values_list('descripcion',flat=True)
    dependencia=datos.dependencia.descripcion
    unidadreg=datos.dependencia.unidades_regionales.descripcion
    fecha_autorizacion=datos.fecha_autorizacion

    if request.POST.get('grabarciu')=="Grabar":
         formc = CiudadesForm(request.POST, request.FILES)
         descripcion = request.POST.get('descripcion')
         pais = request.POST.get('pais')
         #provincia = request.POST.get('provincia')
         #departamento = request.POST.get('departamento')

         if not descripcion or not pais:
                 errors.append('Ingrese una referencia(Pais) a la que pertenece la ciudad')
         else:
                         if not(len(descripcion)>=4 and len(descripcion)< 80):
                                         errors.append('El dato ingresado debe tener entre 4 y 80 caracteres')
                         else:
                                                 formc = CiudadesForm(request.POST, request.FILES)

                                                 if formc.is_valid():
                                                         formc.save()

                                                 else:
                                                            errors.append('La ciudad que UD. desea Guardar ya Existe. Verifique')
    #gaba domicilio
    if request.POST.get('grabadomi')=='Guardar':
         formp = PersonasForm(request.POST, request.FILES)
         dom = DomiciliosForm(request.POST)
         personas = Personas.objects.get(id = idper)

         reside=RefCiudades.objects.get(id=request.POST.get('ciudad_res'))
         if dom.is_valid():
                try:
                 domicilios                      = Domicilios()
                 domicilios.barrio_codigo        = dom.cleaned_data['barrio_codigo']
                 domicilios.calle                = dom.cleaned_data['calle']
                 domicilios.altura               = dom.cleaned_data['altura']
                 domicilios.entre                = dom.cleaned_data['entre']
                 domicilios.fecha_desde          = date.today()
                 domicilios.fecha_hasta          = dom.cleaned_data['fecha_hasta']
                 domicilios.fecha_actualizacion  = dom.cleaned_data['fecha_actualizacion']
                 domicilios.tipos_domicilio      = dom.cleaned_data['tipos_domicilio']
                 domicilios.ref_zona             = dom.cleaned_data['ref_zona']
                 domicilios.departamento         = dom.cleaned_data['departamento']
                 domicilios.piso                 = dom.cleaned_data['piso']
                 domicilios.lote                 = dom.cleaned_data['lote']
                 domicilios.sector               = dom.cleaned_data['sector']
                 domicilios.manzana              = dom.cleaned_data['manzana']
                 domicilios.ref_ciudades         = reside
                 domicilios.personas             = personas
                 domicilios.save()
                except IntegrityError:
                    errors.append('Datos duplicados.-')
         else:
                mostrar="no"

    #graba barrios segun la ciudad
    if request.POST.get('grabab')=='Grabar':
        formp=PersonasForm(request.POST, request.FILES)
        formbarrios = BarriadasForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        ciudad = request.POST.get('ciudad')
        mostrar="no"
        if not descripcion:
            errors.append('Ingrese un Barrio')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<100):
                    errors.append('El dato ingresado debe tener entre 4 y 100 caracteres')
             else:
                    if ciudad == 'Seleccione ciudad':
                        errors.append('Debe seleccionar una ciudad')
                    else:


                            if formbarrios.is_valid():
                             formbarrios.save()



    ##graba calles segun idciu
    if request.POST.get('grabac')=='Grabar':
        formcalles = AddressForm(request.POST, request.FILES)
        descripcion = request.POST.get('descripcion')
        ciudad = request.POST.get('ciudad')

        mostrar="no"
        if not descripcion:
            errors.append('Ingrese el nombre de la calle')
        else:
             if not(len(descripcion)>=4 and len(descripcion)<150):
                    errors.append('El dato ingresado debe tener entre 4 y 150 caracteres')
             else:
                    if not ciudad:
                        errors.append('Debe seleccionar una ciudad')
                    else:
                         if formcalles.is_valid():
                                formcalles.save()





    #busca para mostrar botones aqui hacer lo logico para mostrar botones de violencia familiar relato del hecho y personas

    if request.POST.get('nuevo')=="Nuevo" :
         formp=PersonasForm()
         domicilios = Domicilios()
         dom = DomiciliosForm()
         formr = PerInvolViolenfliarForm()
         formpa = PadresForm()
         idper=0
         mostrar="no"


    else:
        if request.POST.get('search')=="Buscar":
         texto=request.POST.get('texto')

         if texto:
            query_string=texto
            entry_query = get_query(query_string, ['nro_doc', 'apellidos','nombres',])
         else:
            query_string='%'
            entry_query = get_query(query_string, ['nro_doc', 'apellidos','nombres',])



         filtro=Personas.objects.filter(entry_query)

         #}.values('id','tipo_doc','apellidos','nro_doc','nombres','ocupacion','fecha_nac',)

         if filtro:
                if filtro not in todos:
                        todos.append(filtro)
                        mostrar="es"
         else:
                errors.append('No existen Personas con la referencia de búsqueda ingresada')
                mostrar="0"

        if idper != '0' and mostrar!='es':
            mostrar="si"
            filtros=Personas.objects.filter(id = idper)
            if filtros not in todos:
                 todos.append(filtros)

            personas = Personas.objects.get(id = idper)

            formp = PersonasForm(instance=personas)
            domicilios = Domicilios()
            dom = DomiciliosForm()

            formr = PerInvolViolenfliarForm()
            filt= Padres.objects.filter(persona = personas.id)
            if filt:
                 idpapis= Padres.objects.get(persona = personas.id)
                 formpa = PadresForm(instance=idpapis)
            else:
                 formpa=PadresForm()

            if len(Domicilios.objects.filter(personas = idper)) > 0:
             domicilios = Domicilios.objects.filter(personas = idper)[0]
             dom = DomiciliosForm(instance = domicilios)
             dom.fields['barrio_codigo'].queryset = RefBarrios.objects.filter(ciudad=personas.ciudad_res)
             dom.fields['calle'].queryset = dom.fields['entre'].queryset= RefCalles.objects.filter(ciudad=personas.ciudad_res)
            else:

             dom = DomiciliosForm()

             dom.fields['barrio_codigo'].queryset = RefBarrios.objects.filter(ciudad=personas.ciudad_res)
             dom.fields['calle'].queryset = dom.fields['entre'].queryset= RefCalles.objects.filter(ciudad=personas.ciudad_res)

        else:
            if (request.POST.get('grabab') is None and request.POST.get('grabac') is None):
                 formp = PersonasForm()
                 domicilios = Domicilios()
                 dom = DomiciliosForm()
                 formr = PerInvolViolenfliarForm()
                 formpa = PadresForm()



        if request.POST.get('dele'):
            persoinvoluc=PersInvolucradas.objects.filter(id=request.POST.get('dele'))

            if persoinvoluc:
                persoinvolu=PersInvolucradas.objects.get(id=request.POST.get('dele'))
                if 'si' in persoinvolu.detenido:
                    if request.user.userprofile.depe==depe or request.user.userprofile.depe.descripcion == 'INVESTIGACIONES' or 'RADIO' in request.user.userprofile.depe.descripcion:
                        Detenidos.objects.filter(persona_id = persoinvolu.persona_id).update(borrado='S',observaciones=request.user.username)
                        PersInvolucradas.objects.get(id=request.POST.get('dele')).delete()
                    else:
                        errors.append('No se puede modificar preventivos de otras dependencias.')
                else:
                    PersInvolucradas.objects.get(id=request.POST.get('dele')).delete()

            else:
                        errors.append('No se existe persona involucrada.')
            mostrar="0"
        if request.POST.get('grabar')=="Guardar":

                 formp = PersonasForm(request.POST, request.FILES)          #obtiene los datos de la persona en un formulario persona
                 dom = DomiciliosForm(request.POST,request.FILES)           #obtiene los datos del domicilio en un formulario domicilio
                 formr = PerInvolViolenfliarForm(request.POST,request.FILES)   #obtiene los datos de persona involucrada en un formulario persona involucrada
                 formpa = PadresForm(request.POST,request.FILES)
                 #fecha_detencion=datetime.datetime.strptime(request.POST.get('fechahoradetencion'), '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y')

                 if idper!='0':

                     perso=Personas.objects.get(id=idper)
                     fil=Padres.objects.filter(persona=perso.id)
                     if fil:
                            papis= Padres.objects.get(persona = perso.id)
                            formpa = PadresForm(instance=papis)
                     else:
                            formpa=PadresForm()
                            papis=Padres()
                     #buscar en personal
                     findpoli=Personal.objects.filter(persona_id=perso.id)
                     if findpoli:
                            siexistepoli=True
                 else:

                     perso=Personas()
                     papis=Padres()

                     iddom='1'

                 if len(Domicilios.objects.filter(personas = idper)) > 0:
                     domicilios = Domicilios.objects.filter(personas = idper)[0]
                     iddom=domicilios.id
                 else:
                     iddom='1'
                     domicilios=Domicilios()



                 if formp.is_valid() or idper!='0':

                        if formp.is_valid():

                         perso.apellidos  = formp.cleaned_data['apellidos']
                         perso.nombres    = formp.cleaned_data['nombres']
                         perso.tipo_doc   = formp.cleaned_data['tipo_doc']
                         perso.nro_doc    = formp.cleaned_data['nro_doc']
                         perso.fecha_nac  = formp.cleaned_data['fecha_nac']
                         perso.sexo_id    = formp.cleaned_data['sexo_id']
                         perso.pais_nac   = formp.cleaned_data['pais_nac']
                         perso.ciudad_nac = formp.cleaned_data['ciudad_nac']
                         perso.pais_res   = formp.cleaned_data['pais_res']
                         perso.ciudad_res = formp.cleaned_data['ciudad_res']
                         perso.ocupacion  = formp.cleaned_data['ocupacion']
                         perso.alias      = formp.cleaned_data['alias']



                         perso.estado_civil = formp.cleaned_data['estado_civil']
                         idpoli=formp.cleaned_data['ocupacion']
                         refpoli=RefOcupacion()
                         if idpoli:
                            refpoli=RefOcupacion.objects.get(descripcion=idpoli)


                        else:

                             if request.POST.get('ocupacion')=='None' or request.POST.get('ocupacion')=='':
                                 refpoli=RefOcupacion.objects.get(descripcion='EMPLEADO')
                                 texto='EMPLEADO'
                             else:
                                 idpoli=request.POST.get('ocupacion')
                                 refpoli=RefOcupacion.objects.get(id=idpoli)
                                 texto=refpoli.descripcion

                        if dom.is_valid():


                                 domicilios.barrio_codigo        = dom.cleaned_data['barrio_codigo']
                                 domicilios.calle                = dom.cleaned_data['calle']
                                 domicilios.altura               = dom.cleaned_data['altura']
                                 domicilios.entre                = dom.cleaned_data['entre']
                                 domicilios.fecha_desde          = dom.cleaned_data['fecha_desde']
                                 domicilios.fecha_hasta          = dom.cleaned_data['fecha_hasta']
                                 domicilios.fecha_actualizacion  = dom.cleaned_data['fecha_actualizacion']
                                 domicilios.tipos_domicilio      = dom.cleaned_data['tipos_domicilio']
                                 domicilios.ref_zona             = dom.cleaned_data['ref_zona']
                                 domicilios.departamento         = dom.cleaned_data['departamento']
                                 domicilios.piso                 = dom.cleaned_data['piso']
                                 domicilios.lote                 = dom.cleaned_data['lote']
                                 domicilios.sector               = dom.cleaned_data['sector']
                                 domicilios.manzana              = dom.cleaned_data['manzana']


                                 if idper!='0':
                                        persom=Personas.objects.get(id=idper)


                                        if formp.is_valid() or idper!='0':
                                            persom.ciudad_res = formp.cleaned_data['ciudad_res']
                                            persom.ocupacion  = formp.cleaned_data['ocupacion']
                                            persom.alias      = formp.cleaned_data['alias']
                                            persom.estado_civil = formp.cleaned_data['estado_civil']
                                            try:


                                                persom.save()

                                            except IntegrityError:
                                                errors.append('')
                                        else:

                                             mostrar='si'
                                             errors.append(formp.errors.as_text)

                                 else:

                                         perso.save()




                                 if idper!='0':

                                                 idpersu=Personas.objects.get(id=idper)

                                                 personas=idpersu
                                 else:
                                                 idpersu=Personas.objects.get(id=perso.id)

                                                 personas=idpersu

                                 domicilios.personas             = personas
                                 domicilios.ref_ciudades         = formp.cleaned_data['ciudad_res']

                                 domicilios.save()

                                 if refpoli:
                                    if refpoli.descripcion.find('POLICI')>=0:
                                                policia=Personal()
                                                policia.persona_id = personas
                                                policia.credencial=0
                                                try:
                                                    policia.save()
                                                except IntegrityError:
                                                    errors.append('')
                                    else:
                                             if siexistepoli:
                                                    #borro esa persona en personal
                                                    borrar=Personal.objects.get(persona_id=personas).delete()

                                 papis.persona=personas
                                 papis.padre_apellidos=request.POST.get('padre_apellidos')
                                 papis.padre_nombres=request.POST.get('padre_nombres')
                                 papis.madre_apellidos=request.POST.get('madre_apellidos')
                                 papis.madre_nombres=request.POST.get('madre_nombres')
                                 try:
                                        papis.save()
                                 except IntegrityError:
                                     errors.append('Datos existente en Padres')

                                 if formr.is_valid():

                                     persoin=PerInvolViolenfliar()
                                     #detenidos = Detenidos()
                                     persoin.persona=personas
                                     #detenidos.persona = personas
                                     persoin.violencia = buscar
                                     persoin.roles = formr.cleaned_data['roles']
                                     #persoin.menor = formr.cleaned_data['menor']
                                     #persoin.detenido = formr.cleaned_data['detenido']
                                     persoin.cargado_viol=True


                                     if 'DENUNCIANTE' in str(formr.cleaned_data['roles']) or 'Denunciante' in str(formr.cleaned_data['roles']).capitalize():
                                        persoin.vinculovictima = formr.cleaned_data['vinculovictima']
                                        persoin.pidereserva = formr.cleaned_data['pidereserva']
                                        persoin.juridica = formr.cleaned_data['juridica']
                                        if persoin.juridica=='si':
                                            persoin.razon_social = formr.cleaned_data['razon_social']
                                            persoin.cargo = formr.cleaned_data['cargo']
                                        else:
                                            persoin.razon_social = 'SIN DESCRIPCION'
                                            persoin.cargo = 'SIN CARGO'

                                     else:
                                        if 'VICTIMA' in str(formr.cleaned_data['roles']) or 'Victima' in str(formr.cleaned_data['roles']).capitalize():
                                            persoin.composiciongrupofliar = formr.cleaned_data['composiciongrupofliar']
                                            persoin.cgfconviven = formr.cleaned_data['cgfconviven']
                                            persoin.vinculovictima = formr.cleaned_data['vinculovictima']
                                            persoin.vdconviven = formr.cleaned_data['vdconviven']
                                            persoin.teldomalternativos = formr.cleaned_data['teldomalternativos']
                                            persoin.teldomfliaprimaria = formr.cleaned_data['teldomfliaprimaria']
                                            persoin.telconfigurasreferentes = formr.cleaned_data['telconfigurasreferentes']


                                     try:
                                        mostrar='0'

                                        persoin.save()

                                     except IntegrityError:

                                         errors.append('Datos inexistente en Carateristicas segun el Rol')
                                         mostrar='si'
                                     #mostrar='0'
                                 else:
                                     #persi=Personas.objects.get(id=idper)
                                     #formp=PersonasForm(instance=persi)
                                     errors.append('Error faltan datos en seccion de Rol de la Persona')
                                     #return HttpResponseRedirect('./',)
                                     #mostrar='es'
                                     filtros=Personas.objects.filter(id = personas.id)
                                     if filtros not in todos:
                                         todos.append(filtros)

                        else:
                                     mostrar='no'


                 else:
                    mostrar='no'

    try:
        noposee = Personas.objects.get(id=idper).tipo_doc.descripcion
    except Exception as e:
        noposee=""


    listap=Personas.objects.all()
    #envio de datos al template updatehechos.html
    formcalles= AddressForm()
    formbarrios = BarriadasForm()
    formciu=RefCiudades.objects.all()
    datosinvo=buscar.involuvif.all()
    formr = PerInvolViolenfliarForm()
    info={'nro':nro,'anio':anio,'fecha_denuncia':fecha_denuncia,'fecha_carga':fecha_carga,'datosinteres':datosinteres,'formp':formp,'formpa':formpa,'domicilios':domicilios,
    'caratula':caratula,'boton':boton,'descripcion':descripcion,'fecha_desde':fecha_desde,'fecha_hasta':fecha_hasta,'motivo':motivo,
    'actuante':actuante,'fecha_autorizacion':fecha_autorizacion,'errors':errors,'grabo':grabo,'formciu':formciu,'mostrar':mostrar,
    'preventor':preventor,'idprev':idprev,'delito':delito,'formvif':formvif,'formr':formr,'formcalles':formcalles,'formbarrios':formbarrios,
    'autoridades':autoridades,'errors': errors, 'dependencia':dependencia,'unidadreg':unidadreg,'listap':listap,'todos':todos,'datosinvo':datosinvo,
    'state':state,'delitos':delitos,'destino': destino,'idhec':idhec}
    return render(request,'./formpersovif.html',info)

@login_required
def obtener_no_enviados(request):
    depe = request.user.userprofile.depe
    no_enviados = EnvioPreJudicial.objects.filter(enviado = 0,dependencia=depe)
    data = serializers.serialize("json",no_enviados)
    return HttpResponse(data, content_type='application/json')


@login_required
def obtener_cantidad_no_enviados(request):
    depe = request.user.userprofile.depe
    return  Preventivos.objects.filter(fecha_autorizacion__isnull=False, fecha_envio__isnull=True).count()

@login_required
def obtener_cantidad_no_autorizados(request):
    user = request.user
    if Actuantes.objects.filter(funcion__gt=1,documento=user.username):
        id_preventor = Actuantes.objects.get(documento=user.username).id
        return Preventivos.objects.filter(fecha_autorizacion__isnull=True,preventor=id_preventor).count()
    return False

@login_required
def pendientes_envio(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    user = request.user
    depe = user.userprofile.depe
    pendientes = Preventivos.objects.filter(fecha_autorizacion__isnull=False, fecha_envio__isnull=True)
    info = {
        'state'         : state,
        'destino'       : destino,
        'pendientes'    : pendientes,
    }
    return render(request,'./pendientes_envio.html',info)

def verificar_enviado(prev):
    sin_enviar = []
    try:
        envio = EnvioPreJudicial.objects.get(preventivo=prev)
        return bool(envio.enviado),sin_enviar
    except  MultipleObjectsReturned:
        try:
            EnvioPreJudicial.objects.get(preventivo=prev,enviado=1)
            for envio in EnvioPreJudicial.objects.filter(preventivo=prev):
                if not envio.enviado:
                    sin_enviar.append(envio.id)

            return bool(EnvioPreJudicial.objects.get(preventivo=prev,enviado=1).enviado),sin_enviar
        except ObjectDoesNotExist:
            return False,sin_enviar
        except MultipleObjectsReturned:
            for envio in EnvioPreJudicial.objects.filter(preventivo=prev):
                if not envio.enviado:
                    sin_enviar.append(envio.id)
            return True,sin_enviar
    except ObjectDoesNotExist:
        return False, sin_enviar

def marcar_enviados(sin_enviar):
    for elemento in sin_enviar:
        envio = EnvioPreJudicial.objects.get(id=elemento)
        envio.enviado = 1
        envio.save()

@login_required
def pendientes_autorizacion(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    user = request.user
    preventor = Actuantes.objects.get(documento=user.username)
    pendientes = Preventivos.objects.filter(fecha_autorizacion__isnull=True,preventor=preventor)
    info = {
        'state'         :state,
        'destino'       : destino,
        'pendientes'    :pendientes,
    }
    return render(request,'./pendientes_autorizacion.html',info)

def enviarp2(idprev):
    preventivo      = Preventivos.objects.get(id=idprev)
    hecho           = preventivo.hechos.all()[0]
    delitos         = hechos.hechos.all()
    personas        = hecho.involu.all()
    lugar           = hecho.lugar_hecho 
    elementos       = hecho.eleinvolu.all()

    pass

def buscar_calle(localidad,calle):
    if 'AVENIDA' in calle:
        calle = calle.strip('AVENIDA ')
    elif 'AV.' in calle:
        calle = calle.strip('AV. ')
    calles = Calles.objects.filter(idLocalidad=localidad.idLocalidad)
    if calles.filter(descripcion__icontains=calle):
        return calles.filter(descripcion__icontains=calle)[0].pk
    return 0

def buscar_barrio(localidad,barrio):
    barrios = Barrio.objects.filter(idLocalidad=localidad.idLocalidad)
    if 'BARRIO' in barrio:
        barrio = barrio.strip('BARRIO ')
    elif u'B° '.encode('utf8') in barrio:
        barrio = barrio.strip(u'B° '.encode('utf8'))
    elif u'B ° '.encode('utf8') in barrio:
        barrio = barrio.strip(u'B ° '.encode('utf8'))
    if barrios.filter(descripcion__icontains = barrio):
        return barrios.filter(descripcion__icontains = barrio)[0].pk
    return 0

def buscar_tipos_doc(tipo_doc):
    if Tipodocumentos.objects.filter(descripcion__icontains=tipo_doc):
        return Tipodocumentos.objects.filter(descripcion__icontains=tipo_doc)[0].pk
    return 'DESC'

def actualizar_ultimo_ingreso():
    usuarios = User.objects.all()
    for user in usuarios:
        profile = user.userprofile
        profile.ultimo_ingreso = user.last_login
        profile.save()

@login_required
@permission_required('user.is_staff')
def cambiar_password(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    user = request.user
    usuarios = User.objects.all()
    usuarios = usuarios.exclude(username = user)
    usuarios = usuarios.exclude(username = 'fernando')
    form = CambiarContraseniaForm()
    info = {
        'state':state,
        'destino':destino,
        'usuarios':usuarios,
        'form':form,
    }


    if request.method == 'POST':
        form = CambiarContraseniaForm(request.POST)
        if form.is_valid():
            cambio                              = CambiarContrasenia()
            try:
                usuario                             = User.objects.get(id = request.POST.get('usuario'))
                cambio.motivo                       = form.cleaned_data['motivo']
                cambio.detalle_motivo               = form.cleaned_data['detalle_motivo']
                cambio.usuario_que_cambia           = user.username
                cambio.usuario                      = usuario.username
                cambio.fecha_cambio                 = datetime.datetime.now()
                profile_usuario                     = usuario.userprofile
                profile_usuario.last_login          = True
                profile_usuario.solicitud_cambio    = True
                profile_usuario.fecha_solicitud     = datetime.datetime.now()
                profile_usuario.clave_anterior      = usuario.password
                try:
                    cambio.save()
                    profile_usuario.save()
                    usuario.set_password(usuario.username)
                    usuario.save()
                    info['msg'] = 'El cambio de contraseña se realizo con exito.'
                except Exception as e:
                    info['error'] = 'Hubo un error en al intentar cambiar la contraseña, por favor vuelva a intentarlo mas tarde.'
            except Exception as e:
                info['error'] = 'Por favor revise haber ingresado todos los datos del formulario muchas gracias.'
        else:
            info['error'] = 'Por favor revise haber ingresado todos los datos del formulario muchas gracias.'
    return render(request,'./cambiar_password.html',info)


@login_required
def preventivos_autorizados_n_dias(request,dependencia):
    usuario = request.user
    actual_date = datetime.datetime.now()
    initial_date = actual_date - timedelta(days=15)
    preventivos = Preventivos.objects.filter(fecha_autorizacion__range = (initial_date,actual_date),dependencia = dependencia,reenviado=False)
    data = serializers.serialize("json", preventivos)
    return HttpResponse(data, content_type='application/json')

@login_required
def reenvio(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    info = {
        'state':state,
        'destino':destino,
    }
    usuario = request.user
    unidades=False
    if '-' in usuario.userprofile.depe.descripcion:
        ur = usuario.userprofile.depe.descripcion.split('RADIO CABECERA-')[1]
        if ur == 'ESQ':
            ur = UnidadesRegionales.objects.get(descripcion__startswith = 'ESQ')
        elif ur == 'PM':
            ur = UnidadesRegionales.objects.get(descripcion__startswith = 'PUER')
        elif ur == 'CR':
            ur = UnidadesRegionales.objects.get(descripcion__startswith = 'COM')
        else:
            ur = UnidadesRegionales.objects.get(descripcion__startswith = 'TRE')
    else:
        ur = UnidadesRegionales.objects.get(descripcion__startswith = 'AREA COM')

    if 'COMUNICACIONES' in ur.descripcion:
        unidades = UnidadesRegionales.objects.all()
    if unidades:
        dependencias = {}
    else:
        dependencias = Dependencias.objects.filter(ciudad = ur.ciudad,unidades_regionales = ur)
    actual_date = datetime.datetime.now()
    initial_date = actual_date - timedelta(days=15)
    preventivos = {}
    for dependencia in dependencias:
        info[dependencia.id] = Preventivos.objects.filter(fecha_autorizacion__range=(initial_date,actual_date),dependencia=dependencia)
    info['dependencias'] = dependencias
    info['unidades'] = unidades
    request.session['reenvio'] = True
    try:
        if request.session['msg']:
            info['msg'] = request.session['msg']
            del request.session['msg']
    except KeyError as e:
        print( e)

    return render(request,'./reenvio.html',info)

@login_required
def reenviar(request,idprev):
    hecho = Preventivos.objects.get(id=idprev).hecho.id
    return HttpResponseRedirect(reverse('informa',args=[hecho,idprev,0]))

def obtener_preventivo(request,depe,numero,anio):
    preventivo = Preventivos.objects.get(dependencia=depe,nro=numero,anio=anio)
    data = serializers.serialize("json",[preventivo,])
    return HttpResponse(data, content_type='application/json')

@login_required
def autorizados_envio(request):
    state= request.session.get('state')
    destino= request.session.get('destino')
    info = {
        'state':state,
        'destino':destino,
    }
    usuario = request.user
    dependencias = Dependencias.objects.filter(ciudad = usuario.userprofile.depe.ciudad )
    preventivos = Preventivos.objects.filter(dependencia__in=dependencias,fecha_autorizacion__isnull=False,fecha_envio__isnull = True).order_by('-id')           #para esas dependencias obtiene los preven
    depes = preventivos.values('dependencia').distinct()
    dependencias = dependencias.filter(id__in=depes)
    info['preventivos'] = preventivos
    info['dependencias'] = dependencias
    return render(request,'./enviar.html',info)

@login_required
def envio(request,idprev):
    state= request.session.get('state')
    destino= request.session.get('destino')
    info = {
        'state':state,
        'destino':destino,
    }
    usuario = request.user
    info['preventivo'] = Preventivos.objects.get(id = idprev)
    info['hecho'] = info['preventivo'].hecho
    info['autoridades'] = info['preventivo'].autoridades.all()

    return render(request,'./verificar_envio.html',info)

def eliminarduplicados(arreglo):
    resultado = []
    for elemento in arreglo:
        if elemento not in resultado:
            resultado.append(elemento)
    return resultado


def verificar_persona(request,dni):
    msg=""
    try:
        User.objects.get(username=dni)
        msg="El usuario ya existe."
        return HttpResponseBadRequest(msg)
    except:
        try:
            cursor = connections['default'].cursor()
            cursor.execute("select p.nro_doc as documento, p.nombres as nombre, p.apellidos as apellido, p.fecha_nac as fecha_nacimiento, cn.descripcion as ciudad_nacimiento,"\
            "rec.descripcion as estado_civil, cr.descripcion as ciudad_residencia, rpn.descripcion as provincia_nacimiento, rpr.descripcion as provincia_residencia "\
            "from personas p right join personal pp ON p.id = pp.persona_id_id "\
            "join ref_ciudades cn ON p.ciudad_nac_id = cn.id "\
            "join ref_estadociv rec ON p.estado_civil_id = rec.id "\
            "join ref_ciudades cr ON p.ciudad_res_id = cr.id "\
            "join ref_provincia rpn ON cn.provincia_id = rpn.id "\
            "join ref_provincia rpr ON cr.provincia_id = rpr.id "\
            "where nro_doc = %s" % dni)
            columns = [col[0] for col in cursorRh.description]
            personas = [
            	dict(zip(columns, row))
                for row in cursorRh.fetchall()
            ]
            if len(personas<1):
                cursor = connections['default'].cursor()
                cursor.execute("select p.nro_doc as documento, p.nombres as nombre, p.apellidos as apellido, p.fecha_nac as fecha_nacimiento, cn.descripcion as ciudad_nacimiento,"\
                "rec.descripcion as estado_civil, cr.descripcion as ciudad_residencia, rpn.descripcion as provincia_nacimiento, rpr.descripcion as provincia_residencia "\
                "from personas p right join personal pp ON p.id = pp.persona_id_id "\
                "left join ref_ciudades cn ON p.ciudad_nac_id = cn.id "\
                "left join ref_estadociv rec ON p.estado_civil_id = rec.id "\
                "left join ref_ciudades cr ON p.ciudad_res_id = cr.id "\
                "left join ref_provincia rpn ON cn.provincia_id = rpn.id "\
                "left join ref_provincia rpr ON cr.provincia_id = rpr.id "\
                "where nro_doc = %s "\
                "union all "\
                "select p.nro_doc as documento, p.nombres as nombre, p.apellidos as apellido, p.fecha_nac as fecha_nacimiento, cn.descripcion as ciudad_nacimiento,"\
                "rec.descripcion as estado_civil, cr.descripcion as ciudad_residencia, rpn.descripcion as provincia_nacimiento, rpr.descripcion as provincia_residencia "\
                "from personas p right join personal pp ON p.id = pp.persona_id_id "\
                "right join ref_ciudades cn ON p.ciudad_nac_id = cn.id "\
                "right join ref_estadociv rec ON p.estado_civil_id = rec.id "\
                "right join ref_ciudades cr ON p.ciudad_res_id = cr.id "\
                "right join ref_provincia rpn ON cn.provincia_id = rpn.id "\
                "right join ref_provincia rpr ON cr.provincia_id = rpr.id "\
                "where nro_doc = %s" % (dni,dni))
                columns = [col[0] for col in cursorRh.description]
                personas = [
                	dict(zip(columns, row))
                    for row in cursorRh.fetchall()
                ]
                if len(personas) < 1:
                    raise Exception('sin datos')
            obtenerReferencias(personas)
            data = json.dumps(personas, default=date_handler)
            content_type = 'application/json'
            return HttpResponse(data,content_type)
        except:
            try:
                cursorRh = connections['rrhh'].cursor()
                cursorRh.execute("select p.documento as documento, p.nombre as nombre, p.apellido as apellido, p.fecha_nacimiento as fecha_nacimiento, cn.descripcion as ciudad_nacimiento,"\
                "rec.descripcion as estado_civil, cr.descripcion as ciudad_residencia, rpn.descripcion as provincia_nacimiento, rpr.descripcion as provincia_residencia, p.sexo_id as sexo "\
                "from personas p right join personal_policial pp ON p.id = pp.persona_id join referencias.ref_ciudades cn ON p.ciudad_nacimiento_id = cn.id "\
                "join referencias.ref_estado_civil rec ON p.estado_civil_id = rec.id join referencias.ref_ciudades cr ON p.ciudad_domicilio_id = cr.id "\
                "join referencias.ref_provincia rpn ON cn.provincia_id = rpn.id join referencias.ref_provincia rpr ON cr.provincia_id = rpr.id "\
                "where documento = %s" % dni)
                columns = [col[0] for col in cursorRh.description]
                personas = [
                	dict(zip(columns, row))
                    for row in cursorRh.fetchall()
                ]
                if len(personas) < 1:
                    cursorRh = connections['rrhh'].cursor()
                    cursorRh.execute("select p.documento as documento, p.nombre as nombre, p.apellido as apellido, p.fecha_nacimiento as fecha_nacimiento, "\
                    "cn.descripcion as ciudad_nacimiento, rec.descripcion as estado_civil, cr.descripcion as ciudad_residencia, rpn.descripcion as provincia_nacimiento, "\
                    "rpr.descripcion as provincia_residencia, p.sexo_id as sexo from personas p right join personal_policial pp ON p.id = pp.persona_id "\
                    "left join referencias.ref_ciudades cn ON p.ciudad_nacimiento_id = cn.id left join referencias.ref_estado_civil rec ON p.estado_civil_id = rec.id "\
                    "left join referencias.ref_ciudades cr ON p.ciudad_domicilio_id = cr.id left join referencias.ref_provincia rpn ON cn.provincia_id = rpn.id "\
                    "left join referencias.ref_provincia rpr ON cr.provincia_id = rpr.id where documento = %s "\
                    "union all select p.documento as documento, p.nombre as nombre, p.apellido as apellido, p.fecha_nacimiento as fecha_nacimiento, cn.descripcion as ciudad_nacimiento,"\
                    "rec.descripcion as estadoCivil, cr.descripcion as ciudad_residencia, rpn.descripcion as provincia_nacimiento, rpr.descripcion as provincia_residencia, p.sexo_id as sexo from personas p "\
                    "right join personal_policial pp ON p.id = pp.persona_id right join referencias.ref_ciudades cn ON p.ciudad_nacimiento_id = cn.id "\
                    "right join referencias.ref_estado_civil rec ON p.estado_civil_id = rec.id right join referencias.ref_ciudades cr ON p.ciudad_domicilio_id = cr.id "\
                    "right join referencias.ref_provincia rpn ON cn.provincia_id = rpn.id right join referencias.ref_provincia rpr ON cr.provincia_id = rpr.id "\
                    "where documento = %s" % (dni,dni))
                    columns = [col[0] for col in cursorRh.description]
                    personas = [
                    	dict(zip(columns, row))
                        for row in cursorRh.fetchall()
                    ]
                    if len(personas) < 1:
                        raise Exception('sin datos')

                obtenerReferencias(personas)
                data = json.dumps(personas, default=date_handler)
                content_type = 'application/json'
                return HttpResponse(data,content_type)

            except Exception as e:
                msg = 'La busqueda no arrojó ninugn resultado, deberá cargar los datos manualmente'
                return HttpResponseBadRequest(msg)

@login_required
def ciudades_ajax(request):
    if request.is_ajax():
        q = request.GET.get('term','')

        ciudades = RefCiudades.objects.filter(descripcion__icontains = q)[:20]
        results = []
        for ciudad in ciudades:
            ciudad_json = {}
            ciudad_json['id'] = ciudad.id
            if ciudad.provincia:
                ciudad_json['label'] =  ciudad.descripcion +' - '+ ciudad.provincia.descripcion +' - '+ ciudad.pais.descripcion
            else:
                ciudad_json['label'] = ciudad.descripcion +' - '+ ciudad.pais.descripcion
            ciudad_json['value'] = ciudad.descripcion
            ciudad_json['pais']  = ciudad.pais.descripcion
            ciudad_json['pais_id']  = ciudad.pais.id
            results.append(ciudad_json)
        data = json.dumps(results)
    else:
        data = 'fail'
    content_type = 'application/json'
    return HttpResponse(data,content_type)

@login_required
def ciudades_ajax_provincia(request,id):
    if request.is_ajax():
        ciudades = RefCiudades.objects.filter(provincia = id)
        data = serializers.serialize("json",ciudades)
        return HttpResponse(data,content_type='application/json')
    return HttpResponseBadRequest()


@login_required
def paises_ajax(request):
    if request.is_ajax():
        q = request.GET.get('term','')

        paises = RefPaises.objects.filter(descripcion__icontains = q)[:20]
        results = []
        for pais in paises:
            pais_json = {}
            pais_json['id'] = pais.id
            pais_json['label'] = pais.descripcion
            pais_json['value'] = pais.descripcion
            results.append(pais_json)
        data = json.dumps(results)
    else:
        data = 'fail'
    content_type = 'application/json'
    return HttpResponse(data,content_type)


@login_required
def dependencias_ajax(request):
    if request.is_ajax():
        q = request.GET.get('term','')
        dependencias = None
        if q.isnumeric():
            dependencias = Dependencias.objects.filter(id= int(q))
        else:
            dependencias = Dependencias.objects.filter(descripcion__icontains = q)[:20]
        results = []
        for dependencia in dependencias:
            dependencia_json = {}
            dependencia_json['id'] = dependencia.id
            dependencia_json['label'] = str(dependencia.id) +'-' +dependencia.descripcion + ' - ' + dependencia.unidades_regionales.descripcion
            dependencia_json['value'] = str(dependencia.id) +'-' +dependencia.descripcion + ' - ' + dependencia.unidades_regionales.descripcion
            results.append(dependencia_json)
        data = json.dumps(results)
    else:
        data = 'error'
    content_type = 'application/json'
    return HttpResponse(data,content_type)

@login_required
def estados_civiles(request):
    if request.is_ajax():
        data = request.POST
        estados = RefEstadosciv.objects.all()
        data = serializers.serialize("json", estados)
        return HttpResponse(data, content_type='application/json')
    else:
        return HttpResponseBadRequest()

@login_required
def jerarquias_ajax(request):
    if request.is_ajax():
        q = request.GET.get('term','')
        jerarquias = RefJerarquias.objects.filter(descripcion__icontains=q)[:20]
        results = []
        for jerarquia in jerarquias:
            jerarquia_json = {}
            jerarquia_json['id'] = jerarquia.id
            jerarquia_json['label'] = jerarquia.descripcion
            jerarquia_json['value'] = jerarquia.descripcion
            results.append(jerarquia_json)
        data = json.dumps(results)
    else:
        data = 'error'
    content_type = 'application/json'
    return HttpResponse(data,content_type)


@login_required
def usuarios_ajax(request):
    if request.is_ajax():
        q = request.GET.get('term','')
        usuarios = User.objects.filter(Q(username__icontains=q)|Q(first_name__icontains=q)|Q(last_name__icontains=q))[:20]
        results = []
        for usuario in usuarios:
            usuario_json = {}
            usuario_json['id'] = usuario.id
            usuario_json['label'] = "%s - %s, %s" %(usuario.username, usuario.last_name,usuario.first_name)
            usuario_json['value'] = "%s - %s, %s" %(usuario.username, usuario.last_name,usuario.first_name)
            results.append(usuario_json)
        data = json.dumps(results)
    else:
        data = 'error'
    content_type = 'application/json'
    return HttpResponse(data,content_type)

def date_handler(obj):
    """
    Definicion para transformar el formato date en la serializacion de fechas"""
    return obj.isoformat() if hasattr(obj, 'isoformat') else obj

def obtenerReferencias(personas):
    for persona in personas:
        if persona['ciudad_nacimiento'] and persona['provincia_nacimiento']:
            persona['ciudad_nacimiento_id'] = RefCiudades.objects.get(descripcion = persona['ciudad_nacimiento'], provincia = RefProvincia.objects.get(descripcion = persona['provincia_nacimiento']).id ).id
        else:
            persona['ciudad_nacimiento_id'] = ""
        if persona['ciudad_residencia'] and persona['provincia_residencia']:
            persona['ciudad_residencia_id'] = RefCiudades.objects.get(descripcion = persona['ciudad_residencia'], provincia = RefProvincia.objects.get(descripcion = persona['provincia_residencia']).id).id
        else:
            persona['ciudad_residencia_id'] = ""
        if persona['estado_civil']:
            estadoCivil = RefEstadosciv.objects.get(descripcion = persona['estado_civil']).id
            persona['estado_civil'] = estadoCivil
        else:
            persona['estado_civil'] = ""

def buscar_usuario(request,id):
    if request.is_ajax():
        usuario = User.objects.get(id=id)
        data = serializers.serialize("json", [usuario,])
        return HttpResponse(data, content_type='application/json')
    else:
        return HttpResponseBadRequest("El usuario no existe.")
